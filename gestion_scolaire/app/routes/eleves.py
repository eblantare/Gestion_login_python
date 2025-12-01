from flask import Blueprint, render_template, request, jsonify, send_file, abort,current_app
from extensions import db
from ..models import Eleve, Classe, Ecole
from sqlalchemy import cast, String
from flask_login import current_user, login_required
import uuid
import io
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib import colors
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from ..utils import ecole_required, get_current_ecole_id, is_system_admin

eleves_bp = Blueprint("eleves", __name__)

# Après les imports, ajoutez :
FILTER_ALIASES = {
    "sexe": {
        "male": ["M", "Masculin", "Garçon"],
        "femelle": ["F", "Féminin", "Fille"]
    },
    "status": {
        "nouveau": ["Nouveau"],
        "redoublant": ["Redoublant"],
        "transfert": ["Transfert"]
    },
    "etat": {
        "actif": ["Actif"],
        "inactif": ["Inactif"], 
        "suspendu": ["Suspendu"],
        "valide": ["Validé"],
        "sorti": ["Sorti"]
    },
    "classe": {
        # Ceci sera dynamiquement rempli par vos classes
    }
}

def is_admin():
    """Vérifie si l'utilisateur est admin - CORRECTION POUR ADMIN SYSTÈME"""
    user_role = getattr(current_user, "role", "")
    print(f"🔐 Vérification admin - Rôle: {user_role}, SystemAdmin: {is_system_admin()}")
    
    # L'admin système a tous les droits
    if is_system_admin():
        return True
    
    # Pour les autres utilisateurs, vérifier les rôles admin normaux
    return user_role and user_role.lower() in ["admin", "administrateur"]

def get_eleve_secure(eleve_id, ecole_id):
    """Récupère un élève en vérifiant son appartenance à l'école - CORRECTION POUR ADMIN SYSTÈME"""
    try:
        # L'admin système peut accéder à tous les élèves
        if is_system_admin():
            eleve = Eleve.query.filter_by(id=eleve_id).first()
        else:
            # Pour les autres utilisateurs, vérifier l'appartenance à l'école
            eleve = Eleve.query.filter_by(id=eleve_id, ecole_id=ecole_id).first()
        
        if not eleve:
            print(f"❌ Élève {eleve_id} non trouvé pour école {ecole_id}")
            abort(404, "Élève non trouvé")
        
        # Pour les non-admin système, vérification supplémentaire via la relation Classe
        if not is_system_admin():
            # Vérifier que la classe existe et appartient à l'école
            if not eleve.classe or eleve.classe.ecole_id != ecole_id:
                print(f"❌ Accès refusé: élève {eleve_id} n'appartient pas à l'école {ecole_id}")
                abort(403, "Accès non autorisé à cet élève")
        
        print(f"✅ Élève {eleve_id} récupéré - État: {eleve.etat}")
        return eleve
        
    except Exception as e:
        print(f"❌ Erreur récupération élève {eleve_id}: {str(e)}")
        abort(500, "Erreur serveur lors de la récupération de l'élève")

# ========== ENDPOINTS CRUD ÉLÈVES ==========
@eleves_bp.route("/add", methods=["POST"])
@login_required
@ecole_required
def add_eleve():
    ecole_id = get_current_ecole_id()
    
    try:
        data = request.form
        matricule = data.get('matricule')
        nom = data.get('nom', '').strip().upper()
        prenoms = data.get('prenoms', '').strip().title()
        date_naissance = data.get('date_naissance')
        sexe = data.get('sexe')
        status = data.get('status', 'Nouveau')
        classe_id = data.get('classe_id')
        
        print(f"🎯 AJOUT ÉLÈVE: {nom} {prenoms} - Classe: {classe_id} - Ecole: {ecole_id}")

        # Validations de base
        if not nom or not prenoms or not classe_id:
            return jsonify({"error": "Nom, prénoms et classe sont obligatoires"}), 400

        # Vérifier la classe
        classe = Classe.query.filter_by(id=classe_id, ecole_id=ecole_id).first()
        if not classe:
            return jsonify({"error": "Classe invalide"}), 400

        # VÉRIFICATION DOUBLON AMÉLIORÉE
        print(f"🔍 Recherche de doublons: nom='{nom}', prenoms='{prenoms}', classe_id='{classe_id}', ecole_id='{ecole_id}'")
        
        existing = Eleve.query.filter(
            Eleve.nom.ilike(nom),
            Eleve.prenoms.ilike(prenoms),
            Eleve.classe_id == classe_id,
            Eleve.ecole_id == ecole_id
        ).first()
        
        if existing:
            print(f"🚫 DOUBLON TROUVÉ: {existing.matricule}")
            return jsonify({
                "error": f"Élève déjà existant: {existing.matricule} - {existing.nom} {existing.prenoms}"
            }), 400

        # Générer matricule si besoin
        if not matricule:
            matricule = f"EL-{uuid.uuid4().hex[:8].upper()}"
            # S'assurer que le matricule est unique
            while Eleve.query.filter_by(matricule=matricule).first():
                matricule = f"EL-{uuid.uuid4().hex[:8].upper()}"

        print(f"✅ Création de l'élève avec matricule: {matricule}")

        # Créer l'élève
        eleve = Eleve(
            matricule=matricule,
            nom=nom,
            prenoms=prenoms,
            date_naissance=date_naissance,
            sexe=sexe,
            status=status,
            classe_id=classe_id,
            ecole_id=ecole_id,
            etat="Inactif"
        )

        db.session.add(eleve)
        db.session.commit()

        # VÉRIFICATION FINALE
        eleve_verif = Eleve.query.filter_by(matricule=matricule).first()
        if eleve_verif:
            print(f"✅ ÉLÈVE CRÉÉ ET VÉRIFIÉ: {matricule} - ID: {eleve_verif.id}")
        else:
            print("❌ ÉLÈVE NON TROUVÉ APRÈS CRÉATION")

        return jsonify({
            "success": True,
            "message": "Élève ajouté avec succès",
            "matricule": matricule,
            "id": str(eleve.id)
        }), 200

    except Exception as e:
        db.session.rollback()
        print(f"❌ ERREUR: {str(e)}")
        import traceback
        print(f"📜 STACKTRACE: {traceback.format_exc()}")
        return jsonify({"error": f"Erreur serveur: {str(e)}"}), 500

def generate_unique_matricule(nom, prenoms, ecole_id):
    """Génère un matricule unique de manière atomique"""
    base_initials = ""
    if nom and prenoms:
        base_initials = f"{nom[0]}{prenoms[0]}".upper()
    elif nom:
        base_initials = nom[0].upper()
    else:
        base_initials = "EL"
    
    # Essayer avec timestamp pour plus d'unicité
    timestamp = datetime.now().strftime("%H%M%S")
    
    for attempt in range(5):
        matricule = f"{base_initials}-{timestamp}-{uuid.uuid4().hex[:4].upper()}"
        if not Eleve.query.filter_by(matricule=matricule).first():
            return matricule
    
    # Fallback ultime
    return f"EL-{uuid.uuid4().hex[:8].upper()}"

@eleves_bp.route("/")
@login_required
def liste_eleves():
    ecole_id = get_current_ecole_id()
    user_is_system_admin = is_system_admin()
    
    print(f"🔍 LISTE ELEVES - Ecole: {ecole_id}, SystemAdmin: {user_is_system_admin}")
    
    # Logique différente selon admin système ou non
    if user_is_system_admin:
        if ecole_id:
            query = Eleve.query.filter_by(ecole_id=ecole_id)
            classes = Classe.query.filter_by(ecole_id=ecole_id).all()
        else:
            query = Eleve.query
            classes = Classe.query.all()
    else:
        if not ecole_id:
            return render_template("error.html", message="Aucune école sélectionnée"), 400
        query = Eleve.query.filter_by(ecole_id=ecole_id)
        classes = Classe.query.filter_by(ecole_id=ecole_id).all()

    # TRI EXPLICITE ET DISTINCT pour éviter les doublons
    query = query.join(Classe).order_by(Eleve.nom.asc(), Eleve.prenoms.asc(), Eleve.matricule.asc())

    # Recherche texte
    search = request.args.get("search", "", type=str)
    if search:
        query = query.filter(
            (Eleve.nom.ilike(f"%{search}%")) |
            (Eleve.prenoms.ilike(f"%{search}%")) |
            (cast(Eleve.date_naissance, String).ilike(f"%{search}%")) |
            (Eleve.sexe.ilike(f"%{search}%")) |
            (Eleve.status.ilike(f"%{search}%")) |
            (Classe.nom.ilike(f"%{search}%")) |
            (Eleve.etat.ilike(f"%{search}%"))
        )

    # Filtrage spécifique
    filter_value = request.args.get("filter", "", type=str)
    if filter_value:
        print(f"🔍 Filtre appliqué: {filter_value}")
        
        if ":" in filter_value:
            key, val = filter_value.split(":", 1)
            print(f"🔍 Clé: {key}, Valeur: {val}")
            
            if key == "classe":
                query = query.filter(Classe.nom == val)
            elif key == "sexe":
                query = query.filter(Eleve.sexe == val)
            elif key == "status":
                query = query.filter(Eleve.status == val)
            elif key == "etat":
                query = query.filter(Eleve.etat == val)

    # Pagination
    per_page = request.args.get("per_page", 5, type=int)
    page = request.args.get("page", 1, type=int)
    
    # Log final
    eleves_count = query.count()
    print(f"🔍 Requête finale - Élèves trouvés: {eleves_count}")
    
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    eleves = pagination.items

    print(f"📊 Pagination: page {page}, {len(eleves)} élèves affichés")
    
    user_role = (getattr(current_user, "role", "guest") or "guest").lower()
    
    return render_template(
        "eleves/e_liste.html",
        eleves=eleves,
        pagination=pagination,
        search=search,
        per_page=per_page,
        user_role=user_role,
        filter_value=filter_value,
        classes=classes,
        current_ecole_id=ecole_id,
        is_system_admin=user_is_system_admin
    )

@eleves_bp.route("/check_duplicate", methods=["POST"])
@login_required
@ecole_required
def check_duplicate():
    """Vérifie si un élève existe déjà"""
    ecole_id = get_current_ecole_id()
    data = request.json
    
    nom = data.get('nom', '').strip().upper()
    prenoms = data.get('prenoms', '').strip().title()
    classe_id = data.get('classe_id')
    
    if not nom or not prenoms or not classe_id:
        return jsonify({"exists": False})
    
    # Vérification stricte de doublon
    existing = Eleve.query.filter(
        db.func.lower(db.func.trim(Eleve.nom)) == nom.lower().strip(),
        db.func.lower(db.func.trim(Eleve.prenoms)) == prenoms.lower().strip(),
        Eleve.ecole_id == ecole_id,
        Eleve.classe_id == classe_id
    ).first()
    
    return jsonify({
        "exists": existing is not None,
        "existing_eleve": {
            "matricule": existing.matricule if existing else None,
            "nom": existing.nom if existing else None,
            "prenoms": existing.prenoms if existing else None
        } if existing else None
    })

@eleves_bp.route("/update/<string:id>", methods=["POST"])
@login_required
@ecole_required
def update_eleve(id):
    print(f"🔐 Tentative modification élève {id} - Admin: {is_admin()}, SystemAdmin: {is_system_admin()}")
    
    if not is_admin():
        return jsonify({"error": "Accès refusé - Droits administrateur requis"}), 403
    
    ecole_id = get_current_ecole_id()
    eleve = get_eleve_secure(id, ecole_id)
    
    # Contrôle sur l'état
    if eleve.etat.lower() != 'inactif':
        return jsonify({"error": "Impossible de modifier cet élève, état non inactif"}), 403
    
    data = request.form
    eleve.matricule = data.get("matricule")
    eleve.nom = data.get("nom")
    eleve.prenoms = data.get("prenoms")
    eleve.date_naissance = data.get("date_naissance")
    eleve.sexe = data.get("sexe")
    eleve.status = data.get("status")
    
    # Vérification de la classe pour les non-admin système
    if is_system_admin():
        # Admin système peut assigner à n'importe quelle classe
        eleve.classe_id = data.get("classe_id")
    else:
        # Pour les autres, vérifier que la classe appartient à l'école
        classe_id = data.get("classe_id")
        classe = Classe.query.filter_by(id=classe_id, ecole_id=ecole_id).first()
        if not classe:
            return jsonify({"error": "Classe non trouvée ou non autorisée"}), 403
        eleve.classe_id = classe_id
    
    db.session.commit()

    classe = Classe.query.get(eleve.classe_id)
    return jsonify({
        "id": eleve.id,
        "matricule": eleve.matricule,
        "nom": eleve.nom,
        "prenoms": eleve.prenoms,
        "date_naissance": str(eleve.date_naissance) if eleve.date_naissance else None,
        "sexe": eleve.sexe,
        "status": eleve.status,
        "classe_id": eleve.classe_id,
        "classe_nom": classe.nom if classe else "",
        "etat": eleve.etat
    })

@eleves_bp.route("/delete/<string:id>", methods=["POST"])
@login_required
@ecole_required
def delete_eleve(id):
    print(f"🔐 Tentative suppression élève {id} - Admin: {is_admin()}, SystemAdmin: {is_system_admin()}")
    
    if not is_admin():
        return jsonify({"error": "Accès refusé - Droits administrateur requis"}), 403
    
    ecole_id = get_current_ecole_id()
    eleve = get_eleve_secure(id, ecole_id)
    
    # Contrôle sur l'état
    if eleve.etat.lower() != 'inactif':
        return jsonify({"error": "Impossible de supprimer cet élève, état non inactif"}), 403
    
    db.session.delete(eleve)
    db.session.commit()
    return jsonify({"success": True})

@eleves_bp.route("/detail/<string:id>", methods=["GET"])
@login_required
@ecole_required
def detail_eleve(id):
    print(f"🔐 Tentative détail élève {id} - SystemAdmin: {is_system_admin()}")
    
    ecole_id = get_current_ecole_id()
    eleve = get_eleve_secure(id, ecole_id)
    
    return jsonify({
        "matricule": eleve.matricule,
        "nom": eleve.nom,
        "prenoms": eleve.prenoms,
        "date_naissance": eleve.date_naissance.strftime("%Y-%m-%d") if eleve.date_naissance else None,
        "sexe": eleve.sexe,
        "status": eleve.status,
        "classe": eleve.classe.nom if eleve.classe else None,
        "etat": eleve.etat
    })

@eleves_bp.route("/get/<string:id>", methods=["GET"])
@login_required
@ecole_required
def get_eleve(id):
    print(f"🔐 Tentative récupération élève {id} - SystemAdmin: {is_system_admin()}")
    
    ecole_id = get_current_ecole_id()
    eleve = get_eleve_secure(id, ecole_id)
    
    return jsonify({
        "id": str(eleve.id),
        "matricule": eleve.matricule,
        "nom": eleve.nom,
        "prenoms": eleve.prenoms,
        "date_naissance": eleve.date_naissance.strftime("%Y-%m-%d") if eleve.date_naissance else None,
        "sexe": eleve.sexe,
        "status": eleve.status,
        "classe": eleve.classe.nom if eleve.classe else None,
        "etat": eleve.etat
    })

# ========== GESTION DES ÉTATS ==========

WORKFLOW = {
    "Activer": {"from": "Inactif", "to": "Actif"},
    "Suspendre": {"from": "Actif", "to": "Suspendu"},
    "Valider": {"from": "Suspendu", "to": "Validé"},
    "Sortir": {"from": "Validé", "to": "Sorti"}
}

@eleves_bp.route("/<string:id>/changer_etat", methods=["POST"])
@login_required
@ecole_required
def changer_etat(id):
    print(f"🔐 Tentative changement état élève {id} - Admin: {is_admin()}, SystemAdmin: {is_system_admin()}")
    
    if not is_admin():
        return jsonify({"error": "Accès refusé - Droits administrateur requis"}), 403
    
    ecole_id = get_current_ecole_id()
    
    try:
        # Récupérer l'élève normalement
        if is_system_admin():
            eleve = Eleve.query.filter_by(id=id).first()
        else:
            eleve = Eleve.query.filter_by(id=id, ecole_id=ecole_id).first()
        
        if not eleve:
            abort(404, "Élève non trouvé")
        
        # Vérification supplémentaire via la relation Classe pour les non-admin système
        if not is_system_admin() and eleve.classe.ecole_id != ecole_id:
            abort(403, "Accès non autorisé à cet élève")
        
        action = request.json.get("action")
        print(f"🔄 Action demandée: {action}, État actuel dans DB: '{eleve.etat}'")

        if action not in WORKFLOW:
            return jsonify({"error": "Action non valide!"}), 400
        
        trans = WORKFLOW[action]
        
        # Vérification robuste de l'état (case insensitive)
        current_state = eleve.etat.strip().lower() if eleve.etat else ""
        expected_state = trans["from"].strip().lower()
        
        print(f"🔍 Vérification état: DB='{current_state}' vs Attendu='{expected_state}'")
        
        if current_state != expected_state:
            error_msg = f"L'état actuel est '{eleve.etat}', impossible d'appliquer '{action}' (attendait '{trans['from']}')"
            print(f"❌ {error_msg}")
            return jsonify({"error": error_msg}), 400
        
        # Appliquer le changement d'état
        ancien_etat = eleve.etat
        eleve.etat = trans["to"]
        print(f"✅ Changement état: {ancien_etat} → {eleve.etat}")
        
        db.session.commit()
        
        # Recharger depuis la base pour confirmer
        db.session.refresh(eleve)
        print(f"📝 État après commit et refresh: '{eleve.etat}'")
        
        return jsonify({
            "success": True,
            "etat": eleve.etat,
            "ancien_etat": ancien_etat,
            "message": f"État changé avec succès: {ancien_etat} → {eleve.etat}"
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Erreur lors du changement d'état: {str(e)}")
        import traceback
        print(f"📜 STACKTRACE: {traceback.format_exc()}")
        return jsonify({"error": f"Erreur serveur: {str(e)}"}), 500

# ---------------- EXPORTATIONS ÉLÈVES AVEC EN-TÊTE UNIFORMISÉ ----------------
def get_logo_path_eleves(ecole):
    """Retourne le chemin du logo de l'école - VERSION DÉFINITIVE"""
    if not ecole or not ecole.logo_filename:
        return None
    
    # CHEMIN DÉFINITIF confirmé
    logo_path = os.path.join(
        current_app.root_path, 
        'gestion_scolaire', 
        'app', 
        'static', 
        'logos', 
        ecole.logo_filename
    )
    
    if os.path.exists(logo_path) and os.path.getsize(logo_path) > 0:
        print(f"✅ Logo trouvé: {logo_path}")
        return logo_path
    
    print(f"❌ Logo non trouvé: {logo_path}")
    return None

@eleves_bp.route("/export/pdf/<string:classe_id>")
@login_required
def export_eleves_pdf(classe_id):
    """Export PDF de la liste des élèves par classe - Version avec en-tête uniformisé"""
    try:
        # CORRECTION 3: Vérifier que l'admin système a sélectionné une école
        ecole_id = get_current_ecole_id()
        if not ecole_id:
            return jsonify({"error": "Veuillez d'abord sélectionner une école"}), 400
        
        # CORRECTION 1: Vérifier que la classe est valide
        if not classe_id or classe_id == 'null' or classe_id == 'undefined':
            return jsonify({"error": "Veuillez sélectionner une classe avant d'exporter"}), 400
        
        # Récupérer la classe spécifique
        classe = Classe.query.filter_by(id=classe_id, ecole_id=ecole_id).first_or_404()
        
        # Récupérer les élèves de cette classe
        eleves = Eleve.query.filter_by(classe_id=classe_id, ecole_id=ecole_id)\
                          .order_by(Eleve.nom, Eleve.prenoms).all()
        
        # Récupérer les informations de l'école COURANTE
        ecole = Ecole.query.get(ecole_id)
        if not ecole:
            return jsonify({"error": "École non trouvée"}), 404
        
        # DEBUG: Informations sur le logo
        print(f"🔍 RECHERCHE LOGO - Ecole: {ecole.nom}, Logo_filename: {ecole.logo_filename}")
        logo_path = get_logo_path_eleves(ecole)
        print(f"📁 Chemin logo final: {logo_path}")
        
        # Calculer les statistiques
        effectif_total = len(eleves)
        garcons = sum(1 for e in eleves if e.sexe and e.sexe.upper() == 'M')
        filles = effectif_total - garcons

        # UTILISER LES VALEURS DYNAMIQUES DE L'ÉCOLE COURANTE
        nom_ecole = ecole.nom if ecole.nom else "École non renseignée"
        dre = ecole.dre if ecole.dre else "DRE non renseignée"
        inspection = ecole.inspection if ecole.inspection else "Inspection non renseignée"
        telephone = ecole.telephone1 if ecole.telephone1 else "Téléphone non renseigné"
        devise_ecole = ecole.devise if ecole.devise else "Travail - Liberté - Patrie"

        # Créer le PDF en mémoire
        buffer = io.BytesIO()
        
        def add_page_number(canvas, doc):
            canvas.saveState()
            page_num = canvas.getPageNumber()
            total_pages = doc.page
            
            footer_text = f"Page {page_num}/{total_pages}"
            canvas.setFont('Helvetica', 8)
            canvas.setFillColor(colors.gray)
            canvas.drawRightString(doc.pagesize[0] - 20*mm, 10*mm, footer_text)
            
            info_text = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
            canvas.drawString(20*mm, 10*mm, info_text)
            canvas.restoreState()
        
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4, 
            topMargin=15*mm, 
            bottomMargin=20*mm,
            leftMargin=10*mm,
            rightMargin=10*mm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # ========== EN-TÊTE UNIFORMISÉ ==========
        logo = None
        if logo_path:
            try:
                # Vérifier que le fichier existe et n'est pas vide
                if os.path.exists(logo_path) and os.path.getsize(logo_path) > 0:
                    logo = Image(logo_path, width=35*mm, height=35*mm)
                    logo.hAlign = 'CENTER'
                    print("✅ Logo chargé avec succès dans le PDF")
                else:
                    print("❌ Logo trouvé mais fichier vide ou inexistant")
            except Exception as e:
                print(f"❌ Erreur chargement logo: {e}")
                logo = None
        else:
            print("❌ Aucun chemin de logo fourni")
        
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Normal'],
            fontSize=9,
            alignment=1,
            spaceAfter=0,
            leading=10
        )
        
        small_header_style = ParagraphStyle(
            'SmallHeader',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            spaceAfter=0,
            leading=8
        )
        
        # Structure avec valeurs DYNAMIQUES
        left_col_content = [
            Paragraph("<b>MINISTÈRE DE L'EDUCATION NATIONALE</b>", header_style),
            Paragraph("-----------", small_header_style),
            Paragraph(f"DIRECTION RÉGIONALE DE L'ÉDUCATION - {dre}", header_style),
            Paragraph("-----------", small_header_style),
            Paragraph(f"INSPECTION DE L'ENSEIGNEMENT SECONDAIRE GÉNÉRAL - {inspection}", header_style)
        ]
        
        right_col_content = [
            Paragraph("<b>RÉPUBLIQUE TOGOLAISE</b>", header_style),
            Paragraph("-----------", small_header_style),
            Paragraph("Travail - Liberté - Patrie", ParagraphStyle(
                'DeviseStyle',
                parent=styles['Normal'],
                fontSize=7,
                alignment=1,
                spaceAfter=0,
                leading=8
            ))
        ]
        
        # Colonne centrale avec logo
        center_col_content = []
        if logo:
            center_col_content.append(logo)
        else:
            # Si pas de logo, ajouter un espace vide pour l'équilibre
            center_col_content.append(Spacer(1, 35*mm))
            
        center_col_content.extend([
            Paragraph(f"<b>{nom_ecole}</b>", header_style),
            Paragraph(f"Tél: {telephone}", small_header_style),
            Paragraph(f"{devise_ecole}", small_header_style)
        ])

        header_table = Table([[
            left_col_content, 
            center_col_content, 
            right_col_content
        ]], colWidths=[doc.width/3, doc.width/3, doc.width/3])
        
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ]))
        
        elements.append(header_table)
        elements.append(Spacer(1, 6*mm))
        
        separation_line = Table([['']], colWidths=[doc.width])
        separation_line.setStyle(TableStyle([
            ('LINEABOVE', (0, 0), (0, 0), 1, colors.black),
        ]))
        elements.append(separation_line)
        elements.append(Spacer(1, 6*mm))
        # ========== INFORMATIONS DE LA CLASSE AVEC STATISTIQUES ==========
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=14,
            alignment=1,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=6
        )
        
        # Titre principal avec nom de classe
        title = Paragraph(f"LISTE DES ÉLÈVES - CLASSE DE {classe.nom.upper()}", title_style)
        elements.append(title)
        
        # Statistiques de la classe - MISE EN AVANT
        stats_style = ParagraphStyle(
            'StatsStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=8
        )
        
        stats_text = f"Effectif: {effectif_total} élèves | Garçons: {garcons} | Filles: {filles}"
        stats_paragraph = Paragraph(f"<b>{stats_text}</b>", stats_style)
        elements.append(stats_paragraph)
        
        # Date de génération
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            textColor=colors.gray,
            spaceAfter=12
        )
        
        date_text = Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style)
        elements.append(date_text)
        
        # ========== TABLEAU DES ÉLÈVES ==========
        headers = [
            'N°',  # NOUVELLE COLONNE NUMÉRO
            'Matricule',
            'Nom & Prénoms', 
            'Sexe',
            'Date de naissance',
            'Statut'
        ]
        
        data = [headers]
        
        # AJOUT DU NUMÉRO D'ORDRE
        for index, eleve in enumerate(eleves, 1):
            date_naissance = eleve.date_naissance.strftime("%d/%m/%Y") if eleve.date_naissance else "Non définie"
            
            row = [
                str(index),  # NUMÉRO D'ORDRE
                eleve.matricule,
                f"{eleve.nom} {eleve.prenoms}",
                eleve.sexe,
                date_naissance,
                eleve.status
            ]
            data.append(row)
        
        # Création du tableau avec largeurs optimisées
        table = Table(data, colWidths=[
            doc.width * 0.05,  # N°
            doc.width * 0.15,  # Matricule
            doc.width * 0.35,  # Nom & Prénoms
            doc.width * 0.08,  # Sexe
            doc.width * 0.17,  # Date naissance
            doc.width * 0.15   # Statut
        ], repeatRows=1)
        
        # Style du tableau
        table_style = TableStyle([
            # En-têtes
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 4),
            
            # Lignes de données
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),   # N° centré
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),     # Matricule à gauche
            ('ALIGN', (2, 1), (2, -1), 'LEFT'),     # Nom à gauche
            ('ALIGN', (3, 1), (3, -1), 'CENTER'),   # Sexe centré
            ('ALIGN', (4, 1), (4, -1), 'CENTER'),   # Date centrée
            ('ALIGN', (5, 1), (5, -1), 'CENTER'),   # Statut centré
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Bordures
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            
            # Alternance des couleurs des lignes
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
            
            # Padding réduit
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ])
        
        table.setStyle(table_style)
        elements.append(table)
        
        # ========== STATISTIQUES FINALES ==========
        elements.append(Spacer(1, 8*mm))
        
        final_stats_style = ParagraphStyle(
            'FinalStatsStyle',
            parent=styles['Normal'],
            fontSize=9,
            alignment=0,
            textColor=colors.gray,
            spaceAfter=1
        )
        
        pourcentage_garcons = (garcons / effectif_total * 100) if effectif_total > 0 else 0
        pourcentage_filles = (filles / effectif_total * 100) if effectif_total > 0 else 0
        
        final_stats_text = [
            Paragraph(f"<b>RÉCAPITULATIF CLASSE {classe.nom.upper()} :</b>", final_stats_style),
            Paragraph(f"Effectif total : {effectif_total} élèves", final_stats_style),
            Paragraph(f"Garçons : {garcons} ({pourcentage_garcons:.1f}%) | Filles : {filles} ({pourcentage_filles:.1f}%)", final_stats_style),
        ]
        
        for element in final_stats_text:
            elements.append(element)
        
        # ========== GÉNÉRATION DU PDF ==========
        doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
        buffer.seek(0)
        
        # Retourner le PDF
        filename = f"liste_eleves_{classe.nom}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        
        response = send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
        
    except Exception as e:
        print(f"Erreur génération PDF: {str(e)}")
        return jsonify({"error": f"Erreur lors de la génération du PDF: {str(e)}"}), 500

@eleves_bp.route("/export/excel/<string:classe_id>")
@login_required
def export_eleves_excel(classe_id):
    """Export Excel de la liste des élèves par classe avec en-tête uniformisé"""
    try:
        # CORRECTION 3: Vérifier que l'admin système a sélectionné une école
        ecole_id = get_current_ecole_id()
        if not ecole_id:
            return jsonify({"error": "Veuillez d'abord sélectionner une école"}), 400
        
        # CORRECTION 1: Vérifier que la classe est valide
        if not classe_id or classe_id == 'null' or classe_id == 'undefined':
            return jsonify({"error": "Veuillez sélectionner une classe avant d'exporter"}), 400
        
        # Récupérer la classe spécifique
        classe = Classe.query.filter_by(id=classe_id, ecole_id=ecole_id).first_or_404()
        
        # Récupérer les élèves de cette classe
        eleves = Eleve.query.filter_by(classe_id=classe_id, ecole_id=ecole_id)\
                          .order_by(Eleve.nom, Eleve.prenoms).all()
        
        # Récupérer les informations de l'école COURANTE
        ecole = Ecole.query.get(ecole_id)
        if not ecole:
            return jsonify({"error": "École non trouvée"}), 404
        
        # CORRECTION 4: Récupérer le logo pour Excel
        logo_path = get_logo_path_eleves(ecole)
        print(f"📊 Excel - Chemin logo: {logo_path}")
        
        # Calculer les statistiques
        effectif_total = len(eleves)
        garcons = sum(1 for e in eleves if e.sexe and e.sexe.upper() == 'M')
        filles = effectif_total - garcons

        # UTILISER LES VALEURS DYNAMIQUES DE L'ÉCOLE COURANTE
        nom_ecole = ecole.nom if ecole.nom else "École non renseignée"
        dre = ecole.dre if ecole.dre else "DRE non renseignée"
        inspection = ecole.inspection if ecole.inspection else "Inspection non renseignée"
        telephone = ecole.telephone1 if ecole.telephone1 else "Téléphone non renseigné"
        devise_ecole = ecole.devise if ecole.devise else "Travail - Liberté - Patrie"
        
        # Créer le fichier Excel
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = f"Liste élèves {classe.nom}"
        
        # ========== EN-TÊTE EXCEL UNIFORMISÉ AVEC LOGO ==========
        center_align = Alignment(horizontal="center", vertical="center")
        
        # CORRECTION 3: Structure d'en-tête optimisée avec logo
        current_row = 1
        
        # Ligne 1-2: Espace réservé pour le logo (ne pas mettre le logo directement)
        if logo_path and os.path.exists(logo_path):
            # Nous allons ajouter le logo plus tard pour ne pas déformer l'en-tête
            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 30
            current_row = 3
        else:
            current_row = 1
        
        # Structure d'en-tête textuelle
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = "MINISTÈRE DE L'EDUCATION NATIONALE"
        ws[f'A{current_row}'].font = Font(bold=True, size=10)
        ws[f'A{current_row}'].alignment = center_align
        
        ws.merge_cells(f'G{current_row}:J{current_row}')
        ws[f'G{current_row}'] = "RÉPUBLIQUE TOGOLAISE"
        ws[f'G{current_row}'].font = Font(bold=True, size=10)
        ws[f'G{current_row}'].alignment = center_align
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = "-----------"
        ws[f'A{current_row}'].alignment = center_align
        
        ws.merge_cells(f'G{current_row}:J{current_row}')
        ws[f'G{current_row}'] = "-----------"
        ws[f'G{current_row}'].alignment = center_align
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = f"DIRECTION RÉGIONALE DE L'ÉDUCATION - {dre}"
        ws[f'A{current_row}'].font = Font(size=9)
        ws[f'A{current_row}'].alignment = center_align
        
        ws.merge_cells(f'G{current_row}:J{current_row}')
        ws[f'G{current_row}'] = "Travail - Liberté - Patrie"
        ws[f'G{current_row}'].font = Font(bold=True, size=8)
        ws[f'G{current_row}'].alignment = center_align
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = "-----------"
        ws[f'A{current_row}'].alignment = center_align
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = f"INSPECTION DE L'ENSEIGNEMENT SECONDAIRE GÉNÉRAL - {inspection}"
        ws[f'A{current_row}'].font = Font(size=9)
        ws[f'A{current_row}'].alignment = center_align
        
        # Nom de l'école et informations
        current_row += 1
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'] = nom_ecole
        ws[f'A{current_row}'].font = Font(bold=True, size=14, color="2C3E50")
        ws[f'A{current_row}'].alignment = center_align
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'] = f"Tél: {telephone} - {devise_ecole}"
        ws[f'A{current_row}'].alignment = center_align
        
        # Titre du document avec classe
        current_row += 1
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'] = f"LISTE DES ÉLÈVES - CLASSE DE {classe.nom.upper()}"
        ws[f'A{current_row}'].font = Font(bold=True, size=16, color="2C3E50")
        ws[f'A{current_row}'].alignment = center_align
        
        # STATISTIQUES DE LA CLASSE - MISE EN AVANT
        current_row += 1
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'] = f"Effectif: {effectif_total} élèves | Garçons: {garcons} | Filles: {filles}"
        ws[f'A{current_row}'].font = Font(size=11, bold=True, color="2C3E50")
        ws[f'A{current_row}'].alignment = center_align
        
        # Date de génération
        current_row += 1
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws[f'A{current_row}'].font = Font(size=8, italic=True, color="666666")
        ws[f'A{current_row}'].alignment = center_align
        
        # Ligne vide
        current_row += 1
        ws[f'A{current_row}'] = ""
        
        # CORRECTION 3: Ajouter le logo APRÈS la structure d'en-tête
        if logo_path and os.path.exists(logo_path):
            try:
                from openpyxl.drawing.image import Image as ExcelImage
                img = ExcelImage(logo_path)
                
                # Redimensionner le logo pour qu'il soit discret
                img.width = 60
                img.height = 60
                
                # Placer le logo en haut à gauche sans déformer l'en-tête
                img.anchor = 'A1'
                ws.add_image(img)
                print("✅ Logo ajouté discrètement dans Excel")
                
            except Exception as e:
                print(f"❌ Erreur ajout logo Excel: {e}")
        
        # ========== TABLEAU DES DONNÉES ==========
        # AJOUT DE LA COLONNE N° D'ORDRE
        headers = [
            'N°',  # NOUVELLE COLONNE NUMÉRO
            'Matricule', 'Nom', 'Prénoms', 'Sexe', 
            'Date de naissance', 'Statut'
        ]
        
        # Commencer après l'en-tête
        start_row = current_row + 1
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = center_align
        
        # AJOUT DU NUMÉRO D'ORDRE POUR CHAQUE ÉLÈVE
        for index, eleve in enumerate(eleves, 1):
            row_num = start_row + index
            date_naissance = eleve.date_naissance.strftime("%d/%m/%Y") if eleve.date_naissance else "Non définie"
            
            row_data = [
                index,  # NUMÉRO D'ORDRE
                eleve.matricule,
                eleve.nom,
                eleve.prenoms,
                eleve.sexe,
                date_naissance,
                eleve.status
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = value
        # ========== MISE EN FORME FINALE ==========
        column_widths = {
            'A': 6,    # N° - plus étroite
            'B': 15,   # Matricule
            'C': 20,   # Nom
            'D': 20,   # Prénoms
            'E': 8,    # Sexe
            'F': 15,   # Date naissance
            'G': 12    # Statut
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Centrer les colonnes
        for row in range(start_row + 1, ws.max_row + 1):
            ws[f'A{row}'].alignment = center_align  # N°
            ws[f'E{row}'].alignment = center_align  # Sexe
            ws[f'F{row}'].alignment = center_align  # Date naissance
            ws[f'G{row}'].alignment = center_align  # Statut
        
        # Ajouter des bordures au tableau
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, ws.max_row + 1):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        # ========== STATISTIQUES FINALES ==========
        stats_row = ws.max_row + 2
        pourcentage_garcons = (garcons / effectif_total * 100) if effectif_total > 0 else 0
        pourcentage_filles = (filles / effectif_total * 100) if effectif_total > 0 else 0
        
        ws.merge_cells(f'A{stats_row}:G{stats_row}')
        ws[f'A{stats_row}'] = "RÉCAPITULATIF"
        ws[f'A{stats_row}'].font = Font(bold=True, size=10)
        
        ws.merge_cells(f'A{stats_row + 1}:G{stats_row + 1}')
        ws[f'A{stats_row + 1}'] = f"Effectif total : {effectif_total} élèves | Garçons : {garcons} ({pourcentage_garcons:.1f}%) | Filles : {filles} ({pourcentage_filles:.1f}%)"
        ws[f'A{stats_row + 1}'].font = Font(size=9)
        
        # Sauvegarder
        wb.save(buffer)
        buffer.seek(0)
        
        # Retourner le fichier
        filename = f"liste_eleves_{classe.nom}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        response = send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
        
    except Exception as e:
        print(f"Erreur génération Excel: {str(e)}")
        return jsonify({"error": f"Erreur lors de la génération du Excel: {str(e)}"}), 500
    
def find_logo_file(logo_filename):
    """Fonction utilitaire pour trouver où se trouve un fichier logo"""
    import glob
    
    print(f"🔍 RECHERCHE COMPLÈTE DU LOGO: {logo_filename}")
    
    # Recherche récursive dans tout le projet
    search_patterns = [
        f"**/{logo_filename}",
        f"**/logos/{logo_filename}",
        f"**/static/**/{logo_filename}",
        f"**/uploads/**/{logo_filename}"
    ]
    
    found_paths = []
    
    for pattern in search_patterns:
        for file_path in glob.glob(pattern, recursive=True):
            if os.path.isfile(file_path):
                found_paths.append(file_path)
                print(f"✅ Logo trouvé: {file_path}")
    
    if not found_paths:
        print(f"❌ Logo {logo_filename} introuvable dans le projet")
        
        # Lister le contenu des dossiers logos pour debug
        logo_dirs = [
            r"C:\projets\python\gestion_scolaire\app\static\logos",
            r"C:\projets\python\static\uploads\logos", 
            r"C:\projets\python\app\static\uploads\logos"
        ]
        
        for logo_dir in logo_dirs:
            if os.path.exists(logo_dir):
                print(f"📁 Contenu de {logo_dir}:")
                try:
                    files = os.listdir(logo_dir)
                    for f in files:
                        print(f"   - {f}")
                except Exception as e:
                    print(f"   ❌ Erreur lecture: {e}")
            else:
                print(f"📁 Dossier inexistant: {logo_dir}")
    
    return found_paths[0] if found_paths else None

