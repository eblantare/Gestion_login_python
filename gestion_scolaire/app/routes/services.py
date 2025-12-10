from flask import Blueprint, render_template, request, jsonify, send_file, session
from flask_login import login_required, current_user
import os
from datetime import datetime
import io
import traceback
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import mm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
from sqlalchemy.orm import joinedload
# Import des modèles
from ..models import Classe, Eleve, Ecole, Enseignant, Note, Matiere
from gestion_login.gestion_login.models import Utilisateur
from extensions import db
from ..utils import get_current_ecole_id

services_bp = Blueprint('services', __name__)

# ========== FONCTION UTILITAIRE POUR RÉCUPÉRER L'ÉCOLE ==========

def get_ecole_id_courante():
    """Récupère l'ID de l'école courante avec gestion admin système CORRIGÉE"""
    # ✅ CORRECTION : Vérifier d'abord les paramètres de requête pour le changement d'école
    ecole_id_from_request = request.args.get('ecole_id')
    if ecole_id_from_request and ecole_id_from_request not in ['', 'null', 'undefined']:
        ecole_id = ecole_id_from_request
        print(f"🎯 École sélectionnée via paramètre: {ecole_id}")
        
        # ✅ Mettre à jour la session pour les requêtes suivantes
        session['selected_ecole_id'] = ecole_id
        return ecole_id
    
    # Ensuite, utiliser la fonction standard
    ecole_id = get_current_ecole_id()
    
    # ✅ CORRECTION : Vérifier la session pour l'admin système
    # L'admin système utilise 'selected_ecole_id' dans la session
    if ecole_id is None and session.get('selected_ecole_id'):
        ecole_id = session.get('selected_ecole_id')
        print(f"🔧 ADMIN SYSTÈME avec école sélectionnée (selected_ecole_id): {ecole_id}")
    
    # ✅ CORRECTION : Vérifier aussi les autres clés possibles pour compatibilité
    if ecole_id is None and session.get('ecole_selectionnee_id'):
        ecole_id = session.get('ecole_selectionnee_id')
        print(f"🔧 ADMIN SYSTÈME avec ecole_selectionnee_id: {ecole_id}")
    
    if ecole_id is None and session.get('current_ecole_id'):
        ecole_id = session.get('current_ecole_id')
        print(f"🔧 ADMIN SYSTÈME avec current_ecole_id: {ecole_id}")
    
    print(f"🏫 get_ecole_id_courante() retourne: {ecole_id}")
    print(f"🔍 Session - selected_ecole_id: {session.get('selected_ecole_id')}")
    print(f"🔍 Session - selected_ecole_nom: {session.get('selected_ecole_nom')}")
    
    return ecole_id


def recreate_missing_logo(ecole_id):
    """Recrée un logo manquant pour une école avec le BON chemin"""
    try:
        ecole = Ecole.query.get(ecole_id)
        if not ecole:
            return None
            
        print(f"🔄 Recréation du logo pour {ecole.nom}")
        
        from PIL import Image, ImageDraw, ImageFont
        import random
        
        # Dimensions
        width, height = 200, 200
        
        # Couleurs
        colors = [
            ('#2C3E50', '#3498DB'),  # Bleu
            ('#27AE60', '#2ECC71'),  # Vert
            ('#8E44AD', '#9B59B6'),  # Violet
        ]
        
        bg_color, text_color = random.choice(colors)
        
        # Créer l'image
        img = Image.new('RGB', (width, height), color=bg_color)
        draw = ImageDraw.Draw(img)
        
        # Cadre
        draw.rectangle([5, 5, width-5, height-5], outline=text_color, width=3)
        
        # Initiales
        initials = ''.join([word[0].upper() for word in ecole.nom.split()[:2]]) or "ECOLE"
        
        try:
            font_size = 60 if len(initials) <= 3 else 40
            font = ImageFont.truetype("arial.ttf", font_size)
        except:
            font = ImageFont.load_default()
        
        # Centrer le texte
        bbox = draw.textbbox((0, 0), initials, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
        x = (width - text_width) // 2
        y = (height - text_height) // 2
        
        draw.text((x, y), initials, fill=text_color, font=font)
        
        # ✅ UTILISER LE BON DOSSIER : static/logos
        upload_dir = os.path.join('gestion_scolaire', 'app', 'static', 'logos')
        os.makedirs(upload_dir, exist_ok=True)
        
        logo_filename = f"{ecole_id}.png"
        logo_path = os.path.join(upload_dir, logo_filename)
        
        img.save(logo_path)
        print(f"✅ Logo recréé: {logo_path}")
        
        # ✅ METTRE À JOUR LA BD avec le nom cohérent
        ecole.logo_filename = logo_filename
        db.session.commit()
        print(f"📝 BD mise à jour: {logo_filename}")
        
        return logo_path
        
    except Exception as e:
        print(f"❌ Erreur recréation logo: {e}")
        traceback.print_exc()
        return None
# ========== FONCTIONS POUR L'EN-TÊTE ==========

def get_logo_path(ecole_id):
    """Retourne le chemin du logo de l'école - MÊME LOGIQUE QUE DANS eleves.py"""
    try:
        print(f"🔍 Recherche logo pour école_id: {ecole_id}")
        
        if not ecole_id:
            return get_default_logo_path()
        
        ecole = Ecole.query.get(ecole_id)
        if not ecole:
            print(f"❌ École {ecole_id} non trouvée")
            return get_default_logo_path()
        
        print(f"🏫 École: {ecole.nom}")
        print(f"📁 Logo en BD: {ecole.logo_filename}")
        
        # ✅ UTILISER LE MÊME CHEMIN QUE DANS eleves.py
        if ecole.logo_filename:
            # CHEMIN DÉFINITIF confirmé qui fonctionne dans eleves.py
            logo_path = os.path.join(
                'gestion_scolaire', 
                'app', 
                'static', 
                'logos', 
                ecole.logo_filename
            )
            
            # Vérifier si le fichier existe
            if os.path.exists(logo_path) and os.path.isfile(logo_path):
                file_size = os.path.getsize(logo_path)
                print(f"✅ Logo TROUVÉ: {logo_path} ({file_size} octets)")
                return logo_path
            else:
                print(f"❌ Logo non trouvé à: {logo_path}")
        
        # ✅ Si pas trouvé, chercher avec l'ID de l'école
        print("🔍 Recherche par ID d'école...")
        id_based_paths = [
            os.path.join('gestion_scolaire', 'app', 'static', 'logos', f"{ecole_id}.png"),
            os.path.join('gestion_scolaire', 'app', 'static', 'logos', f"{ecole_id}.jpg"),
            os.path.join('gestion_scolaire', 'app', 'static', 'logos', f"{ecole_id}.jpeg"),
            os.path.join('app', 'static', 'logos', f"{ecole_id}.png"),
            os.path.join('static', 'logos', f"{ecole_id}.png"),
        ]
        
        for path in id_based_paths:
            if os.path.exists(path) and os.path.isfile(path):
                file_size = os.path.getsize(path)
                print(f"✅ Logo TROUVÉ (par ID): {path} ({file_size} octets)")
                
                # ✅ METTRE À JOUR LA BASE DE DONNÉES avec le bon nom
                new_filename = os.path.basename(path)
                if not ecole.logo_filename or ecole.logo_filename != new_filename:
                    ecole.logo_filename = new_filename
                    db.session.commit()
                    print(f"📝 BD mise à jour avec: {ecole.logo_filename}")
                
                return path
        
        # Finalement, logo par défaut
        print("❌ Aucun logo trouvé, utilisation du défaut")
        return get_default_logo_path()
        
    except Exception as e:
        print(f"❌ Erreur get_logo_path: {e}")
        traceback.print_exc()
        return get_default_logo_path()

def get_default_logo_path():
    """Retourne le chemin du logo par défaut - MÊME LOGIQUE QUE DANS eleves.py"""
    default_paths = [
        os.path.join('gestion_scolaire', 'app', 'static', 'images', 'default_logo.png'),
        os.path.join('app', 'static', 'images', 'default_logo.png'),
        os.path.join('static', 'images', 'default_logo.png'),
    ]
    
    for path in default_paths:
        if os.path.exists(path):
            print(f"✅ Logo par défaut trouvé: {path}")
            return path
    
    # Créer un logo par défaut si nécessaire
    print("🔄 Création du logo par défaut...")
    return create_default_logo()

def find_logo_recursive(ecole_id):
    """Recherche récursive d'un logo pour une école"""
    search_dirs = [
        os.path.join('app', 'static'),
        os.path.join('static'),
        os.path.join(os.getcwd(), 'app', 'static'),
    ]
    
    # Patterns de recherche
    patterns = [
        f"{ecole_id}.*",  # Par ID
        f"*{ecole_id}*",  # Contenant l'ID
    ]
    
    for search_dir in search_dirs:
        if not os.path.exists(search_dir):
            continue
            
        for root, dirs, files in os.walk(search_dir):
            for file in files:
                if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                    # Vérifier si le fichier correspond à l'école
                    if ecole_id in file or any(pattern.replace('*', '') in file for pattern in patterns):
                        full_path = os.path.join(root, file)
                        print(f"🔍 Logo potentiel trouvé: {full_path}")
                        return full_path
    
    return None

def find_logo_by_filename(logo_filename):
    """Recherche le logo par son nom de fichier dans les dossiers communs"""
    # Obtenir le nom sans extension et l'extension
    name_without_ext, ext = os.path.splitext(logo_filename)
    
    # Rechercher dans plusieurs variantes
    search_patterns = [logo_filename, name_without_ext + ext.lower(), name_without_ext + ext.upper()]
    
    for pattern in search_patterns:
        # Chemins de recherche élargis
        search_paths = [
            os.path.join('app', 'static', 'uploads', 'logos', pattern),
            os.path.join('app', 'static', 'uploads', pattern),
            os.path.join('app', 'static', 'images', pattern),
            os.path.join('app', 'static', 'images', 'logos', pattern),
            os.path.join('app', 'static', pattern),
            os.path.join('static', 'uploads', 'logos', pattern),
            os.path.join('static', 'uploads', pattern),
            os.path.join('static', 'images', pattern),
        ]
        
        for path in search_paths:
            if os.path.exists(path):
                return path
    
    return None

@services_bp.route("/admin/regenerate-logos")
@login_required
def regenerate_logos():
    """Route admin pour régénérer tous les logos manquants"""
    try:
        ecoles = Ecole.query.all()
        results = []
        
        for ecole in ecoles:
            logo_path = get_logo_path(ecole.id)
            status = "✅ Existant" if logo_path and "default_logo" not in logo_path else "🔄 Recréé"
            results.append({
                'ecole': ecole.nom,
                'logo_filename': ecole.logo_filename,
                'status': status,
                'path': logo_path
            })
        
        return jsonify({
            'message': f'Logos vérifiés pour {len(ecoles)} écoles',
            'results': results
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def add_pdf_header(elements, doc, ecole_id, titre_principal):
    """Ajoute l'en-tête uniformisé à tous les PDF - VERSION SIMPLIFIÉE"""
    styles = getSampleStyleSheet()
    
    # Récupération des informations de l'école
    ecole = None
    if ecole_id:
        ecole = Ecole.query.get(ecole_id)
    
    # Configuration de l'en-tête
    logo_path = get_logo_path(ecole_id)
    print(f"🎯 add_pdf_header - logo_path: {logo_path}")
    
    # ✅ APPROCHE SIMPLIFIÉE : Utiliser la même logique que dans eleves.py
    logo_display = None
    if logo_path and os.path.exists(logo_path):
        try:
            print(f"🔄 Chargement du logo depuis: {logo_path}")
            
            # Méthode simple et fiable
            logo_display = Image(logo_path, width=35*mm, height=35*mm)
            logo_display.hAlign = 'CENTER'
            
            print(f"✅ Logo chargé avec succès - Taille: 35x35mm")
            
        except Exception as e:
            print(f"❌ Erreur chargement logo: {e}")
            # Méthode de secours
            logo_display = create_text_logo(ecole.nom if ecole else "ECOLE")
    else:
        logo_display = create_text_logo(ecole.nom if ecole else "ECOLE")
    
    # Styles pour l'en-tête (inchangé)
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Normal'],
        fontSize=9,
        alignment=1,
        spaceAfter=0,
        leading=10
    )
    
    small_header_style = ParagraphStyle(
        'SmallHeader',
        parent=styles['Normal'],
        fontSize=8,
        alignment=1,
        spaceAfter=0,
        leading=8
    )
    
    # Contenu des colonnes
    nom_ecole = ecole.nom if ecole else "ADMINISTRATION SYSTÈME"
    dre = ecole.dre if ecole else "TOUTES LES DRE"
    inspection = ecole.inspection if ecole else "TOUTES LES INSPECTIONS"
    telephone = ecole.telephone1 if ecole else "N/A"
    devise_ecole = ecole.devise if ecole else "Travail - Liberté - Patrie"
    
    # Colonne GAUCHE
    left_col_content = [
        Paragraph("<b>MINISTÈRE DE L'EDUCATION NATIONALE</b>", header_style),
        Paragraph("-----------", small_header_style),
        Paragraph(f"DIRECTION RÉGIONALE DE L'ÉDUCATION - {dre}", header_style),
        Paragraph("-----------", small_header_style),
        Paragraph(f"INSPECTION DE L'ENSEIGNEMENT SECONDAIRE GÉNÉRAL - {inspection}", header_style)
    ]
    
    # Colonne DROITE
    right_col_content = [
        Paragraph("<b>RÉPUBLIQUE TOGOLAISE</b>", header_style),
        Paragraph("-----------", small_header_style),
        Paragraph("Travail - Liberté - Patrie", ParagraphStyle(
            'DeviseStyle',
            parent=styles['Normal'],
            fontSize=7,
            alignment=1,
            spaceAfter=0,
            leading=8
        ))
    ]
    
    # Colonne CENTRALE
    center_col_content = []
    if logo_display:
        center_col_content.append(logo_display)
        center_col_content.append(Spacer(1, 2*mm))
    else:
        center_col_content.append(Spacer(1, 35*mm))
        center_col_content.append(Paragraph("LOGO", header_style))
    
    center_col_content.extend([
        Paragraph(f"<b>{nom_ecole}</b>", header_style),
        Paragraph(f"Tél: {telephone}", small_header_style),
        Paragraph(f"{devise_ecole}", small_header_style)
    ])

    # Création du tableau d'en-tête
    header_table = Table([[
        left_col_content, 
        center_col_content, 
        right_col_content
    ]], colWidths=[doc.width/3, doc.width/3, doc.width/3])
    
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
    ]))
    
    elements.append(header_table)
    elements.append(Spacer(1, 6*mm))
    
    # Ligne de séparation
    separation_line = Table([['']], colWidths=[doc.width])
    separation_line.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (0, 0), 1, colors.black),
    ]))
    elements.append(separation_line)
    elements.append(Spacer(1, 6*mm))
    
    # Titre principal du document
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=1,
        textColor=colors.HexColor('#2C3E50'),
        spaceAfter=8
    )
    
    elements.append(Paragraph(titre_principal, title_style))
    
    return elements

def create_text_logo(ecole_nom):
    """Crée un logo texte comme fallback"""
    try:
        styles = getSampleStyleSheet()
        logo_style = ParagraphStyle(
            'TextLogo',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,
            textColor=colors.white,
            backColor=colors.HexColor('#2C3E50'),
            borderPadding=10,
            borderColor=colors.HexColor('#34495E'),
            borderWidth=1,
            spaceAfter=12
        )
        
        initials = ''.join([word[0].upper() for word in ecole_nom.split()[:2]]) or "ECOLE"
        return Paragraph(f"<b>{initials}</b>", logo_style)
    except:
        return None

def create_simple_logo_placeholder(ecole_nom):
    """Crée un placeholder simple si le logo ne peut pas être chargé"""
    try:
        # Créer un canvas simple avec du texte
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib import colors
        
        style = ParagraphStyle(
            'LogoPlaceholder',
            fontSize=8,
            alignment=1,
            textColor=colors.gray,
            borderColor=colors.gray,
            borderWidth=1,
            borderPadding=5,
            backColor=colors.lightgrey
        )
        
        initials = ''.join([word[0].upper() for word in ecole_nom.split()[:2]]) or "ECOLE"
        placeholder = Paragraph(f"<b>{initials}</b>", style)
        
        print(f"🔄 Placeholder créé pour: {ecole_nom}")
        return placeholder
        
    except Exception as e:
        print(f"❌ Erreur création placeholder: {e}")
        return None

def create_default_logo():
    """Crée un logo par défaut simple si aucun n'existe"""
    try:
        default_logo_path = os.path.join('app', 'static', 'images', 'default_logo.png')
        
        # Créer le dossier si nécessaire
        os.makedirs(os.path.dirname(default_logo_path), exist_ok=True)
        
        # Créer une image simple avec PIL
        from PIL import Image, ImageDraw
        
        # Créer une image 200x200 pixels
        img = Image.new('RGB', (200, 200), color='lightblue')
        draw = ImageDraw.Draw(img)
        
        # Dessiner un cercle et du texte simple
        draw.ellipse([50, 50, 150, 150], fill='white', outline='darkblue', width=3)
        draw.text((100, 100), "ECOLE", fill='darkblue', anchor='mm')
        
        # Sauvegarder
        img.save(default_logo_path)
        print(f"✅ Logo par défaut créé: {default_logo_path}")
        return default_logo_path
        
    except Exception as e:
        print(f"❌ Erreur création logo par défaut: {e}")
        return None

def add_excel_header(worksheet, ecole_id, titre_principal, infos_supplementaires=None):
    """Ajoute l'en-tête uniformisé à tous les Excel"""
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # Récupération des informations de l'école
    ecole = None
    if ecole_id:
        ecole = Ecole.query.get(ecole_id)
    
    nom_ecole = ecole.nom if ecole else "ADMINISTRATION SYSTÈME"
    dre = ecole.dre if ecole else "TOUTES LES DRE"
    inspection = ecole.inspection if ecole else "TOUTES LES INSPECTIONS"
    telephone = ecole.telephone1 if ecole else "N/A"
    devise_ecole = ecole.devise if ecole else "Travail - Liberté - Patrie"
    
    # En-tête institutionnel
    worksheet.merge_cells('A1:G1')
    worksheet['A1'] = "MINISTÈRE DE L'EDUCATION NATIONALE"
    worksheet['A1'].font = Font(bold=True, size=12)
    worksheet['A1'].alignment = center_align
    
    worksheet.merge_cells('A2:G2')
    worksheet['A2'] = f"DIRECTION RÉGIONALE DE L'ÉDUCATION - {dre}"
    worksheet['A2'].font = Font(size=10)
    worksheet['A2'].alignment = center_align
    
    worksheet.merge_cells('A3:G3')
    worksheet['A3'] = f"INSPECTION DE L'ENSEIGNEMENT SECONDAIRE GÉNÉRAL - {inspection}"
    worksheet['A3'].font = Font(size=10)
    worksheet['A3'].alignment = center_align
    
    # Nom de l'école
    worksheet.merge_cells('A4:G4')
    worksheet['A4'] = nom_ecole
    worksheet['A4'].font = Font(bold=True, size=14, color="2C3E50")
    worksheet['A4'].alignment = center_align
    
    worksheet.merge_cells('A5:G5')
    worksheet['A5'] = f"Tél: {telephone} - {devise_ecole}"
    worksheet['A5'].alignment = center_align
    
    # Titre du document
    worksheet.merge_cells('A6:G6')
    worksheet['A6'] = titre_principal
    worksheet['A6'].font = Font(bold=True, size=16, color="2C3E50")
    worksheet['A6'].alignment = center_align
    
    # Informations supplémentaires
    if infos_supplementaires:
        worksheet.merge_cells('A7:G7')
        worksheet['A7'] = infos_supplementaires
        worksheet['A7'].font = Font(size=11, bold=True, color="2C3E50")
        worksheet['A7'].alignment = center_align
    
    # Date de génération
    worksheet.merge_cells('A8:G8')
    worksheet['A8'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    worksheet['A8'].font = Font(size=9, italic=True, color="666666")
    worksheet['A8'].alignment = center_align
    
    return 9  # Retourne la ligne de départ des données

# ========== ROUTES PRINCIPALES ==========

@services_bp.route("/exportations")
@login_required
def exportations():
    """Page principale des services d'exportation"""
    return render_template("services/services.html")

# ========== ROUTES API POUR LES FILTRES ==========

@services_bp.route("/eleves/classes")
@login_required
def get_eleves_classes():
    """API pour récupérer les classes (pour export élèves)"""
    try:
        ecole_id = get_ecole_id_courante()
        
        print(f"🔍 GET /eleves/classes - ecole_id: {ecole_id}")
        print(f"🔍 Paramètres request: {request.args}")
        
        # ✅ CORRECTION : Toujours filtrer par école si disponible
        if ecole_id:
            # Admin école, utilisateur normal, OU admin système avec école sélectionnée
            classes = Classe.query.filter_by(ecole_id=ecole_id).order_by(Classe.nom).all()
            print(f"🏫 Récupération des classes de l'école {ecole_id} - {len(classes)} classes trouvées")
        else:
            # Admin système sans école sélectionnée - toutes les classes
            classes = Classe.query.order_by(Classe.nom).all()
            print("🔧 ADMIN SYSTÈME SANS ÉCOLE - Récupération de TOUTES les classes")
        
        classes_data = [{"id": str(c.id), "nom": c.nom} for c in classes if c and c.nom]
        print(f"✅ Classes retournées: {len(classes_data)}")
        
        return jsonify(classes_data)
        
    except Exception as e:
        print(f"❌ Erreur get_eleves_classes: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ========== EXPORT ÉLÈVES ==========
@services_bp.route("/export/eleves")
@login_required
def export_eleves():
    """Export RÉEL des élèves avec gestion des permissions"""
    try:
        export_type = request.args.get('type', 'pdf')
        classe_id = request.args.get('classe_id')
        ecole_id = get_ecole_id_courante()
        
        print(f"🔍 EXPORT ÉLÈVES - type:{export_type}, classe:{classe_id}, école:{ecole_id}")
        print(f"🔍 Paramètres complets: {dict(request.args)}")
        
        # Construction de la requête selon le type d'utilisateur
        query = Eleve.query
        
        # ✅ CORRECTION : Toujours filtrer par école si disponible
        if ecole_id:
            # Admin école, utilisateur normal, OU admin système avec école sélectionnée
            query = query.join(Classe).filter(Classe.ecole_id == ecole_id)
            print(f"🏫 Export des élèves de l'école {ecole_id}")
            
            # ✅ Vérifier que l'école existe
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                print(f"🏫 Nom école: {ecole.nom}")
        else:
            # Admin système sans école sélectionnée - tous les élèves
            query = query.join(Classe)
            print("🔧 ADMIN SYSTÈME SANS ÉCOLE - Export de TOUS les élèves")
        
        # Filtre par classe si spécifié
        if classe_id and classe_id not in ['', 'null', 'undefined']:
            query = query.filter(Eleve.classe_id == classe_id)
            print(f"✅ Filtre classe appliqué: {classe_id}")
        
        # Récupération des élèves
        eleves = query.order_by(Classe.nom, Eleve.nom, Eleve.prenoms).all()
        
        if not eleves:
            return jsonify({"error": "Aucun élève trouvé avec les critères sélectionnés"}), 404
        
        print(f"✅ {len(eleves)} élève(s) trouvé(s)")
        
        # Génération de l'export
        if export_type == 'pdf':
            return generate_eleves_pdf(eleves, classe_id, ecole_id)
        else:
            return generate_eleves_excel(eleves, classe_id, ecole_id)
            
    except Exception as e:
        print(f"❌ ERREUR export_eleves: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": f"Erreur lors de l'export: {str(e)}"}), 500

# ========== FONCTIONS DE GÉNÉRATION PDF ==========

def generate_eleves_pdf(eleves, classe_id, ecole_id):
    """Génère un PDF avec la liste des élèves"""
    try:
        print(f"📊 DEBUT generate_eleves_pdf - École: {ecole_id}, Classe: {classe_id}")
        
        # Création du buffer pour le PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4,
            topMargin=15*mm,
            bottomMargin=15*mm,
            leftMargin=10*mm,
            rightMargin=10*mm
        )
        elements = []
        
        # Titre principal
        ecole_nom = "TOUTES LES ÉCOLES"
        if ecole_id:
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                ecole_nom = ecole.nom
        
        if classe_id:
            classe = Classe.query.get(classe_id)
            titre_principal = f"LISTE DES ÉLÈVES - {classe.nom.upper()}" if classe else f"LISTE DES ÉLÈVES - {ecole_nom.upper()}"
        else:
            titre_principal = f"LISTE DES ÉLÈVES - {ecole_nom.upper()}"
        
        print(f"🎯 Titre PDF: {titre_principal}")
        
        # Ajouter l'en-tête
        elements = add_pdf_header(elements, doc, ecole_id, titre_principal)
        
        # Styles
        styles = getSampleStyleSheet()
        
        # Informations
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,
            spaceAfter=12,
            textColor=colors.HexColor('#34495E')
        )
        
        effectif_total = len(eleves)
        garcons = sum(1 for e in eleves if e.sexe and e.sexe.upper() == 'M')
        filles = effectif_total - garcons
        
        info_text = f"Effectif total: {effectif_total} élèves | Garçons: {garcons} | Filles: {filles}"
        elements.append(Paragraph(f"<b>{info_text}</b>", info_style))
        
        # Date de génération
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            spaceAfter=8,
            textColor=colors.gray
        )
        
        elements.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style))
        elements.append(Spacer(1, 8))
        
        # Tableau des élèves
        if eleves:
            # ✅ CORRECTION : Nouveaux en-têtes avec matricule et statut
            headers = ['N°', 'Matricule', 'Nom', 'Prénoms', 'Sexe', 'Date Naiss.', 'Statut']
            table_data = [headers]
            
            for index, eleve in enumerate(eleves, 1):
                date_naissance = eleve.date_naissance.strftime("%d/%m/%Y") if eleve.date_naissance else "N/D"
                
                table_data.append([
                    str(index),
                    eleve.matricule or '',
                    eleve.nom or '',
                    eleve.prenoms or '',
                    eleve.sexe or '',
                    date_naissance,
                    eleve.status or ''
                ])
            
            # Largeurs des colonnes optimisées
            column_widths = [
                15*mm,  # N°
                25*mm,  # Matricule
                40*mm,  # Nom
                45*mm,  # Prénoms
                20*mm,  # Sexe
                25*mm,  # Date Naiss.
                25*mm   # Statut
            ]
            
            table = Table(table_data, colWidths=column_widths, repeatRows=1)
            
            # Style du tableau
            table.setStyle(TableStyle([
                # En-têtes
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                
                # Données
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # N°
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),    # Matricule
                ('ALIGN', (2, 1), (3, -1), 'LEFT'),    # Nom, Prénoms
                ('ALIGN', (4, 1), (6, -1), 'CENTER'),  # Sexe, Date, Statut
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('WORDWRAP', (1, 1), (3, -1), True),   # Matricule, Nom, Prénoms
                
                # Bordures
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                
                # Alternance des couleurs
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
                    colors.HexColor('#FFFFFF'), 
                    colors.HexColor('#F8F9FA')
                ]),
                
                # Padding
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            
            elements.append(table)
        else:
            # Aucun élève trouvé
            no_data_style = ParagraphStyle(
                'NoDataStyle',
                parent=styles['Normal'],
                fontSize=12,
                alignment=1,
                textColor=colors.gray,
                spaceAfter=20
            )
            elements.append(Paragraph("Aucun élève trouvé", no_data_style))
        
        # Génération du PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Nom du fichier
        if classe_id:
            classe = Classe.query.get(classe_id)
            filename = f"eleves_{classe.nom.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        else:
            ecole_suffix = ecole_nom.replace(' ', '_') if ecole_id else "toutes_ecoles"
            filename = f"eleves_{ecole_suffix}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        print(f"❌ Erreur génération PDF élèves: {str(e)}")
        traceback.print_exc()
        return f"Erreur lors de la génération du PDF: {str(e)}", 500

# ========== FONCTIONS DE GÉNÉRATION EXCEL ==========

def generate_eleves_excel(eleves, classe_id, ecole_id):
    """Génère un Excel avec la liste des élèves"""
    try:
        # Création du workbook
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        
        # ✅ CORRECTION : Nom de l'onglet avec gestion de l'école
        ecole_nom = "Toutes écoles"
        if ecole_id:
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                ecole_nom = ecole.nom
        
        if classe_id:
            classe = Classe.query.get(classe_id)
            ws.title = f"Élèves {classe.nom}"[:31] if classe else f"Élèves {ecole_nom}"[:31]
        else:
            ws.title = f"Élèves {ecole_nom}"[:31]
        
        # Informations supplémentaires pour l'en-tête
        infos_supp = f"Effectif: {len(eleves)} élèves"
        if classe_id:
            classe = Classe.query.get(classe_id)
            if classe:
                infos_supp = f"Classe: {classe.nom} | {len(eleves)} élèves"
        
        # Ajouter l'en-tête
        start_row = add_excel_header(ws, ecole_id, "LISTE DES ÉLÈVES", infos_supp)
        
        # Style des bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ✅ CORRECTION : Nouveaux en-têtes avec matricule et statut
        headers = ['N°', 'Matricule', 'Nom', 'Prénoms', 'Sexe', 'Date de naissance', 'Statut']
        
        # En-têtes du tableau
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Données des élèves
        for index, eleve in enumerate(eleves, 1):
            row_num = start_row + index
            date_naissance = eleve.date_naissance.strftime("%d/%m/%Y") if eleve.date_naissance else "Non définie"
            
            row_data = [
                index,
                eleve.matricule or '',
                eleve.nom or '',
                eleve.prenoms or '',
                eleve.sexe or '',
                date_naissance,
                eleve.status or ''
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = value
                cell.font = Font(size=9)
                cell.border = thin_border
        
        # Largeurs des colonnes
        column_widths = {
            'A': 6,    # N°
            'B': 12,   # Matricule
            'C': 18,   # Nom
            'D': 20,   # Prénoms
            'E': 8,    # Sexe
            'F': 12,   # Date de naissance
            'G': 10    # Statut
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Alignement et retour à la ligne
        for row in range(start_row, ws.max_row + 1):
            for col in range(1, len(headers) + 1):
                if col in [3, 4]:  # Nom et Prénoms
                    ws.cell(row=row, column=col).alignment = Alignment(
                        horizontal='left', vertical='center', wrap_text=True
                    )
                elif col in [1, 5, 6, 7]:  # Colonnes à centrer
                    ws.cell(row=row, column=col).alignment = Alignment(
                        horizontal='center', vertical='center'
                    )
                else:
                    ws.cell(row=row, column=col).alignment = Alignment(
                        horizontal='left', vertical='center'
                    )
        
        # Alternance des couleurs des lignes
        for row in range(start_row + 1, ws.max_row + 1):
            if row % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
                    )
        
        # Réglage de la hauteur des lignes
        for row in range(start_row, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
        
        # Sauvegarde
        wb.save(buffer)
        buffer.seek(0)
        
        # Nom du fichier
        if classe_id:
            classe = Classe.query.get(classe_id)
            filename = f"eleves_{classe.nom.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        else:
            ecole_suffix = ecole_nom.replace(' ', '_') if ecole_id else "toutes_ecoles"
            filename = f"eleves_{ecole_suffix}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"❌ Erreur génération Excel élèves: {str(e)}")
        traceback.print_exc()
        return f"Erreur lors de la génération du Excel: {str(e)}", 500

def recreate_logos_simple_format():
    """Recrée tous les logos avec un format JPEG simple"""
    try:
        ecoles = Ecole.query.all()
        results = []
        
        for ecole in ecoles:
            try:
                print(f"🔄 Recréation logo pour {ecole.nom}")
                
                # Créer une image simple
                from PIL import Image, ImageDraw, ImageFont
                
                # Dimensions
                width, height = 200, 200
                
                # Couleur de fond
                img = Image.new('RGB', (width, height), color=(41, 128, 185))  # Bleu
                draw = ImageDraw.Draw(img)
                
                # Cadre blanc
                draw.rectangle([10, 10, width-10, height-10], outline='white', width=3)
                
                # Initiales
                initials = ''.join([word[0].upper() for word in ecole.nom.split()[:2]]) or "ECOLE"
                
                # Essayer différentes polices
                try:
                    font_size = 60 if len(initials) <= 3 else 40
                    font = ImageFont.truetype("arial.ttf", font_size)
                except:
                    try:
                        font = ImageFont.truetype("arialbd.ttf", font_size)
                    except:
                        font = ImageFont.load_default()
                
                # Centrer le texte
                bbox = draw.textbbox((0, 0), initials, font=font)
                text_width = bbox[2] - bbox[0]
                text_height = bbox[3] - bbox[1]
                x = (width - text_width) // 2
                y = (height - text_height) // 2
                
                draw.text((x, y), initials, fill='white', font=font)
                
                # Sauvegarder en JPEG (plus compatible)
                upload_dir = os.path.join('app', 'static', 'uploads', 'logos')
                os.makedirs(upload_dir, exist_ok=True)
                
                logo_filename = f"{ecole.id}.jpg"  # ✅ Utiliser JPG au lieu de PNG
                logo_path = os.path.join(upload_dir, logo_filename)
                
                img.save(logo_path, 'JPEG', quality=95)
                
                # Mettre à jour la BD
                ecole.logo_filename = logo_filename
                db.session.commit()
                
                results.append({
                    'ecole': ecole.nom,
                    'new_logo': logo_filename,
                    'status': '✅ Recréé'
                })
                
                print(f"✅ Logo recréé: {logo_path}")
                
            except Exception as e:
                results.append({
                    'ecole': ecole.nom,
                    'status': f'❌ Erreur: {str(e)}'
                })
        
        return jsonify({
            'message': f'Logos recréés pour {len(ecoles)} écoles',
            'results': results
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
# ========== EXPORT ENSEIGNANTS ==========

@services_bp.route("/export/enseignants")
@login_required
def export_enseignants():
    """Export RÉEL des enseignants avec gestion des permissions"""
    try:
        export_type = request.args.get('type', 'pdf')
        ecole_id = get_ecole_id_courante()
        
        print(f"🔍 EXPORT ENSEIGNANTS - type:{export_type}, école:{ecole_id}")
        print(f"🔍 Paramètres complets: {dict(request.args)}")
        
        
        query = Enseignant.query.options(
            joinedload(Enseignant.utilisateur),
            joinedload(Enseignant.matieres)
        )
        
        # ✅ CORRECTION : Toujours filtrer par école si disponible
        if ecole_id:
            # Admin école, utilisateur normal, OU admin système avec école sélectionnée
            query = query.filter(Enseignant.ecole_id == ecole_id)
            print(f"🏫 Export des enseignants de l'école {ecole_id}")
            
            # ✅ Vérifier que l'école existe
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                print(f"🏫 Nom école: {ecole.nom}")
        else:
            # Admin système sans école sélectionnée - tous les enseignants
            print("🔧 ADMIN SYSTÈME SANS ÉCOLE - Export de TOUS les enseignants")
        
        # ✅ CORRECTION : Utiliser l'alias correct pour le tri
        # Au lieu de Utilisateur.nom, utiliser la relation chargée
        enseignants = query.all()
        
        # ✅ CORRECTION : Trier manuellement après récupération
        enseignants = sorted(enseignants, key=lambda e: (
            e.utilisateur.nom or '', 
            e.utilisateur.prenoms or ''
        ))
        
        if not enseignants:
            return jsonify({"error": "Aucun enseignant trouvé avec les critères sélectionnés"}), 404
        
        print(f"✅ {len(enseignants)} enseignant(s) trouvé(s)")
        
        # Génération de l'export
        if export_type == 'pdf':
            return generate_enseignants_pdf(enseignants, ecole_id)
        else:
            return generate_enseignants_excel(enseignants, ecole_id)
            
    except Exception as e:
        print(f"❌ ERREUR export_enseignants: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": f"Erreur lors de l'export: {str(e)}"}), 500

def generate_enseignants_pdf(enseignants, ecole_id):
    """Génère un PDF avec la liste des enseignants"""
    try:
        print(f"📊 DEBUT generate_enseignants_pdf - École: {ecole_id}")
        
        # Création du buffer pour le PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4,
            topMargin=15*mm,
            bottomMargin=15*mm,
            leftMargin=8*mm,  # ✅ Réduit les marges latérales
            rightMargin=8*mm
        )
        elements = []
        
        # Titre principal
        ecole_nom = "TOUTES LES ÉCOLES"
        if ecole_id:
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                ecole_nom = ecole.nom
        
        titre_principal = f"LISTE DES ENSEIGNANTS - {ecole_nom.upper()}"
        print(f"🎯 Titre PDF enseignants: {titre_principal}")
        
        # Ajouter l'en-tête
        elements = add_pdf_header(elements, doc, ecole_id, titre_principal)
        
        # Styles
        styles = getSampleStyleSheet()
        
        # Informations
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,
            spaceAfter=12,
            textColor=colors.HexColor('#34495E')
        )
        
        effectif_total = len(enseignants)
        hommes = sum(1 for e in enseignants if e.utilisateur.sexe and e.utilisateur.sexe.upper() == 'M')
        femmes = effectif_total - hommes
        
        info_text = f"Effectif total: {effectif_total} enseignants | Hommes: {hommes} | Femmes: {femmes}"
        elements.append(Paragraph(f"<b>{info_text}</b>", info_style))
        
        # Date de génération
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            spaceAfter=8,
            textColor=colors.gray
        )
        
        elements.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style))
        elements.append(Spacer(1, 8))
        
        # Tableau des enseignants
        if enseignants:
            headers = ['N°', 'Nom', 'Prénoms', 'Sexe', 'Téléphone', 'Email', 'Matières']
            table_data = [headers]
            
            for index, enseignant in enumerate(enseignants, 1):
                utilisateur = enseignant.utilisateur
                # ✅ CORRECTION : Vérifier que les matières sont chargées
                matieres = ", ".join([m.libelle for m in enseignant.matieres]) if enseignant.matieres and hasattr(enseignant, 'matieres') else "Non assigné"
                
                # ✅ CORRECTION : Tronquer les textes longs
                email = (utilisateur.email or "")[:20] + "..." if len(utilisateur.email or "") > 20 else (utilisateur.email or "")
                matieres_trunc = matieres[:25] + "..." if len(matieres) > 25 else matieres
                
                table_data.append([
                    str(index),
                    utilisateur.nom or '',
                    utilisateur.prenoms or '',
                    utilisateur.sexe or '',
                    utilisateur.telephone or '',
                    email,
                    matieres_trunc
                ])
            
            # ✅ CORRECTION : Largeurs de colonnes optimisées pour A4
            total_width = doc.width
            column_widths = [
                total_width * 0.05,  # N° (5%)
                total_width * 0.15,  # Nom (15%)
                total_width * 0.15,  # Prénoms (15%)
                total_width * 0.07,  # Sexe (7%)
                total_width * 0.12,  # Téléphone (12%)
                total_width * 0.18,  # Email (18%)
                total_width * 0.28   # Matières (28%)
            ]
            
            table = Table(table_data, colWidths=column_widths, repeatRows=1)
            
            # ✅ CORRECTION : Style amélioré avec gestion du texte
            table.setStyle(TableStyle([
                # En-têtes
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),  # ✅ Taille réduite
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 4),
                
                # Données
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),  # ✅ Taille réduite
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # N°
                ('ALIGN', (1, 1), (2, -1), 'LEFT'),    # Nom, Prénoms
                ('ALIGN', (3, 1), (4, -1), 'CENTER'),  # Sexe, Téléphone
                ('ALIGN', (5, 1), (6, -1), 'LEFT'),    # Email, Matières
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                
                # ✅ CORRECTION : Retour à la ligne pour les colonnes texte
                ('WORDWRAP', (1, 1), (2, -1), True),   # Nom, Prénoms
                ('WORDWRAP', (5, 1), (6, -1), True),   # Email, Matières
                
                # Bordures
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                
                # Alternance des couleurs
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
                    colors.HexColor('#FFFFFF'), 
                    colors.HexColor('#F8F9FA')
                ]),
                
                # ✅ CORRECTION : Padding réduit
                ('LEFTPADDING', (0, 0), (-1, -1), 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]))
            
            elements.append(table)
        else:
            # Aucun enseignant trouvé
            no_data_style = ParagraphStyle(
                'NoDataStyle',
                parent=styles['Normal'],
                fontSize=12,
                alignment=1,
                textColor=colors.gray,
                spaceAfter=20
            )
            elements.append(Paragraph("Aucun enseignant trouvé", no_data_style))
        
        # Génération du PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Nom du fichier
        ecole_suffix = ecole_nom.replace(' ', '_') if ecole_id else "toutes_ecoles"
        filename = f"enseignants_{ecole_suffix}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        print(f"❌ Erreur génération PDF enseignants: {str(e)}")
        traceback.print_exc()
        return f"Erreur lors de la génération du PDF: {str(e)}", 500

def generate_enseignants_excel(enseignants, ecole_id):
    """Génère un Excel avec la liste des enseignants"""
    try:
        # Création du workbook
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        
        # ✅ CORRECTION : Nom de l'onglet avec gestion de l'école
        ecole_nom = "Toutes écoles"
        if ecole_id:
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                ecole_nom = ecole.nom
        
        ws.title = f"Enseignants {ecole_nom}"[:31]
        
        # Informations supplémentaires pour l'en-tête
        infos_supp = f"Effectif: {len(enseignants)} enseignants"
        
        # Ajouter l'en-tête
        start_row = add_excel_header(ws, ecole_id, "LISTE DES ENSEIGNANTS", infos_supp)
        
        # Style des bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = ['N°', 'Nom', 'Prénoms', 'Sexe', 'Téléphone', 'Email', 'Matières']
        
        # En-têtes du tableau
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Données des enseignants
        for index, enseignant in enumerate(enseignants, 1):
            row_num = start_row + index
            utilisateur = enseignant.utilisateur
            matieres = ", ".join([m.libelle for m in enseignant.matieres]) if enseignant.matieres else "Non assigné"
            
            # ✅ CORRECTION : Tronquer les emails longs pour Excel
            email = (utilisateur.email or "")[:30] + "..." if len(utilisateur.email or "") > 30 else (utilisateur.email or "")
            
            row_data = [
                index,
                utilisateur.nom or '',
                utilisateur.prenoms or '',
                utilisateur.sexe or '',
                utilisateur.telephone or '',
                email,
                matieres
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = value
                cell.font = Font(size=9)
                cell.border = thin_border
        
        # ✅ CORRECTION : Largeurs de colonnes optimisées
        column_widths = {
            'A': 5,    # N° (plus étroit)
            'B': 18,   # Nom (élargi)
            'C': 18,   # Prénoms (élargi)
            'D': 8,    # Sexe
            'E': 15,   # Téléphone (élargi)
            'F': 25,   # Email (élargi)
            'G': 35    # Matières (élargi)
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # ✅ CORRECTION : Alignement et retour à la ligne amélioré
        for row in range(start_row, ws.max_row + 1):
            # N°
            ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
            # Nom et Prénoms
            ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            # Sexe et Téléphone
            ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
            # Email et Matières
            ws[f'F{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws[f'G{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # ✅ CORRECTION : Ajuster la hauteur des lignes automatiquement
        for row in range(start_row, ws.max_row + 1):
            max_lines = 1
            # Vérifier les colonnes avec du texte long
            for col in ['B', 'C', 'F', 'G']:
                cell_value = ws[f'{col}{row}'].value
                if cell_value and len(str(cell_value)) > 30:
                    max_lines = max(max_lines, 2)
                if cell_value and len(str(cell_value)) > 60:
                    max_lines = max(max_lines, 3)
            
            ws.row_dimensions[row].height = 15 * max_lines  # 15 points par ligne
        
        # Alternance des couleurs des lignes
        for row in range(start_row + 1, ws.max_row + 1):
            if row % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
                    )
        
        # Sauvegarde
        wb.save(buffer)
        buffer.seek(0)
        
        # Nom du fichier
        ecole_suffix = ecole_nom.replace(' ', '_') if ecole_id else "toutes_ecoles"
        filename = f"enseignants_{ecole_suffix}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"❌ Erreur génération Excel enseignants: {str(e)}")
        traceback.print_exc()
        return f"Erreur lors de la génération du Excel: {str(e)}", 500

# ========== ROUTES API POUR LES FILTRES DES NOTES ==========

@services_bp.route("/notes/classes")
@login_required
def get_notes_classes():
    """API pour récupérer les classes (pour export notes)"""
    try:
        ecole_id = get_ecole_id_courante()
        
        print(f"🔍 GET /notes/classes - ecole_id: {ecole_id}")
        
        if ecole_id:
            classes = Classe.query.filter_by(ecole_id=ecole_id).order_by(Classe.nom).all()
            print(f"🏫 Récupération des classes de l'école {ecole_id}")
        else:
            classes = Classe.query.order_by(Classe.nom).all()
            print("🔧 ADMIN SYSTÈME - Récupération de TOUTES les classes")
        
        classes_data = [{"id": str(c.id), "nom": c.nom} for c in classes if c and c.nom]
        print(f"✅ Classes retournées: {len(classes_data)}")
        
        return jsonify(classes_data)
        
    except Exception as e:
        print(f"❌ Erreur get_notes_classes: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@services_bp.route("/notes/matieres")
@login_required
def get_notes_matieres():
    """API pour récupérer les matières (pour export notes) - VERSION CORRIGÉE"""
    try:
        ecole_id = get_ecole_id_courante()
        
        print(f"🔍 GET /notes/matieres - ecole_id: {ecole_id}")
        
        from ..models import Matiere
        
        # ✅ CORRECTION : Récupérer TOUTES les matières de l'école sans filtre d'état
        if ecole_id:
            matieres = Matiere.query.filter_by(ecole_id=ecole_id).order_by(Matiere.libelle).all()
            print(f"🏫 Récupération des matières de l'école {ecole_id}")
        else:
            matieres = Matiere.query.order_by(Matiere.libelle).all()
            print("🔧 ADMIN SYSTÈME - Récupération de TOUTES les matières")
        
        # ✅ CORRECTION : Log détaillé pour débogage
        print(f"📦 Matières trouvées: {len(matieres)}")
        for matiere in matieres[:5]:  # Afficher les 5 premières pour vérification
            print(f"   - {matiere.libelle} (ID: {matiere.id})")
        
        matieres_data = [{"id": str(m.id), "libelle": m.libelle} for m in matieres if m and m.libelle]
        print(f"✅ Matières retournées: {len(matieres_data)}")
        
        return jsonify(matieres_data)
        
    except Exception as e:
        print(f"❌ Erreur get_notes_matieres: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@services_bp.route("/notes/annees-scolaires")
@login_required
def get_notes_annees_scolaires():
    """API pour récupérer les années scolaires disponibles"""
    try:
        from ..models import Note
        from sqlalchemy import distinct
        
        # Récupérer les années scolaires distinctes depuis les notes
        annees = db.session.query(distinct(Note.annee_scolaire)).filter(
            Note.annee_scolaire.isnot(None)
        ).order_by(Note.annee_scolaire.desc()).all()
        
        annees_data = [{"annee": a[0]} for a in annees if a[0]]
        
        # Si aucune année trouvée, utiliser l'année actuelle
        if not annees_data:
            annee_courante = datetime.now().year
            annees_data = [{"annee": f"{annee_courante}-{annee_courante+1}"}]
        
        print(f"✅ Années scolaires retournées: {len(annees_data)}")
        return jsonify(annees_data)
        
    except Exception as e:
        print(f"❌ Erreur get_notes_annees_scolaires: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    
# ========== EXPORT NOTES AVEC GESTION DES SEMESTRES ==========

def export_notes(type_export, classe_id=None, matiere_id=None, ecole_type="college", 
                periode=None, annee_scolaire=None, ecole_id=None):
    try:
        print(f"🔍 EXPORT NOTES - type:{type_export}, classe:{classe_id}, matière:{matiere_id}")
        print(f"🔍 École type:{ecole_type}, Période:{periode}, Année:{annee_scolaire}, École:{ecole_id}")
        
        # Validation
        if not ecole_id:
            return None, "ID de l'école manquant"
        if not periode:
            return None, "Veuillez sélectionner une période"
        if not annee_scolaire:
            return None, "Veuillez sélectionner une année scolaire"
        
        print(f"🏫 Export des notes de l'école {ecole_id} (type: {ecole_type})")
        
        # Déterminer le nom de la classe
        classe_nom = "Toutes les classes"
        if classe_id and classe_id not in ['', 'null', 'undefined', 'toutes']:
            # Vérifier la classe si spécifiée
            classe = Classe.query.filter_by(id=classe_id, ecole_id=ecole_id).first()
            if not classe:
                return None, "Classe non trouvée dans cette école"
            classe_nom = classe.nom
            print(f"✅ Classe spécifique: {classe_nom}")
        else:
            print(f"✅ Export de toutes les classes")
        
        # Déterminer si c'est un trimestre ou semestre selon le type d'établissement
        type_periode = "trimestre" if ecole_type == "college" else "semestre"
        print(f"📅 Type de période: {type_periode} (école type: {ecole_type})")
        
        # Récupérer le nom de la matière
        matiere_nom = "Toutes les matières"
        matiere_libelle_filtre = None
        
        if matiere_id and matiere_id not in ['', 'null', 'undefined', 'toutes']:
            matiere = Matiere.query.get(matiere_id)
            if not matiere:
                return None, f"Matière non trouvée (ID: {matiere_id})"
            
            matiere_nom = matiere.libelle
            matiere_libelle_filtre = matiere.libelle
            print(f"✅ Filtre matière par NOM: {matiere_nom}")
        else:
            print("✅ Export de toutes les matières")
        
        # Construire la requête
        query = db.session.query(Note)\
            .join(Eleve, Note.eleve_id == Eleve.id)\
            .join(Matiere, Note.matiere_id == Matiere.id)\
            .filter(Note.trimestre == periode)\
            .filter(Note.annee_scolaire == annee_scolaire)
        
        # Filtrer par école (important pour la sécurité)
        if ecole_id:
            query = query.join(Classe).filter(Classe.ecole_id == ecole_id)
        
        # Filtrer par classe seulement si une classe spécifique est choisie
        if classe_id and classe_id not in ['', 'null', 'undefined', 'toutes']:
            query = query.filter(Eleve.classe_id == classe_id)
        
        if matiere_libelle_filtre:
            query = query.filter(Matiere.libelle == matiere_libelle_filtre)
        
        # ✅ IMPORTANT : Précharger la relation classe pour éviter les requêtes N+1
        query = query.options(
            joinedload(Note.eleve).joinedload(Eleve.classe),
            joinedload(Note.matiere)
        )
        
        # Tri amélioré pour toutes les classes
        if classe_id and classe_id not in ['', 'null', 'undefined', 'toutes']:
            # Si classe spécifique : trier par matière puis par élève
            notes = query.order_by(
                Matiere.libelle,
                Eleve.nom,
                Eleve.prenoms
            ).all()
        else:
            # Si toutes les classes : trier par classe, puis matière, puis élève
            notes = query.order_by(
                Classe.nom,
                Matiere.libelle,
                Eleve.nom,
                Eleve.prenoms
            ).all()
        
        print(f"✅ Notes récupérées: {len(notes)}")
        
        if not notes:
            error_msg = f"Aucune note trouvée pour {matiere_nom} "
            if classe_id and classe_id not in ['', 'null', 'undefined', 'toutes']:
                error_msg += f"dans {classe_nom} "
            else:
                error_msg += f"dans toutes les classes "
            
            if ecole_type == "college":
                error_msg += f"(Trimestre {periode}, {annee_scolaire})"
            else:
                error_msg += f"(Semestre {periode}, {annee_scolaire})"
            
            return None, error_msg
        
        # Génération de l'export
        if type_export == 'pdf':
            try:
                pdf_buffer = generate_notes_pdf_buffer_periode(
                    notes, ecole_id, classe_nom, matiere_nom, 
                    ecole_type, periode, annee_scolaire
                )
                nom_fichier = generer_nom_fichier_notes_periode(
                    classe_nom, matiere_nom, ecole_type, periode, annee_scolaire, 'pdf'
                )
                print(f"✅ PDF généré: {nom_fichier}")
                return pdf_buffer, nom_fichier
            except Exception as e:
                print(f"❌ Erreur génération PDF: {str(e)}")
                traceback.print_exc()
                return None, f"Erreur lors de la génération du PDF: {str(e)}"
        
        elif type_export == 'excel':
            try:
                excel_buffer = generate_notes_excel_buffer_periode(
                    notes, ecole_id, classe_nom, matiere_nom,
                    ecole_type, periode, annee_scolaire
                )
                nom_fichier = generer_nom_fichier_notes_periode(
                    classe_nom, matiere_nom, ecole_type, periode, annee_scolaire, 'xlsx'
                )
                print(f"✅ Excel généré: {nom_fichier}")
                return excel_buffer, nom_fichier
            except Exception as e:
                print(f"❌ Erreur génération Excel: {str(e)}")
                traceback.print_exc()
                return None, f"Erreur lors de la génération du Excel: {str(e)}"
        
        else:
            return None, f"Type d'export non supporté: {type_export}"
            
    except Exception as e:
        print(f"❌ ERREUR export_notes: {str(e)}")
        traceback.print_exc()
        return None, f"Erreur lors de l'export: {str(e)}"
    
def generer_nom_fichier_notes_periode(classe_nom, matiere_nom, ecole_type, periode, annee_scolaire, extension):
    """Génère un nom de fichier pour l'export des notes"""
    prefixe_periode = "T" if ecole_type == "college" else "S"
    
    # ✅ CORRECTION : Gestion complète du nom de fichier
    nom_fichier = ""
    
    if classe_nom == "Toutes les classes":
        nom_fichier = f"notes_toutes_classes"
    else:
        nom_fichier = f"notes_{classe_nom.replace(' ', '_')}"
    
    if matiere_nom == "Toutes les matières":
        nom_fichier += "_toutes_matieres"
    else:
        nom_fichier += f"_{matiere_nom.replace(' ', '_')}"
    
    nom_fichier += f"_{prefixe_periode}{periode}"
    nom_fichier += f"_{annee_scolaire.replace('-', '_')}"
    nom_fichier += f"_{datetime.now().strftime('%Y%m%d_%H%M')}.{extension}"
    
    return nom_fichier

def generate_notes_pdf_buffer_periode(notes, ecole_id, classe_nom, matiere_nom, ecole_type, periode, annee_scolaire):
    """Génère un buffer PDF avec gestion correcte des périodes et colonnes dynamiques"""
    try:
        print(f"📊 DEBUT generate_notes_pdf_buffer_periode - Type école: {ecole_type}")
        print(f"   Matière: {matiere_nom}, Classe: {classe_nom}")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4,
            topMargin=15*mm,
            bottomMargin=15*mm,
            leftMargin=8*mm,
            rightMargin=8*mm
        )
        elements = []
        
        # Titre principal adapté
        ecole_nom = "TOUTES LES ÉCOLES"
        if ecole_id:
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                ecole_nom = ecole.nom
        
        # Déterminer le titre de période
        titre_periode = f"TRIMESTRE {periode}" if ecole_type == "college" else f"SEMESTRE {periode}"
        
        # Titre différent selon le type d'export
        if classe_nom == "Toutes les classes" and matiere_nom == "Toutes les matières":
            titre_principal = f"LISTE DES NOTES - TOUTES CLASSES & MATIÈRES"
        elif classe_nom == "Toutes les classes":
            titre_principal = f"LISTE DES NOTES - TOUTES CLASSES"
        elif matiere_nom == "Toutes les matières":
            titre_principal = f"LISTE DES NOTES - TOUTES MATIÈRES"
        else:
            titre_principal = f"LISTE DES NOTES - {matiere_nom.upper()}"
        
        elements = add_pdf_header(elements, doc, ecole_id, titre_principal)
        
        styles = getSampleStyleSheet()
        
        # Informations des filtres avec période correcte
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,
            spaceAfter=8,
            textColor=colors.HexColor('#34495E')
        )
        
        info_parts = [
            f"Classe: {classe_nom}",
            f"Matière: {matiere_nom}",
            f"Période: {titre_periode}",
            f"Année: {annee_scolaire}",
            f"Total: {len(notes)} notes"
        ]
        
        info_text = " | ".join(info_parts)
        elements.append(Paragraph(f"<b>{info_text}</b>", info_style))
        
        # Tableau des notes - AVEC COLONNES DYNAMIQUES
        if notes:
            # ✅ AMÉLIORATION : Déterminer les en-têtes dynamiquement
            headers = ['N°', 'Élève']
            
            # Ajouter colonne "Classe" si on exporte toutes les classes
            if classe_nom == "Toutes les classes":
                headers.append('Classe')
            
            # Ajouter colonne "Matière" si on exporte toutes les matières
            if matiere_nom == "Toutes les matières":
                headers.append('Matière')
            
            # Ajouter les colonnes de notes
            headers.extend(['Note 1', 'Note 2', 'Note 3', 'Note Comp', 'Coeff'])
            
            table_data = [headers]
            
            for index, note in enumerate(notes, 1):
                eleve = note.eleve
                matiere = note.matiere
                
                # Récupérer le nom de la classe de l'élève
                classe_eleve = note.eleve.classe.nom if note.eleve and note.eleve.classe else '-'
                
                note1 = f"{note.note1}/20" if note.note1 is not None else "-"
                note2 = f"{note.note2}/20" if note.note2 is not None else "-"
                note3 = f"{note.note3}/20" if note.note3 is not None else "-"
                note_comp = f"{note.note_comp}/20" if note.note_comp is not None else "-"
                
                # Construire la ligne dynamiquement
                row_data = [str(index), f"{eleve.nom} {eleve.prenoms}" if eleve else '-']
                
                # Ajouter colonne "Classe" si nécessaire
                if classe_nom == "Toutes les classes":
                    row_data.append(classe_eleve)
                
                # Ajouter colonne "Matière" si nécessaire
                if matiere_nom == "Toutes les matières":
                    row_data.append(matiere.libelle if matiere else '-')
                
                # Ajouter les notes
                row_data.extend([note1, note2, note3, note_comp, str(note.coefficient) if note.coefficient else "1"])
                
                table_data.append(row_data)
            
            # ✅ AMÉLIORATION : Calculer les largeurs de colonnes dynamiquement
            total_width = doc.width
            column_widths = []
            
            # Largeur fixe pour N°
            column_widths.append(total_width * 0.04)
            
            # Largeur pour Élève
            if classe_nom == "Toutes les classes" and matiere_nom == "Toutes les matières":
                column_widths.append(total_width * 0.20)  # Élève moins large
            elif classe_nom == "Toutes les classes" or matiere_nom == "Toutes les matières":
                column_widths.append(total_width * 0.25)  # Élève moyennement large
            else:
                column_widths.append(total_width * 0.30)  # Élève large
            
            # Largeur pour Classe (si présente)
            if classe_nom == "Toutes les classes":
                column_widths.append(total_width * 0.15)
            
            # Largeur pour Matière (si présente)
            if matiere_nom == "Toutes les matières":
                column_widths.append(total_width * 0.20)
            
            # Largeurs pour les notes (toujours présentes)
            notes_width = total_width - sum(column_widths)
            note_columns = 5  # Note 1, Note 2, Note 3, Note Comp, Coeff
            note_width = notes_width / note_columns
            
            for _ in range(5):
                column_widths.append(note_width)
            
            table = Table(table_data, colWidths=column_widths, repeatRows=1)
            
            # Styles de base
            style_commands = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 4),
                
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            ]
            
            # Calculer les indices des colonnes
            col_count = len(headers)
            eleve_col_index = 1
            last_text_col_index = eleve_col_index
            
            # Déterminer les colonnes de texte
            text_columns = [1]  # Élève est toujours une colonne texte
            
            if classe_nom == "Toutes les classes":
                text_columns.append(2)
                last_text_col_index = 2
            
            if matiere_nom == "Toutes les matières":
                matiere_col_index = 3 if classe_nom == "Toutes les classes" else 2
                text_columns.append(matiere_col_index)
                last_text_col_index = matiere_col_index
            
            # Alignement des colonnes de texte
            for col in text_columns:
                style_commands.append(('ALIGN', (col, 1), (col, -1), 'LEFT'))
                style_commands.append(('WORDWRAP', (col, 1), (col, -1), True))
            
            # Alignement des colonnes de notes (colonnes numériques)
            notes_start_col = last_text_col_index + 1
            style_commands.append(('ALIGN', (notes_start_col, 1), (-1, -1), 'CENTER'))
            
            # Alternance des couleurs
            style_commands.append(('ROWBACKGROUNDS', (0, 1), (-1, -1), [
                colors.HexColor('#FFFFFF'), 
                colors.HexColor('#F8F9FA')
            ]))
            
            # Padding
            style_commands.extend([
                ('LEFTPADDING', (0, 0), (-1, -1), 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ])
            
            table.setStyle(TableStyle(style_commands))
            elements.append(table)
        else:
            no_data_style = ParagraphStyle(
                'NoDataStyle',
                parent=styles['Normal'],
                fontSize=12,
                alignment=1,
                textColor=colors.gray,
                spaceAfter=20
            )
            elements.append(Paragraph("Aucune note trouvée", no_data_style))
        
        # Date de génération
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            spaceAfter=8,
            textColor=colors.gray
        )
        elements.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style))
        
        doc.build(elements)
        buffer.seek(0)
        
        print(f"✅ Buffer PDF généré avec {len(notes)} notes, {len(headers)} colonnes")
        print(f"   Colonnes: {headers}")
        return buffer
        
    except Exception as e:
        print(f"❌ Erreur génération PDF notes: {str(e)}")
        traceback.print_exc()
        raise

def get_trimestres_for_periode(type_periode, periode, ecole_type):
    """Détermine les trimestres à inclure selon la période et le type d'établissement"""
    if type_periode == "trimestre":
        return [periode]
    elif type_periode == "semestre":
        if ecole_type == "lycee":
            # Lycée: Semestre 1 = T1, Semestre 2 = T2
            if periode == "1":
                return ["1"]
            elif periode == "2":
                return ["2"]
        else:
            # Collège: Semestre 1 = T1+T2, Semestre 2 = T3
            if periode == "1":
                return ["1", "2"]
            elif periode == "2":
                return ["3"]
    return [periode]

def generer_nom_fichier_notes(classe_nom, matiere_nom, type_periode, periode, annee_scolaire, extension):
    """Génère un nom de fichier pour l'export des notes"""
    prefixe = "trimestre" if type_periode == "trimestre" else "semestre"
    nom_fichier = f"notes_{classe_nom.replace(' ', '_')}_{matiere_nom.replace(' ', '_')}_{prefixe}{periode}"
    nom_fichier += f"_{annee_scolaire.replace('-', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.{extension}"
    return nom_fichier

def generate_notes_pdf_buffer_semestre(notes, ecole_id, classe_nom, matiere_nom, 
                                     type_periode, periode, annee_scolaire, ecole_type):
    """Génère un buffer PDF avec gestion des semestres"""
    try:
        print(f"📊 DEBUT generate_notes_pdf_buffer_semestre - Type période: {type_periode}")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4,
            topMargin=15*mm,
            bottomMargin=15*mm,
            leftMargin=8*mm,
            rightMargin=8*mm
        )
        elements = []
        
        # Titre principal avec gestion des périodes
        ecole_nom = "TOUTES LES ÉCOLES"
        if ecole_id:
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                ecole_nom = ecole.nom
        
        titre_prefixe = "SEMESTRE" if type_periode == "semestre" else "TRIMESTRE"
        titre_principal = f"LISTE DES NOTES"
        
        elements = add_pdf_header(elements, doc, ecole_id, titre_principal)
        
        styles = getSampleStyleSheet()
        
        # Informations des filtres avec période adaptée
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,
            spaceAfter=8,
            textColor=colors.HexColor('#34495E')
        )
        
        periode_texte = f"{titre_prefixe} {periode}"
        if type_periode == "semestre" and ecole_type == "college":
            if periode == "1":
                periode_texte += " (regroupe T1 + T2)"
            elif periode == "2":
                periode_texte += " (T3)"
        
        info_parts = [
            f"Classe: {classe_nom}",
            f"Matière: {matiere_nom}",
            f"Période: {periode_texte}",
            f"Année: {annee_scolaire}",
            f"Total: {len(notes)} notes"
        ]
        
        info_text = " | ".join(info_parts)
        elements.append(Paragraph(f"<b>{info_text}</b>", info_style))
        
        # Grouper les notes par élève et trimestre pour affichage
        notes_par_eleve = {}
        for note in notes:
            eleve_id = note.eleve_id
            if eleve_id not in notes_par_eleve:
                notes_par_eleve[eleve_id] = {
                    'eleve': note.eleve,
                    'notes_par_trimestre': {}
                }
            
            trimestre = note.trimestre
            if trimestre not in notes_par_eleve[eleve_id]['notes_par_trimestre']:
                notes_par_eleve[eleve_id]['notes_par_trimestre'][trimestre] = []
            
            notes_par_eleve[eleve_id]['notes_par_trimestre'][trimestre].append(note)
        
        # Tableau des notes
        if notes_par_eleve:
            headers = ['N°', 'Élève', 'Trimestre', 'Note 1', 'Note 2', 'Note 3', 'Note Comp', 'Coeff']
            table_data = [headers]
            
            index = 1
            for eleve_id, data in notes_par_eleve.items():
                eleve = data['eleve']
                
                for trimestre, notes_trimestre in sorted(data['notes_par_trimestre'].items()):
                    for note in notes_trimestre:
                        # Formatage des notes
                        note1 = f"{note.note1}/20" if note.note1 is not None else "-"
                        note2 = f"{note.note2}/20" if note.note2 is not None else "-"
                        note3 = f"{note.note3}/20" if note.note3 is not None else "-"
                        note_comp = f"{note.note_comp}/20" if note.note_comp is not None else "-"
                        
                        table_data.append([
                            str(index),
                            f"{eleve.nom} {eleve.prenoms}" if eleve else '-',
                            f"T{trimestre}",
                            note1,
                            note2,
                            note3,
                            note_comp,
                            str(note.coefficient) if note.coefficient else "1"
                        ])
                        index += 1
                
                # Ajouter une ligne de séparation entre les élèves
                if index < len(notes_par_eleve) * 2:  # Pas après le dernier élève
                    table_data.append(['', '', '', '', '', '', '', ''])
            
            total_width = doc.width
            column_widths = [
                total_width * 0.04,
                total_width * 0.25,
                total_width * 0.08,
                total_width * 0.08,
                total_width * 0.08,
                total_width * 0.08,
                total_width * 0.11,
                total_width * 0.08
            ]
            
            table = Table(table_data, colWidths=column_widths, repeatRows=1)
            
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 4),
                
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 6),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('WORDWRAP', (1, 1), (1, -1), True),
                
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
                    colors.HexColor('#FFFFFF'), 
                    colors.HexColor('#F8F9FA')
                ]),
                
                ('LEFTPADDING', (0, 0), (-1, -1), 1),
                ('RIGHTPADDING', (0, 0), (-1, -1), 1),
                ('TOPPADDING', (0, 0), (-1, -1), 1),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
            ]))
            
            elements.append(table)
        else:
            no_data_style = ParagraphStyle(
                'NoDataStyle',
                parent=styles['Normal'],
                fontSize=12,
                alignment=1,
                textColor=colors.gray,
                spaceAfter=20
            )
            elements.append(Paragraph("Aucune note trouvée", no_data_style))
        
        # Date de génération
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            spaceAfter=8,
            textColor=colors.gray
        )
        elements.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style))
        
        doc.build(elements)
        buffer.seek(0)
        
        print("✅ Buffer PDF généré avec succès")
        return buffer
        
    except Exception as e:
        print(f"❌ Erreur génération PDF notes: {str(e)}")
        traceback.print_exc()
        raise

def generate_notes_excel_buffer_semestre(notes, ecole_id, classe_nom, matiere_nom, 
                                        type_periode, periode, annee_scolaire, ecole_type):
    """Génère un buffer Excel avec gestion des semestres"""
    try:
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        
        ecole_nom = "Toutes écoles"
        if ecole_id:
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                ecole_nom = ecole.nom
        
        ws.title = f"Notes {ecole_nom}"[:31]
        
        # Informations avec période adaptée
        periode_texte = f"Semestre {periode}" if type_periode == "semestre" else f"Trimestre {periode}"
        if type_periode == "semestre" and ecole_type == "college":
            if periode == "1":
                periode_texte += " (T1+T2)"
            elif periode == "2":
                periode_texte += " (T3)"
        
        infos_parts = [
            f"Classe: {classe_nom}",
            f"Matière: {matiere_nom}",
            f"Période: {periode_texte}",
            f"Année: {annee_scolaire}",
            f"Total: {len(notes)} notes"
        ]
        
        infos_supp = " | ".join(infos_parts)
        
        # Ajouter l'en-tête
        start_row = add_excel_header(ws, ecole_id, "LISTE DES NOTES", infos_supp)
        
        # Style des bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-têtes adaptés
        headers = ['N°', 'Élève', 'Trimestre', 'Note 1', 'Note 2', 'Note 3', 'Note Comp', 'Coefficient']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Organiser les notes par élève
        notes_par_eleve = {}
        for note in notes:
            eleve_id = note.eleve_id
            if eleve_id not in notes_par_eleve:
                notes_par_eleve[eleve_id] = {
                    'eleve': note.eleve,
                    'notes': []
                }
            notes_par_eleve[eleve_id]['notes'].append(note)
        
        # Données des notes
        index = 1
        for eleve_id, data in notes_par_eleve.items():
            eleve = data['eleve']
            
            for note in sorted(data['notes'], key=lambda x: x.trimestre):
                row_num = start_row + index
                
                # Formatage des notes
                note1 = note.note1 if note.note1 is not None else "-"
                note2 = note.note2 if note.note2 is not None else "-"
                note3 = note.note3 if note.note3 is not None else "-"
                note_comp = note.note_comp if note.note_comp is not None else "-"
                
                row_data = [
                    index,
                    f"{eleve.nom} {eleve.prenoms}" if eleve else '-',
                    f"T{note.trimestre}",
                    note1,
                    note2,
                    note3,
                    note_comp,
                    note.coefficient if note.coefficient else 1
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.value = value
                    cell.font = Font(size=9)
                    cell.border = thin_border
                
                index += 1
            
            # Ajouter une ligne vide entre les élèves (sauf après le dernier)
            if index <= len(notes_par_eleve) * 2:
                row_num = start_row + index
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = thin_border
                index += 1
        
        # Largeurs de colonnes
        column_widths = {
            'A': 6,
            'B': 28,
            'C': 10,
            'D': 8,
            'E': 8,
            'F': 8,
            'G': 10,
            'H': 10
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Alignement
        for row in range(start_row, ws.max_row + 1):
            for col in ['A', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Alternance des couleurs
        for row in range(start_row + 1, ws.max_row + 1):
            if row % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
                    )
        
        # Hauteur des lignes
        for row in range(start_row, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
        
        wb.save(buffer)
        buffer.seek(0)
        
        print("✅ Buffer Excel notes semestre généré avec succès")
        return buffer
        
    except Exception as e:
        print(f"❌ Erreur génération Excel notes semestre: {str(e)}")
        traceback.print_exc()
        raise

def generate_moyennes_excel_buffer_semestre(donnees_moyennes, ecole_id, classe_nom, matiere_id,
                                           type_periode, periode, annee_scolaire, mention, ecole_type):
    """Génère un buffer Excel avec les moyennes pour semestres"""
    try:
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        
        ecole_nom = "Toutes écoles"
        if ecole_id:
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                ecole_nom = ecole.nom
        
        ws.title = f"Moyennes {ecole_nom}"[:31]
        
        # Informations avec période adaptée
        periode_texte = f"Semestre {periode}" if type_periode == "semestre" else f"Trimestre {periode}"
        if type_periode == "semestre" and ecole_type == "college":
            if periode == "1":
                periode_texte += " (T1+T2)"
            elif periode == "2":
                periode_texte += " (T3)"
        
        infos_parts = [
            f"Classe: {classe_nom}",
            f"Période: {periode_texte}",
            f"Année: {annee_scolaire}",
            f"Effectif: {len(donnees_moyennes)} élèves"
        ]
        
        if matiere_id:
            matiere = Matiere.query.get(matiere_id)
            if matiere:
                infos_parts.insert(1, f"Matière: {matiere.libelle}")
        
        if mention:
            infos_parts.append(f"Mention: {mention}")
        
        infos_supp = " | ".join(infos_parts)
        
        start_row = add_excel_header(ws, ecole_id, "CLASSEMENT DES MOYENNES", infos_supp)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = ['Rang', 'Nom', 'Prénoms', 'Moyenne', 'Mention', 'Appréciation']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        for index, donnee in enumerate(donnees_moyennes, 1):
            eleve = donnee['eleve']
            row_num = start_row + index
            
            is_ex_aequo = donnee.get('ex_aequo', False)
            rang_texte = get_rang_avec_suffixe(
                donnee['classement'], 
                eleve.sexe,
                is_ex_aequo
            )
            
            row_data = [
                rang_texte,
                eleve.nom or '-',
                eleve.prenoms or '-',
                round(donnee['moyenne_generale'], 2),
                donnee['mention'],
                donnee['appreciation']
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = value
                
                if is_ex_aequo:
                    cell.font = Font(size=10, italic=True, color="666666")
                else:
                    cell.font = Font(size=10)
                    
                cell.border = thin_border
                
                if col == 4:
                    cell.fill = PatternFill(
                        start_color="F0F8FF",
                        end_color="F0F8FF", 
                        fill_type="solid"
                    )
        
        column_widths = {
            'A': 14,
            'B': 25,
            'C': 25,
            'D': 12,
            'E': 15,
            'F': 35
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        for row in range(start_row, ws.max_row + 1):
            for col in ['A', 'D', 'E']:
                ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            for col in ['B', 'C', 'F']:
                ws[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        for row in range(start_row + 1, ws.max_row + 1):
            if row % 2 == 0:
                for col in range(1, len(headers) + 1):
                    if col != 4:
                        current_cell = ws.cell(row=row, column=col)
                        if not current_cell.font.italic:
                            current_cell.fill = PatternFill(
                                start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
                            )
        
        for row in range(start_row, ws.max_row + 1):
            ws.row_dimensions[row].height = 22
        
        # Légende
        nb_ex_aequo = sum(1 for d in donnees_moyennes if d.get('ex_aequo', False))
        if nb_ex_aequo > 0:
            legende_row = ws.max_row + 2
            ws.merge_cells(f'A{legende_row}:F{legende_row}')
            ws[f'A{legende_row}'] = "* Les suffixes '-ex' indiquent des élèves ex-aequo (même moyenne), classés par ordre alphabétique du nom."
            ws[f'A{legende_row}'].font = Font(size=9, italic=True, color="666666")
            ws[f'A{legende_row}'].alignment = Alignment(horizontal='left')
            
            if type_periode == "semestre":
                semestre_row = legende_row + 1
                ws.merge_cells(f'A{semestre_row}:F{semestre_row}')
                if ecole_type == "college":
                    if periode == "1":
                        ws[f'A{semestre_row}'] = "** Le semestre 1 regroupe les notes des trimestres 1 et 2."
                    elif periode == "2":
                        ws[f'A{semestre_row}'] = "** Le semestre 2 correspond au trimestre 3."
                else:
                    ws[f'A{semestre_row}'] = "** En lycée, les semestres correspondent aux trimestres 1 et 2."
                
                ws[f'A{semestre_row}'].font = Font(size=9, italic=True, color="2C3E50")
                ws[f'A{semestre_row}'].alignment = Alignment(horizontal='left')
        
        wb.save(buffer)
        buffer.seek(0)
        
        print("✅ Buffer Excel moyennes semestre généré avec succès")
        return buffer
        
    except Exception as e:
        print(f"❌ Erreur génération Excel moyennes semestre: {str(e)}")
        traceback.print_exc()
        raise


@services_bp.route("/export/notes")
@login_required
def export_notes_route():
    """Route pour exporter les notes en PDF ou Excel"""
    try:
        type_export = request.args.get('type', 'pdf')
        classe_id = request.args.get('classe_id')
        matiere_id = request.args.get('matiere_id')
        ecole_type = request.args.get('ecole_type', 'college')
        periode = request.args.get('periode')
        annee_scolaire = request.args.get('annee_scolaire')
        ecole_id = get_ecole_id_courante()
        
        print(f"📥 REQUÊTE EXPORT NOTES - Type école: {ecole_type}, Période: {periode}")
        print(f"📥 Classe: {classe_id}, Matière: {matiere_id}")
        
        # ✅ CORRECTION : Validation modifiée - la classe n'est plus obligatoire
        if not periode:
            return jsonify({"error": "Veuillez sélectionner une période"}), 400
        if not annee_scolaire:
            return jsonify({"error": "Veuillez sélectionner une année scolaire"}), 400
        if not ecole_id:
            return jsonify({"error": "ID de l'école manquant"}), 400
        
        # Nettoyer les paramètres
        if classe_id in ['', 'null', 'undefined', 'toutes']:
            classe_id = None  # Permet l'export de toutes les classes
        
        if matiere_id in ['', 'null', 'undefined', 'toutes']:
            matiere_id = None
        
        # Appel de la fonction d'export
        result, filename_or_error = export_notes(
            type_export=type_export,
            classe_id=classe_id,
            matiere_id=matiere_id,
            ecole_type=ecole_type,
            periode=periode,
            annee_scolaire=annee_scolaire,
            ecole_id=ecole_id
        )
        
        if result is None:
            return jsonify({"error": filename_or_error}), 400
        
        # Retour du fichier
        if type_export == 'pdf':
            return send_file(
                result,
                as_attachment=True,
                download_name=filename_or_error,
                mimetype='application/pdf'
            )
        elif type_export == 'excel':
            return send_file(
                result,
                as_attachment=True,
                download_name=filename_or_error,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            return jsonify({"error": "Type d'export non supporté"}), 400
            
    except Exception as e:
        print(f"❌ ERREUR route export notes: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": f"Erreur serveur: {str(e)}"}), 500

@services_bp.route("/export/moyennes")
@login_required
def export_moyennes_route():
    """Route pour exporter les moyennes en PDF ou Excel"""
    try:
        export_type = request.args.get('type', 'pdf')
        classe_id = request.args.get('classe_id')
        matiere_id = request.args.get('matiere_id')
        ecole_type = request.args.get('ecole_type', 'college')
        periode = request.args.get('periode')
        annee_scolaire = request.args.get('annee_scolaire')
        mention = request.args.get('mention')
        ecole_id = get_ecole_id_courante()
        
        print(f"📥 REQUÊTE EXPORT MOYENNES - Type école: {ecole_type}, Période: {periode}")
        print(f"📥 Classe: {classe_id}, Matière: {matiere_id}")
        
        # Validation
        if not periode:
            return jsonify({"error": "Veuillez sélectionner une période"}), 400
        if not annee_scolaire:
            return jsonify({"error": "Veuillez sélectionner une année scolaire"}), 400
        if not ecole_id:
            return jsonify({"error": "ID de l'école manquant"}), 400
        
        # Nettoyer les paramètres
        if classe_id in ['', 'null', 'undefined', 'toutes']:
            classe_id = 'toutes'  # Permet l'export de toutes les classes
        
        if matiere_id in ['', 'null', 'undefined', 'toutes']:
            matiere_id = None
        
        if mention in ['', 'null', 'undefined', 'toutes']:
            mention = None
        
        # Appel de la fonction d'export
        result, filename_or_error = export_moyennes_data(
            type_export=export_type,
            classe_id=classe_id,
            matiere_id=matiere_id,
            ecole_type=ecole_type,
            periode=periode,
            annee_scolaire=annee_scolaire,
            mention=mention,
            ecole_id=ecole_id
        )
        
        if result is None:
            return jsonify({"error": filename_or_error}), 400
        
        # Retour du fichier
        if export_type == 'pdf':
            return send_file(
                result,
                as_attachment=True,
                download_name=filename_or_error,
                mimetype='application/pdf'
            )
        elif export_type == 'excel':
            return send_file(
                result,
                as_attachment=True,
                download_name=filename_or_error,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            return jsonify({"error": "Type d'export non supporté"}), 400
            
    except Exception as e:
        print(f"❌ ERREUR route export moyennes: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": f"Erreur serveur: {str(e)}"}), 500
    
def generate_notes_pdf(notes, ecole_id, classe_nom, matiere_nom, trimestre, annee_scolaire):
    """Génère un PDF avec la liste des notes"""
    try:
        print(f"📊 DEBUT generate_notes_pdf - École: {ecole_id}")
        
        # Création du buffer pour le PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4,
            topMargin=15*mm,
            bottomMargin=15*mm,
            leftMargin=8*mm,
            rightMargin=8*mm
        )
        elements = []
        
        # Titre principal avec filtres
        ecole_nom = "TOUTES LES ÉCOLES"
        if ecole_id:
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                ecole_nom = ecole.nom
        
        titre_principal = f"LISTE DES NOTES"
        print(f"🎯 Titre PDF notes: {titre_principal}")
        
        # Ajouter l'en-tête
        elements = add_pdf_header(elements, doc, ecole_id, titre_principal)
        
        # Styles
        styles = getSampleStyleSheet()
        
        # Informations des filtres
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,
            spaceAfter=8,
            textColor=colors.HexColor('#34495E')
        )
        
        # Construction du texte d'information
        info_parts = []
        info_parts.append(f"Classe: {classe_nom}")
        info_parts.append(f"Matière: {matiere_nom}")
        if trimestre:
            info_parts.append(f"Trimestre: {trimestre}")
        if annee_scolaire:
            info_parts.append(f"Année: {annee_scolaire}")
        info_parts.append(f"Total: {len(notes)} notes")
        
        info_text = " | ".join(info_parts)
        elements.append(Paragraph(f"<b>{info_text}</b>", info_style))
        
        # Statistiques
        stats_style = ParagraphStyle(
            'StatsStyle',
            parent=styles['Normal'],
            fontSize=9,
            alignment=1,
            spaceAfter=12,
            textColor=colors.HexColor('#2C3E50')
        )
        
        if notes:
            # Calcul des statistiques pour toutes les notes
            all_notes = []
            for note in notes:
                # Ajouter toutes les notes disponibles
                if note.note1 is not None:
                    all_notes.append(note.note1)
                if note.note2 is not None:
                    all_notes.append(note.note2)
                if note.note3 is not None:
                    all_notes.append(note.note3)
                if note.note_comp is not None:
                    all_notes.append(note.note_comp)
            
            if all_notes:
                moyenne = sum(all_notes) / len(all_notes)
                max_note = max(all_notes)
                min_note = min(all_notes)
                stats_text = f"Moyenne: {moyenne:.2f}/20 | Max: {max_note}/20 | Min: {min_note}/20"
                elements.append(Paragraph(f"<i>{stats_text}</i>", stats_style))
        
        # Date de génération
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            spaceAfter=8,
            textColor=colors.gray
        )
        
        elements.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style))
        elements.append(Spacer(1, 8))
        
        # Tableau des notes
        if notes:
            # NOUVEAUX en-têtes avec toutes les notes
            headers = ['N°', 'Élève', 'Matière', 'Note 1', 'Note 2', 'Note 3', 'Note Comp', 'Coeff', 'Trim', 'Année']
            table_data = [headers]
            
            for index, note in enumerate(notes, 1):
                eleve = note.eleve
                matiere = note.matiere
                
                # Formatage des notes (afficher "N/A" si None)
                note1 = f"{note.note1}/20" if note.note1 is not None else "N/A"
                note2 = f"{note.note2}/20" if note.note2 is not None else "N/A"
                note3 = f"{note.note3}/20" if note.note3 is not None else "N/A"
                note_comp = f"{note.note_comp}/20" if note.note_comp is not None else "N/A"
                
                table_data.append([
                    str(index),
                    f"{eleve.nom} {eleve.prenoms}" if eleve else '',
                    matiere.libelle if matiere else '',
                    note1,
                    note2,
                    note3,
                    note_comp,
                    str(note.coefficient) if note.coefficient else "1",
                    note.trimestre or '',
                    note.annee_scolaire or ''
                ])
            
            # NOUVELLES largeurs de colonnes optimisées
            total_width = doc.width
            column_widths = [
                total_width * 0.04,  # N°
                total_width * 0.20,  # Élève
                total_width * 0.15,  # Matière
                total_width * 0.08,  # Note 1
                total_width * 0.08,  # Note 2
                total_width * 0.08,  # Note 3
                total_width * 0.10,  # Note Comp
                total_width * 0.06,  # Coeff
                total_width * 0.06,  # Trim
                total_width * 0.08   # Année
            ]
            
            table = Table(table_data, colWidths=column_widths, repeatRows=1)
            
            # Style du tableau
            table.setStyle(TableStyle([
                # En-têtes
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 4),
                
                # Données
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 6),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # N°
                ('ALIGN', (1, 1), (2, -1), 'LEFT'),    # Élève, Matière
                ('ALIGN', (3, 1), (-1, -1), 'CENTER'), # Toutes les notes et infos
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('WORDWRAP', (1, 1), (2, -1), True),   # Élève, Matière
                
                # Bordures
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                
                # Alternance des couleurs
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
                    colors.HexColor('#FFFFFF'), 
                    colors.HexColor('#F8F9FA')
                ]),
                
                # Ligne de séparation de l'en-tête
                ('LINEABOVE', (0, 1), (-1, 1), 1, colors.HexColor('#2C3E50')),
                
                # Padding
                ('LEFTPADDING', (0, 0), (-1, -1), 1),
                ('RIGHTPADDING', (0, 0), (-1, -1), 1),
                ('TOPPADDING', (0, 0), (-1, -1), 1),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
            ]))
            
            elements.append(table)
        else:
            # Aucune note trouvée
            no_data_style = ParagraphStyle(
                'NoDataStyle',
                parent=styles['Normal'],
                fontSize=12,
                alignment=1,
                textColor=colors.gray,
                spaceAfter=20
            )
            elements.append(Paragraph("Aucune note trouvée", no_data_style))
        
        # Génération du PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Nom du fichier
        ecole_suffix = ecole_nom.replace(' ', '_') if ecole_id else "toutes_ecoles"
        filename = f"notes_{ecole_suffix}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        print(f"❌ Erreur génération PDF notes: {str(e)}")
        traceback.print_exc()
        return f"Erreur lors de la génération du PDF: {str(e)}", 500
    
def generate_notes_excel(notes, ecole_id, classe_nom, matiere_nom, trimestre, annee_scolaire):
    """Génère un Excel avec la liste des notes"""
    try:
        # Création du workbook
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        
        # Nom de l'onglet
        ecole_nom = "Toutes écoles"
        if ecole_id:
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                ecole_nom = ecole.nom
        
        ws.title = f"Notes {ecole_nom}"[:31]
        
        # Informations supplémentaires pour l'en-tête
        infos_parts = []
        infos_parts.append(f"Classe: {classe_nom}")
        infos_parts.append(f"Matière: {matiere_nom}")
        if trimestre:
            infos_parts.append(f"Trimestre: {trimestre}")
        if annee_scolaire:
            infos_parts.append(f"Année: {annee_scolaire}")
        infos_parts.append(f"Total: {len(notes)} notes")
        
        infos_supp = " | ".join(infos_parts)
        
        # Ajouter l'en-tête
        start_row = add_excel_header(ws, ecole_id, "LISTE DES NOTES", infos_supp)
        
        # Style des bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # NOUVEAUX en-têtes avec toutes les notes
        headers = ['N°', 'Élève', 'Matière', 'Note 1', 'Note 2', 'Note 3', 'Note Comp', 'Coefficient', 'Trimestre', 'Année scolaire']
        
        # En-têtes du tableau
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Données des notes
        for index, note in enumerate(notes, 1):
            eleve = note.eleve
            matiere = note.matiere
            
            row_num = start_row + index
            
            # Formatage des notes
            note1 = note.note1 if note.note1 is not None else "N/A"
            note2 = note.note2 if note.note2 is not None else "N/A"
            note3 = note.note3 if note.note3 is not None else "N/A"
            note_comp = note.note_comp if note.note_comp is not None else "N/A"
            
            row_data = [
                index,
                f"{eleve.nom} {eleve.prenoms}" if eleve else '',
                matiere.libelle if matiere else '',
                note1,
                note2,
                note3,
                note_comp,
                note.coefficient if note.coefficient else 1,
                note.trimestre or '',
                note.annee_scolaire or ''
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = value
                cell.font = Font(size=9)
                cell.border = thin_border
        
        # NOUVELLES largeurs de colonnes
        column_widths = {
            'A': 6,    # N°
            'B': 25,   # Élève
            'C': 20,   # Matière
            'D': 8,    # Note 1
            'E': 8,    # Note 2
            'F': 8,    # Note 3
            'G': 10,   # Note Comp
            'H': 10,   # Coefficient
            'I': 10,   # Trimestre
            'J': 12    # Année scolaire
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Alignement et retour à la ligne
        for row in range(start_row, ws.max_row + 1):
            # N° et toutes les notes
            for col in ['A', 'D', 'E', 'F', 'G', 'H', 'I']:
                ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            # Élève, Matière, Année scolaire
            for col in ['B', 'C', 'J']:
                ws[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Alternance des couleurs des lignes
        for row in range(start_row + 1, ws.max_row + 1):
            if row % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
                    )
        
        # Ajuster la hauteur des lignes
        for row in range(start_row, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
        
        # Sauvegarde
        wb.save(buffer)
        buffer.seek(0)
        
        # Nom du fichier
        ecole_suffix = ecole_nom.replace(' ', '_') if ecole_id else "toutes_ecoles"
        filename = f"notes_{ecole_suffix}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"❌ Erreur génération Excel notes: {str(e)}")
        traceback.print_exc()
        return f"Erreur lors de la génération du Excel: {str(e)}", 500
    
def generate_notes_pdf_buffer(notes, ecole_id, classe_nom, matiere_nom, trimestre, annee_scolaire):
    """Génère un buffer PDF avec la liste des notes - VERSION BUFFER"""
    try:
        print(f"📊 DEBUT generate_notes_pdf_buffer - École: {ecole_id}")
        
        # Création du buffer pour le PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4,
            topMargin=15*mm,
            bottomMargin=15*mm,
            leftMargin=8*mm,
            rightMargin=8*mm
        )
        elements = []
        
        # Titre principal avec filtres
        ecole_nom = "TOUTES LES ÉCOLES"
        if ecole_id:
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                ecole_nom = ecole.nom
        
        titre_principal = f"LISTE DES NOTES"
        print(f"🎯 Titre PDF notes: {titre_principal}")
        
        # Ajouter l'en-tête
        elements = add_pdf_header(elements, doc, ecole_id, titre_principal)
        
        # Styles
        styles = getSampleStyleSheet()
        
        # Informations des filtres
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,
            spaceAfter=8,
            textColor=colors.HexColor('#34495E')
        )
        
        # Construction du texte d'information
        info_parts = []
        info_parts.append(f"Classe: {classe_nom}")
        info_parts.append(f"Matière: {matiere_nom}")
        if trimestre:
            info_parts.append(f"Trimestre: {trimestre}")
        if annee_scolaire:
            info_parts.append(f"Année: {annee_scolaire}")
        info_parts.append(f"Total: {len(notes)} notes")
        
        info_text = " | ".join(info_parts)
        elements.append(Paragraph(f"<b>{info_text}</b>", info_style))
        
        # Statistiques
        stats_style = ParagraphStyle(
            'StatsStyle',
            parent=styles['Normal'],
            fontSize=9,
            alignment=1,
            spaceAfter=12,
            textColor=colors.HexColor('#2C3E50')
        )
        
        if notes:
            # Calcul des statistiques pour toutes les notes
            all_notes = []
            for note in notes:
                # Ajouter toutes les notes disponibles
                if note.note1 is not None:
                    all_notes.append(note.note1)
                if note.note2 is not None:
                    all_notes.append(note.note2)
                if note.note3 is not None:
                    all_notes.append(note.note3)
                if note.note_comp is not None:
                    all_notes.append(note.note_comp)
            
            if all_notes:
                moyenne = sum(all_notes) / len(all_notes)
                max_note = max(all_notes)
                min_note = min(all_notes)
                stats_text = f"Moyenne: {moyenne:.2f}/20 | Max: {max_note}/20 | Min: {min_note}/20"
                elements.append(Paragraph(f"<i>{stats_text}</i>", stats_style))
        
        # Date de génération
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            spaceAfter=8,
            textColor=colors.gray
        )
        
        elements.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style))
        elements.append(Spacer(1, 8))
        
        # Tableau des notes
        if notes:
            # NOUVEAUX en-têtes sans colonne 'Trim'
            headers = ['N°', 'Élève', 'Matière', 'Note 1', 'Note 2', 'Note 3', 'Note Comp', 'Coeff', 'Année']
            table_data = [headers]
            
            for index, note in enumerate(notes, 1):
                eleve = note.eleve
                matiere = note.matiere
                
                # Formatage des notes (afficher "-" si None)
                note1 = f"{note.note1}/20" if note.note1 is not None else "-"
                note2 = f"{note.note2}/20" if note.note2 is not None else "-"
                note3 = f"{note.note3}/20" if note.note3 is not None else "-"
                note_comp = f"{note.note_comp}/20" if note.note_comp is not None else "-"
                
                table_data.append([
                    str(index),
                    f"{eleve.nom} {eleve.prenoms}" if eleve else '-',
                    matiere.libelle if matiere else '-',
                    note1,
                    note2,
                    note3,
                    note_comp,
                    str(note.coefficient) if note.coefficient else "1",
                    note.annee_scolaire or '-'
                ])
            
            # NOUVELLES largeurs de colonnes optimisées (sans colonne Trim)
            total_width = doc.width
            column_widths = [
                total_width * 0.04,  # N°
                total_width * 0.22,  # Élève
                total_width * 0.16,  # Matière
                total_width * 0.08,  # Note 1
                total_width * 0.08,  # Note 2
                total_width * 0.08,  # Note 3
                total_width * 0.10,  # Note Comp
                total_width * 0.06,  # Coeff
                total_width * 0.10   # Année
            ]
            
            table = Table(table_data, colWidths=column_widths, repeatRows=1)
            
            # Style du tableau
            table.setStyle(TableStyle([
                # En-têtes
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 4),
                
                # Données
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 6),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # N°
                ('ALIGN', (1, 1), (2, -1), 'LEFT'),    # Élève, Matière
                ('ALIGN', (3, 1), (-1, -1), 'CENTER'), # Toutes les notes et infos
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('WORDWRAP', (1, 1), (2, -1), True),   # Élève, Matière
                
                # Bordures
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                
                # Alternance des couleurs
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
                    colors.HexColor('#FFFFFF'), 
                    colors.HexColor('#F8F9FA')
                ]),
                
                # Ligne de séparation de l'en-tête
                ('LINEABOVE', (0, 1), (-1, 1), 1, colors.HexColor('#2C3E50')),
                
                # Padding
                ('LEFTPADDING', (0, 0), (-1, -1), 1),
                ('RIGHTPADDING', (0, 0), (-1, -1), 1),
                ('TOPPADDING', (0, 0), (-1, -1), 1),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
            ]))
            
            elements.append(table)
        else:
            # Aucune note trouvée
            no_data_style = ParagraphStyle(
                'NoDataStyle',
                parent=styles['Normal'],
                fontSize=12,
                alignment=1,
                textColor=colors.gray,
                spaceAfter=20
            )
            elements.append(Paragraph("Aucune note trouvée", no_data_style))
        
        # Génération du PDF
        doc.build(elements)
        buffer.seek(0)
        
        print("✅ Buffer PDF généré avec succès")
        return buffer
        
    except Exception as e:
        print(f"❌ Erreur génération PDF notes: {str(e)}")
        traceback.print_exc()
        raise

def generate_notes_excel_buffer(notes, ecole_id, classe_nom, matiere_nom, trimestre, annee_scolaire):
    """Génère un buffer Excel avec la liste des notes - VERSION BUFFER"""
    try:
        # Création du workbook
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        
        # Nom de l'onglet
        ecole_nom = "Toutes écoles"
        if ecole_id:
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                ecole_nom = ecole.nom
        
        ws.title = f"Notes {ecole_nom}"[:31]
        
        # Informations supplémentaires pour l'en-tête
        infos_parts = []
        infos_parts.append(f"Classe: {classe_nom}")
        infos_parts.append(f"Matière: {matiere_nom}")
        if trimestre:
            infos_parts.append(f"Trimestre: {trimestre}")
        if annee_scolaire:
            infos_parts.append(f"Année: {annee_scolaire}")
        infos_parts.append(f"Total: {len(notes)} notes")
        
        infos_supp = " | ".join(infos_parts)
        
        # Ajouter l'en-tête
        start_row = add_excel_header(ws, ecole_id, "LISTE DES NOTES", infos_supp)
        
        # Style des bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # NOUVEAUX en-têtes sans colonne 'Trimestre'
        headers = ['N°', 'Élève', 'Matière', 'Note 1', 'Note 2', 'Note 3', 'Note Comp', 'Coefficient', 'Année scolaire']
        
        # En-têtes du tableau
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Données des notes
        for index, note in enumerate(notes, 1):
            eleve = note.eleve
            matiere = note.matiere
            
            row_num = start_row + index
            
            # Formatage des notes (remplacer "N/A" par "-")
            note1 = note.note1 if note.note1 is not None else "-"
            note2 = note.note2 if note.note2 is not None else "-"
            note3 = note.note3 if note.note3 is not None else "-"
            note_comp = note.note_comp if note.note_comp is not None else "-"
            
            row_data = [
                index,
                f"{eleve.nom} {eleve.prenoms}" if eleve else '-',
                matiere.libelle if matiere else '-',
                note1,
                note2,
                note3,
                note_comp,
                note.coefficient if note.coefficient else 1,
                note.annee_scolaire or '-'
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = value
                cell.font = Font(size=9)
                cell.border = thin_border
        
        # NOUVELLES largeurs de colonnes (sans colonne Trimestre)
        column_widths = {
            'A': 6,    # N°
            'B': 28,   # Élève (plus large)
            'C': 22,   # Matière (plus large)
            'D': 8,    # Note 1
            'E': 8,    # Note 2
            'F': 8,    # Note 3
            'G': 10,   # Note Comp
            'H': 10,   # Coefficient
            'I': 12    # Année scolaire
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Alignement et retour à la ligne
        for row in range(start_row, ws.max_row + 1):
            # N° et toutes les notes
            for col in ['A', 'D', 'E', 'F', 'G', 'H']:
                ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            # Élève, Matière, Année scolaire
            for col in ['B', 'C', 'I']:
                ws[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Alternance des couleurs des lignes
        for row in range(start_row + 1, ws.max_row + 1):
            if row % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
                    )
        
        # Ajuster la hauteur des lignes
        for row in range(start_row, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
        
        # Sauvegarde
        wb.save(buffer)
        buffer.seek(0)
        
        print("✅ Buffer Excel généré avec succès")
        return buffer
        
    except Exception as e:
        print(f"❌ Erreur génération Excel notes: {str(e)}")
        traceback.print_exc()
        raise


# ========== ROUTES API POUR LES FILTRES DES MOYENNES ==========

@services_bp.route("/moyennes/classes")
@login_required
def get_moyennes_classes():
    """API pour récupérer les classes (pour export moyennes)"""
    try:
        ecole_id = get_ecole_id_courante()
        
        print(f"🔍 GET /moyennes/classes - ecole_id: {ecole_id}")
        
        if ecole_id:
            classes = Classe.query.filter_by(ecole_id=ecole_id).order_by(Classe.nom).all()
            print(f"🏫 Récupération des classes de l'école {ecole_id}")
        else:
            classes = Classe.query.order_by(Classe.nom).all()
            print("🔧 ADMIN SYSTÈME - Récupération de TOUTES les classes")
        
        classes_data = [{"id": str(c.id), "nom": c.nom} for c in classes if c and c.nom]
        print(f"✅ Classes retournées: {len(classes_data)}")
        
        return jsonify(classes_data)
        
    except Exception as e:
        print(f"❌ Erreur get_moyennes_classes: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@services_bp.route("/moyennes/matieres")
@login_required
def get_moyennes_matieres():
    """API pour récupérer les matières (pour export moyennes)"""
    try:
        ecole_id = get_ecole_id_courante()
        
        print(f"🔍 GET /moyennes/matieres - ecole_id: {ecole_id}")
        
        if ecole_id:
            matieres = Matiere.query.filter_by(ecole_id=ecole_id).order_by(Matiere.libelle).all()
            print(f"🏫 Récupération des matières de l'école {ecole_id}")
        else:
            matieres = Matiere.query.order_by(Matiere.libelle).all()
            print("🔧 ADMIN SYSTÈME - Récupération de TOUTES les matières")
        
        matieres_data = [{"id": str(m.id), "libelle": m.libelle} for m in matieres if m and m.libelle]
        print(f"✅ Matières retournées: {len(matieres_data)}")
        
        return jsonify(matieres_data)
        
    except Exception as e:
        print(f"❌ Erreur get_moyennes_matieres: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@services_bp.route("/moyennes/annees-scolaires")
@login_required
def get_moyennes_annees_scolaires():
    """API pour récupérer les années scolaires disponibles pour les moyennes"""
    try:
        from ..models import Note
        from sqlalchemy import distinct
        
        annees = db.session.query(distinct(Note.annee_scolaire)).filter(
            Note.annee_scolaire.isnot(None)
        ).order_by(Note.annee_scolaire.desc()).all()
        
        annees_data = [{"annee": a[0]} for a in annees if a[0]]
        
        if not annees_data:
            annee_courante = datetime.now().year
            annees_data = [{"annee": f"{annee_courante}-{annee_courante+1}"}]
        
        print(f"✅ Années scolaires retournées: {len(annees_data)}")
        return jsonify(annees_data)
        
    except Exception as e:
        print(f"❌ Erreur get_moyennes_annees_scolaires: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@services_bp.route("/moyennes/mentions")
@login_required
def get_moyennes_mentions():
    """API pour récupérer les mentions disponibles"""
    try:
        mentions = [
            {"code": "EXCELLENT", "libelle": "Excellent"},
            {"code": "TRES_BIEN", "libelle": "Très Bien"},
            {"code": "BIEN", "libelle": "Bien"},
            {"code": "ASSEZ_BIEN", "libelle": "Assez Bien"},
            {"code": "PASSABLE", "libelle": "Passable"},
            {"code": "INSUFFISANT", "libelle": "Insuffisant"},
            {"code": "MEDIOCRE", "libelle": "Médiocre"}
        ]
        
        return jsonify(mentions)
        
    except Exception as e:
        print(f"❌ Erreur get_moyennes_mentions: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ========== EXPORT MOYENNES ==========

def calculer_moyenne_eleve(notes):
    """Calcule la moyenne générale d'un élève à partir de ses notes - AVEC DEBUG"""
    if not notes:
        print("❌ Aucune note fournie")
        return 0.0
    
    total_pondere = 0
    total_coef = 0
    
    print(f"🔍 Calcul moyenne pour {len(notes)} note(s)")
    
    for i, note in enumerate(notes):
        note_list = [note.note1, note.note2, note.note3]
        note_valeurs = [x for x in note_list if x is not None]
        
        print(f"   📝 Note {i+1}: note1={note.note1}, note2={note.note2}, note3={note.note3}, comp={note.note_comp}")
        
        if note_valeurs or note.note_comp:
            base_note = sum(note_valeurs) / len(note_valeurs) if note_valeurs else 0
            val = (base_note + (note.note_comp or 0)) / 2
            coef = note.coefficient or 1
            
            total_pondere += val * coef
            total_coef += coef
            
            print(f"   ➡️ Base: {base_note:.2f}, Final: {val:.2f}, Coef: {coef}")
            print(f"   📊 Total pondéré: {total_pondere:.2f}, Total coef: {total_coef}")
        else:
            print(f"   ⚠️ Aucune note valide")
    
    if total_coef > 0:
        moyenne_finale = round(total_pondere / total_coef, 2)
        print(f"✅ MOYENNE FINALE: {moyenne_finale:.2f}")
        return moyenne_finale
    else:
        print("❌ Aucun coefficient valide")
        return 0.0

def get_mention(moyenne):
    """Retourne la mention correspondant à la moyenne"""
    if moyenne >= 17:
        return "EXCELLENT"
    elif moyenne >= 15:
        return "TRES_BIEN"
    elif moyenne >= 13:
        return "BIEN"
    elif moyenne >= 11:
        return "ASSEZ_BIEN"
    elif moyenne >= 9:
        return "PASSABLE"
    elif moyenne >= 7:
        return "INSUFFISANT"
    else:
        return "MEDIOCRE"

def get_appreciation(moyenne):
    """Retourne l'appréciation correspondant à la moyenne"""
    if moyenne >= 17:
        return "Excellente performance"
    elif moyenne >= 15:
        return "Très bon travail"
    elif moyenne >= 13:
        return "Bon travail"
    elif moyenne >= 11:
        return "Satisfaisant"
    elif moyenne >= 9:
        return "Passable"
    elif moyenne >= 7:
        return "Insuffisant"
    else:
        return "Médiocre"

def get_rang_avec_suffixe(classement, sexe, ex_aequo=False):
    """Retourne le rang avec suffixe approprié (1er, 1ère, 6ème, 6ème-ex, etc.)"""
    if classement == 1:
        if sexe and sexe.upper() == 'F':
            return "1ère" + ("-ex" if ex_aequo else "")
        else:
            return "1er" + ("-ex" if ex_aequo else "")
    else:
        return f"{classement}ème" + ("-ex" if ex_aequo else "")

def classer_eleves_avec_ex_aequo(donnees_moyennes):
    """Classe les élèves avec gestion des ex-aequo - VERSION CORRIGÉE"""
    if not donnees_moyennes:
        return []
    
    # Trier par moyenne décroissante puis par nom alphabétique
    donnees_triees = sorted(donnees_moyennes, 
                          key=lambda x: (-x['moyenne_generale'], x['eleve'].nom or '', x['eleve'].prenoms or ''))
    
    # Appliquer le classement
    resultat = []
    classement_actuel = 1
    
    for i, donnee in enumerate(donnees_triees):
        if i == 0:
            # Premier élève
            donnee['classement'] = 1
            donnee['ex_aequo'] = False
        else:
            donnee_precedente = donnees_triees[i-1]
            
            # Vérifier si même moyenne que le précédent
            if abs(donnee['moyenne_generale'] - donnee_precedente['moyenne_generale']) < 0.01:
                # Même moyenne : même classement mais marqué ex-aequo
                donnee['classement'] = donnee_precedente['classement']
                donnee['ex_aequo'] = True
            else:
                # Nouvelle moyenne : classement normal (position dans la liste)
                donnee['classement'] = i + 1
                donnee['ex_aequo'] = False
        
        resultat.append(donnee)
    
    # DEBUG: Afficher les classements pour vérification
    print("🔍 DEBUG - Classements finaux:")
    for donnee in resultat:
        rang_texte = get_rang_avec_suffixe(donnee['classement'], donnee['eleve'].sexe, donnee.get('ex_aequo', False))
        print(f"   {rang_texte} - {donnee['eleve'].nom} {donnee['eleve'].prenoms} - {donnee['moyenne_generale']:.2f}")
    
    return resultat

@services_bp.route("/export/moyennes")
@login_required
def export_moyennes():
    """Export des moyennes des élèves"""
    try:
        export_type = request.args.get('type', 'pdf')
        classe_id = request.args.get('classe_id')
        matiere_id = request.args.get('matiere_id')
        trimestre = request.args.get('trimestre')
        annee_scolaire = request.args.get('annee_scolaire')
        mention = request.args.get('mention')
        ecole_id = get_ecole_id_courante()
        
        print(f"📥 REQUÊTE EXPORT MOYENNES - Type: {export_type}")
        print(f"📥 Classe: {classe_id}, Matière: {matiere_id}")
        print(f"📥 Trimestre: {trimestre}, Année: {annee_scolaire}")
        print(f"📥 Mention: {mention}, École: {ecole_id}")
        
        # Validation
        if not classe_id:
            return jsonify({"error": "Veuillez sélectionner une classe"}), 400
        if not trimestre:
            return jsonify({"error": "Veuillez sélectionner un trimestre"}), 400
        if not annee_scolaire:
            return jsonify({"error": "Veuillez sélectionner une année scolaire"}), 400
        if not ecole_id:
            return jsonify({"error": "ID de l'école manquant"}), 400
        
        # Nettoyer les paramètres
        if matiere_id in ['', 'null', 'undefined', 'toutes']:
            matiere_id = None
        if mention in ['', 'null', 'undefined', 'toutes']:
            mention = None
        
        # Appel de la fonction d'export
        result, filename_or_error = export_moyennes_data(
            type_export=export_type,
            classe_id=classe_id,
            matiere_id=matiere_id,
            trimestre=trimestre,
            annee_scolaire=annee_scolaire,
            mention=mention,
            ecole_id=ecole_id
        )
        
        if result is None:
            return jsonify({"error": filename_or_error}), 400
        
        # Retour du fichier
        if export_type == 'pdf':
            return send_file(
                result,
                as_attachment=True,
                download_name=filename_or_error,
                mimetype='application/pdf'
            )
        elif export_type == 'excel':
            return send_file(
                result,
                as_attachment=True,
                download_name=filename_or_error,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            return jsonify({"error": "Type d'export non supporté"}), 400
            
    except Exception as e:
        print(f"❌ ERREUR route export moyennes: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": f"Erreur serveur: {str(e)}"}), 500

# ========== EXPORT MOYENNES AVEC GESTION CORRECTE DES PÉRIODES ==========

def export_moyennes_data(type_export, classe_id, matiere_id, ecole_type, periode, 
                        annee_scolaire, mention, ecole_id):
    """Fonction principale pour exporter les moyennes avec gestion correcte"""
    try:
        print(f"🔍 EXPORT MOYENNES - Type école: {ecole_type}, Période: {periode}")
        
        # Validation
        if not ecole_id:
            return None, "ID de l'école manquant"
        if not periode:
            return None, "Veuillez sélectionner une période"
        if not annee_scolaire:
            return None, "Veuillez sélectionner une année scolaire"
        
        # ✅ NOUVELLE LOGIQUE : Gestion de "toutes les classes"
        if classe_id in ['', 'null', 'undefined', 'toutes']:
            # Export pour toutes les classes de l'école
            classes = Classe.query.filter_by(ecole_id=ecole_id).all()
            if not classes:
                return None, "Aucune classe trouvée dans cette école"
            
            classe_nom = "Toutes les classes"
            print(f"✅ Export pour TOUTES les classes ({len(classes)} classes)")
            
            # Récupérer tous les élèves de toutes les classes
            donnees_moyennes = []
            
            for classe in classes:
                eleves = Eleve.query.filter_by(classe_id=classe.id).all()
                print(f"  📊 Classe {classe.nom}: {len(eleves)} élève(s)")
                
                for eleve in eleves:
                    donnee_eleve = calculer_moyenne_eleve_periode(eleve.id, matiere_id, ecole_type, periode, annee_scolaire)
                    if donnee_eleve:
                        # Filtrer par mention si spécifié
                        if mention and mention != donnee_eleve['mention']:
                            continue
                        
                        donnee_eleve['eleve'] = eleve
                        donnee_eleve['classe_nom'] = classe.nom  # Ajouter le nom de la classe
                        donnees_moyennes.append(donnee_eleve)
            
            if not donnees_moyennes:
                return None, f"Aucune moyenne trouvée pour les critères sélectionnés dans toutes les classes"
            
        else:
            # Export pour une classe spécifique (logique existante)
            classe = Classe.query.filter_by(id=classe_id, ecole_id=ecole_id).first()
            if not classe:
                return None, "Classe non trouvée dans cette école"
            
            classe_nom = classe.nom
            print(f"✅ Classe spécifique: {classe_nom}")
            
            # Récupérer les élèves de la classe
            eleves = Eleve.query.filter_by(classe_id=classe_id).order_by(Eleve.nom, Eleve.prenoms).all()
            if not eleves:
                return None, "Aucun élève trouvé dans cette classe"
            
            print(f"✅ {len(eleves)} élève(s) trouvé(s) dans la classe")
            
            # Calcul des moyennes pour chaque élève
            donnees_moyennes = []
            
            for eleve in eleves:
                donnee_eleve = calculer_moyenne_eleve_periode(eleve.id, matiere_id, ecole_type, periode, annee_scolaire)
                if donnee_eleve:
                    # Filtrer par mention si spécifié
                    if mention and mention != donnee_eleve['mention']:
                        continue
                    
                    donnee_eleve['eleve'] = eleve
                    donnee_eleve['classe_nom'] = classe_nom
                    donnees_moyennes.append(donnee_eleve)
        
        if not donnees_moyennes:
            periode_texte = f"Trimestre {periode}" if ecole_type == "college" else f"Semestre {periode}"
            error_msg = f"Aucune moyenne trouvée pour "
            if matiere_id and matiere_id not in ['', 'null', 'undefined', 'toutes']:
                matiere = Matiere.query.get(matiere_id)
                error_msg += f"la matière '{matiere.libelle}' "
            else:
                error_msg += "les matières sélectionnées "
            
            error_msg += f"dans {classe_nom} ({periode_texte}, {annee_scolaire})"
            
            print(f"❌ {error_msg}")
            return None, error_msg
        
        # Classement des élèves
        donnees_moyennes = classer_eleves_avec_ex_aequo(donnees_moyennes)
        
        print(f"✅ {len(donnees_moyennes)} moyenne(s) calculée(s) et classée(s)")
        
        # Génération de l'export
        if type_export == 'pdf':
            try:
                pdf_buffer = generate_moyennes_pdf_buffer_periode(
                    donnees_moyennes, ecole_id, classe_nom, matiere_id,
                    ecole_type, periode, annee_scolaire, mention
                )
                nom_fichier = generer_nom_fichier_moyennes_periode(
                    classe_nom, matiere_id, ecole_type, periode, annee_scolaire, mention, 'pdf'
                )
                print(f"✅ PDF généré: {nom_fichier}")
                return pdf_buffer, nom_fichier
            except Exception as e:
                print(f"❌ Erreur génération PDF: {str(e)}")
                traceback.print_exc()
                return None, f"Erreur lors de la génération du PDF: {str(e)}"
        
        elif type_export == 'excel':
            try:
                excel_buffer = generate_moyennes_excel_buffer_periode(
                    donnees_moyennes, ecole_id, classe_nom, matiere_id,
                    ecole_type, periode, annee_scolaire, mention
                )
                nom_fichier = generer_nom_fichier_moyennes_periode(
                    classe_nom, matiere_id, ecole_type, periode, annee_scolaire, mention, 'xlsx'
                )
                print(f"✅ Excel généré: {nom_fichier}")
                return excel_buffer, nom_fichier
            except Exception as e:
                print(f"❌ Erreur génération Excel: {str(e)}")
                traceback.print_exc()
                return None, f"Erreur lors de la génération du Excel: {str(e)}"
        
        else:
            return None, f"Type d'export non supporté: {type_export}"
            
    except Exception as e:
        print(f"❌ ERREUR export_moyennes_data: {str(e)}")
        traceback.print_exc()
        return None, f"Erreur lors de l'export: {str(e)}"

def calculer_moyenne_eleve_periode(eleve_id, matiere_id, ecole_type, periode, annee_scolaire):
    """Calcule la moyenne d'un élève pour une période spécifique"""
    try:
        # Construire la requête des notes
        query = db.session.query(Note)\
            .join(Eleve, Note.eleve_id == Eleve.id)\
            .join(Matiere, Note.matiere_id == Matiere.id)\
            .filter(Eleve.id == eleve_id)\
            .filter(Note.trimestre == periode)\
            .filter(Note.annee_scolaire == annee_scolaire)
        
        if matiere_id and matiere_id not in ['', 'null', 'undefined', 'toutes']:
            query = query.filter(Matiere.id == matiere_id)
        
        notes_eleve = query.all()
        
        if not notes_eleve:
            return None
        
        # Calculer la moyenne
        moyenne = calculer_moyenne_eleve(notes_eleve)
        mention_eleve = get_mention(moyenne)
        appreciation = get_appreciation(moyenne)
        
        return {
            'moyenne_generale': moyenne,
            'mention': mention_eleve,
            'appreciation': appreciation,
            'notes_count': len(notes_eleve)
        }
        
    except Exception as e:
        print(f"❌ Erreur calcul moyenne élève {eleve_id}: {str(e)}")
        return None

def generate_moyennes_pdf_buffer_periode(donnees_moyennes, ecole_id, classe_nom, matiere_id, 
                                        ecole_type, periode, annee_scolaire, mention):
    """Génère un buffer PDF avec les moyennes pour la période correcte"""
    try:
        print(f"📊 DEBUT generate_moyennes_pdf_buffer_periode - Type école: {ecole_type}")
        print(f"   Classe: {classe_nom}, Toutes classes: {'toutes' in classe_nom.lower()}")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4,
            topMargin=15*mm,
            bottomMargin=15*mm,
            leftMargin=10*mm,
            rightMargin=10*mm
        )
        elements = []
        
        # Titre principal adapté
        ecole_nom = "TOUTES LES ÉCOLES"
        if ecole_id:
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                ecole_nom = ecole.nom
        
        # Déterminer le titre de période
        titre_periode = f"TRIMESTRE {periode}" if ecole_type == "college" else f"SEMESTRE {periode}"
        
        # CORRECTION: Utiliser titre_periode dans le titre principal
        titre_principal = f"CLASSEMENT DES MOYENNES - {titre_periode}"
        
        elements = add_pdf_header(elements, doc, ecole_id, titre_principal)
        
        styles = getSampleStyleSheet()
        
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=11,
            alignment=1,
            spaceAfter=10,
            textColor=colors.HexColor('#34495E')
        )
        
        # Informations avec gestion "toutes classes"
        type_etablissement = "Collège" if ecole_type == "college" else "Lycée"
        
        info_parts = [
            f"Type: {type_etablissement}",
            f"Classe: {classe_nom}",
            f"Période: {titre_periode}",
            f"Année: {annee_scolaire}",
            f"Effectif: {len(donnees_moyennes)} élèves"
        ]
        
        # Ajouter la matière si spécifiée
        if matiere_id and matiere_id not in ['', 'null', 'undefined', 'toutes']:
            matiere = Matiere.query.get(matiere_id)
            if matiere:
                info_parts.insert(2, f"Matière: {matiere.libelle}")
        
        if mention:
            info_parts.append(f"Mention: {mention}")
        
        info_text = " | ".join(info_parts)
        elements.append(Paragraph(f"<b>{info_text}</b>", info_style))
        
        # Statistiques
        if donnees_moyennes:
            moyennes = [d['moyenne_generale'] for d in donnees_moyennes]
            moyenne_classe = sum(moyennes) / len(moyennes) if moyennes else 0
            max_moyenne = max(moyennes) if moyennes else 0
            min_moyenne = min(moyennes) if moyennes else 0
            
            stats_style = ParagraphStyle(
                'StatsStyle',
                parent=styles['Normal'],
                fontSize=10,
                alignment=1,
                spaceAfter=12,
                textColor=colors.HexColor('#2C3E50')
            )
            
            stats_text = f"Moyenne générale: {moyenne_classe:.2f}/20 | Max: {max_moyenne:.2f}/20 | Min: {min_moyenne:.2f}/20"
            elements.append(Paragraph(f"<i>{stats_text}</i>", stats_style))
        
        # Date de génération
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=9,
            alignment=1,
            spaceAfter=8,
            textColor=colors.gray
        )
        
        elements.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style))
        elements.append(Spacer(1, 10))
        
        # Tableau des moyennes avec colonne "Classe" si nécessaire
        if donnees_moyennes:
            # ✅ DÉTERMINER LES EN-TÊTES DYNAMIQUEMENT
            headers = ['Rang', 'Nom et Prénoms']
            
            # Ajouter colonne "Classe" si on exporte toutes les classes
            if 'toutes' in classe_nom.lower():
                headers.append('Classe')
            
            headers.extend(['Moyenne', 'Mention', 'Appréciation'])
            
            table_data = [headers]
            
            style_commands = []
            
            for row_index, donnee in enumerate(donnees_moyennes, 1):
                eleve = donnee['eleve']
                is_ex_aequo = donnee.get('ex_aequo', False)
                
                rang_texte = get_rang_avec_suffixe(
                    donnee['classement'], 
                    eleve.sexe,
                    is_ex_aequo
                )
                
                # Construire la ligne dynamiquement
                row_data = [rang_texte, f"{eleve.nom} {eleve.prenoms}" if eleve else '-']
                
                # Ajouter colonne "Classe" si nécessaire
                if 'toutes' in classe_nom.lower():
                    row_data.append(donnee.get('classe_nom', '-'))
                
                row_data.extend([
                    f"{donnee['moyenne_generale']:.2f}/20",
                    donnee['mention'],
                    donnee['appreciation']
                ])
                
                table_data.append(row_data)
                
                if is_ex_aequo:
                    style_commands.append(('FONTNAME', (0, row_index), (0, row_index), 'Helvetica-Oblique'))
                    style_commands.append(('TEXTCOLOR', (0, row_index), (0, row_index), colors.HexColor('#666666')))
            
            # ✅ CALCULER LES LARGEURS DE COLONNES DYNAMIQUES
            total_width = doc.width
            column_widths = []
            
            # Largeur fixe pour Rang
            column_widths.append(total_width * 0.10)
            
            # Largeur pour Nom et Prénoms
            if 'toutes' in classe_nom.lower():
                column_widths.append(total_width * 0.25)  # Moins large
            else:
                column_widths.append(total_width * 0.35)  # Plus large
            
            # Largeur pour Classe (si présente)
            if 'toutes' in classe_nom.lower():
                column_widths.append(total_width * 0.15)
            
            # Largeurs restantes pour les autres colonnes
            columns_left = len(headers) - len(column_widths)
            remaining_width = total_width - sum(column_widths)
            base_width = remaining_width / columns_left
            
            for _ in range(columns_left):
                column_widths.append(base_width)
            
            table = Table(table_data, colWidths=column_widths, repeatRows=1)
            
            # Styles de base
            base_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Rang
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),     # Nom et Prénoms
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('WORDWRAP', (1, 1), (1, -1), True),    # Nom et Prénoms
                
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
                    colors.HexColor('#FFFFFF'), 
                    colors.HexColor('#F5F5F5')
                ]),
                
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]
            
            # Alignement des colonnes dynamiques
            # Colonne "Classe" si présente
            if 'toutes' in classe_nom.lower():
                base_style.append(('ALIGN', (2, 1), (2, -1), 'CENTER'))  # Classe
                base_style.append(('WORDWRAP', (2, 1), (2, -1), True))   # Classe
                
                # Moyenne, Mention, Appréciation
                base_style.append(('ALIGN', (3, 1), (3, -1), 'CENTER'))  # Moyenne
                base_style.append(('ALIGN', (4, 1), (4, -1), 'CENTER'))  # Mention
                base_style.append(('ALIGN', (5, 1), (5, -1), 'LEFT'))    # Appréciation
                base_style.append(('WORDWRAP', (5, 1), (5, -1), True))   # Appréciation
                
                # Couleur de fond pour la colonne Moyenne
                base_style.append(('BACKGROUND', (3, 1), (3, -1), colors.HexColor('#F8F9FA')))
            else:
                # Pas de colonne "Classe"
                base_style.append(('ALIGN', (2, 1), (2, -1), 'CENTER'))  # Moyenne
                base_style.append(('ALIGN', (3, 1), (3, -1), 'CENTER'))  # Mention
                base_style.append(('ALIGN', (4, 1), (4, -1), 'LEFT'))    # Appréciation
                base_style.append(('WORDWRAP', (4, 1), (4, -1), True))   # Appréciation
                
                # Couleur de fond pour la colonne Moyenne
                base_style.append(('BACKGROUND', (2, 1), (2, -1), colors.HexColor('#F8F9FA')))
            
            table_style = TableStyle(base_style + style_commands)
            table.setStyle(table_style)
            
            elements.append(table)
            
            # Légende pour les ex-aequo
            nb_ex_aequo = sum(1 for d in donnees_moyennes if d.get('ex_aequo', False))
            if nb_ex_aequo > 0:
                legende_style = ParagraphStyle(
                    'LegendeStyle',
                    parent=styles['Normal'],
                    fontSize=8,
                    alignment=0,
                    spaceBefore=8,
                    textColor=colors.gray
                )
                elements.append(Spacer(1, 4))
                elements.append(Paragraph(
                    "<i>* Les suffixes '-ex' indiquent des élèves ex-aequo (même moyenne), classés par ordre alphabétique du nom.</i>", 
                    legende_style
                ))
        
        else:
            no_data_style = ParagraphStyle(
                'NoDataStyle',
                parent=styles['Normal'],
                fontSize=14,
                alignment=1,
                textColor=colors.gray,
                spaceAfter=20
            )
            elements.append(Paragraph("Aucune moyenne trouvée", no_data_style))
        
        doc.build(elements)
        buffer.seek(0)
        
        print("✅ Buffer PDF moyennes période généré avec succès")
        return buffer
        
    except Exception as e:
        print(f"❌ Erreur génération PDF moyennes: {str(e)}")
        traceback.print_exc()
        raise
    
def generate_moyennes_excel_buffer_periode(donnees_moyennes, ecole_id, classe_nom, matiere_id, 
                                          ecole_type, periode, annee_scolaire, mention):
    """Génère un buffer Excel avec les moyennes pour une période spécifique"""
    try:
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        
        ecole_nom = "Toutes écoles"
        if ecole_id:
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                ecole_nom = ecole.nom
        
        ws.title = f"Moyennes {ecole_nom}"[:31]
        
        # Informations avec période adaptée
        titre_periode = f"Semestre {periode}" if ecole_type != "college" else f"Trimestre {periode}"
        
        info_parts = [
            f"Classe: {classe_nom}",
            f"Période: {titre_periode}",
            f"Année: {annee_scolaire}",
            f"Effectif: {len(donnees_moyennes)} élèves"
        ]
        
        if matiere_id and matiere_id not in ['', 'null', 'undefined', 'toutes']:
            matiere = Matiere.query.get(matiere_id)
            if matiere:
                info_parts.insert(1, f"Matière: {matiere.libelle}")
        
        if mention:
            info_parts.append(f"Mention: {mention}")
        
        infos_supp = " | ".join(info_parts)
        
        # Ajouter l'en-tête
        start_row = add_excel_header(ws, ecole_id, "CLASSEMENT DES MOYENNES", infos_supp)
        
        # Style des bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ✅ EN-TÊTES DYNAMIQUES avec colonne "Classe" si nécessaire
        headers = ['Rang', 'Nom', 'Prénoms']
        
        # Ajouter colonne "Classe" si on exporte toutes les classes
        if 'toutes' in classe_nom.lower():
            headers.append('Classe')
        
        headers.extend(['Moyenne', 'Mention', 'Appréciation'])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Données des moyennes avec gestion dynamique
        for index, donnee in enumerate(donnees_moyennes, 1):
            eleve = donnee['eleve']
            row_num = start_row + index
            
            is_ex_aequo = donnee.get('ex_aequo', False)
            
            # Rang avec suffixe approprié
            rang_texte = get_rang_avec_suffixe(
                donnee['classement'], 
                eleve.sexe,
                is_ex_aequo
            )
            
            # Construire la ligne dynamiquement
            col_index = 1
            ws.cell(row=row_num, column=col_index, value=rang_texte)
            col_index += 1
            
            ws.cell(row=row_num, column=col_index, value=eleve.nom or '-')
            col_index += 1
            
            ws.cell(row=row_num, column=col_index, value=eleve.prenoms or '-')
            col_index += 1
            
            # Ajouter colonne "Classe" si nécessaire
            if 'toutes' in classe_nom.lower():
                ws.cell(row=row_num, column=col_index, value=donnee.get('classe_nom', '-'))
                col_index += 1
            
            # Ajouter les autres colonnes
            ws.cell(row=row_num, column=col_index, value=round(donnee['moyenne_generale'], 2))
            col_index += 1
            
            ws.cell(row=row_num, column=col_index, value=donnee['mention'])
            col_index += 1
            
            ws.cell(row=row_num, column=col_index, value=donnee['appreciation'])
            
            # Appliquer les styles
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                
                # Style spécial pour les ex-aequo
                if is_ex_aequo:
                    cell.font = Font(size=10, italic=True, color="666666")
                else:
                    cell.font = Font(size=10)
                    
                cell.border = thin_border
                
                # Couleur de fond pour la colonne Moyenne
                moyenne_col = 4 if 'toutes' in classe_nom.lower() else 3
                if col == moyenne_col:
                    cell.fill = PatternFill(
                        start_color="F0F8FF",
                        end_color="F0F8FF", 
                        fill_type="solid"
                    )
        
        # ✅ LARGEURS DE COLONNES DYNAMIQUES
        column_widths = {
            'A': 14,  # Rang
            'B': 20,  # Nom
            'C': 20,  # Prénoms
        }
        
        current_col = 'D'
        if 'toutes' in classe_nom.lower():
            column_widths[current_col] = 15  # Classe
            current_col = chr(ord(current_col) + 1)
        
        # Les colonnes restantes
        column_widths[current_col] = 12  # Moyenne
        current_col = chr(ord(current_col) + 1)
        
        column_widths[current_col] = 15  # Mention
        current_col = chr(ord(current_col) + 1)
        
        column_widths[current_col] = 35  # Appréciation
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Alignement dynamique
        for row in range(start_row, ws.max_row + 1):
            # Colonnes à centrer : Rang, Moyenne, Mention
            ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Déterminer la colonne Moyenne
            moyenne_col = 'D' if 'toutes' in classe_nom.lower() else 'C'
            ws[f'{moyenne_col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Colonne Mention
            mention_col = 'E' if 'toutes' in classe_nom.lower() else 'D'
            ws[f'{mention_col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Colonnes texte alignées à gauche
            ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Colonne Classe si présente
            if 'toutes' in classe_nom.lower():
                ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Colonne Appréciation
            appreciation_col = 'G' if 'toutes' in classe_nom.lower() else 'F'
            ws[f'{appreciation_col}{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Alternance des couleurs
        for row in range(start_row + 1, ws.max_row + 1):
            if row % 2 == 0:
                for col in range(1, len(headers) + 1):
                    current_cell = ws.cell(row=row, column=col)
                    if not current_cell.font.italic:
                        current_cell.fill = PatternFill(
                            start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
                        )
        
        # Hauteur des lignes
        for row in range(start_row, ws.max_row + 1):
            ws.row_dimensions[row].height = 22
        
        # Légende pour les ex-aequo
        nb_ex_aequo = sum(1 for d in donnees_moyennes if d.get('ex_aequo', False))
        if nb_ex_aequo > 0:
            legende_row = ws.max_row + 2
            ws.merge_cells(f'A{legende_row}:{chr(64 + len(headers))}{legende_row}')
            ws[f'A{legende_row}'] = "* Les suffixes '-ex' indiquent des élèves ex-aequo (même moyenne), classés par ordre alphabétique du nom."
            ws[f'A{legende_row}'].font = Font(size=9, italic=True, color="666666")
            ws[f'A{legende_row}'].alignment = Alignment(horizontal='left')
        
        wb.save(buffer)
        buffer.seek(0)
        
        print("✅ Buffer Excel moyennes période généré avec succès")
        return buffer
        
    except Exception as e:
        print(f"❌ Erreur génération Excel moyennes période: {str(e)}")
        traceback.print_exc()
        raise


def generer_nom_fichier_moyennes_periode(classe_nom, matiere_id, ecole_type, periode, annee_scolaire, mention, extension):
    """Génère un nom de fichier pour l'export des moyennes"""
    prefixe_periode = "T" if ecole_type == "college" else "S"
    nom_fichier = f"moyennes_{classe_nom.replace(' ', '_')}_{prefixe_periode}{periode}"
    
    if matiere_id:
        matiere = Matiere.query.get(matiere_id)
        if matiere:
            nom_fichier += f"_{matiere.libelle.replace(' ', '_')}"
    else:
        nom_fichier += "_general"
    
    if mention:
        nom_fichier += f"_{mention}"
    
    nom_fichier += f"_{annee_scolaire.replace('-', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.{extension}"
    
    return nom_fichier

def generate_moyennes_pdf_buffer_semestre(donnees_moyennes, ecole_id, classe_nom, matiere_id, 
                                         type_periode, periode, annee_scolaire, mention, ecole_type):
    """Génère un buffer PDF avec les moyennes pour semestres"""
    try:
        print(f"📊 DEBUT generate_moyennes_pdf_buffer_semestre - Type période: {type_periode}")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4,
            topMargin=15*mm,
            bottomMargin=15*mm,
            leftMargin=10*mm,
            rightMargin=10*mm
        )
        elements = []
        
        # Titre principal adapté
        ecole_nom = "TOUTES LES ÉCOLES"
        if ecole_id:
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                ecole_nom = ecole.nom
        
        titre_prefixe = "SEMESTRE" if type_periode == "semestre" else "TRIMESTRE"
        titre_principal = f"CLASSEMENT DES MOYENNES"
        
        elements = add_pdf_header(elements, doc, ecole_id, titre_principal)
        
        styles = getSampleStyleSheet()
        
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=11,
            alignment=1,
            spaceAfter=10,
            textColor=colors.HexColor('#34495E')
        )
        
        # Informations spécifiques aux semestres
        periode_texte = f"{titre_prefixe} {periode}"
        if type_periode == "semestre" and ecole_type == "college":
            if periode == "1":
                periode_texte += " (basé sur T1 + T2)"
            elif periode == "2":
                periode_texte += " (basé sur T3)"
        
        info_parts = [
            f"Classe: {classe_nom}",
            f"Période: {periode_texte}",
            f"Année: {annee_scolaire}",
            f"Effectif: {len(donnees_moyennes)} élèves"
        ]
        
        # Ajouter la matière si spécifiée
        if matiere_id:
            matiere = Matiere.query.get(matiere_id)
            if matiere:
                info_parts.insert(1, f"Matière: {matiere.libelle}")
        
        if mention:
            info_parts.append(f"Mention: {mention}")
        
        info_text = " | ".join(info_parts)
        elements.append(Paragraph(f"<b>{info_text}</b>", info_style))
        
        # Statistiques
        if donnees_moyennes:
            moyennes = [d['moyenne_generale'] for d in donnees_moyennes]
            moyenne_classe = sum(moyennes) / len(moyennes) if moyennes else 0
            max_moyenne = max(moyennes) if moyennes else 0
            min_moyenne = min(moyennes) if moyennes else 0
            
            stats_style = ParagraphStyle(
                'StatsStyle',
                parent=styles['Normal'],
                fontSize=10,
                alignment=1,
                spaceAfter=12,
                textColor=colors.HexColor('#2C3E50')
            )
            
            stats_text = f"Moyenne classe: {moyenne_classe:.2f}/20 | Max: {max_moyenne:.2f}/20 | Min: {min_moyenne:.2f}/20"
            elements.append(Paragraph(f"<i>{stats_text}</i>", stats_style))
        
        # Date de génération
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=9,
            alignment=1,
            spaceAfter=8,
            textColor=colors.gray
        )
        
        elements.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style))
        elements.append(Spacer(1, 10))
        
        # Tableau des moyennes (identique à la version précédente mais adapté)
        if donnees_moyennes:
            headers = ['Rang', 'Nom et Prénoms', 'Moyenne', 'Mention', 'Appréciation']
            table_data = [headers]
            
            style_commands = []
            
            for row_index, donnee in enumerate(donnees_moyennes, 1):
                eleve = donnee['eleve']
                is_ex_aequo = donnee.get('ex_aequo', False)
                
                rang_texte = get_rang_avec_suffixe(
                    donnee['classement'], 
                    eleve.sexe,
                    is_ex_aequo
                )
                
                table_data.append([
                    rang_texte,
                    f"{eleve.nom} {eleve.prenoms}" if eleve else '-',
                    f"{donnee['moyenne_generale']:.2f}/20",
                    donnee['mention'],
                    donnee['appreciation']
                ])
                
                if is_ex_aequo:
                    style_commands.append(('FONTNAME', (0, row_index), (0, row_index), 'Helvetica-Oblique'))
                    style_commands.append(('TEXTCOLOR', (0, row_index), (0, row_index), colors.HexColor('#666666')))
            
            total_width = doc.width
            column_widths = [
                total_width * 0.12,
                total_width * 0.30,
                total_width * 0.15,
                total_width * 0.15,
                total_width * 0.28
            ]
            
            table = Table(table_data, colWidths=column_widths, repeatRows=1)
            
            base_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('ALIGN', (2, 1), (2, -1), 'CENTER'),
                ('ALIGN', (3, 1), (3, -1), 'CENTER'),
                ('ALIGN', (4, 1), (4, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('WORDWRAP', (1, 1), (4, -1), True),
                
                ('BACKGROUND', (2, 1), (2, -1), colors.HexColor('#F8F9FA')),
                
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
                    colors.HexColor('#FFFFFF'), 
                    colors.HexColor('#F5F5F5')
                ]),
                
                ('LINEBEFORE', (2, 0), (2, -1), 1, colors.HexColor('#E9ECEF')),
                ('LINEAFTER', (2, 0), (2, -1), 1, colors.HexColor('#E9ECEF')),
                
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]
            
            table_style = TableStyle(base_style + style_commands)
            table.setStyle(table_style)
            
            elements.append(table)
            
            # Légende pour les ex-aequo
            nb_ex_aequo = sum(1 for d in donnees_moyennes if d.get('ex_aequo', False))
            if nb_ex_aequo > 0:
                legende_style = ParagraphStyle(
                    'LegendeStyle',
                    parent=styles['Normal'],
                    fontSize=8,
                    alignment=0,
                    spaceBefore=8,
                    textColor=colors.gray
                )
                elements.append(Spacer(1, 4))
                elements.append(Paragraph(
                    "<i>* Les suffixes '-ex' indiquent des élèves ex-aequo (même moyenne), classés par ordre alphabétique du nom.</i>", 
                    legende_style
                ))
                
                # Info spécifique aux semestres
                if type_periode == "semestre":
                    semestre_info_style = ParagraphStyle(
                        'SemestreInfoStyle',
                        parent=styles['Normal'],
                        fontSize=8,
                        alignment=0,
                        spaceBefore=4,
                        textColor=colors.HexColor('#2C3E50')
                    )
                    if ecole_type == "college":
                        if periode == "1":
                            elements.append(Paragraph(
                                "<i>** Le semestre 1 regroupe les notes des trimestres 1 et 2.</i>", 
                                semestre_info_style
                            ))
                        elif periode == "2":
                            elements.append(Paragraph(
                                "<i>** Le semestre 2 correspond au trimestre 3.</i>", 
                                semestre_info_style
                            ))
                    else:
                        elements.append(Paragraph(
                            "<i>** En lycée, les semestres correspondent aux trimestres 1 et 2.</i>", 
                            semestre_info_style
                        ))
        
        else:
            no_data_style = ParagraphStyle(
                'NoDataStyle',
                parent=styles['Normal'],
                fontSize=14,
                alignment=1,
                textColor=colors.gray,
                spaceAfter=20
            )
            elements.append(Paragraph("Aucune moyenne trouvée", no_data_style))
        
        doc.build(elements)
        buffer.seek(0)
        
        print("✅ Buffer PDF moyennes semestre généré avec succès")
        return buffer
        
    except Exception as e:
        print(f"❌ Erreur génération PDF moyennes semestre: {str(e)}")
        traceback.print_exc()
        raise

def generer_nom_fichier_moyennes(classe_nom, matiere_id, trimestre, annee_scolaire, mention, extension):
    """Génère un nom de fichier pour l'export des moyennes"""
    nom_fichier = f"moyennes_{classe_nom.replace(' ', '_')}_T{trimestre}"
    
    if matiere_id:
        matiere = Matiere.query.get(matiere_id)
        if matiere:
            nom_fichier += f"_{matiere.libelle.replace(' ', '_')}"
    else:
        nom_fichier += "_general"
    
    if mention:
        nom_fichier += f"_{mention}"
    
    nom_fichier += f"_{annee_scolaire.replace('-', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.{extension}"
    
    return nom_fichier

def generate_moyennes_pdf_buffer(donnees_moyennes, ecole_id, classe_nom, matiere_id, trimestre, annee_scolaire, mention):
    """Génère un buffer PDF avec les moyennes - VERSION FINALE CORRIGÉE"""
    try:
        print(f"📊 DEBUT generate_moyennes_pdf_buffer - École: {ecole_id}")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4,
            topMargin=15*mm,
            bottomMargin=15*mm,
            leftMargin=10*mm,
            rightMargin=10*mm
        )
        elements = []
        
        # Titre principal
        ecole_nom = "TOUTES LES ÉCOLES"
        if ecole_id:
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                ecole_nom = ecole.nom
        
        titre_principal = f"CLASSEMENT DES MOYENNES"
        print(f"🎯 Titre PDF moyennes: {titre_principal}")
        
        # Ajouter l'en-tête
        elements = add_pdf_header(elements, doc, ecole_id, titre_principal)
        
        # Styles
        styles = getSampleStyleSheet()
        
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=11,
            alignment=1,
            spaceAfter=10,
            textColor=colors.HexColor('#34495E')
        )
        
        stats_style = ParagraphStyle(
            'StatsStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,
            spaceAfter=12,
            textColor=colors.HexColor('#2C3E50')
        )
        
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=9,
            alignment=1,
            spaceAfter=8,
            textColor=colors.gray
        )
        
        # Récupérer le nom de la matière pour l'affichage
        matiere_nom_affichage = "Toutes matières"
        if matiere_id:
            matiere = Matiere.query.get(matiere_id)
            if matiere:
                matiere_nom_affichage = matiere.libelle
        
        # Informations des filtres
        info_parts = []
        info_parts.append(f"Classe: {classe_nom}")
        info_parts.append(f"Matière: {matiere_nom_affichage}")
        info_parts.append(f"Trimestre: {trimestre}")
        info_parts.append(f"Année: {annee_scolaire}")
        
        if mention:
            info_parts.append(f"Mention: {mention}")
            
        info_parts.append(f"Effectif: {len(donnees_moyennes)} élèves")
        
        info_text = " | ".join(info_parts)
        elements.append(Paragraph(f"<b>{info_text}</b>", info_style))
        
        # Statistiques
        if donnees_moyennes:
            moyennes = [d['moyenne_generale'] for d in donnees_moyennes]
            moyenne_classe = sum(moyennes) / len(moyennes)
            max_moyenne = max(moyennes)
            min_moyenne = min(moyennes)
            
            # COMPTAGE DES EX-AEQUO
            nb_ex_aequo = sum(1 for d in donnees_moyennes if d.get('ex_aequo', False))
            
            stats_text = f"Moyenne classe: {moyenne_classe:.2f}/20 | Max: {max_moyenne:.2f}/20 | Min: {min_moyenne:.2f}/20"
            if nb_ex_aequo > 0:
                stats_text += f" | Ex-aequo: {nb_ex_aequo} élève(s)"
                
            elements.append(Paragraph(f"<i>{stats_text}</i>", stats_style))
        
        # Date de génération
        elements.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style))
        elements.append(Spacer(1, 10))
        
        # Tableau des moyennes
        if donnees_moyennes:
            headers = ['Rang', 'Nom et Prénoms', 'Moyenne', 'Mention', 'Appréciation']
            table_data = [headers]
            
            # Préparer la liste des styles conditionnels
            style_commands = []
            
            for row_index, donnee in enumerate(donnees_moyennes, 1):
                eleve = donnee['eleve']
                is_ex_aequo = donnee.get('ex_aequo', False)
                
                # Rang avec suffixe "-ex" pour les ex-aequo
                rang_texte = get_rang_avec_suffixe(
                    donnee['classement'], 
                    eleve.sexe,
                    is_ex_aequo
                )
                
                table_data.append([
                    rang_texte,
                    f"{eleve.nom} {eleve.prenoms}" if eleve else '-',
                    f"{donnee['moyenne_generale']:.2f}/20",
                    donnee['mention'],
                    donnee['appreciation']
                ])
                
                # Ajouter les styles conditionnels pour les ex-aequo
                if is_ex_aequo:
                    style_commands.append(('FONTNAME', (0, row_index), (0, row_index), 'Helvetica-Oblique'))
                    style_commands.append(('TEXTCOLOR', (0, row_index), (0, row_index), colors.HexColor('#666666')))
            
            # Largeurs de colonnes optimisées
            total_width = doc.width
            column_widths = [
                total_width * 0.12,
                total_width * 0.30,
                total_width * 0.15,
                total_width * 0.15,
                total_width * 0.28
            ]
            
            table = Table(table_data, colWidths=column_widths, repeatRows=1)
            
            # Style du tableau
            base_style = [
                # En-têtes
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                
                # Données
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('ALIGN', (2, 1), (2, -1), 'CENTER'),
                ('ALIGN', (3, 1), (3, -1), 'CENTER'),
                ('ALIGN', (4, 1), (4, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('WORDWRAP', (1, 1), (4, -1), True),
                
                # Couleur de fond pour la colonne Moyenne
                ('BACKGROUND', (2, 1), (2, -1), colors.HexColor('#F8F9FA')),
                
                # Bordures
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                
                # Alternance des couleurs des lignes
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
                    colors.HexColor('#FFFFFF'), 
                    colors.HexColor('#F5F5F5')
                ]),
                
                # Bordure accentuée pour la colonne Moyenne
                ('LINEBEFORE', (2, 0), (2, -1), 1, colors.HexColor('#E9ECEF')),
                ('LINEAFTER', (2, 0), (2, -1), 1, colors.HexColor('#E9ECEF')),
                
                # Padding
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]
            
            # Combiner les styles
            table_style = TableStyle(base_style + style_commands)
            table.setStyle(table_style)
            
            elements.append(table)
            
            # Légende pour les ex-aequo
            nb_ex_aequo = sum(1 for d in donnees_moyennes if d.get('ex_aequo', False))
            if nb_ex_aequo > 0:
                legende_style = ParagraphStyle(
                    'LegendeStyle',
                    parent=styles['Normal'],
                    fontSize=8,
                    alignment=0,
                    spaceBefore=8,
                    textColor=colors.gray
                )
                elements.append(Spacer(1, 4))
                elements.append(Paragraph(
                    "<i>* Les suffixes '-ex' indiquent des élèves ex-aequo (même moyenne), classés par ordre alphabétique du nom.</i>", 
                    legende_style
                ))
                
        else:
            # Aucune donnée trouvée
            no_data_style = ParagraphStyle(
                'NoDataStyle',
                parent=styles['Normal'],
                fontSize=14,
                alignment=1,
                textColor=colors.gray,
                spaceAfter=20
            )
            elements.append(Paragraph("Aucune moyenne trouvée", no_data_style))
        
        # Génération du PDF
        doc.build(elements)
        buffer.seek(0)
        
        print("✅ Buffer PDF moyennes généré avec succès")
        return buffer
        
    except Exception as e:
        print(f"❌ Erreur génération PDF moyennes: {str(e)}")
        traceback.print_exc()
        raise

def generate_notes_excel_buffer_periode(notes, ecole_id, classe_nom, matiere_nom, ecole_type, periode, annee_scolaire):
    """Génère un buffer Excel avec les notes pour une période avec colonnes dynamiques"""
    try:
        # Création du workbook
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        
        ecole_nom = "Toutes écoles"
        if ecole_id:
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                ecole_nom = ecole.nom
        
        ws.title = f"Notes {ecole_nom}"[:31]
        
        # Informations avec période adaptée
        titre_periode = f"Semestre {periode}" if ecole_type != "college" else f"Trimestre {periode}"
        
        # Informations de l'export
        infos_parts = [
            f"Classe: {classe_nom}",
            f"Matière: {matiere_nom}",
            f"Période: {titre_periode}",
            f"Année: {annee_scolaire}",
            f"Total: {len(notes)} notes"
        ]
        
        if classe_nom == "Toutes les classes":
            infos_parts.insert(1, "(avec colonne 'Classe' incluse)")
        if matiere_nom == "Toutes les matières":
            infos_parts.insert(2, "(avec colonne 'Matière' incluse)")
        
        infos_supp = " | ".join(infos_parts)
        
        # Ajouter l'en-tête
        start_row = add_excel_header(ws, ecole_id, "LISTE DES NOTES", infos_supp)
        
        # Style des bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ✅ AMÉLIORATION : En-têtes dynamiques
        headers = ['N°', 'Élève']
        
        # Ajouter colonne "Classe" si on exporte toutes les classes
        if classe_nom == "Toutes les classes":
            headers.append('Classe')
        
        # Ajouter colonne "Matière" si on exporte toutes les matières
        if matiere_nom == "Toutes les matières":
            headers.append('Matière')
        
        # Ajouter les colonnes de notes
        headers.extend(['Note 1', 'Note 2', 'Note 3', 'Note Comp', 'Coefficient'])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # ✅ AMÉLIORATION : Données dynamiques
        for index, note in enumerate(notes, 1):
            eleve = note.eleve
            matiere = note.matiere
            row_num = start_row + index
            
            # Récupérer le nom de la classe de l'élève
            classe_eleve = note.eleve.classe.nom if note.eleve and note.eleve.classe else '-'
            
            # Formatage des notes
            note1 = note.note1 if note.note1 is not None else "-"
            note2 = note.note2 if note.note2 is not None else "-"
            note3 = note.note3 if note.note3 is not None else "-"
            note_comp = note.note_comp if note.note_comp is not None else "-"
            
            # Construire la ligne dynamiquement
            col_index = 1
            ws.cell(row=row_num, column=col_index, value=index)
            col_index += 1
            
            ws.cell(row=row_num, column=col_index, value=f"{eleve.nom} {eleve.prenoms}" if eleve else '-')
            col_index += 1
            
            # Ajouter colonne "Classe" si nécessaire
            if classe_nom == "Toutes les classes":
                ws.cell(row=row_num, column=col_index, value=classe_eleve)
                col_index += 1
            
            # Ajouter colonne "Matière" si nécessaire
            if matiere_nom == "Toutes les matières":
                ws.cell(row=row_num, column=col_index, value=matiere.libelle if matiere else '-')
                col_index += 1
            
            # Ajouter les notes
            ws.cell(row=row_num, column=col_index, value=note1)
            col_index += 1
            
            ws.cell(row=row_num, column=col_index, value=note2)
            col_index += 1
            
            ws.cell(row=row_num, column=col_index, value=note3)
            col_index += 1
            
            ws.cell(row=row_num, column=col_index, value=note_comp)
            col_index += 1
            
            ws.cell(row=row_num, column=col_index, value=note.coefficient if note.coefficient else 1)
            
            # Appliquer les bordures à toute la ligne
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).border = thin_border
                ws.cell(row=row_num, column=col).font = Font(size=9)
        
        # ✅ AMÉLIORATION : Largeurs de colonnes dynamiques
        column_widths = {
            'A': 6,    # N°
            'B': 25,   # Élève
        }
        
        current_col = 'C'
        if classe_nom == "Toutes les classes":
            column_widths[current_col] = 15  # Classe
            current_col = chr(ord(current_col) + 1)
        
        if matiere_nom == "Toutes les matières":
            column_widths[current_col] = 20  # Matière
            current_col = chr(ord(current_col) + 1)
        
        # Notes (toujours présentes)
        for i in range(5):  # Note 1, Note 2, Note 3, Note Comp, Coefficient
            column_widths[current_col] = 10
            current_col = chr(ord(current_col) + 1)
        
        # Ajuster la largeur de la colonne Élève si nécessaire
        if classe_nom == "Toutes les classes" and matiere_nom == "Toutes les matières":
            column_widths['B'] = 20  # Élève moins large
        elif classe_nom == "Toutes les classes" or matiere_nom == "Toutes les matières":
            column_widths['B'] = 22  # Élève moyennement large
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Alignement dynamique
        for row in range(start_row, ws.max_row + 1):
            # Colonnes à centrer : N° et toutes les notes
            ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Déterminer la dernière colonne de texte
            last_text_col = 'B'  # Commence à Élève
            
            if classe_nom == "Toutes les classes":
                ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                last_text_col = 'C'
                
                if matiere_nom == "Toutes les matières":
                    ws[f'D{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    last_text_col = 'D'
            elif matiere_nom == "Toutes les matières":
                ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                last_text_col = 'C'
            
            # Élève aligné à gauche
            ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Notes alignées au centre
            notes_start_col = chr(ord(last_text_col) + 1)
            for col_letter in [notes_start_col, 
                              chr(ord(notes_start_col) + 1),
                              chr(ord(notes_start_col) + 2),
                              chr(ord(notes_start_col) + 3),
                              chr(ord(notes_start_col) + 4)]:
                ws[f'{col_letter}{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Alternance des couleurs
        for row in range(start_row + 1, ws.max_row + 1):
            if row % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
                    )
        
        # Hauteur des lignes
        for row in range(start_row, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
        
        wb.save(buffer)
        buffer.seek(0)
        
        print("✅ Buffer Excel généré avec colonnes dynamiques")
        return buffer
        
    except Exception as e:
        print(f"❌ Erreur génération Excel notes période: {str(e)}")
        traceback.print_exc()
        raise

def generate_moyennes_excel_buffer(donnees_moyennes, ecole_id, classe_nom, matiere_id, trimestre, annee_scolaire, mention):
    """Génère un buffer Excel avec les moyennes - VERSION FINALE CORRIGÉE"""
    try:
        # Création du workbook
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        
        # Nom de l'onglet
        ecole_nom = "Toutes écoles"
        if ecole_id:
            ecole = Ecole.query.get(ecole_id)
            if ecole:
                ecole_nom = ecole.nom
        
        ws.title = f"Moyennes {ecole_nom}"[:31]
        
        # Récupérer le nom de la matière pour l'affichage
        matiere_nom_affichage = "Toutes matières"
        if matiere_id:
            matiere = Matiere.query.get(matiere_id)
            if matiere:
                matiere_nom_affichage = matiere.libelle
        
        # Informations supplémentaires pour l'en-tête
        infos_parts = []
        infos_parts.append(f"Classe: {classe_nom}")
        infos_parts.append(f"Matière: {matiere_nom_affichage}")
        infos_parts.append(f"Trimestre: {trimestre}")
        infos_parts.append(f"Année: {annee_scolaire}")
        
        if mention:
            infos_parts.append(f"Mention: {mention}")
            
        infos_parts.append(f"Effectif: {len(donnees_moyennes)} élèves")
        
        infos_supp = " | ".join(infos_parts)
        
        # Ajouter l'en-tête
        start_row = add_excel_header(ws, ecole_id, "CLASSEMENT DES MOYENNES", infos_supp)
        
        # Style des bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-têtes
        headers = ['Rang', 'Nom', 'Prénoms', 'Moyenne', 'Mention', 'Appréciation']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Données des moyennes
        for index, donnee in enumerate(donnees_moyennes, 1):
            eleve = donnee['eleve']
            row_num = start_row + index
            
            is_ex_aequo = donnee.get('ex_aequo', False)
            
            # Rang avec suffixe "-ex"
            rang_texte = get_rang_avec_suffixe(
                donnee['classement'], 
                eleve.sexe,
                is_ex_aequo
            )
            
            row_data = [
                rang_texte,
                eleve.nom or '-',
                eleve.prenoms or '-',
                round(donnee['moyenne_generale'], 2),
                donnee['mention'],
                donnee['appreciation']
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = value
                
                # Style spécial pour les ex-aequo
                if is_ex_aequo:
                    cell.font = Font(size=10, italic=True, color="666666")
                else:
                    cell.font = Font(size=10)
                    
                cell.border = thin_border
                
                # Couleur de fond pour la colonne Moyenne
                if col == 4:
                    cell.fill = PatternFill(
                        start_color="F0F8FF",
                        end_color="F0F8FF", 
                        fill_type="solid"
                    )
        
        # Largeurs de colonnes
        column_widths = {
            'A': 14,
            'B': 25,
            'C': 25,
            'D': 12,
            'E': 15,
            'F': 35
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Alignement et retour à la ligne
        for row in range(start_row, ws.max_row + 1):
            for col in ['A', 'D', 'E']:
                ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            for col in ['B', 'C', 'F']:
                ws[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Alternance des couleurs des lignes
        for row in range(start_row + 1, ws.max_row + 1):
            if row % 2 == 0:
                for col in range(1, len(headers) + 1):
                    if col != 4:
                        current_cell = ws.cell(row=row, column=col)
                        if not current_cell.font.italic:
                            current_cell.fill = PatternFill(
                                start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
                            )
        
        # Augmenter la hauteur des lignes
        for row in range(start_row, ws.max_row + 1):
            ws.row_dimensions[row].height = 22
        
        # Légende pour les ex-aequo
        nb_ex_aequo = sum(1 for d in donnees_moyennes if d.get('ex_aequo', False))
        if nb_ex_aequo > 0:
            legende_row = ws.max_row + 2
            ws.merge_cells(f'A{legende_row}:F{legende_row}')
            ws[f'A{legende_row}'] = "* Les suffixes '-ex' indiquent des élèves ex-aequo (même moyenne), classés par ordre alphabétique du nom."
            ws[f'A{legende_row}'].font = Font(size=9, italic=True, color="666666")
            ws[f'A{legende_row}'].alignment = Alignment(horizontal='left')
        
        # Sauvegarde
        wb.save(buffer)
        buffer.seek(0)
        
        print("✅ Buffer Excel moyennes généré avec succès")
        return buffer
        
    except Exception as e:
        print(f"❌ Erreur génération Excel moyennes: {str(e)}")
        traceback.print_exc()
        raise

@services_bp.route("/admin/recreate-logos-simple")
@login_required
def admin_recreate_logos_simple():
    """Route pour recréer les logos avec format simple"""
    return recreate_logos_simple_format()
    

