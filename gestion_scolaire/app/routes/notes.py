from flask import Blueprint, request, render_template, jsonify, send_file, abort, current_app
from ..models import Note, Eleve, Matiere, Enseignant, Classe, Enseignement, Ecole
from sqlalchemy import cast, String
from flask_login import current_user, login_required
import uuid
from uuid import UUID
from datetime import date
from extensions import db
from sqlalchemy.orm import joinedload
import io
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import re
import logging
import json
from functools import wraps
from ..utils import ecole_required, get_current_ecole_id

notes_bp = Blueprint('notes', __name__)

# ========== CONFIGURATION DE SÉCURITÉ ==========
class SecurityConfig:
    MAX_PER_PAGE = 1000  # CHANGÉ: de 100 à 1000 pour permettre 250,500,1000
    DEFAULT_PER_PAGE = 10
    MAX_EXPORT_ROWS = 1000
    ALLOWED_TRIMESTRES = [1, 2, 3]
    MAX_SEARCH_LENGTH = 100
    ALLOWED_MATIERE_TYPES = ["Matière scientifique", "Matière littéraire", "Autres"]

def validate_periode(cycle_type, periode_value):
    """Validation de la période selon le cycle"""
    if periode_value is None:
        return False  # None n'est pas valide, on ne doit pas entrer dans le bloc
    if cycle_type == 'lycee':
        return periode_value in [1, 2]
    else:
        return periode_value in [1, 2, 3]

def get_unique_matieres(ecole_id=None):
    """Récupère les matières avec déduplication intelligente"""
    toutes_matieres = Matiere.query.filter(
        Matiere.type.in_(SecurityConfig.ALLOWED_MATIERE_TYPES)
    ).all()
    
    # Nettoyer les libellés pour la comparaison
    def clean_libelle(libelle):
        """Nettoie le libellé pour détecter les doublons"""
        if not libelle:
            return ""
        # Convertir en minuscules
        cleaned = libelle.lower().strip()
        # Remplacer les variantes courantes
        cleaned = cleaned.replace('ies', 'ie')
        cleaned = cleaned.replace(' et ', ' ')
        cleaned = cleaned.replace(' de ', ' ')
        cleaned = cleaned.replace(' la ', ' ')
        cleaned = cleaned.replace(' le ', ' ')
        cleaned = cleaned.replace('les ', '')
        cleaned = cleaned.replace('des ', '')
        # Supprimer les espaces multiples
        cleaned = ' '.join(cleaned.split())
        return cleaned
    
    # Dictionnaire pour stocker les matières uniques
    matieres_uniques = {}
    
    for matiere in toutes_matieres:
        cle_nettoyee = clean_libelle(matiere.libelle)
        
        if cle_nettoyee not in matieres_uniques:
            # Première occurrence - la garder
            matieres_uniques[cle_nettoyee] = {
                "id": str(matiere.id),
                "libelle": matiere.libelle,
                "libelle_original": matiere.libelle,
                "type": matiere.type,
                "is_parent": False
            }
        else:
            # Doublon trouvé - logger l'information
            print(f"⚠️ Doublon détecté: '{matieres_uniques[cle_nettoyee]['libelle']}' et '{matiere.libelle}'")
            # Option: marquer pour suppression ou correction
    
    # Trier par libellé
    matieres_list = sorted(
        matieres_uniques.values(), 
        key=lambda x: x["libelle"]
    )
    
    # Regrouper par type
    matieres_par_type = {}
    for matiere in matieres_list:
        type_name = matiere["type"]
        if type_name not in matieres_par_type:
            matieres_par_type[type_name] = []
        matieres_par_type[type_name].append(matiere)
    
    # Construire la structure hiérarchique
    matieres_hierarchiques = []
    for type_name, matieres_du_type in matieres_par_type.items():
        if matieres_du_type:
            type_groupe = {
                "id": f"type_{type_name.lower().replace(' ', '_')}",
                "libelle": type_name,
                "is_parent": True,
                "children": matieres_du_type
            }
            matieres_hierarchiques.append(type_groupe)
    
    return matieres_hierarchiques
    
# ========== FONCTIONS DE VALIDATION ET SÉCURITÉ ==========
def validate_annee_scolaire(annee):
    """Validation stricte du format année scolaire"""
    if not annee:
        return False
    pattern = r'^\d{4}-\d{4}$'
    if not re.match(pattern, annee):
        return False
    try:
        debut, fin = map(int, annee.split('-'))
        return fin == debut + 1 and 2000 <= debut <= 2100
    except ValueError:
        return False

def validate_trimestre(trimestre):
    """Validation des trimestres"""
    return trimestre in SecurityConfig.ALLOWED_TRIMESTRES

def safe_search_term(term):
    """Nettoyage des termes de recherche"""
    if not term:
        return ""
    term = term.replace('%', '\\%').replace('_', '\\_')
    return term[:SecurityConfig.MAX_SEARCH_LENGTH]

class SecurityConfig:
    MAX_PER_PAGE = 1000  # CHANGÉ: de 100 à 1000 pour permettre 250,500,1000
    DEFAULT_PER_PAGE = 10
    MAX_EXPORT_ROWS = 1000
    ALLOWED_TRIMESTRES = [1, 2, 3]
    MAX_SEARCH_LENGTH = 100
    ALLOWED_MATIERE_TYPES = ["Matière scientifique", "Matière littéraire", "Autres"]


def validate_pagination_params(f):
    """Décorateur pour valider les paramètres de pagination"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        per_page = request.args.get('per_page', SecurityConfig.DEFAULT_PER_PAGE, type=int)
        page = request.args.get('page', 1, type=int)
        
        # CHANGÉ: Permettre les valeurs jusqu'à 1000
        if per_page > SecurityConfig.MAX_PER_PAGE or per_page < 1:
            per_page = SecurityConfig.DEFAULT_PER_PAGE
        if page < 1:
            page = 1
            
        request.validated_per_page = per_page
        request.validated_page = page
            
        return f(*args, **kwargs)
    return decorated_function

def validate_export_params(classe_id, matiere_id, trimestre, annee_scolaire, ecole_id, cycle_type="college"):
    """Validation des paramètres d'export"""
    errors = []
    
    if classe_id and classe_id.lower() != "none":
        try:
            classe_uuid = UUID(classe_id)
            classe = Classe.query.filter_by(id=classe_uuid, ecole_id=ecole_id).first()
            if not classe:
                errors.append("Classe non autorisée ou introuvable")
        except ValueError:
            errors.append("Format de classe invalide")
    
    if matiere_id and matiere_id.lower() != "none":
        try:
            UUID(matiere_id)
        except ValueError:
            errors.append("Format de matière invalide")
    
    if trimestre and not validate_trimestre_for_cycle(trimestre, cycle_type):
        errors.append(f"{'Semestre' if cycle_type == 'lycee' else 'Trimestre'} invalide")
    
    if annee_scolaire and not validate_annee_scolaire(annee_scolaire):
        errors.append("Format d'année scolaire invalide")
    
    return errors

def log_security_event(user_id, action, resource, status, details=None):
    """Journalisation simplifiée"""
    try:
        print(f"🔐 SECURITY: {action} - {status} - User: {user_id}")
    except Exception:
        pass

# ========== HEADERS DE SÉCURITÉ ==========
@notes_bp.after_request
def set_security_headers(response):
    """Définir les headers de sécurité HTTP"""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return response

# ========== FONCTIONS UTILITAIRES SÉCURISÉES ==========
def get_note_secure(note_id, ecole_id):
    """Récupère une note en vérifiant l'appartenance"""
    try:
        note_uuid = UUID(note_id)
    except ValueError:
        abort(404, "Note non trouvée")
        
    note = Note.query.join(Eleve).join(Classe)\
                    .filter(Note.id == note_uuid, Classe.ecole_id == ecole_id).first()
    if not note:
        abort(404, "Note non trouvée")
    return note

def get_eleve_secure(eleve_id, ecole_id):
    """Récupère un élève en vérifiant son appartenance à l'école"""
    try:
        eleve_uuid = UUID(eleve_id)
    except ValueError:
        return None
        
    return Eleve.query.join(Classe).filter(
        Eleve.id == eleve_uuid, 
        Classe.ecole_id == ecole_id
    ).first()

def validate_note_value(note_value):
    """Valide une valeur de note"""
    if note_value is None or note_value == "":
        return None
    try:
        note_float = float(note_value)
        if 0 <= note_float <= 20:
            return round(note_float, 2)  # Arrondir à 2 décimales
        return None
    except (ValueError, TypeError):
        return None

def validate_trimestre_for_cycle(trimestre, cycle_type):
    """Validation des trimestres/semestres selon le cycle"""
    if cycle_type == 'lycee':
        return trimestre in [1, 2]  # Semestres
    return trimestre in [1, 2, 3]  # Trimestres
#Fonction d'export
def get_logo_path(ecole):
    """Retourne le chemin du logo de l'école - VERSION CORRIGÉE"""
    if not ecole or not ecole.logo_filename:
        print(f"❌ Logo non disponible - Ecole: {ecole}, Logo_filename: {getattr(ecole, 'logo_filename', None)}")
        return None
    
    # CHEMIN DÉFINITIF confirmé (identique à eleves.py et enseignants.py)
    logo_path = os.path.join(
        current_app.root_path, 
        'gestion_scolaire', 
        'app', 
        'static', 
        'logos', 
        ecole.logo_filename
    )
    
    if os.path.exists(logo_path) and os.path.getsize(logo_path) > 0:
        print(f"✅ Logo trouvé pour notes: {logo_path}")
        return logo_path
    
    print(f"❌ Logo non trouvé pour notes: {logo_path}")
    return None

def generate_pdf_export(notes, classe_nom=None):
    """Génère un export PDF des notes avec en-tête uniformisé"""
    try:
        ecole_id = get_current_ecole_id()
        ecole = Ecole.query.get(ecole_id)
        
        if not ecole:
            raise ValueError("École non trouvée")
        
        # DEBUG: Informations sur le logo
        print(f"🔍 RECHERCHE LOGO NOTES - Ecole: {ecole.nom}, Logo_filename: {ecole.logo_filename}")
        logo_path = get_logo_path(ecole)
        print(f"📁 Chemin logo final notes: {logo_path}")

        # UTILISER LES VALEURS DYNAMIQUES DE L'ÉCOLE COURANTE
        nom_ecole = ecole.nom if ecole.nom else "École non renseignée"
        dre = ecole.dre if ecole.dre else "DRE non renseignée"
        inspection = ecole.inspection if ecole.inspection else "Inspection non renseignée"
        telephone = ecole.telephone1 if ecole.telephone1 else "Téléphone non renseigné"
        devise_ecole = ecole.devise if ecole.devise else "Travail - Liberté - Patrie"
        
        buffer = io.BytesIO()
        
        def add_page_number(canvas, doc):
            canvas.saveState()
            page_num = canvas.getPageNumber()
            total_pages = doc.page
            
            footer_text = f"Page {page_num}/{total_pages}"
            canvas.setFont('Helvetica', 8)
            canvas.setFillColor(colors.gray)
            canvas.drawRightString(doc.pagesize[0] - 20*mm, 10*mm, footer_text)
            
            info_text = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
            canvas.drawString(20*mm, 10*mm, info_text)
            canvas.restoreState()
        
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4, 
            topMargin=15*mm, 
            bottomMargin=20*mm,
            leftMargin=10*mm,
            rightMargin=10*mm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # ========== EN-TÊTE UNIFORMISÉ ==========
        logo = None
        if logo_path:
            try:
                # Vérifier que le fichier existe et n'est pas vide
                if os.path.exists(logo_path) and os.path.getsize(logo_path) > 0:
                    logo = Image(logo_path, width=35*mm, height=35*mm)
                    logo.hAlign = 'CENTER'
                    print("✅ Logo chargé avec succès dans le PDF notes")
                else:
                    print("❌ Logo trouvé mais fichier vide ou inexistant")
            except Exception as e:
                print(f"❌ Erreur chargement logo notes: {e}")
                logo = None
        else:
            print("❌ Aucun chemin de logo fourni pour notes")
        
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Normal'],
            fontSize=9,
            alignment=1,
            spaceAfter=0,
            leading=10
        )
        
        small_header_style = ParagraphStyle(
            'SmallHeader',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            spaceAfter=0,
            leading=8
        )
        
        # Structure avec valeurs DYNAMIQUES
        left_col_content = [
            Paragraph("<b>MINISTÈRE DE L'EDUCATION NATIONALE</b>", header_style),
            Paragraph("-----------", small_header_style),
            Paragraph(f"DIRECTION RÉGIONALE DE L'ÉDUCATION - {dre}", header_style),
            Paragraph("-----------", small_header_style),
            Paragraph(f"INSPECTION DE L'ENSEIGNEMENT SECONDAIRE GÉNÉRAL - {inspection}", header_style)
        ]
        
        right_col_content = [
            Paragraph("<b>RÉPUBLIQUE TOGOLAISE</b>", header_style),
            Paragraph("-----------", small_header_style),
            Paragraph("Travail - Liberté - Patrie", ParagraphStyle(
                'DeviseStyle',
                parent=styles['Normal'],
                fontSize=7,
                alignment=1,
                spaceAfter=0,
                leading=8
            ))
        ]
        
        # Colonne centrale avec logo
        center_col_content = []
        if logo:
            center_col_content.append(logo)
        else:
            # Si pas de logo, ajouter un espace vide pour l'équilibre
            center_col_content.append(Spacer(1, 35*mm))
            
        center_col_content.extend([
            Paragraph(f"<b>{nom_ecole}</b>", header_style),
            Paragraph(f"Tél: {telephone}", small_header_style),
            Paragraph(f"{devise_ecole}", small_header_style)
        ])

        header_table = Table([[
            left_col_content, 
            center_col_content, 
            right_col_content
        ]], colWidths=[doc.width/3, doc.width/3, doc.width/3])
        
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ]))
        
        elements.append(header_table)
        elements.append(Spacer(1, 6*mm))
        
        separation_line = Table([['']], colWidths=[doc.width])
        separation_line.setStyle(TableStyle([
            ('LINEABOVE', (0, 0), (0, 0), 1, colors.black),
        ]))
        elements.append(separation_line)
        elements.append(Spacer(1, 6*mm))
        
        # ========== TITRE DU DOCUMENT ==========
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=14,
            alignment=1,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=6
        )
        
        titre_principal = "LISTE DES NOTES"
        title = Paragraph(titre_principal, title_style)
        elements.append(title)
        
        if classe_nom:
            classe_style = ParagraphStyle(
                'ClasseStyle',
                parent=styles['Normal'],
                fontSize=12,
                alignment=1,
                textColor=colors.HexColor('#2C3E50'),
                spaceAfter=8
            )
            classe_text = Paragraph(f"<b>Classe: {classe_nom}</b>", classe_style)
            elements.append(classe_text)
        
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            textColor=colors.gray,
            spaceAfter=8
        )
        
        date_text = Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style)
        elements.append(date_text)
        
        # ========== TABLEAU DES NOTES ==========
        if notes:
            # AJOUT DE LA COLONNE N° D'ORDRE
            headers = [
                'N°',  # NOUVELLE COLONNE
                'Élève',
                'Matière',
                'Note 1',
                'Note 2', 
                'Note 3',
                'Note Comp.',
                'Coefficient',
                'Trimestre',
                'Année'
            ]
            
            data = [headers]
            # Style pour le wrapping du texte
            cell_style = ParagraphStyle(
                'CellStyle',
                parent=styles['Normal'],
                fontSize=7,
                leading=8,
                wordWrap='CJK'  # Permet le wrapping du texte
            )
            # AJOUT DU NUMÉRO D'ORDRE POUR CHAQUE NOTE
            for index, note in enumerate(notes, 1):
                # Utiliser Paragraph pour permettre le wrapping
                row = [
                    Paragraph(str(index), cell_style),  # NUMÉRO D'ORDRE
                    Paragraph(f"{note.eleve.nom} {note.eleve.prenoms}" if note.eleve else "N/A", cell_style),
                    Paragraph(note.matiere.libelle if note.matiere else "N/A", cell_style),
                    Paragraph(str(note.note1) if note.note1 is not None else "-", cell_style),
                    Paragraph(str(note.note2) if note.note2 is not None else "-", cell_style),
                    Paragraph(str(note.note3) if note.note3 is not None else "-", cell_style),
                    Paragraph(str(note.note_comp) if note.note_comp is not None else "-", cell_style),
                    Paragraph(str(note.coefficient) if note.coefficient else "1", cell_style),
                    Paragraph(str(note.trimestre), cell_style),
                    Paragraph(note.annee_scolaire or "-", cell_style)
                ]
                data.append(row)
            
            # Convertir aussi les en-têtes en Paragraph pour la cohérence
            data[0] = [Paragraph(header, cell_style) for header in headers]
            
            # Largeurs de colonnes optimisées
            col_widths = [
                12*mm,  # N° 
                45*mm,  # Élève - plus large
                40*mm,  # Matière - plus large
                15*mm,  # Note 1
                15*mm,  # Note 2
                15*mm,  # Note 3
                18*mm,  # Note Comp.
                18*mm,  # Coefficient
                15*mm,  # Trimestre
                20*mm   # Année
            ]
            
            # Créer le tableau avec hauteur de ligne automatique
            table = Table(data, repeatRows=1, colWidths=col_widths)
            
            table_style = TableStyle([
                # Style pour l'en-tête
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                
                # Style pour les données
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Alignement en haut pour le wrapping
                
                # Alignement spécifique
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # N°
                ('ALIGN', (1, 0), (2, -1), 'LEFT'),    # Élève et Matière
                ('ALIGN', (3, 0), (9, -1), 'CENTER'),  # Toutes les autres colonnes
                
                # Bordures
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                
                # Alternance des couleurs
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
                
                # Padding
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ])
            
            table.setStyle(table_style)
            elements.append(table)
        else:
            no_data_style = ParagraphStyle(
                'NoDataStyle',
                parent=styles['Normal'],
                fontSize=12,
                alignment=1,
                textColor=colors.HexColor('#666666'),
                spaceAfter=20
            )
            no_data = Paragraph("Aucune note à exporter", no_data_style)
            elements.append(no_data)
        
        elements.append(Spacer(1, 8*mm))
        
        stats_style = ParagraphStyle(
            'StatsStyle',
            parent=styles['Normal'],
            fontSize=8,
            alignment=0,
            textColor=colors.gray,
            spaceAfter=1
        )
        
        total_notes = len(notes)
        
        stats_text = [
            Paragraph(f"<b>STATISTIQUES :</b>", stats_style),
            Paragraph(f"Total notes : {total_notes}", stats_style),
        ]
        
        for element in stats_text:
            elements.append(element)
        
        doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
        buffer.seek(0)
        
        filename = f"liste_notes_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        
        response = send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
        
    except Exception as e:
        print(f"Erreur génération PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        raise

def generate_excel_export(notes, classe_nom=None):
    """Génère un export Excel des notes avec en-tête uniformisé - VERSION COMPLÈTE CORRIGÉE"""
    try:
        ecole_id = get_current_ecole_id()
        ecole = Ecole.query.get(ecole_id)
        
        if not ecole:
            raise ValueError("École non trouvée")
        
        # CORRECTION: Récupérer le logo pour Excel
        logo_path = get_logo_path(ecole)
        print(f"📊 Excel notes - Chemin logo: {logo_path}")

        nom_ecole = ecole.nom if ecole.nom else "École non renseignée"
        dre = ecole.dre if ecole.dre else "DRE non renseignée"
        inspection = ecole.inspection if ecole.inspection else "Inspection non renseignée"
        telephone = ecole.telephone1 if ecole.telephone1 else "Téléphone non renseigné"
        devise_ecole = ecole.devise if ecole.devise else "Travail - Liberté - Patrie"
        
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Liste des notes"
        
        center_align = Alignment(horizontal="center", vertical="center")
        
        # ========== EN-TÊTE EXCEL UNIFORMISÉ AVEC LOGO ==========
        current_row = 1
        
        # Structure d'en-tête textuelle
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = "MINISTÈRE DE L'EDUCATION NATIONALE"
        ws[f'A{current_row}'].font = Font(bold=True, size=10)
        ws[f'A{current_row}'].alignment = center_align
        
        ws.merge_cells(f'G{current_row}:J{current_row}')
        ws[f'G{current_row}'] = "RÉPUBLIQUE TOGOLAISE"
        ws[f'G{current_row}'].font = Font(bold=True, size=10)
        ws[f'G{current_row}'].alignment = center_align
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = "-----------"
        ws[f'A{current_row}'].alignment = center_align
        
        ws.merge_cells(f'G{current_row}:J{current_row}')
        ws[f'G{current_row}'] = "-----------"
        ws[f'G{current_row}'].alignment = center_align
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = f"DIRECTION RÉGIONALE DE L'ÉDUCATION - {dre}"
        ws[f'A{current_row}'].font = Font(size=9)
        ws[f'A{current_row}'].alignment = center_align
        
        ws.merge_cells(f'G{current_row}:J{current_row}')
        ws[f'G{current_row}'] = "Travail - Liberté - Patrie"
        ws[f'G{current_row}'].font = Font(bold=True, size=8)
        ws[f'G{current_row}'].alignment = center_align
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = "-----------"
        ws[f'A{current_row}'].alignment = center_align
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = f"INSPECTION DE L'ENSEIGNEMENT SECONDAIRE GÉNÉRAL - {inspection}"
        ws[f'A{current_row}'].font = Font(size=9)
        ws[f'A{current_row}'].alignment = center_align
        
        # Nom de l'école et informations
        current_row += 1
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'] = nom_ecole
        ws[f'A{current_row}'].font = Font(bold=True, size=14, color="2C3E50")
        ws[f'A{current_row}'].alignment = center_align
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'] = f"Tél: {telephone} - {devise_ecole}"
        ws[f'A{current_row}'].alignment = center_align
        
        # Titre du document
        titre_principal = "LISTE DES NOTES"
        if classe_nom:
            titre_principal += f" - {classe_nom}"
            
        current_row += 1
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'] = titre_principal
        ws[f'A{current_row}'].font = Font(bold=True, size=16, color="2C3E50")
        ws[f'A{current_row}'].alignment = center_align
        
        # Date de génération
        current_row += 1
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws[f'A{current_row}'].font = Font(size=8, italic=True, color="666666")
        ws[f'A{current_row}'].alignment = center_align
        
        # Ligne vide
        current_row += 1
        ws[f'A{current_row}'] = ""
        
        # CORRECTION: Ajouter le logo APRÈS la structure d'en-tête
        if logo_path and os.path.exists(logo_path):
            try:
                from openpyxl.drawing.image import Image as ExcelImage
                img = ExcelImage(logo_path)
                
                # Redimensionner le logo pour qu'il soit discret
                img.width = 50
                img.height = 50
                
                # Placer le logo en dehors de la zone de texte (colonne K, ligne 1)
                img.anchor = 'K1'
                ws.add_image(img)
                print("✅ Logo ajouté discrètement dans Excel notes (colonne K)")
                
                # Ajuster la largeur de la colonne K pour accommoder le logo
                ws.column_dimensions['K'].width = 15
                
            except Exception as e:
                print(f"❌ Erreur ajout logo Excel notes: {e}")
        
        # ========== TABLEAU DES DONNÉES ==========
        # AJOUT DE LA COLONNE N° D'ORDRE
        headers = [
            'N°',  # NOUVELLE COLONNE
            'Élève', 
            'Matière', 
            'Note 1', 
            'Note 2', 
            'Note 3',
            'Note Composition', 
            'Coefficient', 
            'Trimestre', 
            'Année Scolaire'
        ]
        
        # Commencer après l'en-tête
        start_row = current_row + 1
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = center_align
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
        
        # AJOUT DU NUMÉRO D'ORDRE POUR CHAQUE NOTE
        for note_index, note in enumerate(notes, start=1):
            row_num = start_row + note_index
            row_data = [
                note_index,  # NUMÉRO D'ORDRE
                f"{note.eleve.nom} {note.eleve.prenoms}" if note.eleve else "N/A",
                note.matiere.libelle if note.matiere else "N/A",
                note.note1 if note.note1 is not None else "",
                note.note2 if note.note2 is not None else "",
                note.note3 if note.note3 is not None else "",
                note.note_comp if note.note_comp is not None else "",
                note.coefficient if note.coefficient else 1,
                note.trimestre,
                note.annee_scolaire or ""
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = value
                cell.border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
                
                # Centrer les colonnes numériques et la colonne N°
                if col in [1, 4, 5, 6, 7, 8, 9]:
                    cell.alignment = center_align
        
        # Ajuster les largeurs de colonnes avec la nouvelle colonne N°
        column_widths = {
            'A': 6,    # Colonne N° - plus étroite
            'B': 25,   # Élève
            'C': 20,   # Matière
            'D': 8,    # Note 1
            'E': 8,    # Note 2
            'F': 8,    # Note 3
            'G': 12,   # Note Composition
            'H': 12,   # Coefficient
            'I': 10,   # Trimestre
            'J': 12    # Année Scolaire
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # ========== STATISTIQUES ==========
        stats_row = len(notes) + start_row + 2
        total_notes = len(notes)
        
        ws.merge_cells(f'A{stats_row}:J{stats_row}')
        ws[f'A{stats_row}'] = "STATISTIQUES"
        ws[f'A{stats_row}'].font = Font(bold=True, size=10)
        
        ws.merge_cells(f'A{stats_row + 1}:J{stats_row + 1}')
        ws[f'A{stats_row + 1}'] = f"Total notes exportées : {total_notes}"
        ws[f'A{stats_row + 1}'].font = Font(size=9)
        
        # Sauvegarder
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"liste_notes_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        response = send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
        
    except Exception as e:
        print(f"❌ Erreur génération Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        raise

# ========== ROUTES ==========

@notes_bp.route('/')
@login_required
@ecole_required
@validate_pagination_params
def liste_notes():
    ecole_id = get_current_ecole_id()
    data = request.args

    per_page = getattr(request, 'validated_per_page', SecurityConfig.DEFAULT_PER_PAGE)
    page = getattr(request, 'validated_page', 1)
    
    search = safe_search_term(data.get("search", "", type=str))
    eleve_id = data.get("eleve_id", type=str)
    matiere_id = data.get("matiere_id", type=str)
    enseignant_id = data.get("enseignant_id", type=str)
    classe_id = data.get("classe_id", type=str)
    
    # CORRECTION: Récupérer trimestre correctement
    trimestre_str = data.get("trimestre", "")
    if trimestre_str and trimestre_str.strip():
        try:
            trimestre = int(trimestre_str)
        except ValueError:
            trimestre = None
    else:
        trimestre = None  # "Tous" est sélectionné
    
    annee_scolaire = data.get("annee_scolaire", type=str)
    cycle_type = data.get("cycle_type", "college")

    print(f"🔍 Paramètres reçus - cycle_type: {cycle_type}, trimestre: {trimestre}, annee_scolaire: {annee_scolaire}")

    if not annee_scolaire or not validate_annee_scolaire(annee_scolaire):
        today = date.today()
        if today.month >= 8:
            annee_scolaire = f"{today.year}-{today.year + 1}"
        else:
            annee_scolaire = f"{today.year - 1}-{today.year}"

    log_security_event(
        current_user.id,
        "LIST_NOTES",
        "notes",
        "SUCCESS",
        {
            "search_term": search[:50] if search else None,
            "filters_applied": {
                "has_eleve_filter": bool(eleve_id),
                "has_matiere_filter": bool(matiere_id),
                "has_classe_filter": bool(classe_id),
                "trimestre": trimestre,
                "annee_scolaire": annee_scolaire
            }
        }
    )

    query = Note.query.join(Eleve).join(Classe).options(
        joinedload(Note.eleve),
        joinedload(Note.matiere),
        joinedload(Note.enseignement).joinedload(Enseignement.enseignant).joinedload(Enseignant.utilisateur)
    ).filter(Classe.ecole_id == ecole_id)

    # Filtrage par classe
    if classe_id and classe_id.strip() and classe_id.lower() != "none":
        try:
            classe_uuid = UUID(classe_id)
            classe = Classe.query.filter_by(id=classe_uuid, ecole_id=ecole_id).first()
            if classe:
                query = query.filter(Eleve.classe_id == classe_uuid)
        except ValueError:
            pass

    # Recherche texte sécurisée
    if search:
        query = query.filter(
            (Note.note1.cast(String).ilike(f"%{search}%")) |
            (Note.note2.cast(String).ilike(f"%{search}%")) |
            (Note.note3.cast(String).ilike(f"%{search}%")) |
            (Note.note_comp.cast(String).ilike(f"%{search}%")) |
            (cast(Note.coefficient, String).ilike(f"%{search}%")) |
            (cast(Note.date_saisie, String).ilike(f"%{search}%"))|
            (Eleve.nom.ilike(f"%{search}%")) |
            (Eleve.prenoms.ilike(f"%{search}%")) |
            (Note.etat.ilike(f"%{search}%"))
        )

    # Filtrage sécurisé par élève
    if eleve_id and eleve_id.lower() != "none":
        try:
            eleve_uuid = UUID(eleve_id)
            eleve = get_eleve_secure(eleve_id, ecole_id)
            if eleve:
                query = query.filter(Note.eleve_id == eleve_uuid)
        except ValueError:
            pass

    # Filtrage par matière
    if matiere_id and matiere_id.lower() != "none":
        try:
            matiere_uuid = UUID(matiere_id)
            query = query.filter(Note.matiere_id == matiere_uuid)
        except ValueError:
            pass

    # Filtrage par enseignant
    if enseignant_id and enseignant_id.lower() != "none":
        try:
            enseignant_uuid = UUID(enseignant_id)
            query = query.join(Note.enseignement).filter(Enseignement.enseignant_id == enseignant_uuid)
        except ValueError:
            pass
    
    if trimestre is not None:
        query = query.filter(Note.trimestre == trimestre)
    
    if annee_scolaire:
        query = query.filter(Note.annee_scolaire == annee_scolaire)
    
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    notes = pagination.items

    user_role = (getattr(current_user, "role", "guest") or "guest").lower()
    
    eleves = Eleve.query.join(Classe).filter(Classe.ecole_id == ecole_id).all()
    
    # =============================================
    # CORRECTION : Utiliser get_unique_matieres() pour supprimer les doublons
    # =============================================
    # Cette fonction retourne une structure hiérarchique sans doublons
    matieres_hierarchiques = get_unique_matieres(ecole_id)
    
    # Transformer en liste plate pour le select (car le template attend une liste simple)
    matieres = []
    for groupe in matieres_hierarchiques:
        if groupe.get('children'):
            for enfant in groupe['children']:
                matieres.append({
                    'id': enfant['id'],
                    'libelle': enfant['libelle'],
                    'type': enfant['type']
                })
    # =============================================
    # FIN DE LA CORRECTION
    # =============================================
    
    enseignants = Enseignant.query.filter_by(ecole_id=ecole_id).options(joinedload(Enseignant.utilisateur)).all()
    
    classes = Classe.query.filter_by(ecole_id=ecole_id).all()
    
    types = Matiere.query.filter(
        Matiere.parent_id.is_(None),
        Matiere.type.in_(SecurityConfig.ALLOWED_MATIERE_TYPES)
    ).options(joinedload(Matiere.children)).all()

    annees_scolaires = db.session.query(Note.annee_scolaire).join(Eleve).join(Classe).filter(
        Classe.ecole_id == ecole_id
    ).distinct().all()
    annees_scolaires = [a[0] for a in annees_scolaires if validate_annee_scolaire(a[0])]
    
    return render_template("notes/list_notes.html",
        notes=notes,
        pagination=pagination,
        search=search,
        per_page=per_page,
        user_role=user_role,
        eleves=eleves,
        matieres=matieres,
        enseignants=enseignants,
        classes=classes,
        types=types,
        eleve_id=eleve_id,
        matiere_id=matiere_id,
        enseignant_id=enseignant_id,
        classe_id=classe_id,
        trimestre=trimestre,
        annee_scolaire=annee_scolaire,
        annees_scolaires=annees_scolaires,
        ecole_id=ecole_id,
        cycle_type=cycle_type,
        cycles=get_ecole_cycles(ecole_id)
    )

@notes_bp.route("/add", methods=["POST"])
@login_required
@ecole_required
def add_note():
    try:
        ecole_id = get_current_ecole_id()
        data = request.json
        
        required_fields = ['eleve_id', 'matiere_id', 'enseignant_id', 'trimestre', 'annee_scolaire']
        for field in required_fields:
            if not data.get(field):
                return jsonify({"error": f"Le champ {field} est requis"}), 400

        annee_scolaire_value = data.get("annee_scolaire")
        if not validate_annee_scolaire(annee_scolaire_value):
            return jsonify({"error": "Format d'année scolaire invalide"}), 400

        trimestre_value = data.get("trimestre")
        cycle_type = data.get("cycle_type", "college")  # Nouveau
        
        if trimestre_value is None:
            return jsonify({"error": "Le trimestre/semestre est requis"}), 400

        try:
            trimestre_value = int(trimestre_value)
            if not validate_trimestre_for_cycle(trimestre_value, cycle_type):
                if cycle_type == 'lycee':
                    return jsonify({"error": "Semestre invalide. Doit être 1 ou 2"}), 400
                else:
                    return jsonify({"error": "Trimestre invalide. Doit être 1, 2 ou 3"}), 400
        except (ValueError, TypeError):
            return jsonify({"error": "Format de trimestre/semestre invalide"}), 400

        eleve = get_eleve_secure(data.get("eleve_id"), ecole_id)
        
        if not eleve or not eleve.classe_id:
            return jsonify({"error": "Élève ou classe introuvable"}), 400

        try:
            matiere_uuid = UUID(data.get("matiere_id"))
            enseignant_uuid = UUID(data.get("enseignant_id"))
        except ValueError:
            return jsonify({"error": "IDs invalides"}), 400

        note1 = validate_note_value(data.get("note1"))
        note2 = validate_note_value(data.get("note2"))
        note3 = validate_note_value(data.get("note3"))
        note_comp = validate_note_value(data.get("note_comp"))

        notes_fournies = [note1, note2, note3, note_comp]
        if all(note is None for note in notes_fournies):
            return jsonify({"error": "Au moins une note doit être fournie"}), 400

        note_existante = Note.query.join(Eleve).join(Classe).filter(
            Note.eleve_id == eleve.id,
            Note.matiere_id == matiere_uuid,
            Note.trimestre == trimestre_value,
            Note.annee_scolaire == annee_scolaire_value,
            Classe.ecole_id == ecole_id
        ).first()

        if note_existante:
            conflits = []
            
            if note1 is not None and note_existante.note1 is not None:
                conflits.append("note1")
            if note2 is not None and note_existante.note2 is not None:
                conflits.append("note2")
            if note3 is not None and note_existante.note3 is not None:
                conflits.append("note3")
            if note_comp is not None and note_existante.note_comp is not None:
                conflits.append("note de composition")

            if conflits:
                return jsonify({
                    "error": f"Les notes suivantes existent déjà pour cet élève : {', '.join(conflits)}. Veuillez modifier la note existante."
                }), 400
            
            if note1 is not None:
                note_existante.note1 = note1
            if note2 is not None:
                note_existante.note2 = note2
            if note3 is not None:
                note_existante.note3 = note3
            if note_comp is not None:
                note_existante.note_comp = note_comp
            
            if data.get("coefficient"):
                try:
                    coefficient = int(data.get("coefficient", 1))
                    if coefficient >= 1:
                        note_existante.coefficient = coefficient
                except (ValueError, TypeError):
                    pass
            
            note_existante.date_saisie = date.today()
            
            db.session.commit()

            log_security_event(
                current_user.id,
                "UPDATE_NOTE",
                f"note/{note_existante.id}",
                "SUCCESS",
                {
                    "eleve_id": str(eleve.id),
                    "matiere_id": str(matiere_uuid),
                    "trimestre": trimestre_value,
                    "annee_scolaire": annee_scolaire_value,
                    "notes_ajoutees": [note for note in ['note1', 'note2', 'note3', 'note_comp'] 
                                     if data.get(note) is not None]
                }
            )

            note_finale = Note.query.options(
                joinedload(Note.eleve),
                joinedload(Note.matiere),
                joinedload(Note.enseignement).joinedload(Enseignement.enseignant).joinedload(Enseignant.utilisateur)
            ).get(note_existante.id)

            row_html = render_template("notes/_note_row.html", note=note_finale)
            return jsonify({
                "message": "Note mise à jour avec succès",
                "note_html": row_html,
                "note_id": str(note_finale.id),
                "action": "updated"
            }), 200

        enseignement = Enseignement.query.join(Classe).filter(
            Enseignement.matiere_id == matiere_uuid,
            Enseignement.enseignant_id == enseignant_uuid,
            Enseignement.classe_id == eleve.classe_id,
            Classe.ecole_id == ecole_id
        ).first()

        if not enseignement:
            if not Enseignant.query.get(enseignant_uuid):
                return jsonify({"error": "Enseignant introuvable"}), 400
            
            classe_ecole = Classe.query.filter_by(id=eleve.classe_id).first()
            if not classe_ecole or not classe_ecole.ecole_id:
                return jsonify({"error": "Classe sans école associée"}), 400
                
            enseignement = Enseignement(
                id=uuid.uuid4(),
                matiere_id=matiere_uuid,
                enseignant_id=enseignant_uuid,
                classe_id=eleve.classe_id,
                ecole_id=classe_ecole.ecole_id
            )
            db.session.add(enseignement)
            db.session.flush()

        try:
            coefficient = int(data.get("coefficient", 1))
            if coefficient < 1:
                coefficient = 1
        except (ValueError, TypeError):
            coefficient = 1

        nouvelle_note = Note(
            id=uuid.uuid4(),
            note1=note1,
            note2=note2,
            note3=note3,
            note_comp=note_comp,
            coefficient=coefficient,
            date_saisie=date.today(),
            eleve_id=eleve.id,
            matiere_id=matiere_uuid,
            enseignement_id=enseignement.id,
            trimestre=trimestre_value,
            annee_scolaire=annee_scolaire_value,
            ecole_id=ecole_id,
            etat="Actif"
        )
        db.session.add(nouvelle_note)
        db.session.commit()

        log_security_event(
            current_user.id,
            "CREATE_NOTE",
            f"note/{nouvelle_note.id}",
            "SUCCESS",
            {
                "eleve_id": str(eleve.id),
                "matiere_id": str(matiere_uuid),
                "trimestre": trimestre_value,
                "annee_scolaire": annee_scolaire_value
            }
        )

        note_finale = Note.query.options(
            joinedload(Note.eleve),
            joinedload(Note.matiere),
            joinedload(Note.enseignement).joinedload(Enseignement.enseignant).joinedload(Enseignant.utilisateur)
        ).get(nouvelle_note.id)

        row_html = render_template("notes/_note_row.html", note=note_finale)
        return jsonify({
            "message": "Note enregistrée avec succès",
            "note_html": row_html,
            "note_id": str(note_finale.id),
            "action": "created"
        }), 201

    except Exception as e:
        db.session.rollback()
        log_security_event(
            current_user.id,
            "CREATE_NOTE",
            "notes",
            "FAILED",
            {"error": str(e)}
        )
        return jsonify({"error": f"Erreur lors de la création de la note: {str(e)}"}), 500

@notes_bp.route("/<string:note_id>/edit", methods=["PUT"])
@login_required
@ecole_required
def edit_note(note_id):
    ecole_id = get_current_ecole_id()
    
    note = get_note_secure(note_id, ecole_id)
    
    data = request.json
    if note.cloture:
        return jsonify({"error": "Ce trimestre est clôturé, modification impossible"}), 400

    try:
        def validate_note_value(note_value):
            if note_value is None:
                return None
            try:
                note_float = float(note_value)
                if 0 <= note_float <= 20:
                    return note_float
                return None
            except (ValueError, TypeError):
                return None

        if "note1" in data:
            note.note1 = validate_note_value(data.get("note1"))
        if "note2" in data:
            note.note2 = validate_note_value(data.get("note2"))
        if "note3" in data:
            note.note3 = validate_note_value(data.get("note3"))
        if "note_comp" in data:
            note.note_comp = validate_note_value(data.get("note_comp"))
            
        if "coefficient" in data:
            try:
                coefficient = int(data.get("coefficient"))
                if coefficient >= 1:
                    note.coefficient = coefficient
            except (ValueError, TypeError):
                pass

        if "trimestre" in data and validate_trimestre(data.get("trimestre")):
            note.trimestre = data.get("trimestre")
            
        if "annee_scolaire" in data and validate_annee_scolaire(data.get("annee_scolaire")):
            note.annee_scolaire = data.get("annee_scolaire")
            
        if "etat" in data:
            note.etat = data.get("etat")
            
        if "enseignant_id" in data and data.get("enseignant_id"):
            try:
                enseignant_uuid = UUID(data.get("enseignant_id"))
                enseignant = Enseignant.query.filter_by(id=enseignant_uuid, ecole_id=ecole_id).first()
                if enseignant:
                    enseignement = Enseignement.query.filter_by(
                        matiere_id=note.matiere_id,
                        classe_id=note.eleve.classe_id,
                        ecole_id=ecole_id
                    ).first()
                    
                    if enseignement:
                        enseignement.enseignant_id = enseignant_uuid
                    else:
                        enseignement = Enseignement(
                            id=uuid.uuid4(),
                            matiere_id=note.matiere_id,
                            enseignant_id=enseignant_uuid,
                            classe_id=note.eleve.classe_id,
                            ecole_id=ecole_id
                        )
                        db.session.add(enseignement)
                        db.session.flush()
                    
                    note.enseignement_id = enseignement.id
            except ValueError:
                pass
            
        note.date_saisie = date.today()
        
        db.session.commit()

        log_security_event(
            current_user.id,
            "UPDATE_NOTE",
            f"note/{note_id}",
            "SUCCESS"
        )

        user_role = (getattr(current_user, "role", "guest") or "guest").lower()
        
        row_html = render_template("notes/_note_row.html", note=note, user_role=user_role)
        return jsonify({"message": "Note mise à jour avec succès", "note_html": row_html}), 200
    except Exception as e:
        db.session.rollback()
        log_security_event(
            current_user.id,
            "UPDATE_NOTE",
            f"note/{note_id}",
            "FAILED",
            {"error": str(e)}
        )
        return jsonify({"error": "Erreur lors de la mise à jour de la note"}), 500

@notes_bp.route("/list_elements", methods=["GET"])
@login_required
@ecole_required
def list_elements():
    ecole_id = get_current_ecole_id()
    
    try:
        eleves = Eleve.query.join(Classe).filter(Classe.ecole_id == ecole_id).all()
        classes = Classe.query.filter_by(ecole_id=ecole_id).all()
        enseignants = Enseignant.query.filter_by(ecole_id=ecole_id).options(joinedload(Enseignant.utilisateur)).all()
        
        # UTILISER LA NOUVELLE FONCTION DE DÉDUPLICATION
        matieres_hierarchiques = get_unique_matieres(ecole_id)
        
        data = {
            "eleves": [
                {
                    "id": str(e.id), 
                    "nom": e.nom or "", 
                    "prenoms": e.prenoms or "", 
                    "classe_id": str(e.classe_id) if e.classe_id else ""
                }
                for e in eleves
            ],
            "matieres": matieres_hierarchiques,
            "classes": [
                {
                    "id": str(c.id), 
                    "nom": c.nom or ""
                } 
                for c in classes
            ],
            "enseignants": [
                {
                    "id": str(e.id),
                    "noms": e.utilisateur.nom if e.utilisateur else e.nom or "",
                    "prenoms": e.utilisateur.prenoms if e.utilisateur else e.prenoms or ""
                }
                for e in enseignants
            ]
        }
        
        return jsonify(data)
        
    except Exception as e:
        print(f"❌ ERREUR CRITIQUE dans list_elements: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Erreur lors du chargement des données: {str(e)}"}), 500

@notes_bp.route("/<string:note_id>", methods=["GET"])
@login_required
@ecole_required
def get_notes(note_id):
    ecole_id = get_current_ecole_id()
    
    note = get_note_secure(note_id, ecole_id)
    
    if not note:
        return jsonify({"error": "Note non trouvée"}), 404

    enseignant_nom = ""
    enseignant_prenoms = ""
    
    if note.enseignement and note.enseignement.enseignant:
        if note.enseignement.enseignant.utilisateur:
            enseignant_nom = note.enseignement.enseignant.utilisateur.nom or ""
            enseignant_prenoms = note.enseignement.enseignant.utilisateur.prenoms or ""
        else:
            enseignant_nom = getattr(note.enseignement.enseignant, 'nom', '') or getattr(note.enseignement.enseignant, 'noms', '')
            enseignant_prenoms = getattr(note.enseignement.enseignant, 'prenoms', '')
    
    enseignants_ecole = Enseignant.query.filter_by(ecole_id=ecole_id).options(joinedload(Enseignant.utilisateur)).all()
    
    return jsonify({
        "id": str(note.id),
        "note1": note.note1,
        "note2": note.note2,
        "note3": note.note3,
        "note_comp": note.note_comp,
        "coefficient": note.coefficient,
        "eleve_id": str(note.eleve_id),
        "matiere_id": str(note.matiere_id),
        "enseignant_id": str(note.enseignement.enseignant_id) if note.enseignement else None,
        "etat": note.etat,
        "date": note.date_saisie.isoformat() if note.date_saisie else None,
        "eleve_nom": f"{note.eleve.nom} {note.eleve.prenoms}" if note.eleve else "",
        "matiere_nom": note.matiere.libelle if note.matiere else "",
        "enseignant_nom": enseignant_nom,
        "enseignant_prenoms": enseignant_prenoms,
        "trimestre": note.trimestre,
        "annee_scolaire":note.annee_scolaire,
        "valeur": note.note1 or note.note2 or note.note3 or note.note_comp,
        "enseignants_ecole": [
            {
                "id": str(ens.id),
                "nom_complet": f"{ens.utilisateur.nom} {ens.utilisateur.prenoms}" if ens.utilisateur else f"{ens.nom} {ens.prenoms}"
            }
            for ens in enseignants_ecole
        ]
    })

@notes_bp.route("/<string:note_id>/delete", methods=["DELETE"])
@login_required
@ecole_required
def delete_note(note_id):
    ecole_id = get_current_ecole_id()
    
    note = get_note_secure(note_id, ecole_id)
    
    try:
        db.session.delete(note)
        db.session.commit()
        
        log_security_event(
            current_user.id,
            "DELETE_NOTE",
            f"note/{note_id}",
            "SUCCESS"
        )
        
        return jsonify({"message": "Note supprimée avec succès"})
    except Exception as e:
        db.session.rollback()
        log_security_event(
            current_user.id,
            "DELETE_NOTE",
            f"note/{note_id}",
            "FAILED",
            {"error": str(e)}
        )
        return jsonify({"error": "Erreur lors de la suppression de la note"}), 500

@notes_bp.route("/clotures/actifs", methods=["GET"])
@login_required
@ecole_required
def clotures_actifs():
    ecole_id = get_current_ecole_id()
    
    if getattr(current_user, "role", "guest").lower() != "admin":
        return jsonify({"error": "Non autorisé"}), 403

    actifs = db.session.query(Note.annee_scolaire, Note.trimestre).join(Eleve).join(Classe).filter(
        Note.cloture == False,
        Classe.ecole_id == ecole_id
    ).distinct().all()

    return jsonify([
        {"annee": a, "trimestre": t} for a, t in actifs
    ])

@notes_bp.route("/clotures/close", methods=["POST"])
@login_required
@ecole_required
def cloturer_periode():
    ecole_id = get_current_ecole_id()
    
    if getattr(current_user, "role", "guest").lower() != "admin":
        return jsonify({"error": "Non autorisé"}), 403

    data = request.json
    annee = data.get("annee")
    trimestre = data.get("trimestre")
    cycle_type = data.get("cycle_type", "college")  # Nouveau

    if not annee or not validate_annee_scolaire(annee):
        return jsonify({"error": "Année scolaire invalide"}), 400

    if trimestre and not validate_trimestre_for_cycle(trimestre, cycle_type):
        return jsonify({"error": f"{'Semestre' if cycle_type == 'lycee' else 'Trimestre'} invalide"}), 400

    today = date.today()
    current_annee = f"{today.year if today.month >= 8 else today.year-1}-{today.year+1 if today.month >= 8 else today.year}"
    
    # Déterminer la période courante selon le cycle
    if cycle_type == 'lycee':
        current_periode = 1 if today.month in [9, 10, 11, 12, 1, 2] else 2
    else:
        current_periode = ((today.month - 8) // 3 + 1) if today.month >= 8 else ((today.month + 4) // 3)

    if annee == current_annee and (trimestre is None or trimestre >= current_periode):
        periode_label = "semestre" if cycle_type == 'lycee' else "trimestre"
        return jsonify({"error": f"Impossible de clôturer une période en cours ou future"}), 400

    try:
        if trimestre is None:
            notes = Note.query.join(Eleve).join(Classe).filter_by(
                annee_scolaire=annee, 
                cloture=False
            ).filter(Classe.ecole_id == ecole_id).all()
        else:
            notes = Note.query.join(Eleve).join(Classe).filter_by(
                annee_scolaire=annee, 
                trimestre=trimestre, 
                cloture=False
            ).filter(Classe.ecole_id == ecole_id).all()
            
        for n in notes:
            n.cloture = True
            n.etat = "Clôturé"
        db.session.commit()
        
        periode_label = "semestre" if cycle_type == 'lycee' else "trimestre"
        log_security_event(
            current_user.id,
            "CLOSE_PERIOD",
            "notes",
            "SUCCESS",
            {"annee": annee, "periode": trimestre, "cycle": cycle_type, "notes_affected": len(notes)}
        )
        
        return jsonify({"message": f"Période {annee} - {periode_label.capitalize()} {trimestre or 'Tous'} clôturée avec succès"}), 200
    except Exception as e:
        db.session.rollback()
        log_security_event(
            current_user.id,
            "CLOSE_PERIOD",
            "notes",
            "FAILED",
            {"error": str(e), "annee": annee, "trimestre": trimestre, "cycle": cycle_type}
        )
        return jsonify({"error": "Erreur lors de la clôture de la période"}), 500

@notes_bp.route("/annees/actives", methods=["GET"])
@login_required
@ecole_required
def annees_actives():
    ecole_id = get_current_ecole_id()
    today = date.today()
    current_year = today.year if today.month >= 8 else today.year - 1
    next_year = current_year + 1

    annees_possibles = [f"{current_year}-{current_year+1}", f"{next_year}-{next_year+1}"]

    actifs = db.session.query(Note.annee_scolaire).join(Eleve).join(Classe).filter(
        Note.cloture == False,
        Classe.ecole_id == ecole_id
    ).distinct().all()
    actifs = [a[0] for a in actifs if validate_annee_scolaire(a[0])]

    annees = [a for a in annees_possibles if a in actifs]
    if not annees:
        annees = [annees_possibles[0]]

    return jsonify(annees)

# ========== FONCTIONS UTILITAIRES POUR LES CYCLES ==========
def get_ecole_cycles(ecole_id):
    """Récupère les cycles disponibles pour l'école"""
    ecole = Ecole.query.get(ecole_id)
    if not ecole:
        return {'college': True, 'lycee': False}  # Valeur par défaut
    
    cycles = getattr(ecole, 'cycles_disponibles', {})
    return {
        'college': cycles.get('college', True),
        'lycee': cycles.get('lycee', False)
    }

def get_periodes_for_cycle(cycle_type, trimestre_semestre_value):
    """Retourne les périodes selon le cycle"""
    if cycle_type == 'lycee':
        return {
            'type': 'semestre',
            'options': [1, 2],
            'label': 'Semestre',
            'value': trimestre_semestre_value
        }
    else:  # collège par défaut
        return {
            'type': 'trimestre',
            'options': [1, 2, 3],
            'label': 'Trimestre',
            'value': trimestre_semestre_value
        }



@notes_bp.route('/export/<format>')
@login_required
@ecole_required
def export_notes(format):
    """Export des notes dans différents formats"""
    try:
        ecole_id = get_current_ecole_id()
        
        # Récupérer TOUS les paramètres
        search = request.args.get('search', '')
        eleve_id = request.args.get('eleve_id', '')
        matiere_id = request.args.get('matiere_id', '')
        classe_id = request.args.get('classe_id', '')
        trimestre = request.args.get('trimestre', '')
        annee_scolaire = request.args.get('annee_scolaire', '')
        
        # DEBUG EXPORT
        print(f"🔍 EXPORT - Paramètres reçus:")
        print(f"  - classe_id: {classe_id}")
        print(f"  - trimestre: {trimestre}")
        print(f"  - annee_scolaire: {annee_scolaire}")
        print(f"  - search: {search}")
        
        # Trouver le nom de la classe pour l'affichage
        classe_nom = "Toutes les classes"
        if classe_id and classe_id.strip() and classe_id.lower() != "none":
            try:
                classe_uuid = UUID(classe_id)
                classe = Classe.query.filter_by(id=classe_uuid, ecole_id=ecole_id).first()
                if classe:
                    classe_nom = classe.nom
                    print(f"✅ EXPORT - Classe trouvée: {classe_nom}")
                else:
                    print(f"❌ EXPORT - Classe non trouvée: {classe_id}")
            except ValueError as e:
                print(f"❌ EXPORT - Erreur UUID: {e}")
                classe_nom = "Classe inconnue"
        
        # Construction de la requête
        query = Note.query.join(Eleve).join(Classe).options(
            joinedload(Note.eleve),
            joinedload(Note.matiere),
            joinedload(Note.enseignement).joinedload(Enseignement.enseignant).joinedload(Enseignant.utilisateur)
        ).filter(Classe.ecole_id == ecole_id)
        
        # Appliquer les filtres
        if search and search.strip():
            query = query.filter(
                db.or_(
                    Eleve.nom.ilike(f'%{search}%'),
                    Eleve.prenoms.ilike(f'%{search}%'),
                    Matiere.libelle.ilike(f'%{search}%')
                )
            )
        
        if eleve_id and eleve_id.strip() and eleve_id.lower() != "none":
            try:
                eleve_uuid = UUID(eleve_id)
                query = query.filter(Note.eleve_id == eleve_uuid)
            except ValueError:
                pass
        
        if matiere_id and matiere_id.strip() and matiere_id.lower() != "none":
            try:
                matiere_uuid = UUID(matiere_id)
                query = query.filter(Note.matiere_id == matiere_uuid)
            except ValueError:
                pass
        
        # FILTRE CLASSE - CORRIGÉ
        if classe_id and classe_id.strip() and classe_id.lower() != "none":
            try:
                classe_uuid = UUID(classe_id)
                # Vérifier que la classe appartient à l'école
                classe = Classe.query.filter_by(id=classe_uuid, ecole_id=ecole_id).first()
                if classe:
                    query = query.filter(Eleve.classe_id == classe_uuid)
                    print(f"✅ EXPORT - Filtre classe appliqué: {classe.nom}")
                else:
                    print(f"❌ EXPORT - Classe non autorisée: {classe_id}")
            except ValueError as e:
                print(f"❌ EXPORT - UUID classe invalide: {e}")
        
        # Filtres trimestre et année
        if trimestre and trimestre.strip():
            try:
                trimestre_int = int(trimestre)
                if validate_trimestre(trimestre_int):
                    query = query.filter(Note.trimestre == trimestre_int)
                    print(f"✅ EXPORT - Filtre trimestre: {trimestre_int}")
            except (ValueError, TypeError) as e:
                print(f"❌ EXPORT - Trimestre invalide: {e}")
        
        if annee_scolaire and annee_scolaire.strip() and validate_annee_scolaire(annee_scolaire):
            query = query.filter(Note.annee_scolaire == annee_scolaire)
            print(f"✅ EXPORT - Filtre année: {annee_scolaire}")
        
        notes = query.order_by(Eleve.nom, Eleve.prenoms).limit(SecurityConfig.MAX_EXPORT_ROWS).all()
        
        print(f"📊 EXPORT - {len(notes)} notes à exporter pour la classe: {classe_nom}")
        
        if format == 'pdf':
            return generate_pdf_export(notes, classe_nom)
        elif format == 'excel':
            return generate_excel_export(notes, classe_nom)
        else:
            return jsonify({"error": "Format d'export non supporté"}), 400
            
    except Exception as e:
        print(f"❌ ERREUR EXPORT: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Erreur lors de l'export {format}: {str(e)}"}), 500

@notes_bp.route("/export/filters", methods=["GET"])
@login_required
@ecole_required
def get_export_filters():
    """Renvoie les filtres disponibles pour l'export"""
    ecole_id = get_current_ecole_id()
    
    classes = Classe.query.filter_by(ecole_id=ecole_id).order_by(Classe.nom).all()
    matieres = Matiere.query.filter(Matiere.type.in_(SecurityConfig.ALLOWED_MATIERE_TYPES)).order_by(Matiere.libelle).all()
    annees_scolaires = db.session.query(Note.annee_scolaire).join(Eleve).join(Classe).filter(
        Classe.ecole_id == ecole_id
    ).distinct().all()
    
    annees_scolaires = [a[0] for a in annees_scolaires if validate_annee_scolaire(a[0])]
    
    return jsonify({
        "classes": [{"id": str(c.id), "nom": c.nom} for c in classes],
        "matieres": [{"id": str(m.id), "libelle": m.libelle} for m in matieres],
        "annees_scolaires": annees_scolaires,
        "trimestres": SecurityConfig.ALLOWED_TRIMESTRES
    })