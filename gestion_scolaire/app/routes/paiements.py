# C:\projets\python\gestion_scolaire\app\routes\paiements.py
from flask import Blueprint, render_template, request, jsonify, abort,current_app
from ..models import Classe, Eleve, Paiement,Ecole
from sqlalchemy import cast, String
from flask_login import current_user, login_required
from datetime import datetime
import uuid
from extensions import db
from uuid import UUID
from sqlalchemy.orm import joinedload
import re
import logging
import json
from functools import wraps
from ..utils import ecole_required, get_current_ecole_id
import os
import io
from reportlab.lib.pagesizes import A4, A5
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from flask import send_file

# CORRECTION : Importer le logger de sécurité
from ..logging_config import security_logger


paiements_bp = Blueprint('paiements', __name__)

def is_admin():
    return getattr(current_user, "role", "").lower() in ["admin", "administrateur"]

def get_paiement_secure(paiement_id, ecole_id):
    """Récupère un paiement en vérifiant l'appartenance à l'école"""
    try:
        paiement_uuid = UUID(paiement_id)
    except ValueError:
        abort(404, "Paiement non trouvé")
        
    paiement = Paiement.query.filter(
        Paiement.id == paiement_uuid,
        Paiement.ecole_id == ecole_id
    ).first()
    
    if not paiement:
        abort(404, "Paiement non trouvé")
    return paiement

workflow = {
    "Activer": {"from": "Inactif", "to": "Actif"},
    "Valider": {"from": "Actif", "to": "Validé"}
}

# ========== ROUTES ==========

@paiements_bp.route('/')
@login_required
@ecole_required
def liste_paiements():
    ecole_id = get_current_ecole_id()
    data = request.args
    
    search = data.get("search", "", type=str)
    per_page = data.get("per_page", 10, type=int)
    page = data.get("page", 1, type=int)
    classe_id = data.get("classe_id", type=str)
    
    # CORRECTION : Filtrer par école avec les jointures correctes
    query = Paiement.query.join(Eleve).join(Classe).filter(
        Classe.ecole_id == ecole_id
    ).options(
        joinedload(Paiement.eleve),
        joinedload(Paiement.classe)
    )

    # CORRECTION : Recherche texte - VERSION FONCTIONNELLE
    if search:
        print(f"🔍 RECHERCHE ACTIVÉE: '{search}'")  # Debug
        # Recherche dans les champs du paiement ET de l'élève
        search_filter = (
            Paiement.code.ilike(f"%{search}%") |
            Paiement.libelle.ilike(f"%{search}%") |
            Eleve.nom.ilike(f"%{search}%") |
            Eleve.prenoms.ilike(f"%{search}%") |
            Eleve.matricule.ilike(f"%{search}%") |
            Classe.nom.ilike(f"%{search}%")
        )
        query = query.filter(search_filter)
        print(f"✅ FILTRE RECHERCHE APPLIQUÉ")

    # Filtrage par classe
    if classe_id and classe_id.lower() != "none":
        try:
            classe_uuid = UUID(classe_id)
            query = query.filter(Paiement.classe_id == classe_uuid)
            print(f"🏫 FILTRE CLASSE APPLIQUÉ: {classe_id}")
        except ValueError:
            print(f"❌ CLASSE ID INVALIDE: {classe_id}")
            pass

    # CORRECTION : Tri par date de paiement (plus récent en premier) au lieu de created_at
    query = query.order_by(Paiement.date_payement.desc())

    # Pagination
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    paiements = pagination.items
    
    # Debug des résultats
    print(f"📊 {len(paiements)} paiement(s) trouvé(s) pour la recherche: '{search}'")
    
    # Rôle de l'utilisateur connecté
    user_role = (getattr(current_user, "role", "guest") or "guest").lower()

    # CORRECTION : Récupérer les classes et élèves filtrés par école
    classes = Classe.query.filter_by(ecole_id=ecole_id).all()
    eleves = Eleve.query.join(Classe).filter(Classe.ecole_id == ecole_id).all()
    
    return render_template("paiements/pay_list.html",
                           paiements=paiements,
                           pagination=pagination,
                           search=search,
                           per_page=per_page,
                           user_role=user_role,
                           classes=classes,
                           eleves=eleves,
                           classe_id=classe_id)


@paiements_bp.route("/export/liste/classe", methods=["GET"])
@login_required
@ecole_required
def export_liste_paiements_classe():
    """Export PDF de la liste des paiements par classe - VERSION SIMPLIFIÉE"""
    try:
        ecole_id = get_current_ecole_id()
        classe_id = request.args.get("classe_id")
        
        if not classe_id:
            return jsonify({"error": "Classe non sélectionnée"}), 400
        
        # Vérifier que la classe appartient à l'école
        classe = Classe.query.filter(
            Classe.id == UUID(classe_id),
            Classe.ecole_id == ecole_id
        ).first()
        
        if not classe:
            return jsonify({"error": "Classe non trouvée"}), 404
        
        # Récupérer les paiements de la classe
        paiements = Paiement.query.join(Eleve).join(Classe).filter(
            Classe.ecole_id == ecole_id,
            Paiement.classe_id == UUID(classe_id)
        ).order_by(Paiement.date_payement.desc()).all()
        
        if not paiements:
            return jsonify({"error": "Aucun paiement trouvé pour cette classe"}), 404
        
        # Récupérer les informations de l'école
        ecole = Ecole.query.get(ecole_id)
        if not ecole:
            return jsonify({"error": "École non trouvée"}), 404

        # Créer le PDF en mémoire
        buffer = io.BytesIO()
        
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
        from reportlab.lib.utils import ImageReader
        
        # Variables pour la pagination
        class PageData:
            current_page = 0
            total_pages = 1
        
        # ========== FONCTION POUR L'EN-TÊTE PROFESSIONNEL ==========
        def add_professional_header(canvas, doc):
            canvas.saveState()
            
            # Logo CENTRÉ au-dessus du nom de l'école
            logo_path = get_logo_path_paiements(ecole)
            if logo_path and os.path.exists(logo_path):
                try:
                    logo = ImageReader(logo_path)
                    # Logo centré en haut
                    canvas.drawImage(logo, 275, 770, width=50, height=50, preserveAspectRatio=True)
                except:
                    pass
            
            # Nom de l'école centré sous le logo
            canvas.setFont('Helvetica-Bold', 16)
            canvas.drawCentredString(300, 740, ecole.nom or "ÉCOLE")
            
            # Informations de l'école
            canvas.setFont('Helvetica', 9)
            canvas.drawCentredString(300, 720, f"Tél: {ecole.telephone1 or 'Non renseigné'}")
            
            # Ligne de séparation
            canvas.setStrokeColor(colors.HexColor('#2C3E50'))
            canvas.setLineWidth(1)
            canvas.line(40, 710, 560, 710)
            
            # République à droite
            canvas.setFont('Helvetica-Bold', 10)
            canvas.drawRightString(560, 780, "RÉPUBLIQUE TOGOLAISE")
            canvas.setFont('Helvetica', 8)
            canvas.drawRightString(560, 765, "Travail - Liberté - Patrie")
            
            # Ministère à gauche
            canvas.setFont('Helvetica-Bold', 10)
            canvas.drawString(40, 780, "MINISTÈRE DE L'ÉDUCATION")
            canvas.setFont('Helvetica', 8)
            canvas.drawString(40, 765, f"D.R.E: {ecole.dre or 'Non renseignée'}")
            canvas.drawString(40, 752, f"I.E.P.P: {ecole.inspection or 'Non renseignée'}")
            
            canvas.restoreState()
        
        # ========== FONCTION POUR LE PIED DE PAGE ==========
        def add_footer(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 8)
            canvas.setFillColor(colors.gray)
            
            # Pagination complète (Page 1/3, 2/3, etc.)
            page_info = f"Page {PageData.current_page}/{PageData.total_pages}"
            canvas.drawRightString(560, 30, page_info)
            
            # Information de génération
            canvas.drawString(40, 30, f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
            
            # Signature centrée
            canvas.drawCentredString(300, 30, f"École: {ecole.nom}")
            
            canvas.restoreState()
        
        # Configuration du document
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4,
            topMargin=140,
            bottomMargin=50,
            leftMargin=40,
            rightMargin=40
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # ========== TITRE SPÉCIFIQUE DESCENDU ==========
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=14,
            alignment=1,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=20,
            fontName='Helvetica-Bold'
        )
        
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(f"LISTE DES PAIEMENTS - CLASSE: {classe.nom}", title_style))
        
        # ========== TABLEAU SANS COLONNE CODE ==========
        if paiements:
            # En-tête du tableau SANS la colonne Code
            table_data = [[
                "N°", 
                "NOM-PRÉNOMS ", 
                "DATE", 
                "TOTAL NET", 
                "TOTAL PAYÉ", 
                "NON PAYÉ"
            ]]
            
            # Données des paiements
            total_net = 0
            total_paye = 0
            total_reste = 0
            
            for i, p in enumerate(paiements, 1):
                eleve_nom = f"{p.eleve.nom} {p.eleve.prenoms}"
                # Plus besoin de tronquer avec l'espace libéré
                if len(eleve_nom) > 35:
                    eleve_nom = eleve_nom[:35] + "..."
                
                table_data.append([
                    str(i),
                    eleve_nom,
                    p.date_payement.strftime('%d/%m/%Y') if p.date_payement else 'N/A',
                    f"{p.montant_net:,.0f} FCFA",
                    f"{p.montant_pay:,.0f} FCFA",
                    f"{p.montant_rest:,.0f} FCFA"
                ])
                
                total_net += p.montant_net
                total_paye += p.montant_pay
                total_reste += p.montant_rest
            
            # Ligne des totaux
            table_data.append([
                "", 
                "TOTAUX GÉNÉRAUX", 
                "",
                f"{total_net:,.0f} FCFA",
                f"{total_paye:,.0f} FCFA", 
                f"{total_reste:,.0f} FCFA"
            ])
            
            # CRÉATION DU TABLEAU AVEC LARGEURS OPTIMISÉES
            col_widths = [25, 200, 50, 80, 80, 80]  # Total: 515 points
            
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            # STYLE PROFESSIONNEL ET LISIBLE
            table_style = [
                # En-tête
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),  # Police plus grande
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Données
                ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -2), 10),  # Police plus grande
                ('ALIGN', (0, 1), (0, -2), 'CENTER'),      # N°
                ('ALIGN', (1, 1), (1, -2), 'LEFT'),        # Élève
                ('ALIGN', (2, 1), (2, -2), 'CENTER'),      # Date
                ('ALIGN', (3, 1), (5, -2), 'RIGHT'),       # Montants
                ('VALIGN', (0, 1), (-1, -2), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8F9FA')]),
                
                # Totaux
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E9ECEF')),
                ('ALIGN', (3, -1), (5, -1), 'RIGHT'),
                ('FONTSIZE', (0, -1), (-1, -1), 11),  # Police plus grande
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#2C3E50')),
                
                # Bordures
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DEE2E6')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                
                # Espacement généreux
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]
            
            table.setStyle(TableStyle(table_style))
            
            elements.append(table)
            elements.append(Spacer(1, 25))
            
            # ========== SYNTHÈSE HORIZONTALE EN BAS ==========
            # Créer un tableau horizontal pour la synthèse
            synthèse_data = [
                [
                    Paragraph("<b>SYNTHÈSE DES PAIEMENTS</b>", 
                             ParagraphStyle('SynthèseTitle', parent=styles['Normal'], fontSize=11, fontName='Helvetica-Bold')),
                    Paragraph(f"<b>Nombre:</b> {len(paiements)}", 
                             ParagraphStyle('SynthèseItem', parent=styles['Normal'], fontSize=10)),
                    Paragraph(f"<b>Total Net:</b> {total_net:,.0f} FCFA", 
                             ParagraphStyle('SynthèseItem', parent=styles['Normal'], fontSize=10)),
                    Paragraph(f"<b>Total Payé:</b> {total_paye:,.0f} FCFA", 
                             ParagraphStyle('SynthèseItem', parent=styles['Normal'], fontSize=10)),
                    Paragraph(f"<b>Reste:</b> {total_reste:,.0f} FCFA", 
                             ParagraphStyle('SynthèseItem', parent=styles['Normal'], fontSize=10)),
                    Paragraph(f"<b>Taux:</b> {(total_paye/total_net*100) if total_net > 0 else 0:.1f}%", 
                             ParagraphStyle('SynthèseItem', parent=styles['Normal'], fontSize=10))
                ]
            ]
            
            synthèse_table = Table(synthèse_data, colWidths=[90, 70, 90, 90, 90, 70])
            synthèse_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F8F9FA')),
                ('BOX', (0, 0), (-1, 0), 1, colors.HexColor('#DEE2E6')),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, 0), 10),
                ('RIGHTPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ]))
            
            elements.append(synthèse_table)
                
        else:
            # Message stylisé quand aucun paiement
            empty_style = ParagraphStyle(
                'EmptyStyle',
                parent=styles['Normal'],
                fontSize=12,
                alignment=1,
                textColor=colors.grey
            )
            elements.append(Spacer(1, 50))
            elements.append(Paragraph("AUCUN PAIEMENT TROUVÉ POUR CETTE CLASSE", empty_style))
        
        # Calcul du nombre total de pages
        PageData.total_pages = max(1, (len(elements) // 20) + 1)
        
        # Fonction pour mettre à jour le numéro de page
        def on_page(canvas, doc):
            PageData.current_page += 1
            add_professional_header(canvas, doc)
            add_footer(canvas, doc)
        
        # Génération du PDF
        doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)
        
        buffer.seek(0)
        
        # Retourner le fichier
        filename = f"liste_paiements_{classe.nom}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        
        response = send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"Erreur génération PDF liste paiements: {str(e)}")
        return jsonify({"error": f"Erreur lors de la génération de la liste: {str(e)}"}), 500

@paiements_bp.route("/add", methods=["GET", "POST"])
@login_required
@ecole_required
def ajouter_paiement():
    ecole_id = get_current_ecole_id()
    
    if request.method == "POST":
        # === LOG TEMPOREL POUR IDENTIFIER LES DOUBLONS ===
        import time
        current_time = time.time()
        print(f"🕒 [{current_time}] === NOUVELLE REQUÊTE POST /add ===")
        print(f"🕒 [{current_time}] Headers: {dict(request.headers)}")
        print(f"🕒 [{current_time}] Données: {dict(request.form)}")
        
        # === PROTECTION ANTI-DOUBLON IMMÉDIATE ===
        request_key = f"{request.form.get('eleve_id')}-{request.form.get('libelle')}-{request.form.get('date_payement')}"
        
        if not hasattr(ajouter_paiement, 'recent_requests'):
            ajouter_paiement.recent_requests = {}
        
        # Nettoyer les anciennes requêtes (plus de 5 secondes)
        ajouter_paiement.recent_requests = {
            k: v for k, v in ajouter_paiement.recent_requests.items() 
            if current_time - v < 5
        }
        
        # Vérifier si cette requête a déjà été traitée
        if request_key in ajouter_paiement.recent_requests:
            print(f"🚫 [{current_time}] REQUÊTE DOUBLON BLOQUÉE - Clé: {request_key}")
            return jsonify({
                "status": "warning", 
                "message": "Cette requête a déjà été traitée"
            }), 429
        
        # Marquer cette requête comme traitée
        ajouter_paiement.recent_requests[request_key] = current_time
        
        # === VOTRE CODE EXISTANT ===
        data = request.form
        
        # Validation basique
        required_fields = ['eleve_id', 'classe_id', 'libelle', 'date_payement', 'montant_net', 'montant_pay']
        for field in required_fields:
            if not data.get(field):
                print(f"❌ [{current_time}] Champ manquant: {field}")
                return jsonify({"status": "warning", "message": f"Champ {field} manquant"})
        
        try:
            # Conversion des données
            eleve_id = data.get("eleve_id")
            classe_id = data.get("classe_id") 
            libelle = data.get("libelle", "").strip()
            date_payement = datetime.strptime(data.get("date_payement"), "%Y-%m-%d")
            montant_net = float(data.get("montant_net"))
            montant_pay = float(data.get("montant_pay"))
            montant_rest = montant_net - montant_pay
            
            # VÉRIFICATION ANTI-DOUBLON TRÈS STRICTE - AMÉLIORÉE
            existing = Paiement.query.filter(
                Paiement.eleve_id == eleve_id,
                Paiement.libelle == libelle,
                Paiement.date_payement == date_payement,
                Paiement.ecole_id == ecole_id,
                Paiement.etat != "supprimé"  # Ignorer les paiements supprimés
            ).first()
            
            if existing:
                print(f"❌ [{current_time}] Doublon détecté en base - ID: {existing.id}")
                return jsonify({
                    "status": "warning", 
                    "message": "Cet élève a déjà un paiement '" + libelle + "' pour cette date. Veuillez aller en mode modification pour modifier le paiement existant."
                })
            
            # Vérifier aussi si le montant payé dépasse le montant net
            if montant_pay > montant_net:
                return jsonify({
                    "status": "warning",
                    "message": "Le montant payé ne peut pas dépasser le montant net"
                })
            
            # Création
            code = f"PAY-{uuid.uuid4().hex[:6].upper()}"
            paiement_id = uuid.uuid4()
            
            print(f"✅ [{current_time}] Création nouveau paiement - ID: {paiement_id}, Code: {code}")
            
            paiement = Paiement(
                id=paiement_id,
                code=code,
                libelle=libelle,
                eleve_id=eleve_id,
                classe_id=classe_id,
                ecole_id=ecole_id,
                date_payement=date_payement,
                montant_net=montant_net,
                montant_pay=montant_pay,
                montant_rest=montant_rest,
                etat="Inactif"
            )
            
            db.session.add(paiement)
            db.session.commit()
            
            print(f"✅ [{current_time}] Paiement créé avec succès - ID: {paiement_id}")
            
            return jsonify({
                "status": "success",
                "message": "Paiement ajouté avec succès"
            })
                
        except Exception as e:
            db.session.rollback()
            print(f"❌ [{current_time}] Erreur: {str(e)}")
            return jsonify({"status": "danger", "message": "Erreur lors de la création du paiement"}), 500
    
    # GET request
    eleves = Eleve.query.join(Classe).filter(Classe.ecole_id == ecole_id).all()
    classes = Classe.query.filter_by(ecole_id=ecole_id).all()
    return render_template("paiements/add_pay.html", eleves=eleves, classes=classes)

@paiements_bp.route("/detail/<string:pay_id>", methods=["GET"])
@login_required
@ecole_required
def detail_pay(pay_id):
    ecole_id = get_current_ecole_id()
    
    # CORRECTION : Vérifier que le paiement appartient à l'école
    paiement = get_paiement_secure(pay_id, ecole_id)
    
    return jsonify({
        "id": str(paiement.id),
        "code": paiement.code,
        "libelle": paiement.libelle,
        "eleve_id": str(paiement.eleve.id),
        "eleve": f"{paiement.eleve.nom} {paiement.eleve.prenoms}",
        "classe_id": str(paiement.classe.id),
        "classe": paiement.classe.nom,
        "date_payement": paiement.date_payement.strftime("%Y-%m-%d") if paiement.date_payement else '',
        "montant_net": float(paiement.montant_net),
        "montant_pay": float(paiement.montant_pay),
        "montant_rest": float(paiement.montant_rest)
    })

@paiements_bp.route("/edit/<string:paiement_id>", methods=["POST"])
@login_required
@ecole_required
def edit_paiement(paiement_id):
    ecole_id = get_current_ecole_id()
    
    if not is_admin():
        return jsonify({"error": "Accès refusé"}), 403
    
    # CORRECTION : Vérifier que le paiement appartient à l'école
    paiement = get_paiement_secure(paiement_id, ecole_id)
    
    # Récupérer les données du formulaire
    libelle = request.form.get("libelle")
    date_payement = request.form.get("date_payement")
    montant_net = request.form.get("montant_net")
    montant_pay = request.form.get("montant_pay")
    
    # Validation
    try:
        date_payement_dt = datetime.strptime(date_payement, "%Y-%m-%d")
        montant_net_val = float(montant_net)
        montant_pay_val = float(montant_pay)
        
        if montant_pay_val > montant_net_val:
            return jsonify({"status": "error", "message": "Le montant payé ne peut pas dépasser le montant net"}), 400
            
    except (ValueError, TypeError):
        return jsonify({"status": "error", "message": "Données invalides"}), 400
    
    # Mise à jour
    try:
        paiement.libelle = libelle
        paiement.date_payement = date_payement_dt
        paiement.montant_net = montant_net_val
        paiement.montant_pay = montant_pay_val
        paiement.montant_rest = montant_net_val - montant_pay_val
        
        db.session.commit()
        
        return jsonify({"status": "success", "message": "Paiement modifié avec succès"})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": "Erreur lors de la modification"}), 500

@paiements_bp.route("/delete/<string:id>", methods=["POST"])
@login_required
@ecole_required
def delete_pay(id):
    ecole_id = get_current_ecole_id()
    
    if not is_admin():
        return jsonify({"error": "Accès refusé"}), 403
    
    # CORRECTION : Vérifier que le paiement appartient à l'école
    paiement = get_paiement_secure(id, ecole_id)
    
    try:
        db.session.delete(paiement)
        db.session.commit()
        return jsonify({"status": "success", "message": "Paiement supprimé"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": "Erreur lors de la suppression"}), 500

@paiements_bp.route("/workflow/<string:pay_id>", methods=["POST"])
@login_required
@ecole_required
def change_workflow(pay_id):
    ecole_id = get_current_ecole_id()
    
    if not is_admin():
        return jsonify({"error": "Accès refusé"}), 403
    
    # CORRECTION : Vérifier que le paiement appartient à l'école
    paiement = get_paiement_secure(pay_id, ecole_id)
    
    data = request.get_json()
    if not data:
        return jsonify({"error": "Données manquantes"}), 400
    
    action = data.get("action")

    if not action or action not in workflow:
        return jsonify({"error": "Action non valide!"}), 400
    
    trans = workflow[action]
    if paiement.etat != trans["from"]:
        return jsonify({"error": f"L'état actuel est {paiement.etat}, impossible d'appliquer {action}"}), 400
    
    try:
        paiement.etat = trans["to"]
        db.session.commit()
        
        return jsonify({
            "success": True,
            "etat": paiement.etat,
            "message": f"L'état a été mis à jour vers {paiement.etat}"
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": "Erreur lors de la mise à jour de l'état"}), 500

# ========== ROUTES D'EXPORT (à ajouter si nécessaire) ==========
# Vos routes d'export PDF/Excel restent les mêmes

# ========== EXPORTATIONS RECUS PAIEMENTS ==========

def get_logo_path_paiements(ecole):
    """Retourne le chemin du logo de l'école - POUR PAIEMENTS"""
    if not ecole or not ecole.logo_filename:
        return None
    
    possible_paths = [
        os.path.join(current_app.root_path, 'gestion_scolaire', 'app', 'static', 'logos', ecole.logo_filename),
        os.path.join(current_app.root_path, 'static', 'images', 'logos_ecoles', ecole.logo_filename),
        os.path.join(current_app.root_path, 'static', 'uploads', 'logos', ecole.logo_filename),
    ]
    
    for logo_path in possible_paths:
        if os.path.exists(logo_path) and os.path.getsize(logo_path) > 0:
            return logo_path
    
    return None

def create_recu_table(paiement, ecole, recu_width=None):
    """Crée les éléments d'un reçu individuel professionnel au format A5 - VERSION RÉORGANISÉE"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    
    # CORRECTION : Définir recu_width si non fourni
    if recu_width is None:
        page_width, _ = A4
        recu_width = page_width / 2  # 4 reçus par page A4
    
    styles = getSampleStyleSheet()
    
    # ========== STYLES POUR RECU A5 PROFESSIONNEL ==========
    republique_style = ParagraphStyle(
        'RepubliqueStyle',
        parent=styles['Normal'],
        fontSize=8,  # Réduction de la police
        alignment=2,  # Alignement à droite
        spaceAfter=0,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#2C3E50')
    )
    
    ministere_style = ParagraphStyle(
        'MinistereStyle',
        parent=styles['Normal'],
        fontSize=8,  # Réduction de la police
        alignment=0,  # Alignement à gauche
        spaceAfter=0,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#2C3E50')
    )
    
    devise_style = ParagraphStyle(
        'DeviseStyle',
        parent=styles['Normal'],
        fontSize=6,  # Réduction de la police
        alignment=2,  # Alignement à droite
        spaceAfter=1,
        fontName='Helvetica-Bold',  # Maintenant en gras
        textColor=colors.HexColor('#666666')
    )
    
    dre_style = ParagraphStyle(
        'DREStyle',
        parent=styles['Normal'],
        fontSize=6,  # Réduction de la police
        alignment=0,  # Alignement à gauche
        spaceAfter=0,
        fontName='Helvetica-Bold',  # Maintenant en gras
        textColor=colors.HexColor('#666666')
    )
    
    inspection_style = ParagraphStyle(
        'InspectionStyle',
        parent=styles['Normal'],
        fontSize=6,  # Réduction de la police
        alignment=0,  # Alignement à gauche
        spaceAfter=0,
        fontName='Helvetica-Bold',  # Maintenant en gras
        textColor=colors.HexColor('#666666')
    )
    
    ecole_style = ParagraphStyle(
        'EcoleStyle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=1,  # Centré
        spaceAfter=1,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#2C3E50')
    )
    
    telephone_style = ParagraphStyle(
        'TelephoneStyle',
        parent=styles['Normal'],
        fontSize=7,
        alignment=1,  # Centré
        spaceAfter=0,
        textColor=colors.HexColor('#666666')
    )
    
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=14,
        alignment=1,
        textColor=colors.HexColor('#2C3E50'),
        spaceAfter=3,
        fontName='Helvetica-Bold'
    )
    
    code_style = ParagraphStyle(
        'CodeStyle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=1,
        spaceAfter=4,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#1E90FF')  # Bleu pour le code
    )
    
    info_label_style = ParagraphStyle(
        'InfoLabelStyle',
        parent=styles['Normal'],
        fontSize=9,
        alignment=0,
        spaceAfter=1,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#2C3E50')
    )
    
    info_value_style = ParagraphStyle(
        'InfoValueStyle',
        parent=styles['Normal'],
        fontSize=9,
        alignment=0,
        spaceAfter=1,
        textColor=colors.black
    )
    
    montant_style = ParagraphStyle(
        'MontantStyle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=0,
        spaceAfter=1,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#2C3E50')
    )
    
    signature_style = ParagraphStyle(
        'SignatureStyle',
        parent=styles['Normal'],
        fontSize=8,
        alignment=1,  # CENTRÉ pour aligner avec le trait
        spaceBefore=2,  # Réduction de l'espace
        fontName='Helvetica-Oblique',
        textColor=colors.gray
    )
    
    # Informations de l'école
    nom_ecole = ecole.nom if ecole.nom else "École non renseignée"
    dre = ecole.dre if ecole.dre else "non renseignée"
    inspection = ecole.inspection if ecole.inspection else "non renseignée"
    telephone = ecole.telephone1 if ecole.telephone1 else "Téléphone non renseigné"
    devise_ecole = ecole.devise if ecole.devise else "Travail-Liberté-Patrie"
    
    # Logo - dimensions réduites
    logo = None
    logo_path = get_logo_path_paiements(ecole)
    
    if logo_path:
        try:
            logo = Image(logo_path, width=20*mm, height=20*mm)  # Réduction de 30mm à 20mm
            logo.hAlign = 'CENTER'
        except Exception as e:
            print(f"❌ Erreur chargement logo: {e}")
            logo = None
    
    # ========== EN-TÊTE RÉORGANISÉE ==========
    
    # Ligne 1 : République à droite + Logo et école au centre + Ministère à gauche
    republique_content = [
        Paragraph("RÉPUBLIQUE TOGOLAISE", republique_style),
        Paragraph("Travail-Liberté-Patrie", devise_style)
    ]
    
    # Contenu centré - logo et nom d'école bien alignés
    centre_table_data = []
    if logo:
        centre_table_data.append([logo])
    centre_table_data.extend([
        [Paragraph(f"<b>{nom_ecole}</b>", ecole_style)],
        [Paragraph(f"Tél: {telephone}", telephone_style)]
    ])
    
    centre_table = Table(centre_table_data, colWidths=[recu_width*0.4])
    centre_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
    ]))
    
    ministere_content = [
        Paragraph("MINISTÈRE DE L'EDUCATION NATIONALE", ministere_style),
        Paragraph(f"<b>D.R.E:</b> {dre}", dre_style),
        Paragraph(f"<b>I.E.S.G:</b> {inspection}", inspection_style)
    ]
    
    # Tableau d'en-tête avec 3 colonnes
    header_table = Table([[
        # Colonne gauche : Ministère + DRE + Inspection (aligné à gauche)
        Table([[ministere_content[0]], [ministere_content[1]], [ministere_content[2]]], 
              colWidths=[recu_width*0.3]),
        
        # Colonne centre : Logo + École + Téléphone (centré)
        centre_table,
        
        # Colonne droite : République + Devise (aligné à droite)
        Table([[republique_content[0]], [republique_content[1]]], 
              colWidths=[recu_width*0.3])
    ]], colWidths=[recu_width*0.3, recu_width*0.4, recu_width*0.3])
    
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),    # Ministère à gauche
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),  # Centre centré
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),   # République à droite
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
    ]))
    
    # Ligne de séparation après l'en-tête
    separation_line = Table([['']], colWidths=[recu_width])
    separation_line.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (0, 0), 1, colors.HexColor('#2C3E50')),
        ('BOTTOMPADDING', (0, 0), (0, 0), 4),
        ('TOPPADDING', (0, 0), (0, 0), 2),
    ]))
    
    # ========== TITRE ET CODE ==========
    title = Paragraph("REÇU DE PAIEMENT", title_style)
    
    # Code de paiement centré en couleur bleu
    code_paiement = Paragraph(f"<b>Code: {paiement.code}</b>", code_style)
    
    # ========== INFORMATIONS DU PAIEMENT ==========
    
    # Section Élève
    eleve_section = [
        [Paragraph("ÉLÈVE", info_label_style), Paragraph(f"{paiement.eleve.nom} {paiement.eleve.prenoms}", info_value_style)],
        [Paragraph("CLASSE", info_label_style), Paragraph(paiement.classe.nom, info_value_style)],
        [Paragraph("DATE", info_label_style), Paragraph(paiement.date_payement.strftime("%d/%m/%Y") if paiement.date_payement else "N/A", info_value_style)],
    ]
    
    eleve_table = Table(eleve_section, colWidths=[recu_width*0.3, recu_width*0.7])
    eleve_table.setStyle(TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
    ]))
    
    # Section Détails du paiement
    details_section = [
        [Paragraph("LIBELLÉ", info_label_style), Paragraph(paiement.libelle, info_value_style)],
    ]
    
    details_table = Table(details_section, colWidths=[recu_width*0.3, recu_width*0.7])
    details_table.setStyle(TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    # Section Montants avec mise en forme spéciale
    montants_data = [
        [Paragraph("MONTANT NET", montant_style), Paragraph(f"{paiement.montant_net:,.0f} FCFA", montant_style)],
        [Paragraph("MONTANT PAYÉ", montant_style), Paragraph(f"{paiement.montant_pay:,.0f} FCFA", montant_style)],
    ]
    
    # Ajouter le reste seulement s'il est > 0
    if paiement.montant_rest > 0:
        montants_data.append([
            Paragraph("RESTE À PAYER", ParagraphStyle(
                'ResteStyle',
                parent=styles['Normal'],
                fontSize=10,
                alignment=0,
                fontName='Helvetica-Bold',
                textColor=colors.red
            )), 
            Paragraph(f"{paiement.montant_rest:,.0f} FCFA", ParagraphStyle(
                'ResteStyle',
                parent=styles['Normal'],
                fontSize=10,
                alignment=0,
                fontName='Helvetica-Bold',
                textColor=colors.red
            ))
        ])
    else:
        montants_data.append([
            Paragraph("STATUT", ParagraphStyle(
                'StatutStyle',
                parent=styles['Normal'],
                fontSize=10,
                alignment=0,
                fontName='Helvetica-Bold',
                textColor=colors.green
            )), 
            Paragraph("PAIEMENT COMPLET", ParagraphStyle(
                'StatutStyle',
                parent=styles['Normal'],
                fontSize=10,
                alignment=0,
                fontName='Helvetica-Bold',
                textColor=colors.green
            ))
        ])
    
    montants_table = Table(montants_data, colWidths=[recu_width*0.5, recu_width*0.5])
    montants_table.setStyle(TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8F9FA')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#DEE2E6')),
    ]))
    
    # ========== SIGNATURE ==========
    # Tableau pour aligner le trait et "Responsable" au centre
    signature_table = Table([
        [''],  # Ligne vide pour espacement
        ['_________________________'],  # Trait de signature
        ['Responsable']  # Texte "Responsable"
    ], colWidths=[recu_width*0.6])
    
    signature_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Tout centré
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 1), (0, 1), 0),  # Pas d'espace entre le trait et "Responsable"
        ('TOPPADDING', (0, 2), (0, 2), 0),
        ('FONTSIZE', (0, 2), (0, 2), 8),  # Taille police pour "Responsable"
        ('FONT', (0, 2), (0, 2), 'Helvetica-Oblique'),
        ('TEXTCOLOR', (0, 2), (0, 2), colors.gray),
    ]))
    
    # ========== PIED DE PAGE ==========
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontSize=6,
        alignment=1,
        spaceBefore=5,
        textColor=colors.gray
    )
    
    footer = Paragraph(f"Reçu généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", footer_style)
    
    # ========== RETOUR DES ÉLÉMENTS ==========
    return [
        header_table,
        separation_line,
        Spacer(1, 4*mm),
        title,
        code_paiement,
        Spacer(1, 4*mm),
        eleve_table,
        Spacer(1, 2*mm),
        details_table,
        Spacer(1, 4*mm),
        montants_table,
        Spacer(1, 8*mm),
        signature_table,  # Tableau unifié pour signature
        Spacer(1, 3*mm),
        footer
    ]

@paiements_bp.route("/export/recu/pdf/<string:paiement_id>")
@login_required
@ecole_required
def export_recu_pdf(paiement_id):
    """Export PDF d'un reçu de paiement au format A5 - VERSION CORRIGÉE (1 seul reçu par page)"""
    try:
        ecole_id = get_current_ecole_id()
        
        # Récupérer le paiement avec vérification de sécurité
        paiement = get_paiement_secure(paiement_id, ecole_id)
        
        # Récupérer les informations de l'école
        ecole = Ecole.query.get(ecole_id)
        if not ecole:
            return jsonify({"error": "École non trouvée"}), 404

        # Créer le PDF en mémoire
        buffer = io.BytesIO()
        
        from reportlab.lib.pagesizes import A5  # Changement pour A5
        from reportlab.lib.units import mm
        
        # Utiliser A5 pour un seul reçu par page
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A5,  # Format A5 pour un reçu individuel
            topMargin=10*mm,
            bottomMargin=10*mm,
            leftMargin=10*mm,
            rightMargin=10*mm
        )
        
        elements = []
        
        # ========== CREATION D'UN SEUL RECU ==========
        # Calcul de la largeur pour A5
        page_width, _ = A5
        recu_width = page_width - 20*mm  # Largeur moins les marges
        
        recu_elements = create_recu_table(paiement, ecole, recu_width)
        elements.extend(recu_elements)
        
        # ========== GÉNÉRATION DU PDF ==========
        doc.build(elements)
        buffer.seek(0)
        
        # Retourner le PDF
        filename = f"recu_paiement_{paiement.code}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        
        response = send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"Erreur génération PDF reçu paiement: {str(e)}")
        return jsonify({"error": f"Erreur lors de la génération du reçu PDF: {str(e)}"}), 500

@paiements_bp.route("/export/recu/excel/<string:paiement_id>")
@login_required
@ecole_required
def export_recu_excel(paiement_id):
    """Export Excel d'un reçu de paiement formaté - VERSION UNIFORMISÉE"""
    try:
        ecole_id = get_current_ecole_id()
        
        # Récupérer le paiement avec vérification de sécurité
        paiement = get_paiement_secure(paiement_id, ecole_id)
        
        # Récupérer les informations de l'école
        ecole = Ecole.query.get(ecole_id)
        if not ecole:
            return jsonify({"error": "École non trouvée"}), 404

        # Informations de l'école - UNIFORMISÉES AVEC LE FORMAT PDF
        nom_ecole = ecole.nom if ecole.nom else "École non renseignée"
        dre = ecole.dre if ecole.dre else "non renseignée"
        inspection = ecole.inspection if ecole.inspection else "non renseignée"
        telephone = ecole.telephone1 if ecole.telephone1 else "Téléphone non renseigné"
        devise_ecole = ecole.devise if ecole.devise else "Travail-Liberté-Patrie"

        # Créer le fichier Excel
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = f"Reçu {paiement.code}"
        
        # ========== MISE EN FORME EXCEL ==========
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        
        # ========== EN-TÊTE EXCEL UNIFORMISÉ ==========
        # Ministère à GAUCHE (comme PDF)
        ws.merge_cells('A1:C1')
        ws['A1'] = "MINISTÈRE DE L'EDUCATION NATIONALE"
        ws['A1'].font = Font(bold=True, size=9)
        ws['A1'].alignment = left_align
        
        # République à DROITE (comme PDF)
        ws.merge_cells('D1:F1')
        ws['D1'] = "RÉPUBLIQUE TOGOLAISE"
        ws['D1'].font = Font(bold=True, size=9)
        ws['D1'].alignment = right_align
        
        # D.R.E en gras et petite taille (comme PDF)
        ws.merge_cells('A2:C2')
        ws['A2'] = f"D.R.E: {dre}"
        ws['A2'].font = Font(bold=True, size=7)
        ws['A2'].alignment = left_align
        
        # Devise en gras et petite taille (comme PDF)
        ws.merge_cells('D2:F2')
        ws['D2'] = "Travail-Liberté-Patrie"
        ws['D2'].font = Font(bold=True, size=7)
        ws['D2'].alignment = right_align
        
        # I.E.S.G en gras et petite taille (comme PDF)
        ws.merge_cells('A3:C3')
        ws['A3'] = f"I.E.S.G: {inspection}"
        ws['A3'].font = Font(bold=True, size=7)
        ws['A3'].alignment = left_align
        
        # Ligne de séparation
        ws.merge_cells('A4:F4')
        ws['A4'] = "___________________________________________________________"
        ws['A4'].alignment = center_align
        
        # Nom de l'école CENTRÉ sous le logo (position symbolique)
        ws.merge_cells('A5:F5')
        ws['A5'] = nom_ecole
        ws['A5'].font = Font(bold=True, size=12, color="2C3E50")
        ws['A5'].alignment = center_align
        
        # Téléphone centré
        ws.merge_cells('A6:F6')
        ws['A6'] = f"Tél: {telephone}"
        ws['A6'].alignment = center_align
        
        # Titre du document
        ws.merge_cells('A7:F7')
        ws['A7'] = "REÇU DE PAIEMENT"
        ws['A7'].font = Font(bold=True, size=14, color="2C3E50")
        ws['A7'].alignment = center_align
        
        # Code en bleu (comme PDF)
        ws.merge_cells('A8:F8')
        ws['A8'] = f"Code: {paiement.code}"
        ws['A8'].font = Font(bold=True, size=10, color="1E90FF")
        ws['A8'].alignment = center_align
        
        # Ligne vide
        ws['A9'] = ""
        
        # ========== INFORMATIONS DU PAIEMENT UNIFORMISÉES ==========
        info_data = [
            ("ÉLÈVE:", f"{paiement.eleve.nom} {paiement.eleve.prenoms}"),
            ("CLASSE:", paiement.classe.nom),
            ("DATE:", paiement.date_payement.strftime("%d/%m/%Y") if paiement.date_payement else "N/A"),
            ("LIBELLÉ:", paiement.libelle),
            ("MONTANT NET:", f"{paiement.montant_net:,.0f} FCFA"),
            ("MONTANT PAYÉ:", f"{paiement.montant_pay:,.0f} FCFA"),
        ]
        
        # Ajouter le reste à payer ou statut complet
        if paiement.montant_rest > 0:
            info_data.append(("RESTE À PAYER:", f"{paiement.montant_rest:,.0f} FCFA"))
        else:
            info_data.append(("STATUT:", "PAIEMENT COMPLET"))
        
        start_row = 10
        for i, (label, value) in enumerate(info_data):
            # Label en gras
            ws.merge_cells(f'A{start_row + i}:B{start_row + i}')
            ws[f'A{start_row + i}'] = label
            ws[f'A{start_row + i}'].font = Font(bold=True, size=9)
            ws[f'A{start_row + i}'].alignment = left_align
            
            # Valeur
            ws.merge_cells(f'C{start_row + i}:F{start_row + i}')
            ws[f'C{start_row + i}'] = value
            ws[f'C{start_row + i}'].font = Font(size=9)
            ws[f'C{start_row + i}'].alignment = left_align
        
        # ========== SIGNATURE UNIFORMISÉE ==========
        signature_row = start_row + len(info_data) + 2
        
        # Trait de signature centré
        ws.merge_cells(f'C{signature_row}:F{signature_row}')
        ws[f'C{signature_row}'] = "_________________________"
        ws[f'C{signature_row}'].alignment = center_align
        
        # "Responsable" centré sous le trait
        ws.merge_cells(f'C{signature_row + 1}:F{signature_row + 1}')
        ws[f'C{signature_row + 1}'] = "Responsable"
        ws[f'C{signature_row + 1}'].font = Font(italic=True, size=8)
        ws[f'C{signature_row + 1}'].alignment = center_align
        
        # ========== PIED DE PAGE ==========
        footer_row = signature_row + 3
        ws.merge_cells(f'A{footer_row}:F{footer_row}')
        ws[f'A{footer_row}'] = f"Reçu généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws[f'A{footer_row}'].font = Font(size=6, color="666666")
        ws[f'A{footer_row}'].alignment = center_align
        
        # ========== MISE EN FORME FINALE ==========
        # Ajuster les largeurs de colonnes
        column_widths = {'A': 12, 'B': 8, 'C': 15, 'D': 15, 'E': 15, 'F': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Bordures pour le tableau d'information
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, start_row + len(info_data)):
            for col in range(1, 7):  # Colonnes A à F
                ws.cell(row=row, column=col).border = thin_border
        
        # Fond coloré pour les montants (comme PDF)
        for row in range(start_row + 4, start_row + len(info_data)):
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        # Sauvegarder
        wb.save(buffer)
        buffer.seek(0)
        
        # Retourner le fichier
        filename = f"recu_paiement_{paiement.code}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        response = send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"Erreur génération Excel reçu paiement: {str(e)}")
        return jsonify({"error": f"Erreur lors de la génération du reçu Excel: {str(e)}"}), 500

# Route pour exporter plusieurs reçus en lot (inchangée)
@paiements_bp.route("/export/recus/lot", methods=["POST"])
@login_required
@ecole_required
def export_recus_lot():
    """Export PDF de plusieurs reçus de paiement (4 par page A4) - VERSION CORRIGÉE"""
    try:
        ecole_id = get_current_ecole_id()
        data = request.get_json()
        
        if not data or 'paiement_ids' not in data:
            return jsonify({"error": "Liste des paiements manquante"}), 400
        
        paiement_ids = data['paiement_ids']
        
        # Récupérer les informations de l'école
        ecole = Ecole.query.get(ecole_id)
        if not ecole:
            return jsonify({"error": "École non trouvée"}), 404

        # Créer le PDF en mémoire
        buffer = io.BytesIO()
        
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        
        # Calcul des dimensions pour 4 reçus par page A4
        page_width, page_height = A4
        recu_width = page_width / 2
        
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4,
            topMargin=10*mm,
            bottomMargin=10*mm,
            leftMargin=10*mm,
            rightMargin=10*mm
        )
        
        elements = []
        
        # Récupérer tous les paiements
        paiements = []
        for paiement_id in paiement_ids:
            try:
                paiement = get_paiement_secure(paiement_id, ecole_id)
                paiements.append(paiement)
            except:
                continue  # Ignorer les paiements non trouvés
        
        if not paiements:
            return jsonify({"error": "Aucun paiement valide trouvé"}), 404
        
        # Créer les reçus (4 par page)
        for i, paiement in enumerate(paiements):
            # Ajouter un saut de page tous les 4 reçus (sauf pour le premier)
            if i > 0 and i % 4 == 0:
                elements.append(PageBreak())
            
            # Créer un reçu
            recu_elements = create_recu_table(paiement, ecole, recu_width)
            elements.extend(recu_elements)
            
            # Ajouter un saut de page entre les reçus (sauf pour le dernier reçu de la page)
            if (i + 1) % 4 != 0 and i != len(paiements) - 1:
                elements.append(PageBreak())
        
        # Générer le PDF
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"recus_paiements_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        
        response = send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"Erreur génération PDF reçus lot: {str(e)}")
        return jsonify({"error": f"Erreur lors de la génération des reçus: {str(e)}"}), 500

        