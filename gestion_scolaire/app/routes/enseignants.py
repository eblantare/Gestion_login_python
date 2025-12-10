from flask import Blueprint, request, render_template, jsonify, send_file, current_app,session
from ..models import Enseignant, Matiere, Ecole
from sqlalchemy import cast, String, extract
from gestion_login.gestion_login.models import Utilisateur
from flask_login import current_user, login_required
from extensions import db
from sqlalchemy.orm import joinedload
import io
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib import colors
from datetime import datetime
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from ..utils import ecole_required, get_current_ecole_id, is_system_admin
from ..routes.main import get_user_accessible_ecoles

enseignants_bp = Blueprint('enseignants', __name__)

def is_admin():
    return getattr(current_user, "role", "").lower() in ["admin", "administrateur"]

# Fonction améliorée pour détecter l'admin système
def is_system_admin_enhanced():
    """Vérifie si l'utilisateur est administrateur système"""
    if not current_user.is_authenticated:
        return False
    
    # Vérifier le rôle ET le username pour plus de sécurité
    user_role = getattr(current_user, "role", "").lower()
    username = getattr(current_user, "username", "").lower()
    
    # Admin système = rôle admin ET username 'admin'
    return user_role in ["admin", "administrateur"] and username == "admin"

# ---------------- Filtrage ----------------
FILTER_DIC = {
    "sexe": {
        "M": ["M", "Masculin"],
        "F": ["F", "Féminin"]
    },
    "etat": {
        "Inactif": ["Inactif"],
        "Actif": ["Actif"],
        "Muté": ["Muté"],
        "Retraité": ["Retraité"]
    },
    "titre": {
        "Enseignant": ["Enseignant", "Enseignante"],
        "Directeur": ["Directeur", "Directrice"],
        "Directeur adjoint": ["Directeur adjoint", "Directrice adjointe"],
        "Surveillant": ["Surveillant", "Surveillante"],
        "Bibliothécaire": ["Bibliothécaire"],
        "Économe": ["Économe"]
    }
}

@enseignants_bp.route("/")
@login_required
def liste_enseignants():
    """Liste des enseignants - VERSION CORRIGÉE SANS DOUBLONS"""
    data = request.args
    search = data.get("search", "", type=str)
    per_page = data.get("per_page", 10, type=int)
    page = data.get("page", 1, type=int)

    # Détection admin système
    user_is_system_admin = is_system_admin_enhanced()
    
    # Détection cohérente de l'école
    ecole_id = get_current_ecole_id()
    
    # Admin système peut surcharger avec le paramètre URL
    if user_is_system_admin:
        ecole_url = data.get("ecole", "", type=str)
        if ecole_url and ecole_url != str(ecole_id):
            ecole_id = ecole_url
    
    # Récupérer l'école sélectionnée pour affichage
    ecole_selectionnee = None
    if ecole_id:
        ecole_selectionnee = Ecole.query.get(ecole_id)

    # CORRECTION : Construire la requête sans DISTINCT ON pour PostgreSQL
    if user_is_system_admin and ecole_id:
        # Admin système avec école spécifique
        query = Enseignant.query.filter_by(ecole_id=ecole_id)
    elif user_is_system_admin and not ecole_id:
        # Admin système sans école spécifique = toutes les écoles
        query = Enseignant.query
    else:
        # Utilisateurs normaux : seulement leur école
        if not ecole_id:
            return render_template("error.html", message="Aucune école sélectionnée"), 400
        query = Enseignant.query.filter_by(ecole_id=ecole_id)

    # Charger les relations
    query = query.options(
        joinedload(Enseignant.utilisateur),
        joinedload(Enseignant.matieres)
    )

    # Collecte des données pour filtres
    matieres = []
    annees = []
    
    if ecole_id:
        # Matières de cette école
        matieres = Matiere.query.filter_by(ecole_id=ecole_id).with_entities(Matiere.id, Matiere.libelle).all()
        # Années pour cette école
        annees_query = db.session.query(extract("year", Enseignant.date_fonction).label("annee")).filter(Enseignant.ecole_id == ecole_id).distinct().order_by("annee").all()
        annees = [int(a.annee) for a in annees_query if a.annee is not None]
    
    # Recherche
    if search:
        search_like = f"%{search}%"
        query = query.join(Utilisateur, Enseignant.utilisateur)
        
        query = query.filter(
            db.or_(
                Utilisateur.nom.ilike(search_like),
                Utilisateur.prenoms.ilike(search_like),
                cast(Enseignant.date_fonction, String).ilike(search_like),
                Utilisateur.sexe.ilike(search_like),
                Utilisateur.email.ilike(search_like),
                Utilisateur.telephone.ilike(search_like),
                Enseignant.titre.ilike(search_like)
            )
        )

    # Filtrage
    filter_value = data.get("filter", "", type=str)
    if filter_value:
        key, val = filter_value.split(":", 1)
        
        if key == "sexe":
            query = query.join(Utilisateur).filter(Utilisateur.sexe == val)
        elif key == "date_fonction":
            try:
                year_val = int(val)
                query = query.filter(extract("year", Enseignant.date_fonction) == year_val)
            except ValueError:
                pass
        elif key == "matiere" and val:
            try:
                query = query.filter(Enseignant.matieres.any(Matiere.id == val))
            except ValueError:
                pass
        elif key in FILTER_DIC and val in FILTER_DIC[key]:
            if key == "titre" and val in ["Enseignant", "Enseignante"]:
                query = query.filter(Enseignant.titre.in_(FILTER_DIC[key][val]))
            else:
                query = query.filter(getattr(Enseignant, key) == val)
        else:
            query = query.filter(getattr(Enseignant, key) == val)

    # CORRECTION CRITIQUE : Pour éviter les doublons avec PostgreSQL, utiliser group_by au lieu de distinct
    # D'abord, appliquer le group_by pour éliminer les doublons
    query = query.group_by(Enseignant.id)
    
    # Puis trier
    query = query.order_by(
        Enseignant.date_fonction.desc(),
        Enseignant.utilisateur_id
    )

    # Pagination
    try:
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        enseignants = pagination.items
        
        # Vérification des doublons (débogage)
        seen_ids = set()
        duplicates = []
        for ens in enseignants:
            if ens.id in seen_ids:
                duplicates.append(ens.id)
            seen_ids.add(ens.id)
        
        if duplicates:
            current_app.logger.warning(f"⚠️ DOUBLONS détectés dans la page: {duplicates}")
            
    except Exception as e:
        current_app.logger.error(f"❌ Erreur pagination: {str(e)}")
        return render_template("enseignants/error.html", message="Erreur lors du chargement des enseignants"), 500

    # Récupération des données pour le template
    user_role = (getattr(current_user, "role", "guest") or "guest").lower()
    
    # Variables cohérentes
    ecoles_list = get_user_accessible_ecoles() if user_is_system_admin else [ecole_selectionnee] if ecole_selectionnee else []
    
    return render_template("enseignants/ens_liste.html",
                          enseignants=enseignants,
                          pagination=pagination,
                          search=search,
                          per_page=per_page,
                          user_role=user_role,
                          filter_value=filter_value,
                          matieres=matieres,
                          annees=annees,
                          current_ecole_id=ecole_id,
                          selected_ecole_id=ecole_id,
                          is_system_admin=user_is_system_admin,
                          selected_ecole=ecole_selectionnee,
                          ecoles=ecoles_list)

# --------------CRUD-------------
@enseignants_bp.route("/add", methods=["POST"])
@login_required
@ecole_required
def add_ens():
    ecole_id = get_current_ecole_id()
    data = request.form
    utilisateur_id = data.get("utilisateur_id")
    ignore_warning = data.get("ignore_warning") == "true"  # Nouveau paramètre

    # Vérification PRINCIPALE: si cet utilisateur est déjà enregistré comme enseignant dans cette école
    existing_ens = Enseignant.query.filter_by(utilisateur_id=utilisateur_id, ecole_id=ecole_id).first()
    if existing_ens:
        # Récupérer les détails de l'enseignant existant
        utilisateur = existing_ens.utilisateur if hasattr(existing_ens, 'utilisateur') else None
        nom_complet = f"{utilisateur.nom} {utilisateur.prenoms}" if utilisateur else "Cet enseignant"
        
        return jsonify({
            "error": f"{nom_complet} est déjà enregistré comme enseignant dans cette école.",
            "duplicate_id": str(existing_ens.id) if existing_ens else None
        }), 400

    # Récupérer l'utilisateur
    utilisateur = Utilisateur.query.get(utilisateur_id)
    if not utilisateur:
        return jsonify({"error": "Utilisateur non trouvé"}), 404
    
    # Vérification OPTIONNELLE: si un enseignant avec le même nom et prénom existe déjà dans cette école
    # (mais avec un compte utilisateur différent)
    # CORRECTION : Ajouter la condition ignore_warning
    if not ignore_warning:
        enseignant_existant = db.session.query(Enseignant).join(Utilisateur).filter(
            Utilisateur.nom == utilisateur.nom,
            Utilisateur.prenoms == utilisateur.prenoms,
            Enseignant.ecole_id == ecole_id,
            Enseignant.utilisateur_id != utilisateur_id  # Important: exclure le même utilisateur
        ).first()
        
        if enseignant_existant:
            # Si on arrive ici, c'est qu'un enseignant avec le même nom existe mais avec un autre compte utilisateur
            enseignant_existant_user = enseignant_existant.utilisateur
            return jsonify({
                "error": f"Un enseignant nommé {utilisateur.nom} {utilisateur.prenoms} existe déjà dans cette école "
                        f"(compte utilisateur différent: {enseignant_existant_user.email or 'N/A'}). "
                        f"Voulez-vous quand même l'ajouter ?",
                "duplicate_id": str(enseignant_existant.id),
                "warning": True,
                "can_ignore": True  # Nouveau flag pour permettre d'ignorer
            }), 400

    # Vérification 3: si l'utilisateur a déjà un compte enseignant dans une autre école
    # (uniquement pour admin système)
    if is_system_admin_enhanced() and not ignore_warning:
        enseignant_autre_ecole = Enseignant.query.filter_by(utilisateur_id=utilisateur_id).first()
        if enseignant_autre_ecole:
            ecole_existante = Ecole.query.get(enseignant_autre_ecole.ecole_id)
            nom_ecole = ecole_existante.nom if ecole_existante else "une autre école"
            return jsonify({
                "error": f"Cet utilisateur est déjà enseignant dans {nom_ecole}. "
                        f"Voulez-vous vraiment l'ajouter aussi dans cette école ?",
                "warning": True,
                "ecole_existante": nom_ecole,
                "can_ignore": True  # Nouveau flag pour permettre d'ignorer
            }), 400

    # Créer l'enseignant
    enseignant = Enseignant(
        utilisateur_id=utilisateur_id,
        titre=data.get("titre"),
        date_fonction=data.get("date_fonction"),
        etat="Inactif",
        ecole_id=ecole_id
    )
    db.session.add(enseignant)
    db.session.commit()

    # Ajouter les matières sélectionnées
    matieres_ids = request.form.getlist("matiere_id")
    if matieres_ids:
        matieres = Matiere.query.filter(
            Matiere.id.in_(matieres_ids),
            Matiere.ecole_id == ecole_id
        ).all()
        enseignant.matieres = matieres
        db.session.commit()

    return jsonify({"message": "Ajout réussi"}), 200


@enseignants_bp.route("/update/<string:id>", methods=["POST"])
@login_required
def update_ens(id):
    """Modifier un enseignant - CORRECTION : Sans @ecole_required pour admin système"""
    if not is_admin():
        return jsonify({"error": "Accès refusé"}), 403
    
    user_is_system_admin = is_system_admin_enhanced()
    
    if user_is_system_admin:
        # Admin système peut modifier n'importe quel enseignant
        enseignant = Enseignant.query.filter_by(id=id).first_or_404()
    else:
        # Autres admins : seulement les enseignants de leur école
        ecole_id = get_current_ecole_id()
        enseignant = Enseignant.query.filter_by(id=id, ecole_id=ecole_id).first_or_404()
    
    # Contrôle sur l'état
    if enseignant.etat.lower() != 'inactif':
        return jsonify({"error": "État non inactif"}), 403
    
    try:
        # Récupérer l'utilisateur
        utilisateur = enseignant.utilisateur
        if not utilisateur:
            return jsonify({"error": "Utilisateur non trouvé"}), 404
        
        # VÉRIFICATION: si le nom/prénom change, vérifier s'il existe déjà dans cette école
        nouveau_nom = request.form.get("nom")
        nouveau_prenoms = request.form.get("prenoms")
        
        if (nouveau_nom and nouveau_nom != utilisateur.nom) or (nouveau_prenoms and nouveau_prenoms != utilisateur.prenoms):
            # Vérifier s'il existe déjà un enseignant avec ce nom/prénom dans la même école
            ecole_id = enseignant.ecole_id
            nom_a_verifier = nouveau_nom if nouveau_nom else utilisateur.nom
            prenoms_a_verifier = nouveau_prenoms if nouveau_prenoms else utilisateur.prenoms
            
            enseignant_existant = db.session.query(Enseignant).join(Utilisateur).filter(
                Utilisateur.nom == nom_a_verifier,
                Utilisateur.prenoms == prenoms_a_verifier,
                Enseignant.ecole_id == ecole_id,
                Enseignant.id != id  # Exclure l'enseignant en cours de modification
            ).first()
            
            if enseignant_existant:
                return jsonify({
                    "error": f"Un enseignant nommé {nom_a_verifier} {prenoms_a_verifier} existe déjà dans cette école."
                }), 400
        
        # Mettre à jour les champs de base
        for field in ["titre", "date_fonction"]:
            if field in request.form:
                setattr(enseignant, field, request.form.get(field))
        
        # Mettre à jour l'utilisateur si les champs sont présents
        if "nom" in request.form:
            utilisateur.nom = request.form.get("nom")
        if "prenoms" in request.form:
            utilisateur.prenoms = request.form.get("prenoms")
        if "email" in request.form:
            utilisateur.email = request.form.get("email")
        if "telephone" in request.form:
            utilisateur.telephone = request.form.get("telephone")
        if "sexe" in request.form:
            utilisateur.sexe = request.form.get("sexe")
        
        # CORRECTION CRITIQUE : Gestion de la relation many-to-many
        matiere_ids = request.form.getlist('matiere_id')
        current_app.logger.info(f"🔍 Matières reçues pour mise à jour: {matiere_ids}")
        
        # Filtrer les IDs vides
        matiere_ids = [mid for mid in matiere_ids if mid.strip()]
        
        if matiere_ids:
            # Récupérer les matières depuis la base
            ecole_id = enseignant.ecole_id if not user_is_system_admin else get_current_ecole_id()
            nouvelles_matieres = Matiere.query.filter(
                Matiere.id.in_(matiere_ids),
                Matiere.ecole_id == ecole_id
            ).all()
            
            current_app.logger.info(f"✅ Matières trouvées en base: {[m.libelle for m in nouvelles_matieres]}")
            
            # Mettre à jour la relation many-to-many
            enseignant.matieres = nouvelles_matieres
        else:
            # Aucune matière sélectionnée - vider la relation
            enseignant.matieres = []
            current_app.logger.info("ℹ️ Aucune matière sélectionnée - relation vidée")
        
        db.session.commit()
        
        # Préparer la réponse
        matieres_libelles = [m.libelle for m in enseignant.matieres]
        
        return jsonify({
            "message": "Enseignant modifié avec succès",
            "titre": enseignant.titre,
            "date_fonction": enseignant.date_fonction.strftime("%Y-%m-%d") if enseignant.date_fonction else None,
            "matiere": ', '.join(matieres_libelles) if matieres_libelles else None,
            "matiere_ids": [str(m.id) for m in enseignant.matieres]
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"❌ Erreur lors de la modification: {str(e)}")
        import traceback
        current_app.logger.error(f"📋 Stack trace: {traceback.format_exc()}")
        return jsonify({"error": f"Erreur lors de la modification: {str(e)}"}), 500

@enseignants_bp.route("/delete/<string:id>", methods=["POST"])
@login_required
def delete_ens(id):
    """Supprimer un enseignant avec gestion des enseignements associés"""
    if not is_admin():
        return jsonify({"error": "Accès refusé"}), 403
    
    user_is_system_admin = is_system_admin_enhanced()
    
    try:
        # Récupérer l'enseignant
        if user_is_system_admin:
            enseignant = Enseignant.query.filter_by(id=id).first_or_404()
        else:
            ecole_id = get_current_ecole_id()
            enseignant = Enseignant.query.filter_by(id=id, ecole_id=ecole_id).first_or_404()
        
        # Contrôle sur l'état
        if enseignant.etat.lower() != 'inactif':
            return jsonify({"error": "Impossible de supprimer un enseignant dont l'état n'est pas 'Inactif'. Modifiez d'abord son état."}), 403
        
        # OPTION A : Réassigner les enseignements à un autre enseignant
        # Vous pouvez ajouter un paramètre pour choisir un nouvel enseignant
        
        # OPTION B : Supprimer les enseignements associés (si c'est acceptable)
        if hasattr(enseignant, 'enseignements') and enseignant.enseignements:
            # Vérifier si l'utilisateur veut réassigner ou supprimer
            data = request.get_json() or {}
            action = data.get('action', 'delete')  # 'delete' ou 'reassign'
            new_enseignant_id = data.get('new_enseignant_id')
            
            if action == 'reassign' and new_enseignant_id:
                # Réassigner à un autre enseignant
                nouvel_enseignant = Enseignant.query.filter_by(id=new_enseignant_id).first()
                if nouvel_enseignant:
                    for enseignement in enseignant.enseignements:
                        enseignement.enseignant_id = new_enseignant_id
                    db.session.commit()
                else:
                    return jsonify({"error": "Enseignant de remplacement non trouvé"}), 400
            else:
                # Supprimer les enseignements
                for enseignement in enseignant.enseignements:
                    db.session.delete(enseignement)
                db.session.commit()
        
        # Maintenant supprimer l'enseignant
        db.session.delete(enseignant)
        db.session.commit()
        
        return jsonify({
            "success": True, 
            "message": "Enseignant supprimé avec succès"
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erreur lors de la suppression: {str(e)}")
        return jsonify({"error": f"Erreur lors de la suppression: {str(e)}"}), 500

@enseignants_bp.route("/check-dependencies/<string:id>", methods=["GET"])
@login_required
def check_dependencies(id):
    """Vérifier si un enseignant a des dépendances avant suppression"""
    user_is_system_admin = is_system_admin_enhanced()
    
    if user_is_system_admin:
        enseignant = Enseignant.query.filter_by(id=id).first_or_404()
    else:
        ecole_id = get_current_ecole_id()
        enseignant = Enseignant.query.filter_by(id=id, ecole_id=ecole_id).first_or_404()
    
    # Collecter les dépendances
    dependances = []
    
    # Enseignements
    if hasattr(enseignant, 'enseignements') and enseignant.enseignements:
        enseignements_count = len(enseignant.enseignements)
        dependances.append(f"{enseignements_count} enseignement(s)")
        
        # Détails supplémentaires (optionnel)
        enseignements_details = []
        for ens in enseignant.enseignements[:5]:  # Limiter à 5
            # Vous pouvez récupérer plus d'infos sur chaque enseignement
            enseignements_details.append(f"Enseignement #{ens.id[:8]}...")
        
        if enseignements_details:
            dependances.append("Détails: " + ", ".join(enseignements_details))
    
    # Autres vérifications...
    
    if dependances:
        return jsonify({
            "can_delete": False,
            "dependances": dependances,
            "message": f"Cet enseignant a {len(enseignant.enseignements) if hasattr(enseignant, 'enseignements') else 0} élément(s) dépendant(s)",
            "nom": enseignant.utilisateur.nom if hasattr(enseignant, 'utilisateur') and enseignant.utilisateur else '',
            "prenoms": enseignant.utilisateur.prenoms if hasattr(enseignant, 'utilisateur') and enseignant.utilisateur else ''
        }), 400
    
    return jsonify({
        "can_delete": True,
        "nom": enseignant.utilisateur.nom if hasattr(enseignant, 'utilisateur') and enseignant.utilisateur else '',
        "prenoms": enseignant.utilisateur.prenoms if hasattr(enseignant, 'utilisateur') and enseignant.utilisateur else ''
    })

@enseignants_bp.route("/get/<string:enseignant_id>", methods=["GET"])
@login_required
def get_enseignant(enseignant_id):
    """Récupérer un enseignant - CORRECTION : Sans @ecole_required pour admin système"""
    user_is_system_admin = is_system_admin_enhanced()
    
    if user_is_system_admin:
        # Admin système peut voir n'importe quel enseignant
        enseignant = Enseignant.query.filter_by(id=enseignant_id).first_or_404()
    else:
        # Autres utilisateurs : seulement les enseignants de leur école
        ecole_id = get_current_ecole_id()
        enseignant = Enseignant.query.filter_by(id=enseignant_id, ecole_id=ecole_id).first_or_404()
        
    utilisateur = enseignant.utilisateur if hasattr(enseignant, "utilisateur") else None
    matieres_str = ", ".join([m.libelle for m in enseignant.matieres]) if hasattr(enseignant, "matieres") else ""

    return jsonify({
        "id": str(enseignant.id),
        "nom": utilisateur.nom if utilisateur else "",
        "prenoms": utilisateur.prenoms if utilisateur else "",
        "email": utilisateur.email if utilisateur else "",
        "telephone": utilisateur.telephone if utilisateur else "",
        "sexe": utilisateur.sexe if utilisateur else "",
        "photo_filename": utilisateur.photo_filename if utilisateur else None,
        "matiere": matieres_str,
        "titre": enseignant.titre,
        "etat": enseignant.etat,
        "date_fonction": enseignant.date_fonction.strftime("%d/%m/%Y") if enseignant.date_fonction else ""
    })

@enseignants_bp.route("/detail/<string:id>", methods=["GET"])
@login_required
def get_ens(id):
    """Détail d'un enseignant - CORRECTION : Sans @ecole_required pour admin système"""
    user_is_system_admin = is_system_admin_enhanced()
    
    if user_is_system_admin:
        # Admin système peut voir n'importe quel enseignant
        enseignant = Enseignant.query.filter_by(id=id).first_or_404()
    else:
        # Autres utilisateurs : seulement les enseignants de leur école
        ecole_id = get_current_ecole_id()
        enseignant = Enseignant.query.filter_by(id=id, ecole_id=ecole_id).first_or_404()
        
    utilisateur = enseignant.utilisateur
    # CORRECTION : Récupérer les matières via la relation many-to-many
    matieres_data = []
    matiere_ids = []
    
    for matiere in enseignant.matieres:
        matieres_data.append(matiere.libelle)
        matiere_ids.append(str(matiere.id))

    return jsonify({
        "id": str(enseignant.id),
        "utilisateur_id": str(enseignant.utilisateur_id),
        "nom": utilisateur.nom,
        "prenoms": utilisateur.prenoms,
        "sexe": utilisateur.sexe,
        "email": utilisateur.email,
        "telephone": utilisateur.telephone,
        "matiere": ", ".join(matieres_data) if matieres_data else None,
        "matiere_id": ",".join(matiere_ids) if matiere_ids else None,
        "titre": enseignant.titre,
        "date_fonction": enseignant.date_fonction.strftime("%Y-%m-%d") if enseignant.date_fonction else None,
        "photo_filename": utilisateur.photo_filename,
        "etat": enseignant.etat
    })

@enseignants_bp.route("/<string:ens_id>/changer_etat", methods=["POST"])
@login_required
def changer_etat(ens_id):
    """Changer l'état d'un enseignant - CORRECTION : Sans @ecole_required pour admin système"""
    if not is_admin():
        return jsonify({"error": "Accès refusé"}), 403
    
    user_is_system_admin = is_system_admin_enhanced()
    
    if user_is_system_admin:
        # Admin système peut changer l'état de n'importe quel enseignant
        enseignant = Enseignant.query.filter_by(id=ens_id).first_or_404()
    else:
        # Autres admins : seulement les enseignants de leur école
        ecole_id = get_current_ecole_id()
        enseignant = Enseignant.query.filter_by(id=ens_id, ecole_id=ecole_id).first_or_404()

    data = request.get_json()
    action = (data.get("action") or "").lower()

    # Définition du workflow
    transitions = {
        "inactif": {"activer": "Actif"},
        "actif": {"muter": "Muté"},
        "muté": {"retraiter": "Retraité"}
    }

    current_etat = (enseignant.etat or "Inactif").lower()
    next_etat = transitions.get(current_etat, {}).get(action)

    if not next_etat:
        return jsonify({"error": f"Action '{action}' invalide pour l'état {enseignant.etat}"}), 400

    enseignant.etat = next_etat
    db.session.commit()

    return jsonify({"message": "État mis à jour", "etat": enseignant.etat})

@enseignants_bp.route("/options", methods=["GET"])
@login_required
def enseignants_options():
    """
    Renvoie en une seule requête :
      - liste des utilisateurs (selon le rôle : tous pour admin système, seulement l'école pour les autres)
      - liste des matieres (id, libelle) de l'école courante
    CORRECTION : Gestion cohérente avec la liste principale
    """
    try:
        # RÉCUPÉRATION UNIFORMISÉE DE L'ÉCOLE (identique à liste_enseignants)
        ecole_url = request.args.get('ecole', '', type=str)
        user_is_system_admin = is_system_admin_enhanced()
        
        if user_is_system_admin:
            # ADMIN SYSTÈME : Priorité ABSOLUE au paramètre URL (comme liste_enseignants)
            if ecole_url:
                ecole_id = ecole_url
                current_app.logger.info(f"🔍 ADMIN - Options utilisateurs: École URL: {ecole_id}")
            else:
                # Si pas de paramètre URL, on utilise l'école de session
                ecole_id = get_current_ecole_id()
                current_app.logger.info(f"🔍 ADMIN - Options utilisateurs: École session: {ecole_id}")
        else:
            # Non-admin : toujours l'école courante
            ecole_id = get_current_ecole_id()
            current_app.logger.info(f"🔍 NON-ADMIN - Options utilisateurs: École session: {ecole_id}")

        # ========== CHARGEMENT DES UTILISATEURS ==========
        if user_is_system_admin:
            if ecole_id:
                # CORRECTION CRITIQUE : Admin système avec école sélectionnée
                # -> Uniquement les utilisateurs de CETTE école
                users = Utilisateur.query.filter_by(ecole_id=ecole_id).all()
                current_app.logger.info(f"🔍 Admin système avec école {ecole_id} - {len(users)} utilisateurs de cette école")
            else:
                # Admin système SANS école spécifique -> TOUS les utilisateurs
                users = Utilisateur.query.all()
                current_app.logger.info(f"🔍 Admin système sans école - {len(users)} utilisateurs (tous)")
        else:
            # Utilisateurs normaux : seulement leur école
            if not ecole_id:
                return jsonify({"error": "Aucune école sélectionnée"}), 400
            users = Utilisateur.query.filter_by(ecole_id=ecole_id).all()
            current_app.logger.info(f"🔍 Utilisateur normal - École {ecole_id}, {len(users)} utilisateurs")

        # ========== CHARGEMENT DES MATIÈRES ==========
        # CORRECTION : Toujours filtrer par l'école déterminée ci-dessus
        if ecole_id:
            matieres_query = Matiere.query.filter_by(ecole_id=ecole_id)
            matieres = matieres_query.with_entities(Matiere.id, Matiere.libelle).distinct().all()
            current_app.logger.info(f"🔍 Matières pour école {ecole_id}: {len(matieres)}")
        else:
            # Admin système sans école spécifique -> AUCUNE matière
            matieres = []
            current_app.logger.info(f"🔍 Aucune matière (admin système sans école spécifique)")

        # ========== CONSTRUCTION DE LA RÉPONSE ==========
        users_list = []
        for u in users:
            # Récupérer le nom de l'école
            ecole_nom = None
            if u.ecole_id:
                ecole = Ecole.query.get(u.ecole_id)
                ecole_nom = ecole.nom if ecole else "École inconnue"
            
            user_data = {
                "id": str(u.id),
                "nom": u.nom,
                "prenoms": u.prenoms,
                "sexe": u.sexe,
                "email": u.email,
                "telephone": u.telephone,
                "photo_filename": u.photo_filename,
                "ecole_id": str(u.ecole_id) if u.ecole_id else None,
                "ecole_nom": ecole_nom
            }
            users_list.append(user_data)

        # Déduplication des matières
        matieres_dict = {}
        for m in matieres:
            matieres_dict[str(m.id)] = {"id": str(m.id), "libelle": m.libelle}
        
        matieres_list = list(matieres_dict.values())

        response_data = {
            "utilisateurs": users_list, 
            "matieres": matieres_list,
            "debug": {
                "ecole_id": str(ecole_id) if ecole_id else None,
                "ecole_url": ecole_url,
                "is_system_admin": user_is_system_admin,
                "users_count": len(users_list),
                "matieres_count": len(matieres_list)
            }
        }

        return jsonify(response_data), 200

    except Exception as exc:
        import traceback
        current_app.logger.error(f"❌ Erreur chargement options: {traceback.format_exc()}")
        return jsonify({"error": f"Impossible de charger les options: {str(exc)}"}), 500


# ---------------- EXPORTATIONS CORRIGÉES AVEC EN-TÊTE UNIFORMISÉ ----------------
def get_logo_path_enseignants(ecole):
    """Retourne le chemin du logo de l'école - POUR ENSEIGNANTS - VERSION CORRIGÉE"""
    if not ecole or not ecole.logo_filename:
        print(f"❌ Logo non disponible - Ecole: {ecole}, Logo_filename: {getattr(ecole, 'logo_filename', None)}")
        return None
    
    # CHEMIN DÉFINITIF confirmé (identique à eleves.py)
    logo_path = os.path.join(
        current_app.root_path, 
        'gestion_scolaire', 
        'app', 
        'static', 
        'logos', 
        ecole.logo_filename
    )
    
    if os.path.exists(logo_path) and os.path.getsize(logo_path) > 0:
        print(f"✅ Logo trouvé pour enseignants: {logo_path}")
        return logo_path
    
    print(f"❌ Logo non trouvé pour enseignants: {logo_path}")
    return None

@enseignants_bp.route("/export/pdf")
@login_required
@ecole_required
def export_enseignants_pdf():
    """Export PDF de la liste des enseignants - Version avec en-tête uniformisé"""
    try:
        ecole_id = get_current_ecole_id()
        
        # Récupérer tous les enseignants de l'école courante avec leurs relations
        enseignants = Enseignant.query.filter_by(ecole_id=ecole_id).options(
            joinedload(Enseignant.utilisateur),
            joinedload(Enseignant.matieres)
        ).all()

        # Récupérer les informations de l'école COURANTE
        ecole = Ecole.query.get(ecole_id)
        if not ecole:
            return jsonify({"error": "École non trouvée"}), 404

        # DEBUG: Informations sur le logo
        print(f"🔍 RECHERCHE LOGO ENSEIGNANTS - Ecole: {ecole.nom}, Logo_filename: {ecole.logo_filename}")
        logo_path = get_logo_path_enseignants(ecole)
        print(f"📁 Chemin logo final enseignants: {logo_path}")

        # UTILISER LES VALEURS DYNAMIQUES DE L'ÉCOLE COURANTE
        nom_ecole = ecole.nom if ecole.nom else "École non renseignée"
        dre = ecole.dre if ecole.dre else "DRE non renseignée"
        inspection = ecole.inspection if ecole.inspection else "Inspection non renseignée"
        telephone = ecole.telephone1 if ecole.telephone1 else "Téléphone non renseigné"
        devise_ecole = ecole.devise if ecole.devise else "Travail - Liberté - Patrie"

        # Créer le PDF en mémoire
        buffer = io.BytesIO()
        
        def add_page_number(canvas, doc):
            canvas.saveState()
            page_num = canvas.getPageNumber()
            total_pages = doc.page
            
            footer_text = f"Page {page_num}/{total_pages}"
            canvas.setFont('Helvetica', 8)
            canvas.setFillColor(colors.gray)
            canvas.drawRightString(doc.pagesize[0] - 20*mm, 10*mm, footer_text)
            
            info_text = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
            canvas.drawString(20*mm, 10*mm, info_text)
            canvas.restoreState()
        
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4, 
            topMargin=15*mm, 
            bottomMargin=20*mm,
            leftMargin=10*mm,
            rightMargin=10*mm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # ========== EN-TÊTE UNIFORMISÉ (identique à eleves.py) ==========
        logo = None
        if logo_path:
            try:
                # Vérifier que le fichier existe et n'est pas vide
                if os.path.exists(logo_path) and os.path.getsize(logo_path) > 0:
                    logo = Image(logo_path, width=35*mm, height=35*mm)
                    logo.hAlign = 'CENTER'
                    print("✅ Logo chargé avec succès dans le PDF enseignants")
                else:
                    print("❌ Logo trouvé mais fichier vide ou inexistant")
            except Exception as e:
                print(f"❌ Erreur chargement logo enseignants: {e}")
                logo = None
        else:
            print("❌ Aucun chemin de logo fourni pour enseignants")
        
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Normal'],
            fontSize=9,
            alignment=1,
            spaceAfter=0,
            leading=10
        )
        
        small_header_style = ParagraphStyle(
            'SmallHeader',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            spaceAfter=0,
            leading=8
        )
        
        # Structure avec valeurs DYNAMIQUES (identique à eleves.py)
        left_col_content = [
            Paragraph("<b>MINISTÈRE DE L'EDUCATION NATIONALE</b>", header_style),
            Paragraph("-----------", small_header_style),
            Paragraph(f"DIRECTION RÉGIONALE DE L'ÉDUCATION - {dre}", header_style),
            Paragraph("-----------", small_header_style),
            Paragraph(f"INSPECTION DE L'ENSEIGNEMENT SECONDAIRE GÉNÉRAL - {inspection}", header_style)
        ]
        
        right_col_content = [
            Paragraph("<b>RÉPUBLIQUE TOGOLAISE</b>", header_style),
            Paragraph("-----------", small_header_style),
            Paragraph("Travail - Liberté - Patrie", ParagraphStyle(
                'DeviseStyle',
                parent=styles['Normal'],
                fontSize=7,
                alignment=1,
                spaceAfter=0,
                leading=8
            ))
        ]
        
        # Colonne centrale avec logo (identique à eleves.py)
        center_col_content = []
        if logo:
            center_col_content.append(logo)
        else:
            # Si pas de logo, ajouter un espace vide pour l'équilibre
            center_col_content.append(Spacer(1, 35*mm))
            
        center_col_content.extend([
            Paragraph(f"<b>{nom_ecole}</b>", header_style),
            Paragraph(f"Tél: {telephone}", small_header_style),
            Paragraph(f"{devise_ecole}", small_header_style)
        ])

        header_table = Table([[
            left_col_content, 
            center_col_content, 
            right_col_content
        ]], colWidths=[doc.width/3, doc.width/3, doc.width/3])
        
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ]))
        
        elements.append(header_table)
        elements.append(Spacer(1, 6*mm))
        
        separation_line = Table([['']], colWidths=[doc.width])
        separation_line.setStyle(TableStyle([
            ('LINEABOVE', (0, 0), (0, 0), 1, colors.black),
        ]))
        elements.append(separation_line)
        elements.append(Spacer(1, 6*mm))
        
        # ========== TITRE DU DOCUMENT ==========
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=14,
            alignment=1,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=6
        )
        
        title = Paragraph("LISTE DES ENSEIGNANTS", title_style)
        elements.append(title)
        
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            textColor=colors.gray,
            spaceAfter=8
        )
        
        date_text = Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style)
        elements.append(date_text)
        
        # ========== TABLEAU DES ENSEIGNANTS ==========
        headers = [
            'Nom & Prénoms',
            'Sexe', 
            'Date prise\nfonction',
            'Téléphone',
            'Email',
            'Matière(s)'
        ]
        
        data = [headers]
        
        for enseignant in enseignants:
            utilisateur = enseignant.utilisateur
            matieres = ", ".join([m.libelle for m in enseignant.matieres]) if enseignant.matieres else "Non assigné"
            date_fonction = enseignant.date_fonction.strftime("%d/%m/%Y") if enseignant.date_fonction else "Non définie"
            
            # Tronquer les textes longs pour éviter les débordements
            email = (utilisateur.email or "Non renseigné")[:25] + "..." if len(utilisateur.email or "") > 25 else (utilisateur.email or "Non renseigné")
            matieres_trunc = matieres[:30] + "..." if len(matieres) > 30 else matieres
            
            row = [
                f"{utilisateur.nom} {utilisateur.prenoms}",
                utilisateur.sexe,
                date_fonction,
                utilisateur.telephone or "Non renseigné",
                email,
                matieres_trunc
            ]
            data.append(row)
        
        # Création du tableau avec largeurs optimisées
        table = Table(data, colWidths=[
            doc.width * 0.22,  # Nom & Prénoms
            doc.width * 0.08,  # Sexe
            doc.width * 0.12,  # Date fonction
            doc.width * 0.14,  # Téléphone
            doc.width * 0.22,  # Email
            doc.width * 0.22   # Matières
        ], repeatRows=1)
        
        # Style du tableau optimisé
        table_style = TableStyle([
            # En-têtes
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 4),
            
            # Lignes de données
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 1), (1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Bordures
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            
            # Alternance des couleurs des lignes
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
            
            # Padding réduit
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ])
        
        table.setStyle(table_style)
        elements.append(table)
        
        # ========== STATISTIQUES ==========
        elements.append(Spacer(1, 8*mm))
        
        stats_style = ParagraphStyle(
            'StatsStyle',
            parent=styles['Normal'],
            fontSize=8,
            alignment=0,
            textColor=colors.gray,
            spaceAfter=1
        )
        
        total_enseignants = len(enseignants)
        hommes = sum(1 for e in enseignants if e.utilisateur.sexe and e.utilisateur.sexe.upper() == 'M')
        femmes = total_enseignants - hommes
        
        stats_text = [
            Paragraph(f"<b>STATISTIQUES :</b>", stats_style),
            Paragraph(f"Total enseignants : {total_enseignants}", stats_style),
            Paragraph(f"Hommes : {hommes} | Femmes : {femmes}", stats_style),
        ]
        
        for element in stats_text:
            elements.append(element)
        
        # ========== GÉNÉRATION DU PDF ==========
        doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
        buffer.seek(0)
        
        # Retourner le PDF
        filename = f"liste_enseignants_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        
        response = send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"Erreur génération PDF enseignants: {str(e)}")
        return jsonify({"error": f"Erreur lors de la génération du PDF: {str(e)}"}), 500

@enseignants_bp.route("/export/excel")
@login_required
@ecole_required
def export_enseignants_excel():
    """Export Excel de la liste des enseignants avec en-tête uniformisé"""
    try:
        ecole_id = get_current_ecole_id()
        
        # Récupérer les données de l'école courante
        enseignants = Enseignant.query.filter_by(ecole_id=ecole_id).options(
            joinedload(Enseignant.utilisateur),
            joinedload(Enseignant.matieres)
        ).all()

        # Récupérer les informations de l'école COURANTE
        ecole = Ecole.query.get(ecole_id)
        if not ecole:
            return jsonify({"error": "École non trouvée"}), 404

        # CORRECTION: Récupérer le logo pour Excel
        logo_path = get_logo_path_enseignants(ecole)
        print(f"📊 Excel enseignants - Chemin logo: {logo_path}")

        # UTILISER LES VALEURS DYNAMIQUES DE L'ÉCOLE COURANTE
        nom_ecole = ecole.nom if ecole.nom else "École non renseignée"
        dre = ecole.dre if ecole.dre else "DRE non renseignée"
        inspection = ecole.inspection if ecole.inspection else "Inspection non renseignée"
        telephone = ecole.telephone1 if ecole.telephone1 else "Téléphone non renseigné"
        devise_ecole = ecole.devise if ecole.devise else "Travail - Liberté - Patrie"
        
        # Créer le fichier Excel
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Liste des enseignants"
        
        # ========== EN-TÊTE EXCEL UNIFORMISÉ AVEC LOGO ==========
        center_align = Alignment(horizontal="center", vertical="center")
        
        # CORRECTION: Structure d'en-tête optimisée avec logo
        current_row = 1
        
        # Structure d'en-tête textuelle
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = "MINISTÈRE DE L'EDUCATION NATIONALE"
        ws[f'A{current_row}'].font = Font(bold=True, size=10)
        ws[f'A{current_row}'].alignment = center_align
        
        ws.merge_cells(f'G{current_row}:J{current_row}')
        ws[f'G{current_row}'] = "RÉPUBLIQUE TOGOLAISE"
        ws[f'G{current_row}'].font = Font(bold=True, size=10)
        ws[f'G{current_row}'].alignment = center_align
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = "-----------"
        ws[f'A{current_row}'].alignment = center_align
        
        ws.merge_cells(f'G{current_row}:J{current_row}')
        ws[f'G{current_row}'] = "-----------"
        ws[f'G{current_row}'].alignment = center_align
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = f"DIRECTION RÉGIONALE DE L'ÉDUCATION - {dre}"
        ws[f'A{current_row}'].font = Font(size=9)
        ws[f'A{current_row}'].alignment = center_align
        
        ws.merge_cells(f'G{current_row}:J{current_row}')
        ws[f'G{current_row}'] = "Travail - Liberté - Patrie"
        ws[f'G{current_row}'].font = Font(bold=True, size=8)
        ws[f'G{current_row}'].alignment = center_align
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = "-----------"
        ws[f'A{current_row}'].alignment = center_align
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = f"INSPECTION DE L'ENSEIGNEMENT SECONDAIRE GÉNÉRAL - {inspection}"
        ws[f'A{current_row}'].font = Font(size=9)
        ws[f'A{current_row}'].alignment = center_align
        
        # Nom de l'école et informations
        current_row += 1
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'] = nom_ecole
        ws[f'A{current_row}'].font = Font(bold=True, size=14, color="2C3E50")
        ws[f'A{current_row}'].alignment = center_align
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'] = f"Tél: {telephone} - {devise_ecole}"
        ws[f'A{current_row}'].alignment = center_align
        
        # Titre du document
        current_row += 1
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'] = "LISTE DES ENSEIGNANTS"
        ws[f'A{current_row}'].font = Font(bold=True, size=16, color="2C3E50")
        ws[f'A{current_row}'].alignment = center_align
        
        # Date de génération
        current_row += 1
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws[f'A{current_row}'].font = Font(size=8, italic=True, color="666666")
        ws[f'A{current_row}'].alignment = center_align
        
        # Ligne vide
        current_row += 1
        ws[f'A{current_row}'] = ""
        
        # CORRECTION: Ajouter le logo APRÈS la structure d'en-tête
        if logo_path and os.path.exists(logo_path):
            try:
                from openpyxl.drawing.image import Image as ExcelImage
                img = ExcelImage(logo_path)
                
                # Redimensionner le logo pour qu'il soit discret
                img.width = 50
                img.height = 50
                
                # Placer le logo en dehors de la zone de texte (colonne K, ligne 1)
                img.anchor = 'K1'
                ws.add_image(img)
                print("✅ Logo ajouté discrètement dans Excel enseignants (colonne K)")
                
                # Ajuster la largeur de la colonne K pour accommoder le logo
                ws.column_dimensions['K'].width = 15
                
            except Exception as e:
                print(f"❌ Erreur ajout logo Excel enseignants: {e}")
        
        # ========== TABLEAU DES DONNÉES - MÊMES COLONNES QUE PDF ==========
        # CORRECTION: Utiliser les mêmes colonnes que le PDF
        headers = [
            'Nom & Prénoms',  # Même que PDF
            'Sexe',           # Même que PDF  
            'Date prise fonction',  # Même que PDF
            'Téléphone',      # Même que PDF
            'Email',          # Même que PDF
            'Matière(s)'      # Même que PDF
        ]
        
        # Commencer après l'en-tête
        start_row = current_row + 1
        
        # Ajouter les en-têtes manuellement pour éviter les problèmes d'indexation
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = center_align
        
        # Données - MÊME FORMAT QUE PDF
        for enseignant in enseignants:
            utilisateur = enseignant.utilisateur
            matieres = ", ".join([m.libelle for m in enseignant.matieres]) if enseignant.matieres else "Non assigné"
            date_fonction = enseignant.date_fonction.strftime("%d/%m/%Y") if enseignant.date_fonction else "Non définie"
            
            # Même format que PDF (pas de troncature pour Excel)
            row_data = [
                f"{utilisateur.nom} {utilisateur.prenoms}",  # Nom & Prénoms combinés
                utilisateur.sexe,                            # Sexe
                date_fonction,                               # Date fonction
                utilisateur.telephone or "Non renseigné",    # Téléphone
                utilisateur.email or "Non renseigné",        # Email
                matieres                                     # Matières
            ]
            
            # Ajouter la ligne
            row_num = start_row + len([e for e in enseignants if enseignants.index(e) < enseignants.index(enseignant)]) + 1
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = value
        
        # ========== MISE EN FORME FINALE ==========
        # Ajuster les largeurs de colonnes pour correspondre au PDF
        column_widths = {
            'A': 30,  # Nom & Prénoms (plus large car combiné)
            'B': 10,  # Sexe
            'C': 18,  # Date fonction
            'D': 15,  # Téléphone
            'E': 25,  # Email
            'F': 35   # Matières (plus large pour liste complète)
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Centrer les colonnes comme dans le PDF
        for row in range(start_row + 1, ws.max_row + 1):
            ws[f'B{row}'].alignment = center_align  # Sexe centré
            # Les autres colonnes alignées à gauche comme dans le PDF
        
        # Ajouter des bordures au tableau
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, ws.max_row + 1):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        # ========== STATISTIQUES - MÊME QUE PDF ==========
        stats_row = ws.max_row + 2
        total_enseignants = len(enseignants)
        hommes = sum(1 for e in enseignants if e.utilisateur.sexe and e.utilisateur.sexe.upper() == 'M')
        femmes = total_enseignants - hommes
        
        ws.merge_cells(f'A{stats_row}:F{stats_row}')
        ws[f'A{stats_row}'] = "STATISTIQUES"
        ws[f'A{stats_row}'].font = Font(bold=True, size=10)
        
        ws.merge_cells(f'A{stats_row + 1}:F{stats_row + 1}')
        ws[f'A{stats_row + 1}'] = f"Total enseignants : {total_enseignants} | Hommes : {hommes} | Femmes : {femmes}"
        ws[f'A{stats_row + 1}'].font = Font(size=9)
        
        # Sauvegarder
        wb.save(buffer)
        buffer.seek(0)
        
        # Retourner le fichier
        filename = f"liste_enseignants_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        response = send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"Erreur génération Excel enseignants: {str(e)}")
        return jsonify({"error": f"Erreur lors de la génération du Excel: {str(e)}"}), 500

@enseignants_bp.route("/init-matieres-ecole", methods=["POST"])
@login_required
@ecole_required
def init_matieres_ecole():
    """Initialise les matières pour une école qui n'en a pas"""
    try:
        ecole_id = get_current_ecole_id()
        
        # Vérifier si l'école existe
        ecole = Ecole.query.get(ecole_id)
        if not ecole:
            return jsonify({"error": "École non trouvée"}), 404
        
        # Vérifier si l'école a déjà des matières
        matieres_existantes = Matiere.query.filter_by(ecole_id=ecole_id).count()
        if matieres_existantes > 0:
            return jsonify({
                "message": f"L'école a déjà {matieres_existantes} matière(s)",
                "action": "aucune"
            }), 200
        
        # Liste des matières standard à créer avec des codes UNIQUES
        matieres_standard = [
            {"libelle": "Mathématiques", "code": "MATH"},
            {"libelle": "Français", "code": "FRANCAIS"}, 
            {"libelle": "Physiques Chimies et Technologies", "code": "PHYSIQUE"},
            {"libelle": "Histoire-Géographie", "code": "HISTGEO"},
            {"libelle": "Sciences de la Vie et de la Terre", "code": "SVT"},
            {"libelle": "Anglais", "code": "ANGLAIS"},
            {"libelle": "Education civique et morale", "code": "EDUCIVIQUE"},
            {"libelle": "Agriculture", "code": "AGRICULT"},
            {"libelle": "Ewe", "code": "EWE"},
            {"libelle": "Kabyè", "code": "KABYE"},
            {"libelle": "Education physique et sportive", "code": "EPS"},
            {"libelle": "Musique", "code": "MUSIQUE"},
            {"libelle": "Dessin", "code": "DESSIN"},
            {"libelle": "Enseignement Ménager Couture", "code": "MENAGER"}
        ]
        
        # Créer les matières
        matieres_creees = []
        for matiere_data in matieres_standard:
            # Vérifier si le code existe déjà globalement (au cas où)
            code_existe = Matiere.query.filter_by(code=matiere_data["code"]).first()
            if code_existe:
                # Ajouter un suffixe unique si le code existe déjà
                nouveau_code = f"{matiere_data['code']}_{ecole_id[:4]}"
                matiere_data["code"] = nouveau_code
            
            matiere = Matiere(
                libelle=matiere_data["libelle"],
                code=matiere_data["code"],
                ecole_id=ecole_id,
                etat="Actif"
            )
            db.session.add(matiere)
            matieres_creees.append(matiere_data["libelle"])
        
        db.session.commit()
        
        current_app.logger.info(f"✅ {len(matieres_creees)} matières créées pour l'école {ecole.nom}")
        
        return jsonify({
            "message": f"{len(matieres_creees)} matières créées avec succès",
            "matieres": matieres_creees,
            "action": "créées"
        }), 201
        
    except Exception as exc:
        db.session.rollback()
        current_app.logger.error(f"❌ Erreur création matières: {str(exc)}")
        return jsonify({"error": f"Erreur lors de la création: {str(exc)}"}), 500
        