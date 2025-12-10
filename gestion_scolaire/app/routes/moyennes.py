from flask import Blueprint, request, render_template, redirect, url_for, flash, jsonify, session, current_app, Response, send_file
from extensions import db
from ..models import Moyenne, Eleve, Classe, Note, Appreciations, Enseignement, Ecole,SystemeEvaluation
from flask_login import current_user, login_required
from gestion_scolaire.app.tasks.batchMoyennes import calculer_moyennes
import uuid, time, os, io
from datetime import datetime
import re
import logging
from uuid import UUID
from ..utils import ecole_required, get_current_ecole_id
import csv
import pandas as pd
from io import BytesIO, StringIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch, mm
from reportlab.platypus import Image
from sqlalchemy import text

moyennes_bp = Blueprint('moyennes', __name__, url_prefix='/moyennes')

# ==================== CONFIGURATION DE SÉCURITÉ ====================

# Configuration du logger de sécurité
security_logger = logging.getLogger('security')
if not security_logger.handlers:
    handler = logging.StreamHandler()
    formatter = logging.Formatter(
        '%(asctime)s - SECURITY - %(levelname)s - [%(ip)s] %(user_id)s - %(message)s'
    )
    handler.setFormatter(formatter)
    security_logger.addHandler(handler)
    security_logger.setLevel(logging.INFO)




from flask import Blueprint, request, render_template, redirect, url_for, flash, jsonify, session, current_app, Response, send_file
from extensions import db
from ..models import Moyenne, Eleve, Classe, Note, Appreciations, Enseignement, Ecole, SystemeEvaluation
from flask_login import current_user, login_required
from gestion_scolaire.app.tasks.batchMoyennes import calculer_moyennes
import uuid, time, os, io
from datetime import datetime
import re
import logging
from uuid import UUID
from ..utils import ecole_required, get_current_ecole_id
import csv
import pandas as pd
from io import BytesIO, StringIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch, mm
from reportlab.platypus import Image
from sqlalchemy import text

moyennes_bp = Blueprint('moyennes', __name__, url_prefix='/moyennes')

# ==================== CONFIGURATION DE SÉCURITÉ ====================

# Configuration du logger de sécurité
security_logger = logging.getLogger('security')
if not security_logger.handlers:
    handler = logging.StreamHandler()
    formatter = logging.Formatter(
        '%(asctime)s - SECURITY - %(levelname)s - [%(ip)s] %(user_id)s - %(message)s'
    )
    handler.setFormatter(formatter)
    security_logger.addHandler(handler)
    security_logger.setLevel(logging.INFO)

def log_security_event(action, resource, status="success", details=None):
    """Journalise les événements de sécurité"""
    user_id = str(current_user.id) if current_user.is_authenticated else "anonymous"
    ip = request.remote_addr if request else "unknown"
    
    log_data = {
        "user_id": user_id,
        "action": action,
        "resource": resource,
        "status": status,
        "ip": ip,
        "user_agent": request.headers.get('User-Agent', 'unknown') if request else "unknown"
    }
    
    if details:
        log_data["details"] = details
    
    security_logger.info(f"{action} - {resource} - {status}", extra={
        'ip': ip, 
        'user_id': user_id
    })

def validate_uuid(uuid_string):
    """Valide le format UUID de manière sécurisée"""
    if not uuid_string or not isinstance(uuid_string, str):
        return False
    
    try:
        uuid_obj = UUID(uuid_string)
        return str(uuid_obj) == uuid_string
    except (ValueError, AttributeError):
        return False

def sanitize_input(param, max_length=100, allowed_pattern=None):
    """Nettoie et valide les paramètres d'entrée"""
    if param is None:
        return None
    
    param_str = str(param).strip()
    
    # Vérification de la longueur
    if len(param_str) > max_length:
        log_security_event(
            "input_validation", 
            "parameter_length", 
            "failed", 
            {"parameter": str(param)[:50], "max_length": max_length}
        )
        return None
    
    # Nettoyage des caractères dangereux
    param_str = re.sub(r'[<>"\';()&+]', '', param_str)
    
    # Validation par pattern si spécifié
    if allowed_pattern and not re.match(allowed_pattern, param_str):
        log_security_event(
            "input_validation", 
            "parameter_pattern", 
            "failed", 
            {"parameter": param_str, "pattern": allowed_pattern}
        )
        return None
    
    return param_str

def validate_ecole_access_for_classe(classe_id, ecole_id):
    """Valide l'accès à une classe spécifique pour l'école"""
    try:
        if not classe_id:
            return True  # Aucune classe spécifiée = accès à toutes
        
        classe = Classe.query.filter_by(id=UUID(classe_id), ecole_id=ecole_id).first()
        return classe is not None
    except (ValueError, AttributeError):
        return False

def validate_annee_scolaire_format(annee_scolaire):
    """Valide le format de l'année scolaire"""
    if not annee_scolaire:
        return False
    
    pattern = r'^\d{4}-\d{4}$'  # Format: 2024-2025
    return bool(re.match(pattern, annee_scolaire))

def rate_limit_check():
    """Vérification basique de rate limiting basée sur la session"""
    if 'request_count' not in session:
        session['request_count'] = 1
        session['first_request'] = time.time()
    else:
        session['request_count'] += 1
    
    # Limite: 100 requêtes par minute
    current_time = time.time()
    if current_time - session['first_request'] > 60:  # 1 minute
        session['request_count'] = 1
        session['first_request'] = current_time
    elif session['request_count'] > 100:
        log_security_event(
            "rate_limit", 
            f"ip_{request.remote_addr}", 
            "failed", 
            {"request_count": session['request_count']}
        )
        return False
    
    return True

# ---------------- EN-TÊTE UNIFORMISÉ POUR MOYENNES ----------------
def get_logo_path_moyennes(ecole):
    """Retourne le chemin du logo de l'école - POUR MOYENNES"""
    if not ecole or not ecole.logo_filename:
        return None
    
    possible_paths = [
        os.path.join(current_app.root_path, 'gestion_scolaire', 'app', 'static', 'logos', ecole.logo_filename),
        os.path.join(current_app.root_path, 'static', 'images', 'logos_ecoles', ecole.logo_filename),
        os.path.join(current_app.root_path, 'static', 'uploads', 'logos', ecole.logo_filename),
    ]
    
    for logo_path in possible_paths:
        if os.path.exists(logo_path) and os.path.getsize(logo_path) > 0:
            return logo_path
    
    return None

# ==================== FONCTIONS DE DIAGNOSTIC ET CORRECTION DES DOUBLONS ====================

def check_moyennes_duplicates_detailed(ecole_id, classe_id, annee_scolaire, periode, type_systeme='trimestriel'):
    """Vérifie en détail les doublons de moyennes avec toutes les informations"""
    try:
        # Utiliser moy_trim pour la requête SQL
        duplicates_query = """
        SELECT 
            e.id as eleve_id,
            e.nom,
            e.prenoms,
            COUNT(m.id) as nb_moyennes,
            STRING_AGG(m.id::text, ', ') as ids_moyennes,
            STRING_AGG(ROUND(m.moy_trim::numeric, 2)::text, ', ') as moyennes
        FROM geslog_schema.moyennes m
        JOIN geslog_schema.eleves e ON m.eleve_id = e.id
        JOIN geslog_schema.classes c ON e.classe_id = c.id
        WHERE c.ecole_id = :ecole_id
        AND c.id = :classe_id
        AND m.annee_scolaire = :annee_scolaire
        AND m.periode = :periode
        AND m.type_periode = :type_periode
        GROUP BY e.id, e.nom, e.prenoms
        HAVING COUNT(m.id) > 1
        ORDER BY e.nom, e.prenoms
        """
        
        duplicates = db.session.execute(
            text(duplicates_query),
            {
                'ecole_id': ecole_id,
                'classe_id': classe_id,
                'annee_scolaire': annee_scolaire,
                'periode': periode,
                'type_periode': 'semestre' if type_systeme == 'semestriel' else 'trimestre'
            }
        ).fetchall()
        
        return duplicates
    except Exception as e:
        current_app.logger.error(f"Erreur vérification détaillée doublons: {str(e)}")
        return []

def clean_moyennes_duplicates(ecole_id, classe_id, annee_scolaire, periode, type_systeme='trimestriel'):
    """Nettoie les doublons de moyennes en gardant la plus récente"""
    try:
        # Vérifier d'abord les doublons
        duplicates = check_moyennes_duplicates_detailed(ecole_id, classe_id, annee_scolaire, periode, type_systeme)
        
        if not duplicates:
            return {"status": "success", "message": "Aucun doublon détecté", "deleted": 0}
        
        delete_query = """
        DELETE FROM geslog_schema.moyennes 
        WHERE id IN (
            SELECT m1.id
            FROM geslog_schema.moyennes m1
            JOIN geslog_schema.eleves e ON m1.eleve_id = e.id
            JOIN geslog_schema.classes c ON e.classe_id = c.id
            WHERE c.ecole_id = :ecole_id
            AND c.id = :classe_id
            AND m1.annee_scolaire = :annee_scolaire
            AND m1.periode = :periode
            AND m1.type_periode = :type_periode
            AND m1.id NOT IN (
                SELECT MAX(m2.id)
                FROM geslog_schema.moyennes m2
                JOIN geslog_schema.eleves e2 ON m2.eleve_id = e2.id
                JOIN geslog_schema.classes c2 ON e2.classe_id = c2.id
                WHERE c2.ecole_id = :ecole_id
                AND c2.id = :classe_id
                AND m2.annee_scolaire = :annee_scolaire
                AND m2.periode = :periode
                AND m2.type_periode = :type_periode
                GROUP BY e2.id
            )
        )
        """
        
        result = db.session.execute(
            text(delete_query),
            {
                'ecole_id': ecole_id,
                'classe_id': classe_id,
                'annee_scolaire': annee_scolaire,
                'periode': periode,
                'type_periode': 'semestre' if type_systeme == 'semestriel' else 'trimestre'
            }
        )
        
        db.session.commit()
        
        # Vérifier après suppression
        duplicates_after = check_moyennes_duplicates_detailed(ecole_id, classe_id, annee_scolaire, periode, type_systeme)
        
        return {
            "status": "success", 
            "message": f"{result.rowcount} doublons supprimés",
            "deleted": result.rowcount,
            "duplicates_before": len(duplicates),
            "duplicates_after": len(duplicates_after),
            "details": [dict(dup._mapping) for dup in duplicates]
        }
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erreur nettoyage doublons: {str(e)}")
        return {"status": "error", "message": str(e)}

def get_unique_moyennes_data(ecole_id, classe_id, annee_scolaire, periode, type_systeme='trimestriel'):
    """Récupère les données de moyennes uniques par élève"""
    try:
        # Utiliser moy_trim dans la requête
        unique_query = """
        SELECT DISTINCT ON (e.id)
            e.id,
            e.nom,
            e.prenoms,
            m.moy_trim,
            m.moy_gen,
            m.classement_str,
            a.libelle as appreciation,
            m.moy_class
        FROM geslog_schema.eleves e
        JOIN geslog_schema.classes c ON e.classe_id = c.id
        JOIN geslog_schema.moyennes m ON e.id = m.eleve_id
        LEFT JOIN geslog_schema.appreciations a ON m.appreciation_id = a.id
        WHERE c.ecole_id = :ecole_id
        AND c.id = :classe_id
        AND m.annee_scolaire = :annee_scolaire
        AND m.periode = :periode
        AND m.type_periode = :type_periode
        ORDER BY e.id, m.id DESC
        """
        
        items = db.session.execute(
            text(unique_query),
            {
                'ecole_id': ecole_id,
                'classe_id': classe_id,
                'annee_scolaire': annee_scolaire,
                'periode': periode,
                'type_periode': 'semestre' if type_systeme == 'semestriel' else 'trimestre'
            }
        ).fetchall()
        
        return items
    except Exception as e:
        current_app.logger.error(f"Erreur récupération données uniques: {str(e)}")
        return []

def get_systeme_evaluation(ecole_id):
    """Récupère ou crée la configuration du système d'évaluation de l'école"""
    systeme = SystemeEvaluation.query.filter_by(ecole_id=ecole_id).first()
    if not systeme:
        # Créer une configuration par défaut
        systeme = SystemeEvaluation(
            id=str(uuid.uuid4()),
            ecole_id=ecole_id,
            type_systeme='trimestriel'
        )
        db.session.add(systeme)
        db.session.commit()
    return systeme

def get_periodes_for_systeme(type_systeme):
    """Retourne les périodes disponibles selon le système"""
    if type_systeme == 'semestriel':
        return [1, 2]
    else:
        return [1, 2, 3]

def get_periode_display(type_systeme, periode):
    """Retourne l'affichage de la période"""
    if type_systeme == 'semestriel':
        return f"Semestre {periode}"
    else:
        return f"Trimestre {periode}"

# ==================== ROUTES SÉCURISÉES ====================

@moyennes_bp.route("/")
@login_required
@ecole_required
def liste_moyennes():
    """Liste des moyennes - VERSION CORRIGÉE AVEC EFFECTIFS EXACTS"""
    try:
        if not rate_limit_check():
            flash("Trop de requêtes. Veuillez réessayer plus tard.", "error")
            return redirect(url_for('moyennes.liste_moyennes'))
        
        ecole_id = get_current_ecole_id()
        
        # Récupération des paramètres
        search = request.args.get("search", "").strip()
        classe_id = request.args.get("classe_id", "").strip()
        annee_scolaire = request.args.get("annee_scolaire", "").strip()
        
        # Récupérer la configuration du système d'évaluation
        systeme_eval = SystemeEvaluation.query.filter_by(ecole_id=ecole_id).first()
        type_systeme = systeme_eval.type_systeme if systeme_eval else 'trimestriel'
        
        # Validation de la période
        try:
            periode = int(request.args.get("periode", "1"))
            if type_systeme == 'semestriel' and periode not in [1, 2]:
                periode = 1
            elif type_systeme == 'trimestriel' and periode not in [1, 2, 3]:
                periode = 1
        except (ValueError, TypeError):
            periode = 1

        # Pagination
        try:
            page = max(1, int(request.args.get("page", 1)))
            per_page = min(max(1, int(request.args.get("per_page", 10))), 100)
        except (ValueError, TypeError):
            page = 1
            per_page = 10

        current_app.logger.info(f"🔍 DEBUG liste_moyennes - Système: {type_systeme}, Période: {periode}")

        # Vérification d'accès à la classe
        if classe_id and not validate_ecole_access_for_classe(classe_id, ecole_id):
            log_security_event("unauthorized_classe_access", f"classe_{classe_id}", "failed")
            flash("Accès non autorisé à cette classe.", "error")
            return redirect(url_for('moyennes.liste_moyennes'))

        # Récupérer les années scolaires
        annees_scolaires_query = (
            db.session.query(Moyenne.annee_scolaire)
            .join(Eleve).join(Classe)
            .filter(Classe.ecole_id == ecole_id)
            .distinct()
            .all()
        )
        annees_scolaires = [a[0] for a in annees_scolaires_query if a[0]]
        
        if not annees_scolaires:
            current_year = datetime.now().year
            annees_scolaires = [f"{current_year-1}-{current_year}", f"{current_year}-{current_year+1}"]

        # ========== REQUÊTE PRINCIPALE ==========
        query = (
            db.session.query(
                Eleve.id.label("eleve_id"),
                Eleve.nom,
                Eleve.prenoms,
                Classe.nom.label("classe_nom"),
                Classe.effectif.label("classe_effectif"),
                Moyenne.moy_periode.label("moy_trim"),
                Moyenne.moy_gen,
                Moyenne.classement,
                Moyenne.classement_str,
                Moyenne.classement_gen,
                Moyenne.moy_class,
                Moyenne.moy_forte,
                Moyenne.moy_faible,
                Moyenne.eff_comp.label("effectif_composants"),
                Appreciations.libelle.label("appreciation"),
            )
            .join(Eleve, Eleve.id == Moyenne.eleve_id)
            .join(Classe, Classe.id == Eleve.classe_id)
            .outerjoin(Appreciations, Appreciations.id == Moyenne.appreciation_id)
            .filter(Moyenne.periode == periode)
            .filter(Classe.ecole_id == ecole_id)
        )

        # Application des filtres
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                db.or_(
                    Eleve.nom.ilike(search_term),
                    Eleve.prenoms.ilike(search_term),
                )
            )
        
        if classe_id:
            query = query.filter(Eleve.classe_id == UUID(classe_id))
        
        if annee_scolaire:
            query = query.filter(Moyenne.annee_scolaire == annee_scolaire)

        # Tri par classement
        query = query.order_by(Moyenne.classement.asc())

        # Pagination
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        items = pagination.items
        
        # Récupérer les classes de l'école
        classes = Classe.query.filter_by(ecole_id=ecole_id).order_by(Classe.nom).all()

        # ========== CALCUL EXACT DES EFFECTIFS ==========
        effectif_total_classe = 0
        effectif_composants_classe = 0
        
        if classe_id:
            # Récupérer la classe
            classe = Classe.query.filter_by(id=UUID(classe_id)).first()
            if classe:
                # Effectif total = effectif de la classe
                effectif_total_classe = classe.effectif
                
                # Effectif composé = nombre d'élèves ayant effectivement une moyenne
                # pour cette période et année scolaire
                effectif_composants_classe = (
                    db.session.query(db.func.count(Moyenne.id))
                    .join(Eleve)
                    .filter(
                        Eleve.classe_id == UUID(classe_id),
                        Moyenne.annee_scolaire == annee_scolaire if annee_scolaire else True,
                        Moyenne.periode == periode,
                        Moyenne.moy_periode > 0
                    )
                    .scalar() or 0
                )
        else:
            # Si pas de classe spécifique, calculer sur tous les items
            # Mais dans ce cas, les effectifs n'ont pas de sens
            effectif_total_classe = 0
            effectif_composants_classe = 0

        # Calcul du récapitulatif CLASSE (pas seulement de la page courante)
        if classe_id:
            # Récupérer TOUTES les moyennes de la classe (sans pagination)
            toutes_moyennes = (
                db.session.query(Moyenne.moy_periode)
                .join(Eleve)
                .filter(
                    Eleve.classe_id == UUID(classe_id),
                    Moyenne.annee_scolaire == annee_scolaire if annee_scolaire else True,
                    Moyenne.periode == periode,
                    Moyenne.moy_periode > 0
                )
                .all()
            )
            
            moyennes_valides = [m[0] for m in toutes_moyennes if m[0] is not None and m[0] > 0]
            
            if moyennes_valides:
                moy_forte = max(moyennes_valides)
                moy_faible = min(moyennes_valides)
                moy_class = round((moy_forte + moy_faible) / 2, 2)
            else:
                moy_forte = moy_faible = moy_class = 0
            
            classe_recap = {
                "moy_class": moy_class,
                "moy_forte": moy_forte,
                "moy_faible": moy_faible,
                "effectif_composants": effectif_composants_classe,  # Toute la classe
                "effectif_total": effectif_total_classe,  # Effectif de la classe
            }
        else:
            # Si pas de classe spécifique, calculer sur les items de la page
            moyennes_valides = [i.moy_trim for i in items if i.moy_trim is not None and i.moy_trim > 0]
            
            if moyennes_valides:
                moy_forte = max(moyennes_valides)
                moy_faible = min(moyennes_valides)
                moy_class = round((moy_forte + moy_faible) / 2, 2)
                effectif_composants = len(moyennes_valides)
            else:
                moy_forte = moy_faible = moy_class = 0
                effectif_composants = 0
            
            effectif_total = len(items) if items else 0
            
            classe_recap = {
                "moy_class": moy_class,
                "moy_forte": moy_forte,
                "moy_faible": moy_faible,
                "effectif_composants": effectif_composants,
                "effectif_total": effectif_total,
            }

        return render_template(
            "moyennes/list_moy.html",
            items=items,
            pagination=pagination,
            search=search,
            classe_id=classe_id,
            annee_scolaire=annee_scolaire,
            periode=periode,
            type_systeme=type_systeme,
            per_page=per_page,
            classes=classes,
            annees_scolaires=annees_scolaires,
            classe_recap=classe_recap,
            ecole_id=ecole_id
        )
        
    except Exception as e:
        current_app.logger.error(f"Erreur liste moyennes: {str(e)}", exc_info=True)
        log_security_event("page_error", "liste_moyennes", "failed", {"error": str(e)})
        flash("Erreur lors du chargement des données.", "error")
        return redirect(url_for('moyennes.liste_moyennes'))

@moyennes_bp.route("/verifier-appreciations")
@login_required
@ecole_required
def verifier_appreciations():
    """Vérifie les incohérences dans les appréciations"""
    try:
        ecole_id = get_current_ecole_id()
        
        # Récupérer les incohérences
        query = """
        SELECT 
            m.id,
            m.moy_periode,
            a.libelle as appreciation_actuelle,
            se.baremes_appreciations
        FROM geslog_schema.moyennes m
        LEFT JOIN geslog_schema.appreciations a ON m.appreciation_id = a.id
        LEFT JOIN geslog_schema.systeme_evaluation se ON m.ecole_id = se.ecole_id
        WHERE m.ecole_id = :ecole_id
        AND m.moy_periode IS NOT NULL
        AND m.moy_periode > 0
        ORDER BY m.moy_periode DESC
        """
        
        result = db.session.execute(text(query), {'ecole_id': ecole_id}).fetchall()
        
        incohérences = []
        for row in result:
            moy = row.moy_periode
            app_actuelle = row.appreciation_actuelle
            baremes = row.baremes_appreciations or []
            
            # Vérifier si l'appréciation correspond aux barèmes
            correspond = False
            for bareme in baremes:
                if bareme['min'] <= moy <= bareme['max']:
                    if bareme['libelle'] == app_actuelle:
                        correspond = True
                    break
            
            if not correspond:
                incohérences.append({
                    'moyenne': moy,
                    'appreciation_actuelle': app_actuelle,
                    'devrait_etre': baremes[0]['libelle'] if baremes else 'Inconnu'
                })
        
        return render_template(
            "moyennes/verifier_appreciations.html",
            incohérences=incohérences,
            total=len(result),
            incohérentes=len(incohérences)
        )
        
    except Exception as e:
        current_app.logger.error(f"Erreur vérification appréciations: {str(e)}")
        flash("Erreur lors de la vérification", "error")
        return redirect(url_for('moyennes.liste_moyennes'))

@moyennes_bp.route("/export")
@login_required
@ecole_required
def export_moyennes():
    """Exporte les moyennes dans différents formats"""
    try:
        ecole_id = get_current_ecole_id()
        
        # Récupération et validation des paramètres
        classe_id = sanitize_input(request.args.get("classe_id"), max_length=36)
        annee_scolaire = sanitize_input(request.args.get("annee_scolaire"), max_length=9)
        periode_str = request.args.get("periode", "1")
        format_type = sanitize_input(request.args.get("format", "excel"))
        
        # Récupérer la configuration du système
        systeme_eval = SystemeEvaluation.query.filter_by(ecole_id=ecole_id).first()
        type_systeme = systeme_eval.type_systeme if systeme_eval else 'trimestriel'
        
        current_app.logger.info(f"🔍 DEBUG Export - système: {type_systeme}, periode: {periode_str}")
        
        # Validation des paramètres obligatoires
        if not classe_id or not annee_scolaire:
            flash("Classe et année scolaire requis pour l'export", "error")
            return redirect(url_for('moyennes.liste_moyennes'))
        
        # Vérification d'accès à la classe
        if not validate_ecole_access_for_classe(classe_id, ecole_id):
            log_security_event("unauthorized_export", f"classe_{classe_id}", "failed")
            flash("Accès non autorisé à cette classe.", "error")
            return redirect(url_for('moyennes.liste_moyennes'))

        # Validation de la période
        try:
            periode = int(periode_str)
        except ValueError:
            periode = 1
        
        # Récupérer la classe pour le nom
        classe = Classe.query.filter_by(id=UUID(classe_id), ecole_id=ecole_id).first()
        if not classe:
            flash("Classe non trouvée", "error")
            return redirect(url_for('moyennes.liste_moyennes'))
        
        # Récupérer les informations de l'école COURANTE
        ecole = Ecole.query.get(ecole_id)
        if not ecole:
            flash("École non trouvée", "error")
            return redirect(url_for('moyennes.liste_moyennes'))
        
        current_app.logger.info(f"🔍 DEBUG - Classe trouvée: {classe.nom}, École: {ecole.nom}, Système: {type_systeme}")
        
        # Récupérer les données uniques
        items = get_unique_moyennes_data(ecole_id, classe_id, annee_scolaire, periode, type_systeme)
        
        current_app.logger.info(f"🔍 DEBUG - {len(items)} éléments UNIQUES trouvés pour l'export")

        if not items:
            flash("Aucune donnée à exporter pour les critères sélectionnés", "warning")
            return redirect(url_for('moyennes.liste_moyennes'))

        # Vérifier et nettoyer les doublons AVANT l'export
        duplicates_info = check_moyennes_duplicates_detailed(ecole_id, classe_id, annee_scolaire, periode, type_systeme)
        if duplicates_info:
            current_app.logger.warning(f"⚠️ ATTENTION: {len(duplicates_info)} doublons détectés avant export")
            # Nettoyer automatiquement les doublons
            clean_result = clean_moyennes_duplicates(ecole_id, classe_id, annee_scolaire, periode, type_systeme)
            current_app.logger.info(f"🔧 Nettoyage auto: {clean_result['message']}")
            
            # Recharger les données après nettoyage
            items = get_unique_moyennes_data(ecole_id, classe_id, annee_scolaire, periode, type_systeme)
            current_app.logger.info(f"🔍 DEBUG - {len(items)} éléments après nettoyage")

        # Préparer les données pour l'export
        export_data = []
        for idx, item in enumerate(items, 1):
            if hasattr(item, '_mapping'):
                item_dict = item._mapping
            else:
                item_dict = item._asdict() if hasattr(item, '_asdict') else item.__dict__
            
            row_data = {
                'N°': idx,
                'Nom': item_dict.get('nom', ''),
                'Prénom': item_dict.get('prenoms', ''),
                'Moyenne': f"{item_dict.get('moy_trim', 0):.2f}" if item_dict.get('moy_trim') is not None else "0.00",
                'Rang': item_dict.get('classement_str', ""),
                'Mention': item_dict.get('appreciation', "")
            }
            
            # Ajouter la moyenne générale pour la dernière période
            if (type_systeme == 'semestriel' and periode == 2) or (type_systeme == 'trimestriel' and periode == 3):
                row_data['Moyenne Générale'] = f"{item_dict.get('moy_gen', 0):.2f}" if item_dict.get('moy_gen') is not None else "0.00"
            
            export_data.append(row_data)

        # Titre adapté au système
        titre = f"LISTE DES MOYENNES - {classe.nom.upper()}"
        sous_titre = f"Année scolaire: {annee_scolaire} - {get_periode_display(type_systeme, periode)}"
        
        current_app.logger.info(f"🔍 DEBUG - Génération export {format_type} avec {len(export_data)} lignes UNIQUES")
        
        # Journalisation de l'export
        log_security_event("export_moyennes", f"classe_{classe_id}", "success", {
            "annee_scolaire": annee_scolaire,
            "periode": periode,
            "type_systeme": type_systeme,
            "format": format_type,
            "item_count": len(items)
        })

        # Sélection du format d'export
        if format_type == 'pdf':
            return generate_pdf_export_uniforme(export_data, titre, sous_titre, classe_id, annee_scolaire, periode, type_systeme, ecole, classe)
        elif format_type == 'csv':
            return generate_csv_export_uniforme(export_data, titre, sous_titre, classe_id, annee_scolaire, periode, type_systeme, ecole, classe)
        else:  # excel par défaut
            return generate_excel_export_uniforme(export_data, titre, sous_titre, classe_id, annee_scolaire, periode, type_systeme, ecole, classe)
        
    except Exception as e:
        current_app.logger.error(f"Erreur export moyennes: {str(e)}", exc_info=True)
        log_security_event("export_moyennes", "error", "failed", {"error": str(e)})
        flash("Erreur lors de l'export des données", "error")
        return redirect(url_for('moyennes.liste_moyennes'))

# Les autres routes restent similaires...

# ==================== FONCTIONS D'EXPORT ====================

def generate_excel_export_uniforme(data, titre, sous_titre, classe_id, annee_scolaire, periode, type_systeme, ecole, classe):
    """Génère un export Excel avec en-tête uniformisé - ADAPTÉ POUR TRIMESTRES/SEMESTRES"""
    try:
        current_app.logger.info(f"DEBUG - Début génération Excel uniformisé (Système: {type_systeme}, Période: {periode})")
        
        # Créer un nouveau workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Moyennes"
        
        # UTILISER LES VALEURS DYNAMIQUES DE L'ÉCOLE COURANTE
        nom_ecole = ecole.nom if ecole.nom else "École non renseignée"
        dre = ecole.dre if ecole.dre else "DRE non renseignée"
        inspection = ecole.inspection if ecole.inspection else "Inspection non renseignée"
        telephone = ecole.telephone1 if ecole.telephone1 else "Téléphone non renseigné"
        devise_ecole = ecole.devise if ecole.devise else "Travail - Liberté - Patrie"
        
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        
        # ========== EN-TÊTE EXCEL UNIFORMISÉ ==========
        # Ligne 1: Ministère (gauche) et République (droite)
        ws.merge_cells('A1:E1')
        ws['A1'] = "MINISTÈRE DE L'EDUCATION NATIONALE"
        ws['A1'].font = Font(bold=True, size=10)
        ws['A1'].alignment = left_align
        
        ws.merge_cells('F1:J1')
        ws['F1'] = "RÉPUBLIQUE TOGOLAISE"
        ws['F1'].font = Font(bold=True, size=10)
        ws['F1'].alignment = right_align
        
        # Ligne 2: DRE (gauche) et Devise (droite)
        ws.merge_cells('A2:E2')
        ws['A2'] = f"DIRECTION RÉGIONALE DE L'ÉDUCATION: {dre}"
        ws['A2'].font = Font(size=9)
        ws['A2'].alignment = left_align
        
        ws.merge_cells('F2:J2')
        ws['F2'] = "Travail - Liberté - Patrie"
        ws['F2'].font = Font(bold=True, size=9)
        ws['F2'].alignment = right_align
        
        # Ligne 3: Inspection (gauche)
        ws.merge_cells('A3:E3')
        ws['A3'] = f"INSPECTION DE L'ENSEIGNEMENT SECONDAIRE GÉNÉRAL: {inspection}"
        ws['A3'].font = Font(size=9)
        ws['A3'].alignment = left_align
        
        # Ligne 4: Vide pour séparation
        ws.merge_cells('A4:J4')
        ws['A4'] = ""
        
        # Ligne 5: Logo et informations de l'école (CENTRÉ)
        ws.merge_cells('A5:J5')
        ws['A5'] = nom_ecole
        ws['A5'].font = Font(bold=True, size=14, color="2C3E50")
        ws['A5'].alignment = center_align
        
        # Ligne 6: Téléphone et devise (CENTRÉ)
        ws.merge_cells('A6:J6')
        ws['A6'] = f"Tél: {telephone} - {devise_ecole}"
        ws['A6'].font = Font(size=10)
        ws['A6'].alignment = center_align
        
        # Ligne 7: Séparation
        ws.merge_cells('A7:J7')
        ws['A7'] = "__________________________________________________________"
        ws['A7'].alignment = center_align
        
        # Ligne 8: Vide
        ws['A8'] = ""
        
        # Titre du document avec classe
        ws.merge_cells('A9:J9')
        ws['A9'] = titre
        ws['A9'].font = Font(bold=True, size=16, color="2C3E50")
        ws['A9'].alignment = center_align
        
        # Sous-titre
        ws.merge_cells('A10:J10')
        ws['A10'] = sous_titre
        ws['A10'].font = Font(size=12)
        ws['A10'].alignment = center_align
        
        # STATISTIQUES DE LA CLASSE
        effectif_total = len(data)
        ws.merge_cells('A11:J11')
        ws['A11'] = f"Effectif: {effectif_total} élèves"
        ws['A11'].font = Font(size=11, bold=True, color="2C3E50")
        ws['A11'].alignment = center_align
        
        # Date de génération
        ws.merge_cells('A12:J12')
        ws['A12'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A12'].font = Font(size=8, italic=True, color="666666")
        ws['A12'].alignment = center_align
        
        # Ligne vide
        ws['A13'] = ""
        
        # ========== TABLEAU DES DONNÉES ==========
        headers = list(data[0].keys()) if data else []
        
        # Commencer à la ligne 14 pour les données
        start_row = 14
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = center_align
        
        # Données des élèves
        for index, row_data in enumerate(data, 1):
            row_num = start_row + index
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = row_data[header]
                cell.alignment = center_align
        
        # ========== MISE EN FORME FINALE ==========
        column_widths = {
            'A': 6,    # N°
            'B': 20,   # Nom
            'C': 20,   # Prénom
            'D': 12,   # Moyenne
            'E': 10,   # Rang
            'F': 15    # Mention
        }
        
        # Ajouter la colonne moyenne générale selon le système
        if (type_systeme == 'semestriel' and periode == 2) or (type_systeme == 'trimestriel' and periode == 3):
            column_widths['G'] = 15  # Moyenne Générale
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Ajouter des bordures au tableau
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, ws.max_row + 1):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        # Sauvegarder
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        current_app.logger.info(f"DEBUG - Excel uniformisé généré avec succès pour système {type_systeme}")
        
        # Retourner le fichier
        filename = f"moyennes_{classe_id}_{annee_scolaire}_{'S' if type_systeme == 'semestriel' else 'T'}{periode}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        current_app.logger.error(f"DEBUG - Erreur génération Excel uniformisé: {str(e)}")
        raise e

# Les fonctions generate_pdf_export_uniforme et generate_csv_export_uniforme restent similaires...

@moyennes_bp.route("/configurer-systeme", methods=["GET", "POST"])
@login_required
@ecole_required
def configurer_systeme():
    """Configure le système d'évaluation (trimestriel ou semestriel)"""
    ecole_id = get_current_ecole_id()
    
    if request.method == "POST":
        type_systeme = request.form.get("type_systeme", "trimestriel")
        
        # Récupérer ou créer la configuration
        systeme_eval = SystemeEvaluation.query.filter_by(ecole_id=ecole_id).first()
        if not systeme_eval:
            systeme_eval = SystemeEvaluation(
                id=str(uuid.uuid4()),
                ecole_id=ecole_id,
                type_systeme=type_systeme
            )
            db.session.add(systeme_eval)
        else:
            systeme_eval.type_systeme = type_systeme
        
        db.session.commit()
        flash(f"Système d'évaluation configuré en mode {type_systeme}", "success")
        return redirect(url_for('moyennes.liste_moyennes'))
    
    # GET: Afficher la page de configuration
    systeme_eval = SystemeEvaluation.query.filter_by(ecole_id=ecole_id).first()
    type_systeme = systeme_eval.type_systeme if systeme_eval else 'trimestriel'
    
    return render_template("moyennes/config_systeme.html", type_systeme=type_systeme)

@moyennes_bp.route("/verifier-export")
@login_required
@ecole_required
def verifier_export():
    """Vérifie les données avant export et signale les doublons"""
    try:
        ecole_id = get_current_ecole_id()
        classe_id = request.args.get("classe_id")
        annee_scolaire = request.args.get("annee_scolaire")
        trimestre = int(request.args.get("trimestre", 1))
        
        if not classe_id or not annee_scolaire:
            return jsonify({"available": False, "error": "Classe et année scolaire requis"})
        
        # Vérifier les doublons
        duplicates = check_moyennes_duplicates_detailed(ecole_id, classe_id, annee_scolaire, trimestre)
        
        # Compter les éléments uniques
        unique_items = get_unique_moyennes_data(ecole_id, classe_id, annee_scolaire, trimestre)
        
        return jsonify({
            "available": len(unique_items) > 0,
            "unique_count": len(unique_items),
            "duplicate_count": len(duplicates),
            "has_duplicates": len(duplicates) > 0,
            "duplicates": [dict(dup._mapping) for dup in duplicates] if duplicates else []
        })
        
    except Exception as e:
        current_app.logger.error(f"Erreur vérification export: {str(e)}")
        return jsonify({"available": False, "error": str(e)})

# ==================== ROUTE DE NETTOYAGE MANUEL ====================

@moyennes_bp.route("/nettoyer-doublons", methods=["POST"])
@login_required
@ecole_required
def nettoyer_doublons():
    """Nettoie les doublons de moyennes pour une classe"""
    try:
        ecole_id = get_current_ecole_id()
        classe_id = request.form.get("classe_id")
        annee_scolaire = request.form.get("annee_scolaire")
        trimestre = int(request.form.get("trimestre", 1))
        
        if not classe_id or not annee_scolaire:
            return jsonify({"status": "error", "message": "Classe et année scolaire requis"}), 400
        
        if not validate_ecole_access_for_classe(classe_id, ecole_id):
            return jsonify({"status": "error", "message": "Accès non autorisé"}), 403
        
        result = clean_moyennes_duplicates(ecole_id, classe_id, annee_scolaire, trimestre)
        
        if result["status"] == "success":
            log_security_event("nettoyage_doublons", f"classe_{classe_id}", "success", {
                "doublons_avant": result.get("duplicates_before", 0),
                "doublons_apres": result.get("duplicates_after", 0),
                "lignes_supprimees": result.get("deleted", 0)
            })
        
        return jsonify(result)
        
    except Exception as e:
        current_app.logger.error(f"Erreur nettoyage doublons: {str(e)}")
        log_security_event("nettoyage_doublons", "error", "failed", {"error": str(e)})
        return jsonify({"status": "error", "message": str(e)}), 500
    
@moyennes_bp.route("/diagnostic-doublons")
@login_required
@ecole_required
def diagnostic_doublons():
    """Page de diagnostic des doublons"""
    try:
        ecole_id = get_current_ecole_id()
        classe_id = request.args.get("classe_id")
        annee_scolaire = request.args.get("annee_scolaire")
        trimestre = int(request.args.get("trimestre", 1))
        
        if not classe_id or not annee_scolaire:
            flash("Classe et année scolaire requis pour le diagnostic", "error")
            return redirect(url_for('moyennes.liste_moyennes'))
        
        duplicates = check_moyennes_duplicates_detailed(ecole_id, classe_id, annee_scolaire, trimestre)
        
        return render_template(
            "moyennes/diagnostic_doublons.html",
            duplicates=duplicates,
            classe_id=classe_id,
            annee_scolaire=annee_scolaire,
            trimestre=trimestre
        )
        
    except Exception as e:
        current_app.logger.error(f"Erreur diagnostic doublons: {str(e)}")
        flash("Erreur lors du diagnostic des doublons", "error")
        return redirect(url_for('moyennes.liste_moyennes'))
    
def get_systeme_evaluation(ecole_id):
    """Récupère la configuration du système d'évaluation de l'école"""
    systeme = SystemeEvaluation.query.filter_by(ecole_id=ecole_id).first()
    if not systeme:
        # Créer une configuration par défaut
        systeme = SystemeEvaluation(
            id=str(uuid.uuid4()),
            ecole_id=ecole_id,
            type_systeme='trimestriel'
        )
        db.session.add(systeme)
        db.session.commit()
    return systeme

def get_periodes_for_systeme(type_systeme):
    """Retourne les périodes disponibles selon le système"""
    if type_systeme == 'semestriel':
        return [1, 2]
    else:
        return [1, 2, 3]





def generate_pdf_export_uniforme(data, titre, sous_titre, classe_id, annee_scolaire, periode, type_systeme, ecole, classe):
    """Génère un export PDF avec en-tête uniformisé CORRIGÉ - AVEC TYPE_SYSTEME"""
    try:
        current_app.logger.info(f"DEBUG - Début génération PDF uniformisé (Système: {type_systeme}, Période: {periode})")
        
        # UTILISER LES VALEURS DYNAMIQUES DE L'ÉCOLE COURANTE
        nom_ecole = ecole.nom if ecole.nom else "École non renseignée"
        dre = ecole.dre if ecole.dre else "DRE non renseignée"
        inspection = ecole.inspection if ecole.inspection else "Inspection non renseignée"
        telephone = ecole.telephone1 if ecole.telephone1 else "Téléphone non renseigné"
        devise_ecole = ecole.devise if ecole.devise else "Travail - Liberté - Patrie"

        buffer = io.BytesIO()
        
        def add_page_number(canvas, doc):
            canvas.saveState()
            page_num = canvas.getPageNumber()
            
            footer_text = f"Page {page_num}"
            canvas.setFont('Helvetica', 8)
            canvas.setFillColor(colors.gray)
            canvas.drawRightString(doc.pagesize[0] - 20*mm, 10*mm, footer_text)
            
            info_text = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
            canvas.drawString(20*mm, 10*mm, info_text)
            canvas.restoreState()
        
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4, 
            topMargin=15*mm, 
            bottomMargin=20*mm,
            leftMargin=10*mm,
            rightMargin=10*mm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # ========== EN-TÊTE UNIFORMISÉ CORRIGÉ ==========
        logo = None
        logo_path = get_logo_path_moyennes(ecole)
        
        if logo_path:
            try:
                logo = Image(logo_path, width=35*mm, height=35*mm)
                logo.hAlign = 'CENTER'
            except Exception as e:
                current_app.logger.error(f"❌ Erreur chargement logo: {e}")
                logo = None
        
        # Style pour les en-têtes
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Normal'],
            fontSize=9,
            alignment=0,
            spaceAfter=0,
            leading=10
        )
        
        header_right_style = ParagraphStyle(
            'CustomHeaderRight',
            parent=styles['Normal'],
            fontSize=9,
            alignment=2,
            spaceAfter=0,
            leading=10
        )
        
        # Structure avec 3 colonnes
        left_col_content = [
            Paragraph("<b>MINISTÈRE DE L'EDUCATION NATIONALE</b>", header_style),
            Paragraph(f"DIRECTION RÉGIONALE DE L'ÉDUCATION: {dre}", header_style),
            Paragraph(f"INSPECTION DE L'ENSEIGNEMENT SECONDAIRE GÉNÉRAL: {inspection}", header_style)
        ]
        
        right_col_content = [
            Paragraph("<b>RÉPUBLIQUE TOGOLAISE</b>", header_right_style),
            Paragraph("Travail - Liberté - Patrie", header_right_style)
        ]
        
        # Colonne centrale avec logo et informations de l'école
        center_col_content = []
        if logo:
            center_col_content.append(logo)
        center_col_content.extend([
            Paragraph(f"<b>{nom_ecole}</b>", ParagraphStyle(
                'EcoleStyle',
                parent=styles['Normal'],
                fontSize=12,
                alignment=1,
                spaceAfter=2,
                leading=12
            )),
            Paragraph(f"Tél: {telephone}", ParagraphStyle(
                'TelStyle',
                parent=styles['Normal'],
                fontSize=8,
                alignment=1,
                spaceAfter=0,
                leading=8
            )),
            Paragraph(f"{devise_ecole}", ParagraphStyle(
                'DeviseCenterStyle',
                parent=styles['Normal'],
                fontSize=7,
                alignment=1,
                spaceAfter=0,
                leading=7
            ))
        ])

        # Tableau à 3 colonnes pour l'en-tête
        header_table = Table([
            [left_col_content, center_col_content, right_col_content]
        ], colWidths=[doc.width/3, doc.width/3, doc.width/3])
        
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))
        
        elements.append(header_table)
        elements.append(Spacer(1, 4*mm))
        
        # Ligne de séparation
        separation_line = Table([['']], colWidths=[doc.width])
        separation_line.setStyle(TableStyle([
            ('LINEABOVE', (0, 0), (0, 0), 1, colors.black),
        ]))
        elements.append(separation_line)
        elements.append(Spacer(1, 6*mm))
        
        # ========== INFORMATIONS DES MOYENNES ==========
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=14,
            alignment=1,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=6
        )
        
        # Titre principal
        title = Paragraph(titre, title_style)
        elements.append(title)
        
        # Sous-titre
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Heading2'],
            fontSize=12,
            alignment=1,
            spaceAfter=8
        )
        subtitle = Paragraph(sous_titre, subtitle_style)
        elements.append(subtitle)
        
        # Statistiques
        stats_style = ParagraphStyle(
            'StatsStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=8
        )
        
        effectif_total = len(data)
        stats_text = f"Effectif: {effectif_total} élèves"
        stats_paragraph = Paragraph(f"<b>{stats_text}</b>", stats_style)
        elements.append(stats_paragraph)
        
        # Date de génération
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            textColor=colors.gray,
            spaceAfter=12
        )
        
        date_text = Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style)
        elements.append(date_text)
        
        # ========== TABLEAU DES MOYENNES ==========
        if data:
            headers = list(data[0].keys())
            table_data = [headers]
            
            for row in data:
                table_data.append([str(row[header]) for header in headers])
            
            # Largeurs optimisées
            col_widths = [
                doc.width * 0.05,  # N°
                doc.width * 0.25,  # Nom
                doc.width * 0.25,  # Prénom
                doc.width * 0.15,  # Moyenne
                doc.width * 0.15,  # Rang
                doc.width * 0.15   # Mention
            ]
            
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            # Style du tableau
            table_style = TableStyle([
                # En-têtes
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                
                # Lignes de données
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                ('ALIGN', (1, 1), (2, -1), 'LEFT'),
                ('ALIGN', (3, 1), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                
                # Bordures
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
                
                # Alternance des couleurs des lignes
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
            ])
            
            table.setStyle(table_style)
            elements.append(table)
        
        # ========== STATISTIQUES FINALES ==========
        elements.append(Spacer(1, 8*mm))
        
        # Calcul des statistiques
        moyennes = [float(row['Moyenne']) for row in data if row['Moyenne'] != "0.00"]
        if moyennes:
            moy_forte = max(moyennes)
            moy_faible = min(moyennes)
            moy_class = (moy_forte + moy_faible) / 2
            
            final_stats_style = ParagraphStyle(
                'FinalStatsStyle',
                parent=styles['Normal'],
                fontSize=9,
                alignment=0,
                textColor=colors.gray,
                spaceAfter=1
            )
            
            final_stats_text = [
                Paragraph(f"<b>RÉCAPITULATIF CLASSE {classe.nom.upper()} :</b>", final_stats_style),
                Paragraph(f"Moyenne de la classe : {moy_class:.2f}", final_stats_style),
                Paragraph(f"Moyenne la plus forte : {moy_forte:.2f}", final_stats_style),
                Paragraph(f"Moyenne la plus faible : {moy_faible:.2f}", final_stats_style),
            ]
            
            for element in final_stats_text:
                elements.append(element)
        
        # ========== GÉNÉRATION DU PDF ==========
        doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
        buffer.seek(0)
        
        current_app.logger.info("DEBUG - PDF uniformisé généré avec succès")
        
        # Nom du fichier adapté au système
        prefix = 'S' if type_systeme == 'semestriel' else 'T'
        filename = f"moyennes_{classe_id}_{annee_scolaire}_{prefix}{periode}.pdf"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        current_app.logger.error(f"DEBUG - Erreur génération PDF uniformisé: {str(e)}")
        raise e

def generate_csv_export_uniforme(data, titre, sous_titre, classe_id, annee_scolaire, periode, type_systeme, ecole, classe):
    """Génère un export CSV avec en-tête uniformisé CORRIGÉ - AVEC TYPE_SYSTEME"""
    try:
        current_app.logger.info(f"DEBUG - Début génération CSV uniformisé (Système: {type_systeme}, Période: {periode})")
        
        output = StringIO()
        writer = csv.writer(output, delimiter=';')
        
        # UTILISER LES VALEURS DYNAMIQUES DE L'ÉCOLE COURANTE
        nom_ecole = ecole.nom if ecole.nom else "École non renseignée"
        dre = ecole.dre if ecole.dre else "DRE non renseignée"
        inspection = ecole.inspection if ecole.inspection else "Inspection non renseignée"
        telephone = ecole.telephone1 if ecole.telephone1 else "Téléphone non renseigné"
        devise_ecole = ecole.devise if ecole.devise else "Travail - Liberté - Patrie"
        
        # Entêtes du document avec valeurs dynamiques
        writer.writerow(["MINISTÈRE DE L'EDUCATION NATIONALE", "", "", "", "", "RÉPUBLIQUE TOGOLAISE"])
        writer.writerow([f"DIRECTION RÉGIONALE DE L'ÉDUCATION: {dre}", "", "", "", "", "Travail - Liberté - Patrie"])
        writer.writerow([f"INSPECTION DE L'ENSEIGNEMENT SECONDAIRE GÉNÉRAL: {inspection}"])
        writer.writerow([])
        writer.writerow(["", "", nom_ecole])
        writer.writerow(["", "", f"Tél: {telephone} - {devise_ecole}"])
        writer.writerow([])
        writer.writerow([titre])
        writer.writerow([sous_titre])
        writer.writerow([f"Effectif: {len(data)} élèves"])
        writer.writerow([f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}"])
        writer.writerow([])
        
        # En-têtes du tableau
        if data:
            headers = list(data[0].keys())
            writer.writerow(headers)
            
            # Données
            for row in data:
                writer.writerow([str(row[header]) for header in headers])
        
        output.seek(0)
        
        current_app.logger.info("DEBUG - CSV uniformisé généré avec succès")
        
        # Nom du fichier adapté au système
        prefix = 'S' if type_systeme == 'semestriel' else 'T'
        filename = f"moyennes_{classe_id}_{annee_scolaire}_{prefix}{periode}.csv"
        
        response = Response(
            output.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
        
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"DEBUG - Erreur génération CSV uniformisé: {str(e)}")
        raise e

# ==================== ROUTES BATCH (restent identiques) ====================

@moyennes_bp.route("/lancer-batch", methods=["POST"])
@login_required
@ecole_required
def lancer_batch():
    """Lance le calcul des moyennes pour une classe - VERSION CORRIGÉE"""
    try:
        current_app.logger.info(f"[DEBUG] lancer_batch appelé par l'utilisateur {current_user.id}")
        
        # Journaliser toutes les données reçues
        current_app.logger.info(f"[DEBUG] Données reçues: {request.form}")
        
        if not rate_limit_check():
            current_app.logger.warning("[DEBUG] Limite de requêtes dépassée")
            return jsonify({"status": "error", "message": "Trop de requêtes. Veuillez réessayer plus tard."}), 429
        
        ecole_id = get_current_ecole_id()
        current_app.logger.info(f"[DEBUG] École ID: {ecole_id}")
        
        # CORRECTION : Utiliser request.form.get() au lieu de form.get()
        classe_id = sanitize_input(request.form.get("classe_id"), max_length=36)
        annee_scolaire = sanitize_input(request.form.get("annee_scolaire"), max_length=9)
        periode_str = request.form.get("periode", "1")
        type_systeme = request.form.get("type_systeme", "trimestriel")
        
        # Log des paramètres reçus
        current_app.logger.info(f"[DEBUG] Paramètres: classe_id={classe_id}, annee_scolaire={annee_scolaire}, "
                               f"periode={periode_str}, type_systeme={type_systeme}")
        
        # Validation des paramètres obligatoires
        if not classe_id:
            return jsonify({"status": "error", "message": "Classe requise"}), 400
        
        if not annee_scolaire:
            return jsonify({"status": "error", "message": "Année scolaire requise"}), 400
        
        # Validation de la période
        try:
            periode = int(periode_str)
        except ValueError:
            return jsonify({"status": "error", "message": "Période invalide"}), 400
        
        # Validation selon le système
        if type_systeme == 'semestriel' and periode not in [1, 2]:
            return jsonify({"status": "error", "message": "Période invalide pour système semestriel. Doit être 1 ou 2."}), 400
        elif type_systeme == 'trimestriel' and periode not in [1, 2, 3]:
            return jsonify({"status": "error", "message": "Période invalide pour système trimestriel. Doit être 1, 2 ou 3."}), 400

        # Validation de l'année scolaire
        if not validate_annee_scolaire_format(annee_scolaire):
            return jsonify({"status": "error", "message": "Format d'année scolaire invalide (doit être: 2024-2025)"}), 400

        # Validation d'accès à la classe
        if not validate_ecole_access_for_classe(classe_id, ecole_id):
            log_security_event("unauthorized_batch_access", f"classe_{classe_id}", "failed")
            return jsonify({"status": "error", "message": "Accès non autorisé à cette classe"}), 403

        # Journalisation du démarrage
        log_security_event("batch_calculation", f"classe_{classe_id}", "started", {
            "annee_scolaire": annee_scolaire,
            "periode": periode,
            "type_systeme": type_systeme,
            "ecole_id": str(ecole_id)
        })

        current_app.logger.info(f"[DEBUG] Lancement du calcul pour classe {classe_id}, période {periode}, système {type_systeme}")
        
        # Lancement du calcul des moyennes avec les paramètres adaptés
        result = calculer_moyennes(
            classe_id=classe_id,
            annee_scolaire=annee_scolaire,
            periode=periode,
            ecole_id=ecole_id,
            type_systeme=type_systeme
        )

        current_app.logger.info(f"[DEBUG] Résultat du calcul: {result}")
        
        # Vérification du résultat
        if "errors" in result and result["errors"]:
            current_app.logger.error(f"[DEBUG] Erreurs dans le calcul: {result['errors']}")
            log_security_event("batch_calculation", f"classe_{classe_id}", "failed", {
                "errors": result["errors"][:3],
                "total_errors": len(result["errors"])
            })
            return jsonify({
                "status": "error", 
                "message": f"Erreurs lors du calcul: {', '.join(result['errors'][:3])}" + 
                          (f" (+ {len(result['errors']) - 3} autres)" if len(result['errors']) > 3 else "")
            }), 500

        # Récupération des données mises à jour pour l'affichage
        items_query = (
            db.session.query(
                Eleve.nom, 
                Eleve.prenoms, 
                Classe.nom.label("classe_nom"),
                Moyenne.moy_periode.label("moy_trim"),
                Moyenne.moy_gen,
                Moyenne.classement_str,
                Appreciations.libelle.label("appreciation"), 
                Moyenne.moy_class
            )
            .join(Eleve, Eleve.id == Moyenne.eleve_id)
            .join(Classe, Classe.id == Eleve.classe_id)
            .outerjoin(Appreciations, Appreciations.id == Moyenne.appreciation_id)
            .filter(
                Moyenne.periode == periode, 
                Moyenne.annee_scolaire == annee_scolaire,
                Classe.ecole_id == ecole_id
            )
        )
        
        if classe_id:
            items_query = items_query.filter(Classe.id == UUID(classe_id))
            
        items_data = items_query.all()
        items = [dict(row._mapping) for row in items_data]

        current_app.logger.info(f"[DEBUG] {len(items)} éléments récupérés après calcul")
        
        # Calcul des statistiques de la classe
        valeurs_non_nulles = [i["moy_trim"] for i in items if i["moy_trim"] is not None and i["moy_trim"] > 0]
        if valeurs_non_nulles:
            classe_recap = {
                "moy_class": round((max(valeurs_non_nulles) + min(valeurs_non_nulles))/2, 2),
                "moy_forte": max(valeurs_non_nulles),
                "moy_faible": min(valeurs_non_nulles),
                "effectif_composants": len(valeurs_non_nulles),
                "effectif_total": len(items)
            }
        else:
            classe_recap = {
                "moy_class": 0,
                "moy_forte": 0,
                "moy_faible": 0,
                "effectif_composants": 0,
                "effectif_total": len(items)
            }

        # Journalisation du succès
        log_security_event("batch_calculation", f"classe_{classe_id}", "success", {
            "created": result.get("created", 0),
            "updated": result.get("updated", 0),
            "total": result.get("total", 0),
            "periode": periode,
            "type_systeme": type_systeme
        })

        current_app.logger.info(f"[DEBUG] Batch terminé avec succès: {result.get('created', 0)} créés, {result.get('updated', 0)} mis à jour")
        
        # Réponse avec toutes les informations
        return jsonify({
            "status": "ok", 
            "message": "Calcul terminé avec succès",
            "details": {
                "created": result.get("created", 0),
                "updated": result.get("updated", 0),
                "total": result.get("total", 0),
                "success_count": len(result.get("success", [])),
                "periode": periode,
                "type_systeme": type_systeme
            },
            "items": items, 
            "classe_recap": classe_recap, 
            "periode": periode,
            "type_systeme": type_systeme
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[ERROR] Erreur batch moyennes: {str(e)}", exc_info=True)
        log_security_event("batch_calculation", "global_error", "failed", {
            "error": str(e),
            "ecole_id": str(ecole_id) if 'ecole_id' in locals() else "unknown"
        })
        return jsonify({
            "status": "error", 
            "message": "Erreur lors du calcul des moyennes",
            "technical_error": str(e)[:200]  # Premiers 200 caractères seulement
        }), 500

# Les autres routes batch et middleware restent identiques...

@moyennes_bp.after_request
def add_security_headers(response):
    """Ajoute des headers de sécurité HTTP"""
    security_headers = {
        'X-Content-Type-Options': 'nosniff',
        'X-Frame-Options': 'DENY',
        'X-XSS-Protection': '1; mode=block',
        'Strict-Transport-Security': 'max-age=31536000; includeSubDomains',
        'Cache-Control': 'no-store, no-cache, must-revalidate, max-age=0'
    }
    
    for header, value in security_headers.items():
        response.headers[header] = value
        
    return response

# Gestionnaires d'erreurs...
@moyennes_bp.errorhandler(400)
def bad_request_error(error):
    log_security_event("bad_request", request.path, "failed", {"user_agent": request.headers.get('User-Agent')})
    return jsonify({"error": "Requête invalide"}), 400

@moyennes_bp.errorhandler(403)
def forbidden_error(error):
    log_security_event("forbidden_access", request.path, "failed", {"user_id": current_user.id if current_user.is_authenticated else "anonymous"})
    return jsonify({"error": "Accès non autorisé"}), 403

@moyennes_bp.errorhandler(429)
def rate_limit_error(error):
    log_security_event("rate_limit_exceeded", request.path, "failed", {"ip": request.remote_addr})
    return jsonify({'error': 'Trop de requêtes. Veuillez réessayer plus tard.'}), 429

@moyennes_bp.errorhandler(500)
def internal_error(error):
    log_security_event("internal_server_error", request.path, "failed", {"error": str(error)})
    return jsonify({"error": "Erreur interne du serveur"}), 500