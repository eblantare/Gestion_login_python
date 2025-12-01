from flask import Blueprint, request, render_template, redirect, url_for, flash, jsonify, session, current_app, Response
from extensions import db
from ..models import Moyenne, Eleve, Classe, Note, Appreciations, Enseignement
from flask_login import current_user, login_required
from gestion_scolaire.app.tasks.batchMoyennes import calculer_moyennes
import uuid, time
from datetime import datetime
import re
import logging
from uuid import UUID
from ..utils import ecole_required, get_current_ecole_id
import csv
import pandas as pd
from io import BytesIO, StringIO
import openpyxl
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch

moyennes_bp = Blueprint('moyennes', __name__, url_prefix='/moyennes')

# ==================== CONFIGURATION DE SÉCURITÉ ====================

# Configuration du logger de sécurité
security_logger = logging.getLogger('security')
if not security_logger.handlers:
    handler = logging.StreamHandler()
    formatter = logging.Formatter(
        '%(asctime)s - SECURITY - %(levelname)s - [%(ip)s] %(user_id)s - %(message)s'
    )
    handler.setFormatter(formatter)
    security_logger.addHandler(handler)
    security_logger.setLevel(logging.INFO)

def log_security_event(action, resource, status="success", details=None):
    """Journalise les événements de sécurité"""
    user_id = str(current_user.id) if current_user.is_authenticated else "anonymous"
    ip = request.remote_addr if request else "unknown"
    
    log_data = {
        "user_id": user_id,
        "action": action,
        "resource": resource,
        "status": status,
        "ip": ip,
        "user_agent": request.headers.get('User-Agent', 'unknown') if request else "unknown"
    }
    
    if details:
        log_data["details"] = details
    
    security_logger.info(f"{action} - {resource} - {status}", extra={
        'ip': ip, 
        'user_id': user_id
    })

def validate_uuid(uuid_string):
    """Valide le format UUID de manière sécurisée"""
    if not uuid_string or not isinstance(uuid_string, str):
        return False
    
    try:
        uuid_obj = UUID(uuid_string)
        return str(uuid_obj) == uuid_string
    except (ValueError, AttributeError):
        return False

def sanitize_input(param, max_length=100, allowed_pattern=None):
    """Nettoie et valide les paramètres d'entrée"""
    if param is None:
        return None
    
    param_str = str(param).strip()
    
    # Vérification de la longueur
    if len(param_str) > max_length:
        log_security_event(
            "input_validation", 
            "parameter_length", 
            "failed", 
            {"parameter": str(param)[:50], "max_length": max_length}
        )
        return None
    
    # Nettoyage des caractères dangereux
    param_str = re.sub(r'[<>"\';()&+]', '', param_str)
    
    # Validation par pattern si spécifié
    if allowed_pattern and not re.match(allowed_pattern, param_str):
        log_security_event(
            "input_validation", 
            "parameter_pattern", 
            "failed", 
            {"parameter": param_str, "pattern": allowed_pattern}
        )
        return None
    
    return param_str

def validate_ecole_access_for_classe(classe_id, ecole_id):
    """Valide l'accès à une classe spécifique pour l'école"""
    try:
        if not classe_id:
            return True  # Aucune classe spécifiée = accès à toutes
        
        classe = Classe.query.filter_by(id=UUID(classe_id), ecole_id=ecole_id).first()
        return classe is not None
    except (ValueError, AttributeError):
        return False

def validate_annee_scolaire_format(annee_scolaire):
    """Valide le format de l'année scolaire"""
    if not annee_scolaire:
        return False
    
    pattern = r'^\d{4}-\d{4}$'  # Format: 2024-2025
    return bool(re.match(pattern, annee_scolaire))

def rate_limit_check():
    """Vérification basique de rate limiting basée sur la session"""
    if 'request_count' not in session:
        session['request_count'] = 1
        session['first_request'] = time.time()
    else:
        session['request_count'] += 1
    
    # Limite: 100 requêtes par minute
    current_time = time.time()
    if current_time - session['first_request'] > 60:  # 1 minute
        session['request_count'] = 1
        session['first_request'] = current_time
    elif session['request_count'] > 100:
        log_security_event(
            "rate_limit", 
            f"ip_{request.remote_addr}", 
            "failed", 
            {"request_count": session['request_count']}
        )
        return False
    
    return True

# ==================== ROUTES SÉCURISÉES ====================

@moyennes_bp.route("/")
@login_required
@ecole_required
def liste_moyennes():
    """Liste des moyennes - VERSION SÉCURISÉE"""
    try:
        # Rate limiting
        if not rate_limit_check():
            flash("Trop de requêtes. Veuillez réessayer plus tard.", "error")
            return redirect(url_for('moyennes.liste_moyennes'))
        
        ecole_id = get_current_ecole_id()
        
        # Récupération et validation des paramètres
        search = sanitize_input(request.args.get("search", ""), max_length=50)
        classe_id = sanitize_input(request.args.get("classe_id", ""), max_length=36)
        annee_scolaire = sanitize_input(request.args.get("annee_scolaire", ""), max_length=9)
        
        # Sécurisation du trimestre
        trimestre_str = request.args.get("trimestre", "1")
        try:
            trimestre = int(trimestre_str)
        except ValueError:
            trimestre = 1
        if trimestre not in [1, 2, 3]:
            trimestre = 1

        # Validation de la pagination
        try:
            page = max(1, int(request.args.get("page", 1)))
            per_page = min(max(1, int(request.args.get("per_page", 10))), 100)  # Limite à 100
        except ValueError:
            page = 1
            per_page = 10

        # Vérification d'accès à la classe
        if classe_id and not validate_ecole_access_for_classe(classe_id, ecole_id):
            log_security_event("unauthorized_classe_access", f"classe_{classe_id}", "failed")
            flash("Accès non autorisé à cette classe.", "error")
            return redirect(url_for('moyennes.liste_moyennes'))

        # CORRECTION : Récupérer les années scolaires depuis la table Moyenne avec filtre par école
        annees_scolaires_query = (
            db.session.query(Moyenne.annee_scolaire)
            .join(Eleve).join(Classe)
            .filter(Classe.ecole_id == ecole_id)
            .distinct()
            .all()
        )
        annees_scolaires = [a[0] for a in annees_scolaires_query if a[0]]
        
        # Si aucune année n'est trouvée, fournir des valeurs par défaut
        if not annees_scolaires:
            current_year = datetime.now().year
            annees_scolaires = [f"{current_year-1}-{current_year}", f"{current_year}-{current_year+1}"]

        # Requête principale avec filtre par école
        query = (
            db.session.query(
                Eleve.id.label("eleve_id"),
                Eleve.nom,
                Eleve.prenoms,
                Classe.nom.label("classe_nom"),
                Moyenne.moy_trim,
                Moyenne.moy_gen,
                Moyenne.classement.label("classement"),
                Moyenne.classement_str.label("classement_str"),
                Moyenne.classement_gen.label("classement_gen"),
                Moyenne.moy_class,
                Moyenne.moy_forte,
                Moyenne.moy_faible,
                Moyenne.eff_comp.label("effectif_composants"),
                Appreciations.libelle.label("appreciation"),
            )
            .join(Eleve, Eleve.id == Moyenne.eleve_id)
            .join(Classe, Classe.id == Eleve.classe_id)
            .outerjoin(Appreciations, Appreciations.id == Moyenne.appreciation_id)
            .filter(Moyenne.trimestre == trimestre)
            .filter(Classe.ecole_id == ecole_id)  # Filtre par école
        )

        # Application des filtres avec données sanitizées
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                db.or_(
                    Eleve.nom.ilike(search_term),
                    Eleve.prenoms.ilike(search_term),
                )
            )
        if classe_id:
            query = query.filter(Eleve.classe_id == UUID(classe_id))
        if annee_scolaire:
            query = query.filter(Moyenne.annee_scolaire == annee_scolaire)

        query = query.order_by(Moyenne.classement.asc())
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        items = pagination.items

        # Récupérer les classes de l'école actuelle seulement
        classes = Classe.query.filter_by(ecole_id=ecole_id).all()

        # CORRECTION : Calcul du récapitulatif avec effectif TOTAL de la classe
        if items:
            # Filtrer les valeurs None avant de calculer
            moy_fortes = [i.moy_forte for i in items if i.moy_forte is not None]
            moy_faibles = [i.moy_faible for i in items if i.moy_faible is not None]
            
            moy_forte = max(moy_fortes) if moy_fortes else 0
            moy_faible = min(moy_faibles) if moy_faibles else 0
            moy_class = round((moy_forte + moy_faible) / 2, 2) if moy_fortes and moy_faibles else 0
            
            # CORRECTION : Créer une sous-requête SANS ORDER BY pour le comptage
            count_query = (
                db.session.query(Eleve.id)
                .join(Moyenne, Eleve.id == Moyenne.eleve_id)
                .join(Classe, Classe.id == Eleve.classe_id)
                .filter(Moyenne.trimestre == trimestre)
                .filter(Classe.ecole_id == ecole_id)
            )
            
            # Appliquer les mêmes filtres que la requête principale
            if search:
                count_query = count_query.filter(
                    db.or_(
                        Eleve.nom.ilike(search_term),
                        Eleve.prenoms.ilike(search_term),
                    )
                )
            if classe_id:
                count_query = count_query.filter(Eleve.classe_id == UUID(classe_id))
            if annee_scolaire:
                count_query = count_query.filter(Moyenne.annee_scolaire == annee_scolaire)
            
            # Compter les résultats SANS ORDER BY
            effectif_total = count_query.distinct().count()
            
            # Effectif ayant composé (ceux avec moyenne > 0)
            effectif_composants = sum([1 for i in items if i.moy_trim is not None and i.moy_trim > 0])
            
            classe_recap = {
                "moy_class": moy_class,
                "moy_forte": moy_forte,
                "moy_faible": moy_faible,
                "effectif_composants": effectif_composants,
                "effectif_total": effectif_total,  # Nouveau champ pour l'effectif total
            }
        else:
            classe_recap = {
                "moy_class": 0, 
                "moy_forte": 0, 
                "moy_faible": 0, 
                "effectif_composants": 0,
                "effectif_total": 0
            }

        # Journalisation de l'accès
        log_security_event("page_access", "liste_moyennes", "success", {
            "search_used": bool(search),
            "classe_filter": bool(classe_id),
            "item_count": len(items),
            "effectif_total": classe_recap["effectif_total"]
        })

        return render_template(
            "moyennes/list_moy.html",
            items=items,
            pagination=pagination,
            search=search,
            classe_id=classe_id,
            annee_scolaire=annee_scolaire,
            trimestre=trimestre,
            per_page=per_page,
            classes=classes,
            annees_scolaires=annees_scolaires,
            classe_recap=classe_recap,
            ecole_id=ecole_id
        )
        
    except Exception as e:
        current_app.logger.error(f"Erreur liste moyennes: {str(e)}")
        log_security_event("page_error", "liste_moyennes", "failed", {"error": str(e)})
        flash("Erreur lors du chargement des données.", "error")
        return redirect(url_for('moyennes.liste_moyennes'))


@moyennes_bp.route("/lancer-batch", methods=["POST"])
@login_required
@ecole_required
def lancer_batch():
    """Lance le calcul des moyennes pour une classe - VERSION SÉCURISÉE"""
    try:
        # Rate limiting pour les opérations lourdes
        if not rate_limit_check():
            return jsonify({"status": "error", "message": "Trop de requêtes. Veuillez réessayer plus tard."}), 429
        
        ecole_id = get_current_ecole_id()
        
        # Récupération et validation des paramètres
        classe_id = sanitize_input(request.form.get("classe_id"), max_length=36)
        annee_scolaire = sanitize_input(request.form.get("annee_scolaire"), max_length=9)
        trimestre_str = request.form.get("trimestre", "1")
        
        # Validation du trimestre
        try:
            trimestre = int(trimestre_str)
        except ValueError:
            return jsonify({"status": "error", "message": "Trimestre invalide"}), 400
        
        if trimestre not in [1, 2, 3]:
            return jsonify({"status": "error", "message": "Trimestre invalide"}), 400

        # Validation des paramètres obligatoires
        if not annee_scolaire:
            return jsonify({"status": "error", "message": "Année scolaire requise"}), 400
        
        if not validate_annee_scolaire_format(annee_scolaire):
            return jsonify({"status": "error", "message": "Format d'année scolaire invalide"}), 400

        # Vérification d'accès à la classe
        if classe_id and not validate_ecole_access_for_classe(classe_id, ecole_id):
            log_security_event("unauthorized_batch_access", f"classe_{classe_id}", "failed")
            return jsonify({"status": "error", "message": "Accès non autorisé à cette classe"}), 403

        # Journalisation du lancement du batch
        log_security_event("batch_calculation", f"classe_{classe_id or 'all'}", "started", {
            "annee_scolaire": annee_scolaire,
            "trimestre": trimestre
        })

        # 🔹 Lancer le batch
        result = calculer_moyennes(classe_id, annee_scolaire, trimestre, ecole_id)

        # 🔹 Retourner items + résumé pour le JS
        items_query = (
            db.session.query(
                Eleve.nom, Eleve.prenoms, Classe.nom.label("classe_nom"),
                Moyenne.moy_trim, Moyenne.moy_gen, Moyenne.classement_str,
                Appreciations.libelle.label("appreciation"), Moyenne.moy_class
            )
            .join(Eleve, Eleve.id == Moyenne.eleve_id)
            .join(Classe, Classe.id == Eleve.classe_id)
            .outerjoin(Appreciations, Appreciations.id == Moyenne.appreciation_id)
            .filter(
                Moyenne.trimestre == trimestre, 
                Moyenne.annee_scolaire == annee_scolaire,
                Classe.ecole_id == ecole_id  # Filtre par école
            )
        )
        
        # Filtrer par classe si spécifiée
        if classe_id:
            items_query = items_query.filter(Classe.id == UUID(classe_id))
            
        items_data = items_query.all()
        items = [dict(row._mapping) for row in items_data]

        # Récapitulatif
        valeurs_non_nulles = [i["moy_trim"] for i in items if i["moy_trim"] is not None and i["moy_trim"] > 0]
        if valeurs_non_nulles:
            classe_recap = {
                "moy_class": round((max(valeurs_non_nulles) + min(valeurs_non_nulles))/2, 2),
                "moy_forte": max(valeurs_non_nulles),
                "moy_faible": min(valeurs_non_nulles),
                "effectif_composants": len(valeurs_non_nulles)
            }
        else:
            classe_recap = {
                "moy_class": 0,
                "moy_forte": 0,
                "moy_faible": 0,
                "effectif_composants": 0
            }

        # Journalisation du succès
        log_security_event("batch_calculation", f"classe_{classe_id or 'all'}", "success", {
            "created": result.get("created", 0),
            "updated": result.get("updated", 0),
            "total": result.get("total", 0)
        })

        return jsonify({
            "status": "ok", 
            **result, 
            "items": items, 
            "classe_recap": classe_recap, 
            "trimestre": trimestre
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erreur batch moyennes: {str(e)}")
        log_security_event("batch_calculation", "error", "failed", {"error": str(e)})
        return jsonify({"status": "error", "message": "Erreur lors du calcul des moyennes"}), 500


@moyennes_bp.route("/lancer-batch-toutes-classes", methods=["POST"])
@login_required
@ecole_required
def lancer_batch_toutes_classes():
    """
    Lance le calcul des moyennes pour toutes les classes de l'école - VERSION SÉCURISÉE
    """
    try:
        # Rate limiting pour les opérations lourdes
        if not rate_limit_check():
            return jsonify({"status": "error", "message": "Trop de requêtes. Veuillez réessayer plus tard."}), 429
        
        ecole_id = get_current_ecole_id()
        
        # Récupération et validation des paramètres
        annee_scolaire = sanitize_input(request.form.get("annee_scolaire"), max_length=9)
        trimestre_str = request.form.get("trimestre", "1")
        
        # Validation du trimestre
        try:
            trimestre = int(trimestre_str)
        except ValueError:
            return jsonify({"status": "error", "message": "Trimestre invalide"}), 400
        
        if trimestre not in [1, 2, 3]:
            return jsonify({"status": "error", "message": "Trimestre invalide"}), 400

        # Validation des paramètres obligatoires
        if not annee_scolaire:
            return jsonify({"status": "error", "message": "Année scolaire requise"}), 400
        
        if not validate_annee_scolaire_format(annee_scolaire):
            return jsonify({"status": "error", "message": "Format d'année scolaire invalide"}), 400

        # Journalisation du lancement du batch global
        log_security_event("batch_all_classes", "started", "started", {
            "annee_scolaire": annee_scolaire,
            "trimestre": trimestre,
            "ecole_id": ecole_id
        })

        # Récupérer toutes les classes de l'école
        classes = Classe.query.filter_by(ecole_id=ecole_id).all()
        
        if not classes:
            return jsonify({"status": "error", "message": "Aucune classe trouvée pour cette école"}), 404

        resume_global = {
            "success": [], 
            "errors": [], 
            "created": 0, 
            "updated": 0, 
            "total": 0,
            "classes_traitees": []
        }

        for classe in classes:
            try:
                result = calculer_moyennes(classe.id, annee_scolaire, trimestre, ecole_id)
                resume_global["success"].extend(result.get("success", []))
                resume_global["errors"].extend(result.get("errors", []))
                resume_global["created"] += result.get("created", 0)
                resume_global["updated"] += result.get("updated", 0)
                resume_global["total"] += result.get("total", 0)
                resume_global["classes_traitees"].append(classe.nom)
            except Exception as e:
                error_msg = f"Erreur classe {classe.nom}: {str(e)}"
                resume_global["errors"].append(error_msg)
                current_app.logger.error(error_msg)

        # Journalisation du résultat
        log_security_event("batch_all_classes", "completed", "success", {
            "classes_traitees": len(resume_global["classes_traitees"]),
            "created": resume_global["created"],
            "updated": resume_global["updated"],
            "total": resume_global["total"],
            "erreurs": len(resume_global["errors"])
        })

        return jsonify({"status": "ok", "result": resume_global})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erreur batch toutes classes: {str(e)}")
        log_security_event("batch_all_classes", "error", "failed", {"error": str(e)})
        return jsonify({"status": "error", "message": "Erreur lors du calcul des moyennes pour toutes les classes"}), 500


@moyennes_bp.route("/test-batch")
@login_required
@ecole_required
def test_batch():
    """
    Endpoint de test pour vérifier le batch - VERSION SÉCURISÉE
    """
    try:
        # Rate limiting
        if not rate_limit_check():
            return jsonify({"status": "error", "message": "Trop de requêtes. Veuillez réessayer plus tard."}), 429
        
        ecole_id = get_current_ecole_id()
        
        # Prendre la première classe de l'école actuelle
        classe = Classe.query.filter_by(ecole_id=ecole_id).first()
        if not classe:
            return jsonify({"status": "error", "message": "Aucune classe trouvée pour cette école"}), 404

        # Journalisation du test
        log_security_event("test_batch", f"classe_{classe.id}", "started")

        # Lancer le batch pour cette classe et l'année scolaire courante
        current_year = datetime.now().year
        annee_test = f"{current_year}-{current_year+1}"
        
        result = calculer_moyennes(classe.id, annee_test, 1, ecole_id)  # trimestre 1
        
        # Journalisation du résultat du test
        log_security_event("test_batch", f"classe_{classe.id}", "success", {
            "created": result.get("created", 0),
            "updated": result.get("updated", 0),
            "total": result.get("total", 0)
        })

        return jsonify({
            "status": "ok", 
            "classe": classe.nom, 
            "result": result,
            "message": "Test du batch effectué avec succès"
        })

    except Exception as e:
        current_app.logger.error(f"Erreur test batch: {str(e)}")
        log_security_event("test_batch", "error", "failed", {"error": str(e)})
        return jsonify({"status": "error", "message": f"Erreur lors du test: {str(e)}"}), 500


@moyennes_bp.route("/export")
@login_required
@ecole_required
def export_moyennes():
    """Exporte les moyennes dans différents formats - VERSION SÉCURISÉE"""
    try:
        ecole_id = get_current_ecole_id()
        
        # Récupération et validation des paramètres
        classe_id = sanitize_input(request.args.get("classe_id"), max_length=36)
        annee_scolaire = sanitize_input(request.args.get("annee_scolaire"), max_length=9)
        trimestre_str = request.args.get("trimestre", "1")
        format_type = sanitize_input(request.args.get("format", "excel"))  # excel, pdf, csv
        
        print(f"DEBUG Export - classe_id: {classe_id}, annee: {annee_scolaire}, trimestre: {trimestre_str}, format: {format_type}")
        
        # Validation des paramètres obligatoires
        if not classe_id or not annee_scolaire:
            flash("Classe et année scolaire requis pour l'export", "error")
            return redirect(url_for('moyennes.liste_moyennes'))
        
        # Vérification d'accès à la classe
        if not validate_ecole_access_for_classe(classe_id, ecole_id):
            log_security_event("unauthorized_export", f"classe_{classe_id}", "failed")
            flash("Accès non autorisé à cette classe.", "error")
            return redirect(url_for('moyennes.liste_moyennes'))

        # Validation du trimestre
        try:
            trimestre = int(trimestre_str)
        except ValueError:
            trimestre = 1
        
        # Récupérer la classe pour le nom
        classe = Classe.query.filter_by(id=UUID(classe_id), ecole_id=ecole_id).first()
        if not classe:
            flash("Classe non trouvée", "error")
            return redirect(url_for('moyennes.liste_moyennes'))
        
        print(f"DEBUG - Classe trouvée: {classe.nom}")
        
        # Récupérer les données
        items_query = (
            db.session.query(
                Eleve.nom,
                Eleve.prenoms,
                Moyenne.moy_trim,
                Moyenne.moy_gen,
                Moyenne.classement_str,
                Appreciations.libelle.label("appreciation"),
            )
            .join(Eleve, Eleve.id == Moyenne.eleve_id)
            .join(Classe, Classe.id == Eleve.classe_id)
            .outerjoin(Appreciations, Appreciations.id == Moyenne.appreciation_id)
            .filter(
                Moyenne.trimestre == trimestre,
                Moyenne.annee_scolaire == annee_scolaire,
                Classe.id == UUID(classe_id),
                Classe.ecole_id == ecole_id
            )
            .order_by(Moyenne.classement.asc())
        )
        
        items = items_query.all()
        
        print(f"DEBUG - {len(items)} éléments trouvés pour l'export")
        
        if not items:
            flash("Aucune donnée à exporter pour les critères sélectionnés", "warning")
            return redirect(url_for('moyennes.liste_moyennes'))

        # Préparer les données pour l'export
        export_data = []
        for idx, item in enumerate(items, 1):
            row_data = {
                'N°': idx,
                'Nom': item.nom or '',
                'Prénom': item.prenoms or '',
                'Moyenne': f"{item.moy_trim:.2f}" if item.moy_trim is not None else "0.00",
                'Rang': item.classement_str or "",
                'Mention': item.appreciation or ""
            }
            
            if trimestre == 3:
                row_data['Moyenne Générale'] = f"{item.moy_gen:.2f}" if item.moy_gen is not None else "0.00"
            
            export_data.append(row_data)

        # Titre du document
        titre = f"LISTE DES MOYENNES - {classe.nom.upper()}"
        sous_titre = f"Année scolaire: {annee_scolaire} - Trimestre: {trimestre}"
        
        print(f"DEBUG - Génération export {format_type} avec {len(export_data)} lignes")
        
        # Journalisation de l'export
        log_security_event("export_moyennes", f"classe_{classe_id}", "success", {
            "annee_scolaire": annee_scolaire,
            "trimestre": trimestre,
            "format": format_type,
            "item_count": len(items)
        })

        # Sélection du format d'export
        if format_type == 'pdf':
            return generate_pdf_export(export_data, titre, sous_titre, classe_id, annee_scolaire, trimestre)
        elif format_type == 'csv':
            return generate_csv_export(export_data, titre, sous_titre, classe_id, annee_scolaire, trimestre)
        else:  # excel par défaut
            return generate_excel_export(export_data, titre, sous_titre, classe_id, annee_scolaire, trimestre)
        
    except Exception as e:
        current_app.logger.error(f"Erreur export moyennes: {str(e)}", exc_info=True)
        log_security_event("export_moyennes", "error", "failed", {"error": str(e)})
        flash("Erreur lors de l'export des données", "error")
        return redirect(url_for('moyennes.liste_moyennes'))


def generate_excel_export(data, titre, sous_titre, classe_id, annee_scolaire, trimestre):
    """Génère un export Excel"""
    try:
        print("DEBUG - Début génération Excel")
        
        # Créer un nouveau workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Moyennes"
        
        # Styles d'alignement
        left_align = Alignment(horizontal='left')
        center_align = Alignment(horizontal='center')
        right_align = Alignment(horizontal='right')
        
        # ========== EN-TÊTE EXCEL UNIFORMISÉ ==========
        # Ministère à GAUCHE
        ws.merge_cells('A1:C1')
        ws['A1'] = "MINISTÈRE DE L'EDUCATION NATIONALE"
        ws['A1'].font = Font(bold=True, size=9)
        ws['A1'].alignment = left_align
        
        # République à DROITE
        ws.merge_cells('D1:F1')
        ws['D1'] = "RÉPUBLIQUE TOGOLAISE"
        ws['D1'].font = Font(bold=True, size=9)
        ws['D1'].alignment = right_align
        
        # D.R.E en gras et petite taille
        ws.merge_cells('A2:C2')
        ws['A2'] = "D.R.E: Région Maritime"
        ws['A2'].font = Font(bold=True, size=7)
        ws['A2'].alignment = left_align
        
        # Devise en gras et petite taille
        ws.merge_cells('D2:F2')
        ws['D2'] = "Travail-Liberté-Patrie"
        ws['D2'].font = Font(bold=True, size=7)
        ws['D2'].alignment = right_align
        
        # I.E.S.G en gras et petite taille
        ws.merge_cells('A3:C3')
        ws['A3'] = "I.E.S.G: Inspection Lomé"
        ws['A3'].font = Font(bold=True, size=7)
        ws['A3'].alignment = left_align
        
        # Ligne de séparation
        ws.merge_cells('A4:F4')
        ws['A4'] = "___________________________________________________________"
        ws['A4'].alignment = center_align
        
        # Nom de l'école CENTRÉ
        ws.merge_cells('A5:F5')
        ws['A5'] = "ECOLE SECONDAIRE MODERNE"
        ws['A5'].font = Font(bold=True, size=12, color="2C3E50")
        ws['A5'].alignment = center_align
        
        # Téléphone centré
        ws.merge_cells('A6:F6')
        ws['A6'] = "Tél: +228 22 21 20 19"
        ws['A6'].alignment = center_align
        
        # Espacement
        ws.append([])
        ws.append([])
        
        # Titre principal
        current_row = 9
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = titre
        ws[f'A{current_row}'].font = Font(size=16, bold=True)
        ws[f'A{current_row}'].alignment = center_align
        
        # Sous-titre
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = sous_titre
        ws[f'A{current_row}'].font = Font(size=12)
        ws[f'A{current_row}'].alignment = center_align
        
        # Informations
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = f"Effectif: {len(data)} élèves"
        ws[f'A{current_row}'].font = Font(size=10)
        ws[f'A{current_row}'].alignment = center_align
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws[f'A{current_row}'].font = Font(size=10)
        ws[f'A{current_row}'].alignment = center_align
        
        # Espacement avant le tableau
        current_row += 2
        
        # En-têtes du tableau
        if data:
            headers = list(data[0].keys())
            
            # Écrire les en-têtes
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = Font(bold=True, size=11)
                cell.alignment = center_align
                cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            current_row += 1
            
            # Données
            for row in data:
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = row[header]
                    cell.alignment = center_align
                    cell.font = Font(size=10)
                    
                    # Alternance de couleurs pour les lignes
                    if current_row % 2 == 0:
                        cell.fill = openpyxl.styles.PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                current_row += 1
        
        # ========== AJUSTEMENT DES COLONNES POUR FORMAT A4 ==========
        # Largeurs optimisées pour format A4
        column_widths = {
            'A': 8,   # N°
            'B': 20,  # Nom
            'C': 20,  # Prénom
            'D': 12,  # Moyenne
            'E': 10,  # Rang
            'F': 15   # Mention
        }
        
        # Appliquer les largeurs
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Ajustement automatique supplémentaire pour les colonnes de données
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Ne pas dépasser la largeur maximale définie
            current_width = column_widths.get(column_letter, 15)
            adjusted_width = min(max_length + 2, current_width + 5)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Bordures pour tout le tableau
        if data:
            table_range = f"A{current_row - len(data)}:F{current_row - 1}"
            for row in ws[table_range]:
                for cell in row:
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
        
        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        print("DEBUG - Excel généré avec succès")
        
        # Retourner le fichier
        filename = f"moyennes_{classe_id}_{annee_scolaire}_T{trimestre}.xlsx"
        return Response(
            buffer.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
        
    except Exception as e:
        print(f"DEBUG - Erreur génération Excel: {str(e)}")
        raise e


def generate_pdf_export(data, titre, sous_titre, classe_id, annee_scolaire, trimestre):
    """Génère un export PDF"""
    try:
        print("DEBUG - Début génération PDF")
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        
        # ========== EN-TÊTE PDF UNIFORMISÉ ==========
        # Ligne 1: Ministère et République
        header1_style = styles['Normal']
        header1_style.fontSize = 9
        header1_style.fontName = 'Helvetica-Bold'
        
        ministere = Paragraph("MINISTÈRE DE L'EDUCATION NATIONALE", header1_style)
        republique = Paragraph("RÉPUBLIQUE TOGOLAISE", header1_style)
        
        # Tableau pour aligner gauche-droite
        header_table1 = Table([[ministere, republique]], colWidths=[4*inch, 2*inch])
        header_table1.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ]))
        elements.append(header_table1)
        
        # Ligne 2: D.R.E et Devise
        header2_style = styles['Normal']
        header2_style.fontSize = 7
        header2_style.fontName = 'Helvetica-Bold'
        
        dre = Paragraph("D.R.E: Région Maritime", header2_style)
        devise = Paragraph("Travail-Liberté-Patrie", header2_style)
        
        header_table2 = Table([[dre, devise]], colWidths=[4*inch, 2*inch])
        header_table2.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ]))
        elements.append(header_table2)
        
        # Ligne 3: I.E.S.G
        header3_style = styles['Normal']
        header3_style.fontSize = 7
        header3_style.fontName = 'Helvetica-Bold'
        
        iesg = Paragraph("I.E.S.G: Inspection Lomé", header3_style)
        elements.append(iesg)
        
        # Ligne de séparation
        elements.append(Paragraph("_" * 80, styles['Normal']))
        elements.append(Spacer(1, 0.1*inch))
        
        # Nom de l'école
        ecole_style = styles['Heading1']
        ecole_style.fontSize = 14
        ecole_style.textColor = colors.HexColor('#2C3E50')
        ecole_style.alignment = 1  # Centré
        
        ecole = Paragraph("ECOLE SECONDAIRE MODERNE", ecole_style)
        elements.append(ecole)
        
        # Téléphone
        tel_style = styles['Normal']
        tel_style.alignment = 1  # Centré
        telephone = Paragraph("Tél: +228 22 21 20 19", tel_style)
        elements.append(telephone)
        
        elements.append(Spacer(1, 0.3*inch))
        
        # Titre principal
        title_style = styles['Heading1']
        title_style.alignment = 1  # Centré
        title_style.fontSize = 16
        title = Paragraph(titre, title_style)
        elements.append(title)
        
        # Sous-titre
        subtitle_style = styles['Heading2']
        subtitle_style.alignment = 1  # Centré
        subtitle_style.fontSize = 12
        subtitle = Paragraph(sous_titre, subtitle_style)
        elements.append(subtitle)
        
        elements.append(Spacer(1, 0.2*inch))
        
        # Informations
        info_style = styles['Normal']
        info_style.alignment = 1  # Centré
        elements.append(Paragraph(f"Effectif: {len(data)} élèves", info_style))
        elements.append(Paragraph(f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}", info_style))
        
        elements.append(Spacer(1, 0.3*inch))
        
        # Préparer les données pour le tableau
        if data:
            # En-têtes du tableau
            headers = list(data[0].keys())
            table_data = [headers]
            
            # Données
            for row in data:
                table_data.append([str(row[header]) for header in headers])
            
            # Créer le tableau avec largeurs optimisées pour A4
            col_widths = [0.8*inch, 2*inch, 2*inch, 1.2*inch, 1*inch, 1.5*inch]  # Ajusté pour A4
            
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            # Style du tableau optimisé pour A4
            style = TableStyle([
                # En-têtes
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                
                # Données
                ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
            ])
            
            # Appliquer le style
            table.setStyle(style)
            elements.append(table)
        
        # Générer le PDF
        doc.build(elements)
        buffer.seek(0)
        
        print("DEBUG - PDF généré avec succès")
        
        # Retourner le fichier
        filename = f"moyennes_{classe_id}_{annee_scolaire}_T{trimestre}.pdf"
        return Response(
            buffer.getvalue(),
            mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
        
    except Exception as e:
        print(f"DEBUG - Erreur génération PDF: {str(e)}")
        raise e


def generate_csv_export(data, titre, sous_titre, classe_id, annee_scolaire, trimestre):
    """Génère un export CSV"""
    try:
        print("DEBUG - Début génération CSV")
        
        output = StringIO()
        writer = csv.writer(output, delimiter=';')
        
        # Entêtes du document
        writer.writerow(["MINISTÈRE DE L'EDUCATION NATIONALE", "", "", "RÉPUBLIQUE TOGOLAISE"])
        writer.writerow(["D.R.E: Région Maritime", "", "", "Travail-Liberté-Patrie"])
        writer.writerow(["I.E.S.G: Inspection Lomé"])
        writer.writerow([])
        writer.writerow(["ECOLE SECONDAIRE MODERNE"])
        writer.writerow(["Tél: +228 22 21 20 19"])
        writer.writerow([])
        writer.writerow([titre])
        writer.writerow([sous_titre])
        writer.writerow([f"Effectif: {len(data)} élèves"])
        writer.writerow([f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
        writer.writerow([])  # Ligne vide
        
        # En-têtes du tableau
        if data:
            headers = list(data[0].keys())
            writer.writerow(headers)
            
            # Données
            for row in data:
                writer.writerow([str(row[header]) for header in headers])
        
        output.seek(0)
        
        print("DEBUG - CSV généré avec succès")
        
        # Retourner le fichier
        filename = f"moyennes_{classe_id}_{annee_scolaire}_T{trimestre}.csv"
        return Response(
            output.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
        
    except Exception as e:
        print(f"DEBUG - Erreur génération CSV: {str(e)}")
        raise e


# ==================== MIDDLEWARE DE SÉCURITÉ ====================

@moyennes_bp.after_request
def add_security_headers(response):
    """Ajoute des headers de sécurité HTTP"""
    security_headers = {
        'X-Content-Type-Options': 'nosniff',
        'X-Frame-Options': 'DENY',
        'X-XSS-Protection': '1; mode=block',
        'Strict-Transport-Security': 'max-age=31536000; includeSubDomains',
        'Cache-Control': 'no-store, no-cache, must-revalidate, max-age=0'
    }
    
    for header, value in security_headers.items():
        response.headers[header] = value
        
    return response


# ==================== GESTIONNAIRES D'ERREURS ====================

@moyennes_bp.errorhandler(400)
def bad_request_error(error):
    """Gestionnaire d'erreur 400 sécurisé"""
    log_security_event(
        "bad_request", 
        request.path, 
        "failed", 
        {"user_agent": request.headers.get('User-Agent')}
    )
    return jsonify({"error": "Requête invalide"}), 400

@moyennes_bp.errorhandler(403)
def forbidden_error(error):
    """Gestionnaire d'erreur 403 sécurisé"""
    log_security_event(
        "forbidden_access", 
        request.path, 
        "failed", 
        {"user_id": current_user.id if current_user.is_authenticated else "anonymous"}
    )
    return jsonify({"error": "Accès non autorisé"}), 403

@moyennes_bp.errorhandler(429)
def rate_limit_error(error):
    """Gestionnaire d'erreur rate limiting"""
    log_security_event(
        "rate_limit_exceeded", 
        request.path, 
        "failed", 
        {"ip": request.remote_addr}
    )
    return jsonify({
        'error': 'Trop de requêtes. Veuillez réessayer plus tard.'
    }), 429

@moyennes_bp.errorhandler(500)
def internal_error(error):
    """Gestionnaire d'erreur 500 sécurisé"""
    log_security_event(
        "internal_server_error", 
        request.path, 
        "failed", 
        {"error": str(error)}
    )
    return jsonify({"error": "Erreur interne du serveur"}), 500