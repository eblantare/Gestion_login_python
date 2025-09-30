from flask import Blueprint, request, jsonify, send_file
from ..models import Moyenne, Eleve, Classe, Matiere, Enseignant, Enseignement, Ecole, Appreciations
from extensions import db
from sqlalchemy.orm import joinedload
import io
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
from uuid import UUID

moyennes_export_bp = Blueprint('moyennes_export', __name__)

# ---------------- EXPORTATIONS MOYENNES ----------------
@moyennes_export_bp.route("/export/pdf", methods=["GET"])
def export_moyennes_pdf():
    """Export PDF de la liste des moyennes avec filtres"""
    try:
        # Récupérer les paramètres de filtrage
        classe_id = request.args.get("classe_id", type=str)
        matiere_id = request.args.get("matiere_id", type=str)
        trimestre = request.args.get("trimestre", type=int)
        annee_scolaire = request.args.get("annee_scolaire", type=str)
        mention = request.args.get("mention", type=str)  # Nouveau filtre pour les mentions

        # Construire la requête de base
        query = db.session.query(
            Eleve.id.label("eleve_id"),
            Eleve.nom,
            Eleve.prenoms,
            Classe.nom.label("classe_nom"),
            Moyenne.moy_trim,
            Moyenne.moy_gen,
            Moyenne.classement.label("classement"),
            Moyenne.classement_str.label("classement_str"),
            Moyenne.classement_gen.label("classement_gen"),
            Moyenne.moy_class,
            Moyenne.moy_forte,
            Moyenne.moy_faible,
            Moyenne.eff_comp.label("effectif_composants"),
            Appreciations.libelle.label("appreciation"),
        ).join(Eleve, Eleve.id == Moyenne.eleve_id)\
         .join(Classe, Classe.id == Eleve.classe_id)\
         .outerjoin(Appreciations, Appreciations.id == Moyenne.appreciation_id)

        # Appliquer les filtres
        if classe_id and classe_id.lower() != "none":
            try:
                query = query.filter(Eleve.classe_id == UUID(classe_id))
            except ValueError:
                pass

        if trimestre:
            query = query.filter(Moyenne.trimestre == trimestre)
            
        if annee_scolaire:
            query = query.filter(Moyenne.annee_scolaire == annee_scolaire)

        # Filtre par mention (appréciation)
        if mention and mention.lower() != "none":
            query = query.filter(Appreciations.libelle == mention)

        # Ordonner par classement
        query = query.order_by(Moyenne.classement.asc())
        items = query.all()

        # Récupérer les informations pour l'en-tête
        classe = None
        matiere = None
        enseignant = None
        
        if classe_id and classe_id.lower() != "none":
            classe = Classe.query.get(UUID(classe_id))
            
        if matiere_id and matiere_id.lower() != "none":
            matiere = Matiere.query.get(UUID(matiere_id))
            # Récupérer l'enseignant si une matière est sélectionnée
            if classe and matiere:
                enseignement = Enseignement.query.filter_by(
                    classe_id=classe.id,
                    matiere_id=matiere.id
                ).first()
                if enseignement:
                    enseignant = enseignement.enseignant
            
        # Calculer les statistiques de la classe
        if items:
            moyennes_non_nulles = [i.moy_trim for i in items if i.moy_trim is not None and i.moy_trim > 0]
            if moyennes_non_nulles:
                plus_forte_moyenne = max(moyennes_non_nulles)
                plus_faible_moyenne = min(moyennes_non_nulles)
                moyenne_classe = round(sum(moyennes_non_nulles) / len(moyennes_non_nulles), 2)
                effectif_composant = len(moyennes_non_nulles)
            else:
                plus_forte_moyenne = 0
                plus_faible_moyenne = 0
                moyenne_classe = 0
                effectif_composant = 0
        else:
            plus_forte_moyenne = 0
            plus_faible_moyenne = 0
            moyenne_classe = 0
            effectif_composant = 0

        # Récupérer les informations de l'école
        ecole = Ecole.query.first()
        nom_ecole = ecole.nom if ecole else "COLLÈGE D'ENSEIGNEMENT GÉNÉRAL 'SAINT BLANT'"
        dre = ecole.dre if ecole and ecole.dre else "MARITIME"
        inspection = ecole.inspection if ecole and ecole.inspection else "TSÉVIÉ"

        # Créer le PDF en mémoire
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4, 
            topMargin=15*mm, 
            bottomMargin=15*mm,
            leftMargin=10*mm,
            rightMargin=10*mm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # ========== EN-TÊTE ==========
        logo_path = r"C:\projets\python\gestion_scolaire\app\static\images\logo.png"
        
        # Vérifier si le logo existe
        if os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=35*mm, height=35*mm)
                logo.hAlign = 'CENTER'
            except:
                logo = Paragraph("<b>[LOGO]</b>", styles['Normal'])
        else:
            logo = Paragraph("<b>[LOGO ÉCOLE]</b>", styles['Normal'])
        
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Normal'],
            fontSize=9,
            alignment=1,
            spaceAfter=0,
            leading=10
        )
        
        small_header_style = ParagraphStyle(
            'SmallHeader',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            spaceAfter=0,
            leading=8
        )
        
        # Structure de l'en-tête
        left_col_content = [
            Paragraph("<b>MINISTÈRE DES ENSEIGNEMENTS PRIMAIRES ET SECONDAIRE</b>", header_style),
            Paragraph("-----------", small_header_style),
            Paragraph(f"DIRECTION RÉGIONALE DE L'ÉDUCATION - {dre}", header_style),
            Paragraph("-----------", small_header_style),
            Paragraph(f"INSPECTION DE L'ENSEIGNEMENT SECONDAIRE GÉNÉRAL - {inspection}", header_style)
        ]
        
        right_col_content = [
            Paragraph("<b>RÉPUBLIQUE TOGOLAISE</b>", header_style),
            Paragraph("-----------", small_header_style),
            Paragraph("Travail - Liberté - Patrie", ParagraphStyle(
                'DeviseStyle',
                parent=styles['Normal'],
                fontSize=7,
                alignment=1,
                spaceAfter=0,
                leading=8
            ))
        ]
        
        center_col_content = [
            logo,
            Paragraph(f"<b>{nom_ecole}</b>", header_style)
        ]

        header_table = Table([[
            left_col_content, 
            center_col_content, 
            right_col_content
        ]], colWidths=[doc.width/3, doc.width/3, doc.width/3])
        
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ]))
        
        elements.append(header_table)
        elements.append(Spacer(1, 6*mm))
        
        # Ligne de séparation
        separation_line = Table([['']], colWidths=[doc.width])
        separation_line.setStyle(TableStyle([
            ('LINEABOVE', (0, 0), (0, 0), 1, colors.black),
        ]))
        elements.append(separation_line)
        elements.append(Spacer(1, 6*mm))
        
        # ========== INFORMATIONS DES MOYENNES ==========
        # Titre principal
        title_text = "LISTE DES MOYENNES"
        if matiere:
            title_text += f" - {matiere.libelle}"
        
        title = Paragraph(title_text, ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=14,
            alignment=1,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=12
        ))
        elements.append(title)

        # Informations détaillées
        info_data_line1 = [
            f"Année scolaire: {annee_scolaire or 'Non spécifiée'}",
            f"Trimestre: {trimestre or 'Non spécifié'}", 
            f"Classe: {classe.nom if classe else 'Toutes'}"
        ]
        
        info_table_line1 = Table([info_data_line1], colWidths=[doc.width/3, doc.width/3, doc.width/3])
        info_table_line1.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(info_table_line1)
        
        # Deuxième ligne avec statistiques et enseignant
        info_line2_items = []
        if matiere and enseignant:
            info_line2_items.append(f"Matière: {matiere.libelle}")
            info_line2_items.append(f"Enseignant: {enseignant.utilisateur.nom if enseignant.utilisateur else 'Non assigné'}")
        else:
            info_line2_items.append(f"Matière: {matiere.libelle if matiere else 'Toutes'}")
            info_line2_items.append("")  # Espace vide
        
        info_line2_items.append(f"Mention: {mention or 'Toutes'}")
        
        info_table_line2 = Table([info_line2_items], colWidths=[doc.width/3, doc.width/3, doc.width/3])
        info_table_line2.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(info_table_line2)
        
        # Troisième ligne avec statistiques de la classe
        info_data_line3 = [
            f"Plus forte moyenne: {plus_forte_moyenne:.2f}/20",
            f"Plus faible moyenne: {plus_faible_moyenne:.2f}/20",
            f"Moyenne de la classe: {moyenne_classe:.2f}/20"
        ]
        
        info_table_line3 = Table([info_data_line3], colWidths=[doc.width/3, doc.width/3, doc.width/3])
        info_table_line3.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(info_table_line3)
        
        elements.append(Spacer(1, 8*mm))
        
        # ========== TABLEAU DES MOYENNES ==========
        headers = [
            'N°',
            'Nom & Prénoms',
            'Moyenne Trim',
            'Moyenne Gén',
            'Classement',
            'Appréciation'
        ]
        
        data = [headers]
        
        # Ajouter les données avec numérotation
        for index, item in enumerate(items, 1):
            row = [
                str(index),
                f"{item.nom} {item.prenoms}",
                f"{item.moy_trim:.2f}" if item.moy_trim else "-",
                f"{item.moy_gen:.2f}" if item.moy_gen else "-",
                item.classement_str or "-",
                item.appreciation or "-"
            ]
            data.append(row)
        
        # Création du tableau
        table = Table(data, colWidths=[
            doc.width * 0.05,   # N°
            doc.width * 0.30,   # Nom & Prénoms
            doc.width * 0.15,   # Moyenne Trim
            doc.width * 0.15,   # Moyenne Gén
            doc.width * 0.15,   # Classement
            doc.width * 0.20    # Appréciation
        ], repeatRows=1)
        
        # Style du tableau
        table_style = TableStyle([
            # En-têtes
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 4),
            
            # Lignes de données
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),   # N° centré
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),     # Nom à gauche
            ('ALIGN', (2, 1), (-1, -1), 'CENTER'),  # Données centrées
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Bordures
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            
            # Alternance des couleurs des lignes
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
            
            # Padding réduit
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ])
        
        table.setStyle(table_style)
        elements.append(table)
        
        # ========== PIED DE PAGE ==========
        elements.append(Spacer(1, 10*mm))
        
        # Date de génération
        date_gen = Paragraph(
            f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
            ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=8,
                alignment=2,
                textColor=colors.grey
            )
        )
        elements.append(date_gen)
        
        # ========== GÉNÉRATION DU PDF ==========
        def add_page_number(canvas, doc):
            """Ajoute les numéros de page au PDF"""
            page_num = canvas.getPageNumber()
            text = f"Page {page_num}"
            canvas.setFont("Helvetica", 8)
            canvas.drawRightString(doc.pagesize[0] - 20, 20, text)
            
        doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
        buffer.seek(0)
        
        # Nom du fichier
        filename = f"releve_moyennes"
        if classe:
            filename += f"_{classe.nom}"
        if matiere:
            filename += f"_{matiere.libelle}"
        if mention:
            filename += f"_{mention}"
        filename += f"_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        
        response = send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
        return response
        
    except Exception as e:
        print(f"Erreur génération PDF moyennes: {str(e)}")
        return jsonify({"error": f"Erreur lors de la génération du PDF: {str(e)}"}), 500

@moyennes_export_bp.route("/export/excel", methods=["GET"])
def export_moyennes_excel():
    """Export Excel de la liste des moyennes avec filtres"""
    try:
        # Récupérer les paramètres de filtrage (identique à PDF)
        classe_id = request.args.get("classe_id", type=str)
        matiere_id = request.args.get("matiere_id", type=str)
        trimestre = request.args.get("trimestre", type=int)
        annee_scolaire = request.args.get("annee_scolaire", type=str)
        mention = request.args.get("mention", type=str)

        # Construire la requête (identique à PDF)
        query = db.session.query(
            Eleve.id.label("eleve_id"),
            Eleve.nom,
            Eleve.prenoms,
            Classe.nom.label("classe_nom"),
            Moyenne.moy_trim,
            Moyenne.moy_gen,
            Moyenne.classement.label("classement"),
            Moyenne.classement_str.label("classement_str"),
            Moyenne.classement_gen.label("classement_gen"),
            Moyenne.moy_class,
            Moyenne.moy_forte,
            Moyenne.moy_faible,
            Moyenne.eff_comp.label("effectif_composants"),
            Appreciations.libelle.label("appreciation"),
        ).join(Eleve, Eleve.id == Moyenne.eleve_id)\
         .join(Classe, Classe.id == Eleve.classe_id)\
         .outerjoin(Appreciations, Appreciations.id == Moyenne.appreciation_id)

        if classe_id and classe_id.lower() != "none":
            try:
                query = query.filter(Eleve.classe_id == UUID(classe_id))
            except ValueError:
                pass

        if trimestre:
            query = query.filter(Moyenne.trimestre == trimestre)
            
        if annee_scolaire:
            query = query.filter(Moyenne.annee_scolaire == annee_scolaire)

        if mention and mention.lower() != "none":
            query = query.filter(Appreciations.libelle == mention)

        query = query.order_by(Moyenne.classement.asc())
        items = query.all()

        # Récupérer les informations pour l'en-tête
        classe = None
        matiere = None
        enseignant = None
        
        if classe_id and classe_id.lower() != "none":
            classe = Classe.query.get(UUID(classe_id))
            
        if matiere_id and matiere_id.lower() != "none":
            matiere = Matiere.query.get(UUID(matiere_id))
            if classe and matiere:
                enseignement = Enseignement.query.filter_by(
                    classe_id=classe.id,
                    matiere_id=matiere.id
                ).first()
                if enseignement:
                    enseignant = enseignement.enseignant
            
        # Calculer les statistiques
        if items:
            moyennes_non_nulles = [i.moy_trim for i in items if i.moy_trim is not None and i.moy_trim > 0]
            if moyennes_non_nulles:
                plus_forte_moyenne = max(moyennes_non_nulles)
                plus_faible_moyenne = min(moyennes_non_nulles)
                moyenne_classe = round(sum(moyennes_non_nulles) / len(moyennes_non_nulles), 2)
                effectif_composant = len(moyennes_non_nulles)
            else:
                plus_forte_moyenne = 0
                plus_faible_moyenne = 0
                moyenne_classe = 0
                effectif_composant = 0
        else:
            plus_forte_moyenne = 0
            plus_faible_moyenne = 0
            moyenne_classe = 0
            effectif_composant = 0

        # Récupérer les informations de l'école
        ecole = Ecole.query.first()
        nom_ecole = ecole.nom if ecole else "COLLÈGE D'ENSEIGNEMENT GÉNÉRAL 'SAINT BLANT'"
        dre = ecole.dre if ecole and ecole.dre else "MARITIME"
        inspection = ecole.inspection if ecole and ecole.inspection else "TSÉVIÉ"
        
        # Créer le fichier Excel
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        
        # Titre de l'onglet
        ws.title = "Relevé de moyennes"
        
        # ========== EN-TÊTE EXCEL ==========
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        current_row = 1

        # Ligne 1 - MINISTÈRE à gauche et RÉPUBLIQUE à droite
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = "MINISTÈRE DES ENSEIGNEMENTS PRIMAIRES ET SECONDAIRE"
        ws[f'A{current_row}'].font = Font(bold=True, size=11)
        ws[f'A{current_row}'].alignment = left_align

        ws.merge_cells(f'F{current_row}:I{current_row}')
        ws[f'F{current_row}'] = "RÉPUBLIQUE TOGOLAISE"
        ws[f'F{current_row}'].font = Font(bold=True, size=11)
        ws[f'F{current_row}'].alignment = right_align
        current_row += 1

        # Ligne 2 - DRE à gauche et devise à droite
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = f"DIRECTION RÉGIONALE DE L'ÉDUCATION - {dre}"
        ws[f'A{current_row}'].font = Font(size=9)
        ws[f'A{current_row}'].alignment = left_align

        ws.merge_cells(f'F{current_row}:I{current_row}')
        ws[f'F{current_row}'] = "Travail - Liberté - Patrie"
        ws[f'F{current_row}'].font = Font(bold=True, size=9)
        ws[f'F{current_row}'].alignment = right_align
        current_row += 1

        # Ligne 3 - Inspection à gauche
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = f"INSPECTION DE L'ENSEIGNEMENT SECONDAIRE GÉNÉRAL - {inspection}"
        ws[f'A{current_row}'].font = Font(size=9)
        ws[f'A{current_row}'].alignment = left_align
        current_row += 1

        # Ligne vide
        current_row += 1

        # Ligne 4 - Nom de l'école
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = nom_ecole
        ws[f'A{current_row}'].font = Font(bold=True, size=14, color="2C3E50")
        ws[f'A{current_row}'].alignment = center_align
        current_row += 1

        # Ligne 5 - Titre du document
        title_text = "LISTE DES MOYENNES"
        if matiere:
            title_text += f" - {matiere.libelle}"
            
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = title_text
        ws[f'A{current_row}'].font = Font(bold=True, size=16, color="2C3E50")
        ws[f'A{current_row}'].alignment = center_align
        current_row += 1

        # ========== INFORMATIONS DÉTAILLÉES ==========
        # Première ligne d'informations
        info_row = current_row

        ws.merge_cells(f'A{info_row}:B{info_row}')
        ws[f'A{info_row}'] = f"Année scolaire: {annee_scolaire or 'Non spécifiée'}"
        ws[f'A{info_row}'].font = Font(size=9, bold=True)
        ws[f'A{info_row}'].alignment = left_align

        ws.merge_cells(f'C{info_row}:D{info_row}')
        ws[f'C{info_row}'] = f"Trimestre: {trimestre or 'Non spécifié'}"
        ws[f'C{info_row}'].font = Font(size=9, bold=True)
        ws[f'C{info_row}'].alignment = center_align

        ws.merge_cells(f'E{info_row}:F{info_row}')
        ws[f'E{info_row}'] = f"Classe: {classe.nom if classe else 'Toutes'}"
        ws[f'E{info_row}'].font = Font(size=9, bold=True)
        ws[f'E{info_row}'].alignment = center_align

        ws.merge_cells(f'G{info_row}:I{info_row}')
        ws[f'G{info_row}'] = f"Mention: {mention or 'Toutes'}"
        ws[f'G{info_row}'].font = Font(size=9, bold=True)
        ws[f'G{info_row}'].alignment = right_align

        current_row += 1

        # Deuxième ligne d'informations
        info_row2 = current_row

        if matiere and enseignant:
            ws.merge_cells(f'A{info_row2}:C{info_row2}')
            ws[f'A{info_row2}'] = f"Matière: {matiere.libelle}"
            ws[f'A{info_row2}'].font = Font(size=9, bold=True)
            ws[f'A{info_row2}'].alignment = left_align

            ws.merge_cells(f'D{info_row2}:F{info_row2}')
            ws[f'D{info_row2}'] = f"Enseignant: {enseignant.utilisateur.nom if enseignant.utilisateur else 'Non assigné'}"
            ws[f'D{info_row2}'].font = Font(size=9, bold=True)
            ws[f'D{info_row2}'].alignment = center_align
        else:
            ws.merge_cells(f'A{info_row2}:C{info_row2}')
            ws[f'A{info_row2}'] = f"Matière: {matiere.libelle if matiere else 'Toutes'}"
            ws[f'A{info_row2}'].font = Font(size=9, bold=True)
            ws[f'A{info_row2}'].alignment = left_align

        ws.merge_cells(f'G{info_row2}:I{info_row2}')
        ws[f'G{info_row2}'] = f"Effectif: {effectif_composant} élèves"
        ws[f'G{info_row2}'].font = Font(size=9, bold=True)
        ws[f'G{info_row2}'].alignment = right_align

        current_row += 1

        # Troisième ligne avec statistiques
        info_row3 = current_row

        ws.merge_cells(f'A{info_row3}:C{info_row3}')
        ws[f'A{info_row3}'] = f"Plus forte moyenne: {plus_forte_moyenne:.2f}/20"
        ws[f'A{info_row3}'].font = Font(size=9, bold=True, color="006400")  # Vert foncé
        ws[f'A{info_row3}'].alignment = left_align

        ws.merge_cells(f'D{info_row3}:F{info_row3}')
        ws[f'D{info_row3}'] = f"Plus faible moyenne: {plus_faible_moyenne:.2f}/20"
        ws[f'D{info_row3}'].font = Font(size=9, bold=True, color="8B0000")  # Rouge foncé
        ws[f'D{info_row3}'].alignment = center_align

        ws.merge_cells(f'G{info_row3}:I{info_row3}')
        ws[f'G{info_row3}'] = f"Moyenne de la classe: {moyenne_classe:.2f}/20"
        ws[f'G{info_row3}'].font = Font(size=9, bold=True, color="00008B")  # Bleu foncé
        ws[f'G{info_row3}'].alignment = right_align

        current_row += 2  # Deux lignes vides avant le tableau

        # ========== TABLEAU DES DONNÉES ==========
        headers = [
            'N°',
            'Nom & Prénoms', 
            'Moyenne Trim', 
            'Moyenne Gén', 
            'Classement', 
            'Appréciation'
        ]
        
        # En-têtes du tableau
        table_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        table_header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.fill = table_header_fill
            cell.font = table_header_font
            cell.alignment = center_align
        
        current_row += 1
        
        # Données avec numérotation
        for index, item in enumerate(items, 1):
            ws.cell(row=current_row, column=1).value = index
            ws.cell(row=current_row, column=2).value = f"{item.nom} {item.prenoms}"
            ws.cell(row=current_row, column=3).value = item.moy_trim if item.moy_trim else "-"
            ws.cell(row=current_row, column=4).value = item.moy_gen if item.moy_gen else "-"
            ws.cell(row=current_row, column=5).value = item.classement_str or "-"
            ws.cell(row=current_row, column=6).value = item.appreciation or "-"
            
            current_row += 1
        
        # ========== MISE EN FORME FINALE ==========
        column_widths = {
            'A': 5,   # N°
            'B': 35,  # Nom & Prénoms
            'C': 12,  # Moyenne Trim
            'D': 12,  # Moyenne Gén
            'E': 12,  # Classement
            'F': 20,  # Appréciation
            'G': 12,  # Colonnes supplémentaires
            'H': 8,
            'I': 8
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Centrer les colonnes numériques
        for row in range(current_row - len(items), current_row):
            for col in ['A', 'C', 'D', 'E']:
                ws[f'{col}{row}'].alignment = center_align
        
        # Ajouter des bordures au tableau
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Appliquer les bordures seulement au tableau de données
        table_start_row = current_row - len(items) - 1
        for row in range(table_start_row, current_row):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        # Sauvegarder
        wb.save(buffer)
        buffer.seek(0)
        
        # Nom du fichier
        filename = f"releve_moyennes"
        if classe:
            filename += f"_{classe.nom}"
        if matiere:
            filename += f"_{matiere.libelle}"
        if mention:
            filename += f"_{mention}"
        filename += f"_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        response = send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        return response
        
    except Exception as e:
        print(f"Erreur génération Excel moyennes: {str(e)}")
        return jsonify({"error": f"Erreur lors de la génération du Excel: {str(e)}"}), 500

@moyennes_export_bp.route("/export/filters", methods=["GET"])
def get_export_filters():
    """Renvoie les filtres disponibles pour l'export des moyennes"""
    classes = Classe.query.order_by(Classe.nom).all()
    matieres = Matiere.query.order_by(Matiere.libelle).all()
    annees_scolaires = [m[0] for m in db.session.query(Moyenne.annee_scolaire).distinct()]
    mentions = [a[0] for a in db.session.query(Appreciations.libelle).distinct()]
    
    return jsonify({
        "classes": [{"id": str(c.id), "nom": c.nom} for c in classes],
        "matieres": [{"id": str(m.id), "libelle": m.libelle} for m in matieres],
        "annees_scolaires": annees_scolaires,
        "mentions": mentions,
        "trimestres": [1, 2, 3]
    })